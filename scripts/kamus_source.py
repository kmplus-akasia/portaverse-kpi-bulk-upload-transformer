#!/usr/bin/env python3
"""Resolve canonical Kamus KPI HO source roots and workbook paths."""

from __future__ import annotations

import json
import re
from dataclasses import dataclass
from pathlib import Path

REPO_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_INVENTORY_CONFIG = REPO_ROOT / "configs/kamus_kpi_ho_visible_20260729.json"
DEFAULT_SOURCE_ROOT = (
    REPO_ROOT / "outputs/kamus-ho-config-20260729/source/KAMUS KPI PELINDO GROUP 1 (HO) 5"
)
DOWNLOADS_KAMUS_PATTERN = re.compile(r"/Downloads/KAMUS KPI PELINDO GROUP 1 \(HO\)", re.I)
LEGACY_HO_VERSION_PATTERN = re.compile(r"KAMUS KPI PELINDO GROUP 1 \(HO\) [34]\b", re.I)


@dataclass(frozen=True)
class KamusSourceContext:
    source_root: Path
    inventory_config: Path
    inventory_metadata: dict[str, object]

    @property
    def source_root_display(self) -> str:
        try:
            return str(self.source_root.relative_to(REPO_ROOT))
        except ValueError:
            return str(self.source_root)

    @property
    def inventory_config_display(self) -> str:
        try:
            return str(self.inventory_config.relative_to(REPO_ROOT))
        except ValueError:
            return str(self.inventory_config)


def find_latest_inventory_config(repo_root: Path = REPO_ROOT) -> Path:
    candidates = sorted((repo_root / "configs").glob("kamus_kpi_ho_visible_*.json"))
    if not candidates:
        raise FileNotFoundError("No configs/kamus_kpi_ho_visible_*.json inventory found.")
    return candidates[-1]


def load_inventory_config(path: Path | None = None) -> dict[str, object]:
    inventory_path = (path or DEFAULT_INVENTORY_CONFIG).resolve()
    payload = json.loads(inventory_path.read_text(encoding="utf-8"))
    if "metadata" not in payload:
        raise ValueError(f"Inventory config missing metadata block: {inventory_path}")
    return payload


def _normalize_workbook_key(value: str) -> str:
    text = value.replace("\\", "/").strip().casefold()
    text = re.sub(r"\s+", " ", text)
    text = re.sub(r"\(\d+\)(?=\.xlsx$)", "", text)
    text = text.replace(" dan ", " & ")
    return text


def inventory_workbook_index(payload: dict[str, object]) -> dict[str, str]:
    index: dict[str, str] = {}
    for section in ("kamus_kpi_v2", "kamus_kpi_v1_pre_restructure"):
        for row in payload.get(section, []):
            source_workbook = str(row.get("source_workbook") or "").strip()
            if not source_workbook:
                continue
            index[_normalize_workbook_key(source_workbook)] = source_workbook
            folder = source_workbook.split("/", 1)[0] if "/" in source_workbook else ""
            filename = source_workbook.rsplit("/", 1)[-1]
            index[_normalize_workbook_key(f"{folder}/{filename}")] = source_workbook
    return index


def _filename_tokens(value: str) -> set[str]:
    filename = value.rsplit("/", 1)[-1]
    if filename.lower().endswith(".xlsx"):
        filename = filename[:-5]
    normalized = _normalize_workbook_key(filename)
    tokens = set(re.findall(r"[a-z0-9]+", normalized))
    return {token for token in tokens if not re.fullmatch(r"\d+", token)}


def _filename_similarity(left: str, right: str) -> float:
    left_tokens = _filename_tokens(left)
    right_tokens = _filename_tokens(right)
    if not left_tokens or not right_tokens:
        return 0.0
    return len(left_tokens & right_tokens) / len(left_tokens | right_tokens)


def resolve_kamus_source_root(
    *,
    inventory_config: Path | None = None,
    explicit_root: Path | None = None,
    allow_external: bool = False,
    repo_root: Path = REPO_ROOT,
) -> KamusSourceContext:
    inventory_path = (inventory_config or DEFAULT_INVENTORY_CONFIG).resolve()
    payload = load_inventory_config(inventory_path)
    metadata = dict(payload["metadata"])

    if explicit_root is not None:
        source_root = explicit_root.expanduser().resolve()
    else:
        configured = str(metadata.get("source_root") or "").strip()
        if not configured:
            source_root = DEFAULT_SOURCE_ROOT.resolve()
        else:
            configured_path = Path(configured).expanduser()
            source_root = (
                configured_path.resolve()
                if configured_path.is_absolute()
                else (repo_root / configured_path).resolve()
            )

    if not allow_external:
        source_text = str(source_root)
        try:
            source_root.relative_to(repo_root.resolve())
        except ValueError as exc:
            raise ValueError(
                f"Kamus source root must live inside the repository: {source_root}. "
                "Pass allow_external=True only for an explicitly authorized override."
            ) from exc
        if DOWNLOADS_KAMUS_PATTERN.search(source_text):
            raise ValueError(f"Refusing Downloads Kamus source root: {source_root}")
        if LEGACY_HO_VERSION_PATTERN.search(source_text):
            raise ValueError(f"Refusing legacy HO 3/4 Kamus source root: {source_root}")

    if not source_root.is_dir():
        raise FileNotFoundError(f"Kamus source root does not exist: {source_root}")

    return KamusSourceContext(
        source_root=source_root,
        inventory_config=inventory_path,
        inventory_metadata=metadata,
    )


def canonicalize_source_workbook(
    relative_workbook: str,
    inventory: dict[str, object],
) -> str:
    relative_workbook = relative_workbook.replace("\\", "/").strip()
    if not relative_workbook:
        raise ValueError("source_workbook is blank.")

    index = inventory_workbook_index(inventory)
    exact = index.get(_normalize_workbook_key(relative_workbook))
    if exact:
        return exact

    folder = relative_workbook.split("/", 1)[0] if "/" in relative_workbook else ""
    filename = relative_workbook.rsplit("/", 1)[-1]
    folder_key = _normalize_workbook_key(folder)
    scored_matches: list[tuple[float, str]] = []
    for key, canonical in index.items():
        if folder and not key.startswith(folder_key + "/"):
            continue
        score = _filename_similarity(relative_workbook, canonical)
        if score >= 0.85:
            scored_matches.append((score, canonical))
    if scored_matches:
        scored_matches.sort(key=lambda item: (-item[0], item[1].casefold()))
        best_score = scored_matches[0][0]
        best = {canonical for score, canonical in scored_matches if score == best_score}
        if len(best) == 1:
            return next(iter(best))
        raise ValueError(
            f"Ambiguous source_workbook alias {relative_workbook!r}; candidates: {sorted(best)}"
        )

    if folder_key:
        folder_candidates = sorted(
            {canonical for key, canonical in index.items() if key.startswith(folder_key + "/")}
        )
        if len(folder_candidates) == 1:
            return folder_candidates[0]

    raise FileNotFoundError(f"source_workbook not found in inventory: {relative_workbook}")


def resolve_source_workbook(
    source_root: Path,
    relative_workbook: str,
    inventory: dict[str, object] | None = None,
) -> Path:
    canonical = (
        canonicalize_source_workbook(relative_workbook, inventory)
        if inventory is not None
        else relative_workbook.replace("\\", "/").strip()
    )
    path = (source_root / canonical).resolve()
    if path.is_file():
        return path

    folder = canonical.split("/", 1)[0] if "/" in canonical else ""
    filename = canonical.rsplit("/", 1)[-1]
    if folder:
        matches = sorted(
            (
                candidate
                for candidate in (source_root / folder).glob("*.xlsx")
                if _normalize_workbook_key(candidate.name) == _normalize_workbook_key(filename)
            ),
            key=lambda item: item.name.casefold(),
        )
        if len(matches) == 1:
            return matches[0].resolve()

    raise FileNotFoundError(
        f"Source workbook not found under {source_root}: {canonical} (from {relative_workbook})"
    )


def verify_position_configs(
    source_context: KamusSourceContext,
    configs: list[object],
    *,
    require_sheet_name: bool = True,
) -> list[str]:
    inventory = load_inventory_config(source_context.inventory_config)
    missing: list[str] = []
    for config in configs:
        relative = getattr(config, "source_workbook", None)
        sheet_name = getattr(config, "sheet_name", None)
        position_name = getattr(config, "position_name", None)
        if not relative:
            missing.append(f"{position_name or '?'}: blank source_workbook")
            continue
        try:
            workbook_path = resolve_source_workbook(
                source_context.source_root,
                str(relative),
                inventory,
            )
        except (FileNotFoundError, ValueError) as exc:
            missing.append(f"{position_name or '?'}: {exc}")
            continue
        if require_sheet_name and sheet_name:
            try:
                from openpyxl import load_workbook

                workbook = load_workbook(workbook_path, read_only=True, data_only=True)
                if sheet_name not in workbook.sheetnames:
                    missing.append(
                        f"{position_name or '?'}: sheet {sheet_name!r} missing in {workbook_path.name}"
                    )
                workbook.close()
            except Exception as exc:  # pragma: no cover - surfaced to caller
                missing.append(f"{position_name or '?'}: cannot inspect workbook ({exc})")
    return missing


def attach_config_metadata(
    payload: dict[str, object],
    source_context: KamusSourceContext,
    *,
    mapping_review_artifact: str | None = None,
) -> dict[str, object]:
    metadata = dict(payload.get("metadata") or {})
    metadata.update(
        {
            "source_root": source_context.source_root_display,
            "inventory_config": source_context.inventory_config_display,
            "inventory_generated_at": source_context.inventory_metadata.get("generated_at"),
        }
    )
    if mapping_review_artifact:
        metadata["mapping_review_artifact"] = mapping_review_artifact
    updated = dict(payload)
    updated["metadata"] = metadata
    return updated


def write_readme_source(
    run_dir: Path,
    source_context: KamusSourceContext,
    *,
    extra_lines: list[str] | None = None,
) -> Path:
    readme = run_dir / "README_SOURCE.md"
    lines = [
        "# Source receipt",
        "",
        f"- Kamus inventory: `{source_context.inventory_config_display}`",
        f"- Kamus source root: `{source_context.source_root_display}`",
        f"- Inventory generated at: `{source_context.inventory_metadata.get('generated_at', 'unknown')}`",
        "",
        "Head Office Kamus KPI conversions must read raw workbooks from the repository source root above.",
        "Do not convert from `~/Downloads/KAMUS KPI PELINDO GROUP 1 (HO) *` unless an explicit override is authorized.",
    ]
    if extra_lines:
        lines.extend(["", *extra_lines])
    readme.write_text("\n".join(lines) + "\n", encoding="utf-8")
    return readme
