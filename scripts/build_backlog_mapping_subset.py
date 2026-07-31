#!/usr/bin/env python3
"""Build a position config subset from File Backlog Kamus KPI.xlsx."""

from __future__ import annotations

import json
import re
from collections import OrderedDict
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
BACKLOG = Path("/Users/alfredoteja/Downloads/File Backlog Kamus KPI.xlsx")
KAMUS_MAIN = ROOT / "configs/kamus_kpi_ho_visible_20260729.json"
PENGENDALIAN_INV = ROOT / "outputs/backlog-kamus-kpi-mapping-20260729/support/pengendalian_proyek_visible_20260729.json"
OUT_CONFIG = ROOT / "outputs/backlog-kamus-kpi-mapping-20260729/support/backlog_target_positions.config.json"
OUT_MANIFEST = ROOT / "outputs/backlog-kamus-kpi-mapping-20260729/backlog_target_manifest.json"


def norm(value: object) -> str:
    text = str(value or "").lower().strip()
    text = re.sub(r"\bdh\b", "department head", text)
    text = text.replace("&", " dan ")
    text = re.sub(r"[-_/(),.]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


ALIASES: dict[str, str | list[str]] = {
    "dh pengelolaan keberlanjutan": "department head pengelolaan keberlanjutan",
    "officer pengelolaan keberlanjutan": "officer pengelolaan keberlanjutan",
    "manager implementasi dan pelaporan keberlanjutan": "manager implementasi dan pelaporan keberlanjutan korporasi",
    "officer implementasi dan pelaporan keberlanjutan": "officer implementasi dan pelaporan keberlanjutan korporasi",
    "dh perencanaan dan evaluasi pemasaran": "department head perencanaan dan evaluasi pemasaran",
    "dh pengelolaan pelanggan": "department head pengelolaan pelanggan",
    "department head dan officer aset tetap": ["department head aset tetap", "officer aset tetap"],
    "dh komunikasi korporasi": "department head komunikasi korporasi",
    "group head": "group head",
}


def split_targets(raw: str) -> list[str]:
    if not raw or raw.lower().strip() == "semua posisi":
        return ["__ALL__"]
    parts = re.split(r",\s*", raw.strip().rstrip(","))
    out: list[str] = []
    for part in parts:
        part = part.strip().rstrip(",")
        if not part:
            continue
        key = norm(part)
        alias = ALIASES.get(key, key)
        if isinstance(alias, list):
            out.extend(alias)
        else:
            out.append(alias)
    return out


def load_kamus_items() -> list[dict[str, object]]:
    main = json.loads(KAMUS_MAIN.read_text(encoding="utf-8"))
    peng = json.loads(PENGENDALIAN_INV.read_text(encoding="utf-8"))
    items: list[dict[str, object]] = []
    for row in main.get("kamus_kpi_v2", []):
        if row.get("include_in_position_config"):
            items.append(dict(row))
    peng_by_key = {
        (row.get("source_workbook"), row.get("sheet_name")): row
        for row in peng.get("kamus_kpi_v2", [])
        if row.get("include_in_position_config")
        and "Group Pengendalian Proyek" in str(row.get("source_workbook", ""))
    }
    replaced = 0
    kept: list[dict[str, object]] = []
    for row in items:
        key = (row.get("source_workbook"), row.get("sheet_name"))
        folder = str(row.get("source_folder") or "")
        if folder == "Group Pengendalian Proyek" and key in peng_by_key:
            kept.append(dict(peng_by_key[key]))
            replaced += 1
        elif folder != "Group Pengendalian Proyek":
            kept.append(row)
    for key, row in peng_by_key.items():
        if not any(
            str(r.get("source_workbook")) == str(row.get("source_workbook"))
            and str(r.get("sheet_name")) == str(row.get("sheet_name"))
            for r in kept
        ):
            kept.append(dict(row))
    return kept


def group_items(items: list[dict[str, object]], group_name: str) -> list[dict[str, object]]:
    return [
        row
        for row in items
        if norm(row.get("source_folder") or row.get("group_name") or "") == norm(group_name)
        or norm(group_name) in norm(str(row.get("source_workbook", "")))
    ]


def match_items(group_name: str, target: str, items: list[dict[str, object]]) -> list[dict[str, object]]:
    if group_name == "Group Pengendalian Proyek":
        role_map = {
            "pimpinan proyek": lambda sn: sn.startswith("pimpinan proyek"),
            "deputi pimpinan proyek": lambda sn: sn.startswith("deputi pimpinan proyek") or sn.startswith("deputy pimpinan proyek"),
            "manager proyek": lambda sn: sn.startswith("manager proyek"),
            "officer proyek": lambda sn: sn.startswith("officer proyek"),
        }
        predicate = role_map.get(target)
        if predicate:
            pool = group_items(items, group_name)
            return [row for row in pool if predicate(norm(str(row.get("sheet_name") or "")))]
    pool = group_items(items, group_name)
    matched: list[dict[str, object]] = []
    for row in pool:
        names = [norm(row.get("position_name")), norm(row.get("sheet_name"))]
        for name in names:
            if not name:
                continue
            if name == target or target in name or name in target:
                matched.append(row)
                break
            if target.replace(" korporasi", "") in name or name.replace(" korporasi", "") in target:
                matched.append(row)
                break
    return matched


def to_position_config(row: dict[str, object], backlog_meta: dict[str, str]) -> dict[str, object]:
    folder = str(row.get("source_folder") or backlog_meta.get("group") or "")
    workbook = str(row.get("source_workbook") or "")
    if folder and not workbook.startswith(folder):
        source_workbook = f"{folder}/{workbook}" if workbook else folder
    else:
        source_workbook = workbook
    return {
        "source_workbook": source_workbook,
        "sheet_name": row.get("sheet_name"),
        "position_name": row.get("position_name") or row.get("sheet_name"),
        "position_master_id": None,
        "position_nomenclature_id": None,
        "position_scope": None,
        "portaverse_position_title": None,
        "portaverse_group_name": row.get("group_name") or folder or backlog_meta.get("group"),
        "portaverse_company_name": "PT Pelabuhan Indonesia (Persero)",
        "portaverse_company_code": "PLND",
        "cluster_label": row.get("position_name") or row.get("sheet_name"),
        "mapping_confidence_label": None,
        "mapping_confidence_reason": None,
        "mapping_review_status": None,
        "mapping_override_approved": False,
        "mapping_override_trust_source": None,
        "group_name": row.get("group_name") or folder or backlog_meta.get("group"),
        "directorate_name": backlog_meta.get("direktorat"),
        "expected_impact_count": 10,
        "drop_comment_values": ["Drop"],
        "backlog_nama_file": backlog_meta.get("nama_file"),
        "backlog_posisi_diminta": backlog_meta.get("posisi_raw"),
    }


def main() -> None:
    kamus_items = load_kamus_items()
    wb = load_workbook(BACKLOG, read_only=True, data_only=True)
    ws = wb["Sheet1"]
    rows = list(ws.iter_rows(values_only=True))
    headers = [h for h in rows[0] if h]
    backlog_by_key: dict[tuple[str, str], dict[str, str]] = {}
    for row in rows[1:]:
        if not any(v is not None and str(v).strip() for v in row):
            continue
        data = {headers[i]: row[i] for i in range(len(headers))}
        group = str(data.get("Group") or "").strip()
        nama_file = str(data.get("Nama File") or "").strip()
        backlog_by_key[(group, nama_file)] = {
            "direktorat": str(data.get("Direktorat") or "").strip(),
            "group": group,
            "posisi_raw": str(data.get("Posisi ") or data.get("Posisi") or "").strip().rstrip(","),
            "nama_file": nama_file,
            "status": str(data.get("Status") or "").strip(),
        }

    selected: OrderedDict[tuple[str, str], dict[str, object]] = OrderedDict()
    manifest_groups: list[dict[str, object]] = []
    for entry in backlog_by_key.values():
        group = entry["group"]
        targets = split_targets(entry["posisi_raw"])
        group_selected: list[dict[str, object]] = []
        if targets == ["__ALL__"]:
            for row in group_items(kamus_items, group):
                cfg = to_position_config(row, entry)
                key = (str(cfg["source_workbook"]), str(cfg["sheet_name"]))
                selected[key] = cfg
                group_selected.append(cfg)
        else:
            for target in targets:
                matches = match_items(group, target, kamus_items)
                if not matches:
                    placeholder = {
                        "source_workbook": entry["nama_file"],
                        "sheet_name": None,
                        "position_name": target,
                        "backlog_nama_file": entry["nama_file"],
                        "backlog_posisi_diminta": entry["posisi_raw"],
                        "backlog_target_requested": target,
                        "backlog_group": group,
                        "backlog_direktorat": entry["direktorat"],
                        "mapping_note": "worksheet tidak ditemukan di inventaris kamus",
                    }
                    key = (entry["nama_file"], f"__MISSING__:{target}")
                    selected[key] = placeholder
                    group_selected.append(placeholder)
                    continue
                for row in matches:
                    cfg = to_position_config(row, entry)
                    key = (str(cfg["source_workbook"]), str(cfg["sheet_name"]))
                    selected[key] = cfg
                    group_selected.append(cfg)
        manifest_groups.append(
            {
                "backlog_group": group,
                "backlog_direktorat": entry["direktorat"],
                "backlog_nama_file": entry["nama_file"],
                "backlog_posisi_diminta": entry["posisi_raw"],
                "selected_count": len(group_selected),
                "selected_worksheets": [
                    {
                        "source_workbook": item.get("source_workbook"),
                        "sheet_name": item.get("sheet_name"),
                        "position_name": item.get("position_name"),
                        "mapping_note": item.get("mapping_note"),
                    }
                    for item in group_selected
                ],
            }
        )

    positions = [row for row in selected.values() if row.get("sheet_name")]
    missing = [row for row in selected.values() if not row.get("sheet_name")]
    payload = {
        "backlog_source": str(BACKLOG),
        "kamus_main": str(KAMUS_MAIN),
        "pengendalian_inventory": str(PENGENDALIAN_INV),
        "positions": positions,
        "missing_targets": missing,
    }
    OUT_CONFIG.parent.mkdir(parents=True, exist_ok=True)
    OUT_CONFIG.write_text(json.dumps({"positions": positions}, ensure_ascii=False, indent=2), encoding="utf-8")
    OUT_MANIFEST.write_text(
        json.dumps(
            {
                "backlog_groups": len(backlog_by_key),
                "selected_positions": len(positions),
                "missing_targets": len(missing),
                "groups": manifest_groups,
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )
    print(json.dumps({"positions": len(positions), "missing_targets": len(missing), "config": str(OUT_CONFIG)}, indent=2))


if __name__ == "__main__":
    main()
