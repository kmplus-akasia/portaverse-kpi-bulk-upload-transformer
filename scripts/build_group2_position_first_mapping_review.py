#!/usr/bin/env python3
"""Build a position-first Group 2 mapping review workbook.

Rows are keyed by production identity (PMID structural / PNID non-structural).
Each row proposes zero or more Kamus worksheets; one worksheet may be shared by
many positions. Reviewer columns stay blank for human decision.
"""

from __future__ import annotations

import argparse
import json
import re
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo

import position_mapping as pm

GENERIC_TOKENS = {
    "group",
    "department",
    "head",
    "manager",
    "officer",
    "senior",
    "junior",
    "staf",
    "staff",
    "pt",
    "dan",
    "the",
    "kamus",
    "kpi",
    "cabang",
    "regional",
    "mapping",
    "dengan",
    "kontrak",
    "manajemen",
    "subholding",
    "persero",
    "tbk",
}

COMPANY_GENERIC_TOKENS = GENERIC_TOKENS | {
    "divisi",
    "wilayah",
    "area",
    "branch",
}

HOLDING_COMPANY_CODES = {"PLND"}
HOLDING_COMPANY_NAMES = {"pt pelabuhan indonesia persero", "pt pelabuhan indonesia (persero)"}
COMPANY_MATCH_FLOOR = 0.82
COMPANY_STRONG = 0.90

# Explicit HQ aliases so parent Subholding companies resolve to their own Kamus workbook,
# not to child PT/wilayah workbooks that share generic tokens.
SUBHOLDING_HQ_ALIASES: dict[str, tuple[str, ...]] = {
    "sptp": (
        "sptp",
        "subholding sptp",
        "pelindo terminal petikemas",
        "pt pelindo terminal petikemas",
    ),
    "spmt": (
        "spmt",
        "subholding spmt",
        "pelindo multi terminal",
        "pt pelindo multi terminal",
    ),
    "spjm": (
        "spjm",
        "subholding spjm",
        "pelindo jasa maritim",
        "pt pelindo jasa maritim",
    ),
    "spsl": (
        "spsl",
        "subholding spsl",
        "pelindo sinergi lokaseva",
        "pt pelindo sinergi lokaseva",
    ),
}

REVIEW_COLUMNS = [
    "Reviewer Confirm Mapping",
    "Reviewer Source Workbook",
    "Reviewer Worksheet",
    "Reviewer Notes",
]

REPORT_COLUMNS = [
    "No.",
    "Identity Scope",
    "PMID",
    "PNID",
    "Position Title",
    "Group / Unit",
    "Company",
    "Company Code",
    "Company ID",
    "Active Employees",
    "Active Employee NIPPs",
    "Active Employee Names",
    "Confidence Label",
    "Confidence Reason",
    "Candidate Score",
    "Candidate Source Folder",
    "Candidate Source Workbook",
    "Candidate Worksheet",
    "Candidate Worksheet Title",
    "Title Match Source",
    "Candidate Group",
    "Runner-up Score",
    "Runner-up Workbook",
    "Runner-up Worksheet",
    "Runner-up Title",
    "Shared Worksheet Position Count",
    "Recommended Action",
    *REVIEW_COLUMNS,
]

CONFIDENCE_FILLS = {
    pm.HIGH_CONFIDENCE: PatternFill("solid", fgColor="D9EAD3"),
    pm.LOW_CONFIDENCE: PatternFill("solid", fgColor="FFF2CC"),
    pm.NO_CANDIDATE: PatternFill("solid", fgColor="E7E6E6"),
    pm.MAPPING_CONFLICT: PatternFill("solid", fgColor="F4CCCC"),
    pm.SCOPE_UNCERTAIN: PatternFill("solid", fgColor="CFE2F3"),
}

NAVY = "173651"
TEAL = "138074"
PALE_BLUE = "E9F1F8"
PALE_GRAY = "F3F6F8"
BORDER = "C7D3DC"
BODY = "Aptos"


@dataclass(frozen=True)
class WorksheetEntry:
    source_folder: str
    source_workbook: str
    sheet_name: str
    position_name: str
    group_name: str
    company_hint: str
    workbook_company_key: str
    normalized_title: str
    normalized_position_title: str
    normalized_sheet_title: str
    tokens: frozenset[str]
    company_tokens: frozenset[str]
    inferred_scope: str


@dataclass(frozen=True)
class PositionEntry:
    scope: str
    pmid: str | None
    pnid: str | None
    title: str
    group_name: str | None
    company_name: str | None
    company_code: str | None
    company_id: str | None
    active_employee_count: int
    active_employee_nipps: str
    active_employee_names: str
    normalized_title: str
    tokens: frozenset[str]
    company_tokens: frozenset[str]
    company_key: str


def norm(value: Any) -> str:
    return "" if value is None else str(value).strip()


def significant_tokens(value: str) -> frozenset[str]:
    return frozenset(
        token
        for token in pm.normalize_position_lookup(value).split()
        if len(token) > 2 and token not in GENERIC_TOKENS
    )


def company_tokens(value: str) -> frozenset[str]:
    return frozenset(
        token
        for token in pm.normalize_title(value).split()
        if len(token) > 2 and token not in COMPANY_GENERIC_TOKENS
    )


def company_key(value: str) -> str:
    """Compact company identity used to align position company ↔ workbook path."""
    text = pm.normalize_title(value)
    text = re.sub(r"\b(pt|persero|tbk|kamus|kpi)\b", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def company_hint_from_workbook(source_workbook: str) -> str:
    parts = [part.strip() for part in Path(source_workbook).parts if part.strip()]
    cleaned: list[str] = []
    for part in parts:
        text = re.sub(r"(?i)^kamus kpi\s*", "", part)
        text = re.sub(r"(?i)\s*-\s*mapping.*$", "", text)
        text = re.sub(r"\.xlsx$|\.xlsm$", "", text, flags=re.I)
        if text:
            cleaned.append(text)
    return " / ".join(cleaned[-4:])


def workbook_company_identity(source_workbook: str) -> str:
    """Best company-like segment from a Kamus workbook path."""
    parts = [part.strip() for part in Path(source_workbook).parts if part.strip()]
    ranked: list[tuple[int, str]] = []
    for part in parts:
        raw = re.sub(r"(?i)^kamus kpi\s*", "", part)
        raw = re.sub(r"(?i)\s*-\s*mapping.*$", "", raw)
        raw = re.sub(r"\.xlsx$|\.xlsm$", "", raw, flags=re.I).strip()
        key = company_key(raw)
        if not key:
            continue
        score = 0
        if re.search(r"(?i)\bpt\b", raw):
            score += 5
        if key not in {"regional 1", "regional 2", "regional 3", "regional 4", "cabang", "subholding", "spmt", "spjm", "spsl", "sptp"}:
            score += 2
        score += min(len(key.split()), 6)
        ranked.append((score, raw))
    if not ranked:
        return company_hint_from_workbook(source_workbook)
    ranked.sort(key=lambda item: (-item[0], -len(item[1])))
    return ranked[0][1]


def position_company_aliases(position: PositionEntry) -> tuple[str, ...]:
    aliases: list[str] = []
    for raw in (
        position.company_name,
        position.company_code,
        position.company_key,
    ):
        key = company_key(raw or "")
        if key and key not in aliases:
            aliases.append(key)
    code = norm(position.company_code).casefold()
    name_key = company_key(position.company_name or "")
    for hq_code, hq_aliases in SUBHOLDING_HQ_ALIASES.items():
        if code == hq_code or name_key in {company_key(a) for a in hq_aliases}:
            for alias in hq_aliases:
                key = company_key(alias)
                if key and key not in aliases:
                    aliases.append(key)
    return tuple(aliases)


def _path_segments(source_workbook: str) -> list[str]:
    segments: list[str] = []
    for part in Path(source_workbook).parts:
        raw = re.sub(r"(?i)^kamus kpi\s*", "", part.strip())
        raw = re.sub(r"(?i)\s*-\s*mapping.*$", "", raw)
        raw = re.sub(r"\.xlsx$|\.xlsm$", "", raw, flags=re.I).strip()
        key = company_key(raw)
        if key:
            segments.append(key)
    return segments


def company_affinity(position: PositionEntry, worksheet: WorksheetEntry) -> float:
    """How strongly the position's company points at this workbook/path."""
    aliases = position_company_aliases(position)
    if not aliases:
        return 0.0
    segments = _path_segments(worksheet.source_workbook)
    workbook_key = company_key(worksheet.workbook_company_key)
    hq_codes = set(SUBHOLDING_HQ_ALIASES)
    scores: list[float] = []

    for alias in aliases:
        code = alias.removeprefix("subholding ").strip() if alias.startswith("subholding ") else alias
        is_hq_alias = code in hq_codes or alias.startswith("subholding ")

        if is_hq_alias:
            # HQ aliases may only match the Subholding HQ workbook identity itself.
            if workbook_key in {code, f"subholding {code}"}:
                scores.append(1.0)
            continue

        if alias == workbook_key or alias in segments:
            scores.append(1.0)
            continue

        # Phrase containment against the workbook identity only (not ancestor folders).
        if len(alias) >= 10 and (alias in workbook_key or workbook_key in alias):
            scores.append(0.97)
            continue

    # Controlled token overlap against workbook identity; capped below company floor when
    # the workbook has extra distinctive tokens (child PT / wilayah names).
    pos_tokens = set().union(*(company_tokens(alias) for alias in aliases)) | set(position.company_tokens)
    wb_tokens = company_tokens(workbook_key)
    if pos_tokens and wb_tokens:
        overlap = len(pos_tokens & wb_tokens) / max(len(pos_tokens), 1)
        extra = wb_tokens - pos_tokens - {"subholding"}
        if overlap >= 0.9 and len(extra) <= 1:
            scores.append(0.88)
        elif overlap >= 0.75 and not extra:
            scores.append(0.86)
        elif overlap >= 0.5:
            scores.append(min(0.78, overlap))
        else:
            scores.append(overlap * 0.6)

    return max(scores or [0.0])


def is_holding_company(company_code: str | None, company_name: str | None) -> bool:
    code = norm(company_code).upper()
    if code in HOLDING_COMPANY_CODES:
        return True
    name = pm.normalize_title(company_name)
    return name in HOLDING_COMPANY_NAMES


def load_subholding_tree_company_ids(reference_path: Path) -> set[int]:
    payload = json.loads(reference_path.read_text(encoding="utf-8"))
    companies = {
        int(row["company_in_id"]): row
        for row in payload.get("company_rows", [])
        if isinstance(row, dict) and row.get("company_in_id") not in (None, "")
    }
    subholding_ids = {
        company_id
        for company_id, row in companies.items()
        if (row.get("type_org") or "") == "Subholding"
    }

    def under_subholding(company_id: int, seen: set[int] | None = None) -> bool:
        seen = seen or set()
        if company_id in seen:
            return False
        seen.add(company_id)
        if company_id in subholding_ids:
            return True
        parent = companies.get(company_id, {}).get("parent_id")
        if parent in (None, ""):
            return False
        try:
            return under_subholding(int(parent), seen)
        except (TypeError, ValueError):
            return False

    return {company_id for company_id in companies if under_subholding(company_id)}


def count_unique_active_employees(reference_path: Path, company_ids: set[int]) -> int:
    payload = json.loads(reference_path.read_text(encoding="utf-8"))
    people: set[str] = set()
    for row in payload.get("active_assignment_rows", []):
        if not isinstance(row, dict):
            continue
        company_id = row.get("company_id")
        try:
            if int(company_id) not in company_ids:
                continue
        except (TypeError, ValueError):
            continue
        nipps = row.get("active_employee_nipps")
        if isinstance(nipps, list):
            values = [str(item).strip() for item in nipps if str(item).strip()]
        elif isinstance(nipps, str) and nipps.strip():
            values = [part.strip() for part in nipps.replace(";", ",").split(",") if part.strip()]
        else:
            values = []
        people.update(values)
    return len(people)


def load_worksheets(inventory_path: Path, source_folder: str | None = None) -> list[WorksheetEntry]:
    payload = json.loads(inventory_path.read_text(encoding="utf-8"))
    rows = [
        row
        for row in payload.get("kamus_kpi_v2", [])
        if isinstance(row, dict) and row.get("include_in_position_config")
    ]
    if source_folder:
        rows = [row for row in rows if norm(row.get("source_folder")) == source_folder]
    entries: list[WorksheetEntry] = []
    for row in rows:
        position_name = norm(row.get("position_name"))
        sheet_name = norm(row.get("sheet_name"))
        # Display/scoring primary prefers in-sheet position title; sheet tab is fallback.
        primary_title = position_name or sheet_name
        scope = pm.infer_worksheet_scope(primary_title).scope
        workbook = norm(row.get("source_workbook"))
        workbook_company = workbook_company_identity(workbook)
        hint = company_hint_from_workbook(workbook)
        tokens = significant_tokens(primary_title)
        if position_name and sheet_name and position_name != sheet_name:
            tokens = frozenset(tokens | significant_tokens(sheet_name))
        entries.append(
            WorksheetEntry(
                source_folder=norm(row.get("source_folder")),
                source_workbook=workbook,
                sheet_name=sheet_name,
                position_name=position_name,
                group_name=norm(row.get("group_name")),
                company_hint=hint,
                workbook_company_key=workbook_company,
                normalized_title=pm.normalize_position_lookup(primary_title),
                normalized_position_title=pm.normalize_position_lookup(position_name) if position_name else "",
                normalized_sheet_title=pm.normalize_position_lookup(sheet_name) if sheet_name else "",
                tokens=tokens,
                company_tokens=company_tokens(f"{workbook_company} {hint} {workbook}"),
                inferred_scope=scope,
            )
        )
    return entries


def load_group2_positions(
    reference_path: Path,
    *,
    company_ids: set[int] | None = None,
) -> list[PositionEntry]:
    payload = json.loads(reference_path.read_text(encoding="utf-8"))
    indexes = pm.build_lookup_indexes(payload)
    entries: list[PositionEntry] = []
    for candidate in [*indexes.structural, *indexes.non_structural]:
        if is_holding_company(candidate.company_code, candidate.company_name):
            continue
        title = norm(candidate.title)
        if not title:
            continue
        company_name = norm(candidate.company_name)
        if "(dummy)" in title.casefold() or "(dummy)" in company_name.casefold():
            continue
        if company_ids is not None:
            try:
                company_id = int(candidate.company_id) if candidate.company_id not in (None, "") else None
            except (TypeError, ValueError):
                company_id = None
            if company_id not in company_ids:
                continue
        entries.append(
            PositionEntry(
                scope=candidate.scope,
                pmid=candidate.position_master_id,
                pnid=candidate.position_nomenclature_id,
                title=title,
                group_name=candidate.group_name,
                company_name=candidate.company_name,
                company_code=candidate.company_code,
                company_id=norm(candidate.company_id),
                active_employee_count=candidate.active_employee_count,
                active_employee_nipps="; ".join(candidate.active_employee_nipps),
                active_employee_names="; ".join(candidate.active_employee_names),
                normalized_title=pm.normalize_position_lookup(title),
                tokens=significant_tokens(title),
                company_tokens=company_tokens(
                    " ".join(filter(None, [candidate.company_name, candidate.company_code]))
                ),
                company_key=company_key(company_name or norm(candidate.company_code)),
            )
        )
    entries.sort(
        key=lambda row: (
            norm(row.company_name).casefold(),
            norm(row.group_name).casefold(),
            row.title.casefold(),
            row.pmid or "",
            row.pnid or "",
        )
    )
    return entries


def build_token_index(worksheets: list[WorksheetEntry]) -> dict[str, list[int]]:
    index: dict[str, list[int]] = defaultdict(list)
    for idx, worksheet in enumerate(worksheets):
        for token in worksheet.tokens:
            index[token].append(idx)
        for token in worksheet.company_tokens:
            index[f"__co__:{token}"].append(idx)
        if worksheet.normalized_position_title:
            index[f"__exact_pos__:{worksheet.normalized_position_title}"].append(idx)
            index[f"__exact__:{worksheet.normalized_position_title}"].append(idx)
        if worksheet.normalized_sheet_title:
            index[f"__exact_sheet__:{worksheet.normalized_sheet_title}"].append(idx)
            index[f"__exact__:{worksheet.normalized_sheet_title}"].append(idx)
        elif worksheet.normalized_title:
            index[f"__exact__:{worksheet.normalized_title}"].append(idx)
        wb_key = company_key(worksheet.workbook_company_key)
        if wb_key:
            index[f"__wb__:{wb_key}"].append(idx)
            code = wb_key.removeprefix("subholding ").strip() if wb_key.startswith("subholding ") else wb_key
            if code in SUBHOLDING_HQ_ALIASES:
                index[f"__wb__:{code}"].append(idx)
                index[f"__wb__:subholding {code}"].append(idx)
                for alias in SUBHOLDING_HQ_ALIASES[code]:
                    index[f"__wb__:{company_key(alias)}"].append(idx)
    return index


def candidate_worksheet_indexes(
    position: PositionEntry,
    worksheets: list[WorksheetEntry],
    token_index: dict[str, list[int]],
) -> list[int]:
    hits: Counter[int] = Counter()
    # Company/workbook first — include HQ aliases (SPTP ↔ subholding sptp, etc.).
    for alias in position_company_aliases(position):
        for idx in token_index.get(f"__wb__:{alias}", []):
            hits[idx] += 200
    for token in position.company_tokens:
        for idx in token_index.get(f"__co__:{token}", []):
            hits[idx] += 50
    exact_key = f"__exact__:{position.normalized_title}"
    for idx in token_index.get(exact_key, []):
        hits[idx] += 40
    for token in position.tokens:
        for idx in token_index.get(token, []):
            hits[idx] += 1

    company_local = [idx for idx, score in hits.items() if score >= 50]
    if company_local:
        ranked = sorted(company_local, key=lambda idx: hits[idx], reverse=True)[:120]
        global_title = [idx for idx, _ in hits.most_common(30) if idx not in set(ranked)][:15]
        return ranked + global_title

    if not hits:
        company_tokens_pos = position.company_tokens
        for idx, worksheet in enumerate(worksheets):
            if company_tokens_pos & worksheet.company_tokens:
                hits[idx] += 1
    return [idx for idx, _ in hits.most_common(80)]


_AFFINITY_CACHE: dict[tuple[str, str], float] = {}


def cached_company_affinity(position: PositionEntry, worksheet: WorksheetEntry) -> float:
    key = (position.company_key or norm(position.company_name), worksheet.source_workbook)
    cached = _AFFINITY_CACHE.get(key)
    if cached is not None:
        return cached
    value = company_affinity(position, worksheet)
    _AFFINITY_CACHE[key] = value
    return value


def score_position_against_worksheet(
    position: PositionEntry,
    worksheet: WorksheetEntry,
    *,
    title_source: str = "auto",
) -> pm.ScoredCandidate:
    if title_source == "position_name":
        title_value = worksheet.position_name
        source_label = "position_name"
    elif title_source == "sheet_name":
        title_value = worksheet.sheet_name
        source_label = "sheet_name"
    elif worksheet.position_name:
        title_value = worksheet.position_name
        source_label = "position_name"
    else:
        title_value = worksheet.sheet_name
        source_label = "sheet_name"

    fake = pm.LookupCandidate(
        scope=worksheet.inferred_scope if worksheet.inferred_scope != pm.SCOPE_UNCERTAIN else position.scope,
        position_master_id=None,
        position_nomenclature_id=None,
        title=title_value or worksheet.sheet_name,
        group_name=worksheet.group_name or None,
        company_id=None,
        company_name=worksheet.workbook_company_key or worksheet.company_hint or None,
        company_code=None,
        active_variant_count=1,
        active_employee_count=1,
        definitive_employee_count=0,
        secondary_employee_count=0,
        lookup_keys=pm._candidate_keys({}, title_value, worksheet.sheet_name if source_label == "sheet_name" else title_value),
    )
    title = pm._title_score(position.title, fake)
    group = max(
        pm._context_score(position.group_name, worksheet.group_name),
        pm._context_score(position.group_name, worksheet.company_hint),
    )
    company = cached_company_affinity(position, worksheet)
    # Company/workbook is the primary gate. Title decides the sheet inside that workbook.
    if company >= COMPANY_STRONG:
        if title < 0.75:
            # Workbook is right, but no adequate sheet title inside it.
            score = round(0.45 + (0.25 * title), 4)
        else:
            score = round((company * 0.50) + (title * 0.45) + (group * 0.05), 4)
    elif company >= COMPANY_MATCH_FLOOR:
        if title < 0.75:
            score = round(0.45 + (0.25 * title), 4)
        else:
            score = round((company * 0.48) + (title * 0.45) + (group * 0.07), 4)
    else:
        # Cross-company title-only matches stay weak even if title is exact.
        score = round((title * 0.55) + (company * 0.35) + (group * 0.10), 4)
        if title >= 0.98 and company < COMPANY_MATCH_FLOOR:
            score = min(score, 0.72)
    return pm.ScoredCandidate(
        candidate=fake,
        score=score,
        reason=f"title={title:.2f}; group={group:.2f}; company={company:.2f}; title_source={source_label}",
    )


def _company_from_reason(reason: str) -> float:
    match = re.search(r"company=([0-9.]+)", reason or "")
    return float(match.group(1)) if match else 0.0


def _title_from_reason(reason: str) -> float:
    match = re.search(r"title=([0-9.]+)", reason or "")
    return float(match.group(1)) if match else 0.0


def classify_scores(
    scored: list[tuple[WorksheetEntry, pm.ScoredCandidate]],
) -> tuple[str, str, WorksheetEntry | None, float | None, WorksheetEntry | None, float | None]:
    if not scored or scored[0][1].score < 0.65:
        best_ws = scored[0][0] if scored else None
        best_score = scored[0][1].score if scored else None
        best_reason = scored[0][1].reason if scored else ""
        if best_ws is not None and _company_from_reason(best_reason) >= COMPANY_MATCH_FLOOR:
            return (
                pm.NO_CANDIDATE,
                (
                    "Company workbook matched, but no worksheet title inside it cleared the "
                    f"minimum sheet-title threshold. {best_reason}"
                ),
                best_ws,
                best_score,
                None,
                None,
            )
        return (
            pm.NO_CANDIDATE,
            "No Group 2 Kamus worksheet scored above the minimum threshold for this position.",
            None,
            None,
            None,
            None,
        )
    best_ws, best = scored[0]
    runner_ws, runner = (scored[1] if len(scored) > 1 else (None, None))
    duplicate_strong = bool(
        runner
        and runner.score >= 0.80
        and best.score - runner.score < 0.15
        and (
            best_ws.source_workbook != runner_ws.source_workbook
            or best_ws.sheet_name != runner_ws.sheet_name
        )
    )
    if duplicate_strong:
        return (
            pm.MAPPING_CONFLICT,
            f"Multiple strong worksheet candidates. {best.reason}",
            best_ws,
            best.score,
            runner_ws,
            runner.score if runner else None,
        )
    exact_title = any(f"title={x}" in best.reason for x in ("1.00", "0.98", "0.99", "0.88", "0.90", "0.92", "0.94", "0.96"))
    company_value = _company_from_reason(best.reason)
    company_strong = company_value >= COMPANY_STRONG
    high = (
        best.score >= 0.88
        and company_strong
        and _title_from_reason(best.reason) >= 0.85
        and (runner is None or best.score - runner.score >= 0.05 or company_strong)
    )
    label = pm.HIGH_CONFIDENCE if high else pm.LOW_CONFIDENCE
    reason = (
        f"Company-first worksheet match. {best.reason}"
        if high
        else f"Worksheet candidate exists, but company/title evidence is incomplete. {best.reason}"
    )
    return (
        label,
        reason,
        best_ws,
        best.score,
        runner_ws,
        runner.score if runner else None,
    )


def recommended_action(label: str) -> str:
    return {
        pm.HIGH_CONFIDENCE: "Konfirmasi YES jika worksheet draft benar; satu worksheet boleh dipakai banyak posisi.",
        pm.LOW_CONFIDENCE: "Review draft worksheet sebelum YES.",
        pm.MAPPING_CONFLICT: "Pilih salah satu worksheet di kolom reviewer.",
        pm.NO_CANDIDATE: "Cari/assign worksheet manual atau tandai belum ada kamus.",
        pm.SCOPE_UNCERTAIN: "Pastikan scope identitas lalu pilih worksheet.",
    }.get(label, "Review sebelum konversi.")


def resolve_all(
    positions: list[PositionEntry], worksheets: list[WorksheetEntry]
) -> tuple[list[dict[str, Any]], Counter[str], dict[tuple[str, str], int]]:
    global _AFFINITY_CACHE
    _AFFINITY_CACHE = {}
    token_index = build_token_index(worksheets)
    rows: list[dict[str, Any]] = []
    labels: Counter[str] = Counter()
    shared_best: Counter[tuple[str, str]] = Counter()

    for position in positions:
        candidate_idxs = candidate_worksheet_indexes(position, worksheets, token_index)
        seen: set[tuple[str, str]] = set()
        primary: list[tuple[WorksheetEntry, pm.ScoredCandidate]] = []
        for idx in candidate_idxs:
            worksheet = worksheets[idx]
            key = (worksheet.source_workbook, worksheet.sheet_name)
            if key in seen:
                continue
            seen.add(key)
            if not worksheet.position_name:
                continue
            scored = score_position_against_worksheet(position, worksheet, title_source="position_name")
            primary.append((worksheet, scored))

        company_local_primary = [
            item
            for item in primary
            if cached_company_affinity(position, item[0]) >= COMPANY_MATCH_FLOOR
        ]
        primary_pool = company_local_primary or primary
        primary_pool.sort(
            key=lambda item: (item[1].score, pm._title_score(position.title, item[1].candidate)),
            reverse=True,
        )
        use_fallback = (
            not primary_pool
            or primary_pool[0][1].score < 0.65
            or _title_from_reason(primary_pool[0][1].reason) < 0.75
        )

        title_match_source = "position_name"
        if use_fallback:
            title_match_source = "sheet_name"
            seen_fb: set[tuple[str, str]] = set()
            fallback: list[tuple[WorksheetEntry, pm.ScoredCandidate]] = []
            for idx in candidate_idxs:
                worksheet = worksheets[idx]
                key = (worksheet.source_workbook, worksheet.sheet_name)
                if key in seen_fb:
                    continue
                seen_fb.add(key)
                scored = score_position_against_worksheet(position, worksheet, title_source="sheet_name")
                fallback.append((worksheet, scored))
            company_local_fb = [
                item
                for item in fallback
                if cached_company_affinity(position, item[0]) >= COMPANY_MATCH_FLOOR
            ]
            candidates = company_local_fb or fallback
        else:
            candidates = primary_pool

        candidates.sort(
            key=lambda item: (item[1].score, pm._title_score(position.title, item[1].candidate)),
            reverse=True,
        )
        label, reason, best_ws, best_score, runner_ws, runner_score = classify_scores(candidates)
        labels[label] += 1
        if best_ws is not None:
            shared_best[(best_ws.source_workbook, best_ws.sheet_name)] += 1
            if "title_source=" in (candidates[0][1].reason if candidates else ""):
                match = re.search(r"title_source=([a-z_]+)", candidates[0][1].reason)
                if match:
                    title_match_source = match.group(1)
        rows.append(
            {
                "Identity Scope": position.scope,
                "PMID": position.pmid or "",
                "PNID": position.pnid or "",
                "Position Title": position.title,
                "Group / Unit": position.group_name or "",
                "Company": position.company_name or "",
                "Company Code": position.company_code or "",
                "Company ID": position.company_id or "",
                "Active Employees": position.active_employee_count,
                "Active Employee NIPPs": position.active_employee_nipps,
                "Active Employee Names": position.active_employee_names,
                "Confidence Label": label,
                "Confidence Reason": reason,
                "Candidate Score": best_score if best_score is not None else "",
                "Candidate Source Folder": best_ws.source_folder if best_ws else "",
                "Candidate Source Workbook": best_ws.source_workbook if best_ws else "",
                "Candidate Worksheet": best_ws.sheet_name if best_ws else "",
                "Candidate Worksheet Title": best_ws.position_name if best_ws else "",
                "Title Match Source": title_match_source if best_ws else "",
                "Candidate Group": best_ws.group_name if best_ws else "",
                "Runner-up Score": runner_score if runner_score is not None else "",
                "Runner-up Workbook": runner_ws.source_workbook if runner_ws else "",
                "Runner-up Worksheet": runner_ws.sheet_name if runner_ws else "",
                "Runner-up Title": (runner_ws.position_name if runner_ws else ""),
                "Shared Worksheet Position Count": "",  # filled after pass
                "Recommended Action": recommended_action(label),
                "Reviewer Confirm Mapping": "",
                "Reviewer Source Workbook": "",
                "Reviewer Worksheet": "",
                "Reviewer Notes": "",
            }
        )

    for row in rows:
        key = (row["Candidate Source Workbook"], row["Candidate Worksheet"])
        row["Shared Worksheet Position Count"] = shared_best.get(key, 0) if key[0] else 0
    return rows, labels, shared_best


def style_header(cell, fill: str) -> None:
    cell.fill = PatternFill("solid", fgColor=fill)
    cell.font = Font(name=BODY, bold=True, color="FFFFFF", size=10)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def write_summary(
    ws,
    *,
    exported_at: str,
    inventory_path: str,
    reference_path: str,
    position_count: int,
    worksheet_count: int,
    labels: Counter[str],
    shared_best: dict[tuple[str, str], int],
    scope_label: str,
    unique_active_employees: int | None = None,
    company_count: int | None = None,
) -> None:
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:F1")
    ws["A1"] = f"Position-First Mapping Review — {scope_label}"
    ws["A1"].fill = PatternFill("solid", fgColor=NAVY)
    ws["A1"].font = Font(name=BODY, size=18, bold=True, color="FFFFFF")
    ws.row_dimensions[1].height = 34

    ws.merge_cells("A2:F2")
    ws["A2"] = (
        "Baris = posisi aktif (PMID/PNID) dalam scope di bawah. "
        "Workbook dipilih dari company lokasi posisi; title memilih worksheet di dalam workbook itu. "
        "Satu worksheet boleh dipakai banyak posisi. Kolom reviewer dikosongkan."
    )
    ws["A2"].fill = PatternFill("solid", fgColor=PALE_BLUE)
    ws["A2"].font = Font(name=BODY, size=10, italic=True, color="37556E")

    meta = [
        ("Scope", scope_label),
        ("Reference snapshot", reference_path),
        ("Exported at", exported_at),
        ("Inventory", inventory_path),
        ("Position rows in scope", position_count),
        ("Kamus worksheets available", worksheet_count),
        ("Worksheets proposed as best for ≥2 positions", sum(1 for count in shared_best.values() if count >= 2)),
        ("Generated at", datetime.now().astimezone().isoformat(timespec="seconds")),
    ]
    if company_count is not None:
        meta.insert(5, ("Companies in scope", company_count))
    if unique_active_employees is not None:
        meta.insert(5 if company_count is None else 6, ("Unique active employees in scope", unique_active_employees))
        meta.insert(
            6 if company_count is None else 7,
            ("Sum Active Employees column (not unique)", "lihat kolom Active Employees di Position Coverage"),
        )
    ws["A4"] = "Provenance"
    ws["B4"] = "Value"
    style_header(ws["A4"], TEAL)
    style_header(ws["B4"], TEAL)
    for offset, (label, value) in enumerate(meta, start=5):
        ws.cell(offset, 1, label).fill = PatternFill("solid", fgColor=PALE_GRAY)
        ws.cell(offset, 1).font = Font(name=BODY, bold=True, color=NAVY)
        ws.cell(offset, 2, value).font = Font(name=BODY, color="263746")
        ws.cell(offset, 2).alignment = Alignment(wrap_text=True)

    conf_start = 5 + len(meta) + 2
    ws.cell(conf_start, 1, "Confidence").fill = PatternFill("solid", fgColor=TEAL)
    ws.cell(conf_start, 2, "Count").fill = PatternFill("solid", fgColor=TEAL)
    style_header(ws.cell(conf_start, 1), TEAL)
    style_header(ws.cell(conf_start, 2), TEAL)
    for offset, label in enumerate(
        [pm.HIGH_CONFIDENCE, pm.LOW_CONFIDENCE, pm.MAPPING_CONFLICT, pm.NO_CANDIDATE, pm.SCOPE_UNCERTAIN],
        start=conf_start + 1,
    ):
        ws.cell(offset, 1, label).fill = CONFIDENCE_FILLS.get(label, PatternFill())
        ws.cell(offset, 2, labels.get(label, 0))
        ws.cell(offset, 2).alignment = Alignment(horizontal="center")

    review_start = conf_start + 7
    ws.cell(review_start, 1, "Kolom keputusan reviewer").font = Font(name=BODY, bold=True, color=NAVY)
    for offset, column in enumerate(REVIEW_COLUMNS, start=review_start + 1):
        ws.cell(offset, 1, column)
        ws.cell(
            offset,
            2,
            "YES / NEEDS_CHECK / NO"
            if column == "Reviewer Confirm Mapping"
            else "kosongkan untuk terima draft; isi untuk override",
        )

    ws.column_dimensions["A"].width = 48
    ws.column_dimensions["B"].width = 88


def write_report_sheet(ws, title: str, rows: list[dict[str, Any]], table_name: str) -> None:
    ws.sheet_view.showGridLines = False
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(REPORT_COLUMNS))
    ws.cell(1, 1, title).fill = PatternFill("solid", fgColor=NAVY)
    ws.cell(1, 1).font = Font(name=BODY, size=16, bold=True, color="FFFFFF")
    ws.row_dimensions[1].height = 30

    nipp_col = REPORT_COLUMNS.index("Active Employee NIPPs") + 1
    for col, header in enumerate(REPORT_COLUMNS, start=1):
        cell = ws.cell(3, col, header)
        style_header(cell, TEAL)
    for row_idx, row in enumerate(rows, start=1):
        excel_row = 3 + row_idx
        values = [row_idx, *[row.get(column, "") for column in REPORT_COLUMNS[1:]]]
        for col, value in enumerate(values, start=1):
            cell = ws.cell(excel_row, col, value)
            cell.font = Font(name=BODY, size=9, color="263746")
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if col == nipp_col:
                # NIPP is an identifier; keep long/leading-zero values intact in Excel.
                cell.number_format = "@"
        label = row.get("Confidence Label")
        if label in CONFIDENCE_FILLS:
            ws.cell(excel_row, REPORT_COLUMNS.index("Confidence Label") + 1).fill = CONFIDENCE_FILLS[label]
        ws.row_dimensions[excel_row].height = 22

    end_row = 3 + max(len(rows), 1)
    if rows:
        table = Table(displayName=table_name, ref=f"A3:{get_column_letter(len(REPORT_COLUMNS))}{end_row}")
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
        ws.add_table(table)
        confirm_col = get_column_letter(REPORT_COLUMNS.index("Reviewer Confirm Mapping") + 1)
        dv = DataValidation(type="list", formula1='"YES,NEEDS_CHECK,NO"', allow_blank=True)
        ws.add_data_validation(dv)
        dv.add(f"{confirm_col}4:{confirm_col}{end_row}")

    widths = [6, 14, 12, 12, 34, 28, 28, 12, 10, 26, 28, 16, 36, 10, 18, 42, 24, 28, 22, 10, 36, 20, 28, 10, 36, 14, 36, 20, 28]
    for idx, width in enumerate(widths[: len(REPORT_COLUMNS)], start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    ws.freeze_panes = "E4"


def write_shared_sheet(ws, shared_best: dict[tuple[str, str], int], worksheets: list[WorksheetEntry]) -> None:
    by_key = {(item.source_workbook, item.sheet_name): item for item in worksheets}
    ws.sheet_view.showGridLines = False
    ws["A1"] = "Worksheets dipakai sebagai draft terbaik untuk ≥2 posisi"
    ws["A1"].font = Font(name=BODY, size=14, bold=True, color="FFFFFF")
    ws.merge_cells("A1:E1")
    ws["A1"].fill = PatternFill("solid", fgColor=NAVY)
    headers = ["Source Workbook", "Worksheet", "Worksheet Title", "Source Folder", "Position Count (best draft)"]
    for col, header in enumerate(headers, start=1):
        style_header(ws.cell(3, col, header), TEAL)
    shared_rows = sorted(
        ((key, count) for key, count in shared_best.items() if count >= 2),
        key=lambda item: (-item[1], item[0][0], item[0][1]),
    )
    for idx, ((workbook, sheet), count) in enumerate(shared_rows, start=4):
        entry = by_key.get((workbook, sheet))
        ws.cell(idx, 1, workbook)
        ws.cell(idx, 2, sheet)
        ws.cell(idx, 3, entry.position_name if entry else "")
        ws.cell(idx, 4, entry.source_folder if entry else "")
        ws.cell(idx, 5, count)
    for col, width in enumerate([55, 28, 34, 22, 18], start=1):
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.freeze_panes = "A4"


def write_unmatched_worksheets(
    ws, worksheets: list[WorksheetEntry], shared_best: dict[tuple[str, str], int]
) -> None:
    ws.sheet_view.showGridLines = False
    ws["A1"] = "Worksheets Kamus yang tidak pernah jadi draft terbaik posisi manapun"
    ws.merge_cells("A1:E1")
    ws["A1"].fill = PatternFill("solid", fgColor=NAVY)
    ws["A1"].font = Font(name=BODY, size=14, bold=True, color="FFFFFF")
    headers = ["Source Folder", "Source Workbook", "Worksheet", "Worksheet Title", "Group"]
    for col, header in enumerate(headers, start=1):
        style_header(ws.cell(3, col, header), TEAL)
    unused = [
        item
        for item in worksheets
        if shared_best.get((item.source_workbook, item.sheet_name), 0) == 0
    ]
    for idx, item in enumerate(unused, start=4):
        ws.cell(idx, 1, item.source_folder)
        ws.cell(idx, 2, item.source_workbook)
        ws.cell(idx, 3, item.sheet_name)
        ws.cell(idx, 4, item.position_name)
        ws.cell(idx, 5, item.group_name)
    for col, width in enumerate([22, 55, 28, 34, 28], start=1):
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.freeze_panes = "A4"


def build_workbook(
    *,
    rows: list[dict[str, Any]],
    labels: Counter[str],
    shared_best: dict[tuple[str, str], int],
    worksheets: list[WorksheetEntry],
    exported_at: str,
    inventory_path: str,
    reference_path: str,
    output_path: Path,
    scope_label: str,
    coverage_title: str,
    unique_active_employees: int | None = None,
    company_count: int | None = None,
) -> None:
    wb = Workbook()
    summary = wb.active
    summary.title = "Ringkasan"
    write_summary(
        summary,
        exported_at=exported_at,
        inventory_path=inventory_path,
        reference_path=reference_path,
        position_count=len(rows),
        worksheet_count=len(worksheets),
        labels=labels,
        shared_best=shared_best,
        scope_label=scope_label,
        unique_active_employees=unique_active_employees,
        company_count=company_count,
    )
    write_report_sheet(
        wb.create_sheet("Position Coverage"),
        coverage_title,
        rows,
        "PositionCoverageTable",
    )
    queue = [row for row in rows if row["Confidence Label"] != pm.HIGH_CONFIDENCE]
    write_report_sheet(
        wb.create_sheet("Review Queue"),
        "Antrian review (bukan high_confidence)",
        queue,
        "ReviewQueueTable",
    )
    write_shared_sheet(wb.create_sheet("Shared Worksheets"), shared_best, worksheets)
    write_unmatched_worksheets(wb.create_sheet("Unmatched Worksheets"), worksheets, shared_best)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--inventory",
        type=Path,
        default=Path("configs/kamus_kpi_group2_visible_20260807.json"),
    )
    parser.add_argument(
        "--reference",
        type=Path,
        default=Path("configs/production_position_reference.json"),
    )
    parser.add_argument(
        "--scope",
        choices=("all", "subholding-tree"),
        default="all",
        help="all=Group 2 non-Holding; subholding-tree=companies under Subholding org parents only",
    )
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=None,
    )
    args = parser.parse_args()
    if args.output_dir is None:
        args.output_dir = (
            Path("outputs/kamus-group2-subholding-position-mapping-20260805")
            if args.scope == "subholding-tree"
            else Path("outputs/kamus-group2-position-mapping-20260805")
        )

    reference = json.loads(args.reference.read_text(encoding="utf-8"))
    exported_at = norm(reference.get("source", {}).get("exported_at"))

    company_ids: set[int] | None = None
    unique_employees: int | None = None
    if args.scope == "subholding-tree":
        company_ids = load_subholding_tree_company_ids(args.reference)
        unique_employees = count_unique_active_employees(args.reference, company_ids)
        worksheets = load_worksheets(args.inventory, source_folder="KAMUS KPI SUBHOLDING")
        positions = load_group2_positions(args.reference, company_ids=company_ids)
        scope_label = "Subholding org-tree only (unique active employees baseline)"
        coverage_title = "Posisi pohon Subholding → draft worksheet Kamus Subholding"
        json_name = "group2_subholding_position_first_mapping_draft_20260805.json"
        xlsx_name = "Position_First_Mapping_Review_Subholding_20260805.xlsx"
    else:
        worksheets = load_worksheets(args.inventory)
        positions = load_group2_positions(args.reference)
        scope_label = "Kamus KPI Group 2 (Regional/Cabang/Subholding non-Holding)"
        coverage_title = "Seluruh posisi Group 2 → draft worksheet Kamus"
        json_name = "group2_position_first_mapping_draft_20260805.json"
        xlsx_name = "Position_First_Mapping_Review_Group2_20260805.xlsx"

    rows, labels, shared_best = resolve_all(positions, worksheets)

    args.output_dir.mkdir(parents=True, exist_ok=True)
    json_path = args.output_dir / json_name
    xlsx_path = args.output_dir / xlsx_name
    receipt_path = args.output_dir / "MAPPING_RECEIPT.md"

    payload = {
        "metadata": {
            "orientation": "position_first",
            "scope": args.scope,
            "allows_shared_worksheet": True,
            "reference": str(args.reference),
            "reference_exported_at": exported_at,
            "inventory": str(args.inventory),
            "generated_at": datetime.now().astimezone().isoformat(timespec="seconds"),
            "position_count": len(rows),
            "worksheet_count": len(worksheets),
            "company_count": len(company_ids) if company_ids is not None else None,
            "unique_active_employees": unique_employees,
            "sum_active_employees_column": sum(int(row.get("Active Employees") or 0) for row in rows),
            "confidence_counts": dict(labels),
            "shared_worksheet_count": sum(1 for count in shared_best.values() if count >= 2),
            "reviewer_columns": REVIEW_COLUMNS,
            "notes": (
                "Identity is PMID/PNID on each row. Company/location is the primary clue for workbook "
                "selection; worksheet title is matched inside that workbook. One worksheet may be "
                "assigned to many positions. For scope=subholding-tree, companies follow production "
                "parent_id under type_org=Subholding (the ~2911 unique-active-employee baseline)."
            ),
        },
        "rows": rows,
    }
    json_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    build_workbook(
        rows=rows,
        labels=labels,
        shared_best=shared_best,
        worksheets=worksheets,
        exported_at=exported_at,
        inventory_path=str(args.inventory),
        reference_path=str(args.reference),
        output_path=xlsx_path,
        scope_label=scope_label,
        coverage_title=coverage_title,
        unique_active_employees=unique_employees,
        company_count=len(company_ids) if company_ids is not None else None,
    )

    receipt = f"""# Mapping Receipt — {scope_label}

## Reference
- Kind: active production export
- Path: `{args.reference}`
- Exported at: `{exported_at}`
- Inventory: `{args.inventory}`
- Scope: `{args.scope}`

## Orientation
- Company/location is the **primary** clue for choosing the Kamus workbook; title selects the sheet inside it
- One worksheet may map to many positions
- Reviewer columns blank: {', '.join(REVIEW_COLUMNS)}
{f"- Unique active employees in Subholding org-tree: **{unique_employees}**" if unique_employees is not None else ""}
{f"- Companies in Subholding org-tree: **{len(company_ids)}**" if company_ids is not None else ""}

## Counts
| Measure | Count |
| --- | ---: |
| Positions reviewed | {len(rows)} |
| Kamus worksheets available | {len(worksheets)} |
| Sum Active Employees column | {payload["metadata"]["sum_active_employees_column"]} |
| Unique active employees (scope baseline) | {unique_employees if unique_employees is not None else "n/a"} |
| high_confidence | {labels.get(pm.HIGH_CONFIDENCE, 0)} |
| low_confidence | {labels.get(pm.LOW_CONFIDENCE, 0)} |
| mapping_conflict | {labels.get(pm.MAPPING_CONFLICT, 0)} |
| no_candidate | {labels.get(pm.NO_CANDIDATE, 0)} |
| Worksheets shared by ≥2 positions (as best draft) | {sum(1 for count in shared_best.values() if count >= 2)} |

## Artifacts
- `{xlsx_path}`
- `{json_path}`

## Next
Reviewer fills decision columns. Apply must be adapted for position-first rows before conversion.
"""
    receipt_path.write_text(receipt, encoding="utf-8")
    print(json.dumps({k: payload["metadata"][k] for k in (
        "scope", "position_count", "worksheet_count", "company_count",
        "unique_active_employees", "sum_active_employees_column", "confidence_counts",
    )}, ensure_ascii=False, indent=2))
    print(f"wrote {xlsx_path}")
    print(f"wrote {json_path}")
    print(f"wrote {receipt_path}")


if __name__ == "__main__":
    main()
