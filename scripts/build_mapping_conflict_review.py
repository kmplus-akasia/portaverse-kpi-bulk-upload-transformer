#!/usr/bin/env python3
"""Build a review workbook for unresolved KPI position mappings."""

from __future__ import annotations

import argparse
import json
from collections import Counter
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

import kpi_bulk_transform as transform
import position_mapping

GENERIC_TOKENS = {
    "group",
    "department",
    "head",
    "manager",
    "officer",
    "senior",
    "junior",
    "staf",
    "staff",
    "pt",
    "dan",
    "the",
}

REVIEW_INPUT_COLUMNS = [
    "Reviewer Confirm Mapping",
    "Reviewer Actual PMID",
    "Reviewer Actual PNID",
    "Reviewer Notes",
]

REMOVED_MAPPING_COLUMNS = {
    "Runner-up Scope",
    "Runner-up PMID",
    "Runner-up PNID",
    "Runner-up Title",
    "Runner-up Score",
}

CONFIDENCE_FILLS = {
    position_mapping.HIGH_CONFIDENCE: PatternFill("solid", fgColor="D9EAD3"),
    position_mapping.LOW_CONFIDENCE: PatternFill("solid", fgColor="FFF2CC"),
    position_mapping.SCOPE_UNCERTAIN: PatternFill("solid", fgColor="CFE2F3"),
    position_mapping.NO_CANDIDATE: PatternFill("solid", fgColor="E7E6E6"),
    position_mapping.MAPPING_CONFLICT: PatternFill("solid", fgColor="F4CCCC"),
}


def norm(value: Any) -> str:
    return "" if value is None else str(value).strip()


def tokens(value: str) -> set[str]:
    return {
        token
        for token in transform.normalize_position_lookup(value).split()
        if len(token) > 2 and token not in GENERIC_TOKENS
    }


def ratio(left: str, right: str) -> float:
    if not left or not right:
        return 0.0
    if left == right:
        return 1.0
    if left in right or right in left:
        return 0.85
    left_tokens = {token for token in left.split() if len(token) > 2 and token not in GENERIC_TOKENS}
    right_tokens = {token for token in right.split() if len(token) > 2 and token not in GENERIC_TOKENS}
    if not left_tokens or not right_tokens:
        return 0.0
    overlap = len(left_tokens & right_tokens)
    if not overlap:
        return 0.0
    return overlap / max(len(left_tokens), len(right_tokens))


def load_conflicts(config_path: Path) -> list[dict[str, Any]]:
    payload = json.loads(config_path.read_text(encoding="utf-8"))
    conflicts: list[dict[str, Any]] = []
    for row in payload.get("positions", []):
        scope = transform.normalize_position_scope(row.get("position_scope"))
        if scope == "mapping_conflict" or not (row.get("position_master_id") or row.get("position_nomenclature_id")):
            conflicts.append(row)
    return conflicts


def load_positions(config_path: Path) -> list[dict[str, Any]]:
    payload = json.loads(config_path.read_text(encoding="utf-8"))
    return [row for row in payload.get("positions", []) if isinstance(row, dict)]


def resolution_from_config(row: dict[str, Any]) -> position_mapping.MappingResolution:
    label = norm(row.get("mapping_confidence_label")) or (
        position_mapping.HIGH_CONFIDENCE
        if row.get("position_master_id") or row.get("position_nomenclature_id")
        else position_mapping.MAPPING_CONFLICT
    )
    scope = transform.normalize_position_scope(row.get("position_scope")) or position_mapping.SCOPE_UNCERTAIN
    candidate_pmid = norm(row.get("position_master_id")) or norm(row.get("candidate_position_master_id")) or None
    candidate_pnid = norm(row.get("position_nomenclature_id")) or norm(row.get("candidate_position_nomenclature_id")) or None
    return position_mapping.MappingResolution(
        source_workbook=norm(row.get("source_workbook")) or None,
        worksheet=norm(row.get("sheet_name")),
        raw_worksheet_title=norm(row.get("position_name")) or norm(row.get("sheet_name")),
        normalized_worksheet_title=position_mapping.normalize_position_lookup(row.get("position_name") or row.get("sheet_name")),
        inferred_scope=scope,
        confidence_label=label,
        confidence_reason=norm(row.get("mapping_confidence_reason")) or "Loaded from discovered config.",
        position_master_id=candidate_pmid if scope == position_mapping.STRUCTURAL else None,
        position_nomenclature_id=candidate_pnid if scope == position_mapping.NON_STRUCTURAL else None,
        candidate_title=norm(row.get("portaverse_position_title")) or None,
        candidate_group=norm(row.get("portaverse_group_name")) or None,
        candidate_company=norm(row.get("portaverse_company_name")) or None,
        candidate_company_code=norm(row.get("portaverse_company_code")) or None,
        candidate_score=row.get("candidate_score"),
        runner_up_position_master_id=norm(row.get("runner_up_position_master_id")) or None,
        runner_up_position_nomenclature_id=norm(row.get("runner_up_position_nomenclature_id")) or None,
        runner_up_score=row.get("runner_up_score"),
        active_variant_count=row.get("active_variant_count"),
        active_employee_count=row.get("active_employee_count"),
        active_employee_name=norm(row.get("active_employee_name")) or None,
        active_employee_nipp=norm(row.get("active_employee_nipp")) or None,
        upload_allowed=label == position_mapping.HIGH_CONFIDENCE,
    )


def build_mapping_report_rows(
    positions: list[dict[str, Any]],
    reference_path: Path,
    target_company_id: str | None,
) -> list[dict[str, Any]]:
    payload = json.loads(reference_path.read_text(encoding="utf-8"))
    indexes = position_mapping.build_lookup_indexes(payload, target_company_id)
    rows: list[dict[str, Any]] = []
    for item in positions:
        if item.get("mapping_confidence_label"):
            resolution = resolution_from_config(item)
        else:
            resolution = position_mapping.resolve_mapping(
                worksheet=norm(item.get("sheet_name")),
                worksheet_title=norm(item.get("position_name")) or norm(item.get("sheet_name")),
                group_name=norm(item.get("group_name")) or None,
                source_workbook=norm(item.get("source_workbook")) or None,
                indexes=indexes,
            )
        rows.append(simplify_mapping_report_row(position_mapping.mapping_report_row(resolution)))
    return rows


def simplify_mapping_report_row(row: dict[str, Any]) -> dict[str, Any]:
    simplified = {
        key: value
        for key, value in row.items()
        if key not in REMOVED_MAPPING_COLUMNS
    }
    output: dict[str, Any] = {}
    for key, value in simplified.items():
        output[key] = value
        if key == "Recommended Action":
            for review_key in REVIEW_INPUT_COLUMNS:
                output[review_key] = ""
    return output


def reference_candidates(reference_path: Path, target_company_id: str | None) -> list[dict[str, Any]]:
    payload = json.loads(reference_path.read_text(encoding="utf-8"))
    candidates: list[dict[str, Any]] = []
    target = norm(target_company_id)
    for row in payload.get("position_master_rows", []):
        company_id = norm(row.get("company_id"))
        if target and company_id != target:
            continue
        pmid = norm(row.get("position_master_id"))
        title = norm(row.get("position_name"))
        if not pmid or not title:
            continue
        candidates.append(
            {
                "scope": "structural",
                "id": pmid,
                "title": title,
                "group_name": norm(row.get("active_group_name")) or norm(row.get("group_name")),
                "company_name": norm(row.get("active_company_name")) or norm(row.get("company_name")),
                "company_code": norm(row.get("active_company_code")) or norm(row.get("company_code")),
                "active_position_org": norm(row.get("is_position_organization_active")),
                "source": "position_master_rows",
            }
        )
    seen_pnid: set[str] = set()
    for row in payload.get("rows", []):
        company_id = norm(row.get("company_id"))
        if target and company_id != target:
            continue
        pnid = norm(row.get("cluster_id"))
        if not pnid or pnid in seen_pnid:
            continue
        seen_pnid.add(pnid)
        title = norm(row.get("cluster_label")) or norm(row.get("position_name"))
        if not title:
            continue
        candidates.append(
            {
                "scope": "non_structural",
                "id": pnid,
                "title": title,
                "group_name": norm(row.get("active_group_name")) or norm(row.get("group_name")),
                "company_name": norm(row.get("active_company_name")) or norm(row.get("company_name")),
                "company_code": norm(row.get("active_company_code")) or norm(row.get("company_code")),
                "active_position_org": norm(row.get("is_position_organization_active")),
                "source": "rows",
            }
        )
    for candidate in candidates:
        candidate["_title_norm"] = transform.normalize_position_lookup(candidate.get("title"))
        candidate["_group_norm"] = transform.normalize_position_lookup(candidate.get("group_name"))
        candidate["_company_norm"] = transform.normalize_company_context(
            f"{candidate.get('company_name', '')} {candidate.get('company_code', '')}"
        )
        candidate["_tokens"] = tokens(
            f"{candidate.get('title', '')} {candidate.get('group_name', '')} {candidate.get('company_name', '')}"
        )
    return candidates


def candidate_index(candidates: list[dict[str, Any]]) -> dict[str, list[dict[str, Any]]]:
    index: dict[str, list[dict[str, Any]]] = {}
    for candidate in candidates:
        for token in candidate.get("_tokens", set()):
            index.setdefault(token, []).append(candidate)
    return index


def conflict_context(conflict: dict[str, Any]) -> dict[str, Any]:
    raw_position = transform.normalize_position_lookup(conflict.get("position_name"))
    raw_group = transform.normalize_position_lookup(conflict.get("group_name"))
    raw_sheet = transform.normalize_position_lookup(conflict.get("sheet_name"))
    lookup_names = [
        raw_position,
        raw_sheet,
        transform.normalize_position_lookup(f"{conflict.get('position_name', '')} {conflict.get('group_name', '')}"),
        transform.normalize_position_lookup(f"{conflict.get('sheet_name', '')} {conflict.get('group_name', '')}"),
    ]
    source_hints = transform.source_workbook_context_hints(conflict.get("source_workbook"))
    lookup_tokens = tokens(" ".join(lookup_names))
    return {
        "raw_group": raw_group,
        "lookup_names": [lookup for lookup in lookup_names if lookup],
        "source_hints": source_hints,
        "tokens": lookup_tokens,
    }


def score_candidate(context: dict[str, Any], candidate: dict[str, Any]) -> tuple[float, str]:
    candidate_title = candidate.get("_title_norm", "")
    candidate_group = candidate.get("_group_norm", "")
    candidate_company = candidate.get("_company_norm", "")
    title_score = max(ratio(lookup, candidate_title) for lookup in context["lookup_names"])
    group_score = ratio(context["raw_group"], candidate_group)
    source_hints = context["source_hints"]
    company_score = 1.0 if any(hint and hint in candidate_company for hint in source_hints) else 0.0
    score = (title_score * 0.70) + (group_score * 0.20) + (company_score * 0.10)
    reason_parts = [
        f"title={title_score:.2f}",
        f"group={group_score:.2f}",
    ]
    if company_score:
        reason_parts.append("company_hint=match")
    return score, "; ".join(reason_parts)


def build_review_rows(
    conflicts: list[dict[str, Any]],
    candidates: list[dict[str, Any]],
    per_conflict: int,
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    index = candidate_index(candidates)
    for conflict in conflicts:
        context = conflict_context(conflict)
        candidate_pool: dict[int, dict[str, Any]] = {}
        for token in context["tokens"]:
            for candidate in index.get(token, []):
                candidate_pool[id(candidate)] = candidate
        if not candidate_pool:
            candidate_pool = {id(candidate): candidate for candidate in candidates[:2000]}
        if len(candidate_pool) > 200:
            context_tokens = context["tokens"]
            ranked_pool = sorted(
                candidate_pool.values(),
                key=lambda candidate: (
                    len(context_tokens & candidate.get("_tokens", set())),
                    1 if any(hint and hint in candidate.get("_company_norm", "") for hint in context["source_hints"]) else 0,
                ),
                reverse=True,
            )
            candidate_pool = {id(candidate): candidate for candidate in ranked_pool[:200]}
        scored: list[tuple[float, str, dict[str, Any]]] = []
        for candidate in candidate_pool.values():
            score, reason = score_candidate(context, candidate)
            if score >= 0.35:
                scored.append((score, reason, candidate))
        scored.sort(key=lambda item: item[0], reverse=True)
        if not scored:
            rows.append(review_row(conflict, None, 0.0, "no candidate above threshold", 0))
            continue
        unique_scored: list[tuple[float, str, dict[str, Any]]] = []
        seen_candidate_keys: set[tuple[str, str]] = set()
        for score, reason, candidate in scored:
            scope = norm(candidate.get("scope"))
            candidate_key = (scope, norm(candidate.get("id")))
            if candidate_key in seen_candidate_keys:
                continue
            seen_candidate_keys.add(candidate_key)
            unique_scored.append((score, reason, candidate))
            if len(unique_scored) >= per_conflict:
                break
        for rank, (score, reason, candidate) in enumerate(unique_scored, start=1):
            rows.append(review_row(conflict, candidate, score, reason, rank))
    return rows


def review_row(
    conflict: dict[str, Any],
    candidate: dict[str, Any] | None,
    score: float,
    reason: str,
    rank: int,
) -> dict[str, Any]:
    candidate = candidate or {}
    scope = norm(candidate.get("scope"))
    return {
        "Source Workbook": norm(conflict.get("source_workbook")),
        "Sheet": norm(conflict.get("sheet_name")),
        "Raw Group": norm(conflict.get("group_name")),
        "Raw Position": norm(conflict.get("position_name")),
        "Direktorat": norm(conflict.get("directorate_name")),
        "Candidate Rank": rank or "",
        "Candidate Score": round(score, 4),
        "Candidate Scope": scope,
        "Candidate PMID": norm(candidate.get("id")) if scope == "structural" else "",
        "Candidate PNID": norm(candidate.get("id")) if scope == "non_structural" else "",
        "Candidate Title": norm(candidate.get("title")),
        "Candidate Group": norm(candidate.get("group_name")),
        "Candidate Company": norm(candidate.get("company_name")),
        "Candidate Company Code": norm(candidate.get("company_code")),
        "Candidate Source": norm(candidate.get("source")),
        "Match Reason": reason,
        "Recommended Action": "Review candidate and copy PMID/PNID into manual config override if correct.",
    }


def write_sheet(wb: Workbook, title: str, rows: list[dict[str, Any]]) -> None:
    ws = wb.create_sheet(title[:31])
    ws.sheet_properties.tabColor = "1F4E78"
    if not rows:
        ws.append(["No data"])
        return
    headers = list(rows[0].keys())
    ws.append(headers)
    for row in rows:
        ws.append([row.get(header, "") for header in headers])
    style_sheet(ws)


def style_sheet(ws: Any) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    thin_gray = Side(style="thin", color="D9E2EC")
    border = Border(left=thin_gray, right=thin_gray, top=thin_gray, bottom=thin_gray)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if cell.row % 2 == 0:
                cell.fill = PatternFill("solid", fgColor="F8FAFC")
    header_by_name = {norm(cell.value): cell.column for cell in ws[1]}
    confidence_col = header_by_name.get("Confidence Label")
    if confidence_col:
        for row_index in range(2, ws.max_row + 1):
            cell = ws.cell(row_index, confidence_col)
            fill = CONFIDENCE_FILLS.get(norm(cell.value))
            if fill:
                cell.fill = fill
                cell.font = Font(bold=True, color="1F2937")
    confirm_col = header_by_name.get("Reviewer Confirm Mapping")
    if confirm_col and ws.max_row >= 2:
        validation = DataValidation(type="list", formula1='"YES,NO,NEEDS_CHECK"', allow_blank=True)
        ws.add_data_validation(validation)
        validation.add(f"{get_column_letter(confirm_col)}2:{get_column_letter(confirm_col)}{ws.max_row}")
    preferred_widths = {
        "Source Workbook": 58,
        "Confidence Label": 22,
        "Confidence Reason": 60,
        "Raw Group": 36,
        "Raw Position": 32,
        "Candidate PMID": 16,
        "Candidate PNID": 16,
        "Candidate Title": 42,
        "Candidate Group": 38,
        "Candidate Company": 38,
        "Active Employee Name": 42,
        "Active Employee NIPP": 30,
        "Match Reason": 28,
        "Recommended Action": 46,
        "Reviewer Confirm Mapping": 24,
        "Reviewer Actual PMID": 18,
        "Reviewer Actual PNID": 18,
        "Reviewer Notes": 42,
    }
    for column_index in range(1, ws.max_column + 1):
        header = norm(ws.cell(1, column_index).value)
        width = preferred_widths.get(header, 18)
        ws.column_dimensions[get_column_letter(column_index)].width = width


def build_workbook(positions: list[dict[str, Any]], review_rows: list[dict[str, Any]], reference_path: Path) -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)
    by_label = Counter(norm(row.get("Confidence Label")) for row in review_rows)
    by_source = Counter(norm(row.get("source_workbook")).split("/")[0] or "<unknown>" for row in positions)
    summary_rows = [
        {"Metric": "Worksheet Count", "Value": len(positions)},
        {"Metric": "Mapping Report Rows", "Value": len(review_rows)},
        {"Metric": "Reference", "Value": str(reference_path)},
    ]
    summary_rows.extend({"Metric": f"Confidence: {label}", "Value": count} for label, count in sorted(by_label.items()))
    summary_rows.extend({"Metric": f"Conflicts in {source}", "Value": count} for source, count in by_source.most_common())
    write_sheet(wb, "Summary", summary_rows)
    write_sheet(wb, "Position Mapping Report", review_rows)
    blocked_rows = [
        row
        for row in review_rows
        if norm(row.get("Confidence Label")) != position_mapping.HIGH_CONFIDENCE
    ]
    write_sheet(wb, "Review Queue", blocked_rows)
    return wb


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--config", required=True, type=Path)
    parser.add_argument("--reference", required=True, type=Path)
    parser.add_argument("--output", required=True, type=Path)
    parser.add_argument("--target-company-id", default="")
    args = parser.parse_args()

    positions = load_positions(args.config)
    review_rows = build_mapping_report_rows(positions, args.reference, args.target_company_id)
    workbook = build_workbook(positions, review_rows, args.reference)
    args.output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(args.output)
    blocked_count = sum(1 for row in review_rows if norm(row.get("Confidence Label")) != position_mapping.HIGH_CONFIDENCE)
    print(f"worksheets={len(positions)}")
    print(f"review_rows={len(review_rows)}")
    print(f"blocked_rows={blocked_count}")
    print(f"output={args.output}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
