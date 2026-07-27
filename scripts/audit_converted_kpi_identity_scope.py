#!/usr/bin/env python3
"""Audit converted KPI upload workbooks for PMID/PNID scope inversions."""

from __future__ import annotations

import argparse
import csv
import json
import re
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


EXPECTED_HEADERS = [
    "IDKPI",
    "Group",
    "Direktorat",
    "Posisi",
    "Position Master ID (Required)",
    "Position Master Variant ID (Optional)",
    "BSC Perspective",
    "KPI Type",
    "Parent KPI ID",
    "Parent KPI Title",
    "Title",
    "Description",
    "Unit",
    "Polarity",
    "Period",
    "Formula",
    "Weight (%)",
    "Cascading",
    "Nature Of Work (KAI Only)",
    "External ID (PKPI)",
    "System KPI ID",
    "Ownership Type",
    "Position Nomenklatur ID",
    "RKM Code ID",
]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--roots",
        nargs="+",
        type=Path,
        default=[Path("output"), Path("outputs")],
        help="Directories to scan recursively for converted .xlsx files.",
    )
    parser.add_argument(
        "--reference",
        type=Path,
        default=Path("configs/production_position_reference.json"),
        help="Position reference JSON containing position_master_rows and rows.",
    )
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=Path("output/identity_scope_audit_20260623"),
        help="Directory for audit CSV/JSON outputs.",
    )
    return parser.parse_args()


def norm(value: object) -> str:
    text = str(value or "").lower()
    text = re.sub(r"\bdh\b", "department head", text)
    text = re.sub(r"\bdept\b", "department", text)
    text = re.sub(r"\badmin\b", "administrasi", text)
    text = text.replace("&", " dan ")
    text = re.sub(r"[-_/(),.]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def identity_tokens(value: object) -> set[str]:
    stopwords = {
        "dan",
        "group",
        "unit",
        "pendukung",
        "department",
        "dept",
        "company",
        "pt",
        "persero",
        "pelabuhan",
        "indonesia",
    }
    tokens = set()
    for token in norm(value).split():
        if not token or token in stopwords:
            continue
        if token.isdigit():
            continue
        tokens.add(token)
    return tokens


def token_match(left: object, right: object, min_tokens: int = 3) -> bool:
    left_tokens = identity_tokens(left)
    right_tokens = identity_tokens(right)
    if len(left_tokens) < min_tokens or len(right_tokens) < min_tokens:
        return False
    return left_tokens.issubset(right_tokens) or right_tokens.issubset(left_tokens)


def load_reference(reference_path: Path) -> dict[str, Any]:
    reference = json.loads(reference_path.read_text(encoding="utf-8"))
    masters_by_id: dict[str, list[dict[str, Any]]] = {}
    pnid_rows_by_id: dict[str, list[dict[str, Any]]] = {}
    for row in reference.get("position_master_rows", []):
        pmid = row.get("position_master_id")
        if pmid not in (None, "", 0, "0"):
            masters_by_id.setdefault(str(pmid), []).append(row)
    for row in reference.get("rows", []):
        pnid = row.get("cluster_id")
        if pnid not in (None, "", 0, "0"):
            pnid_rows_by_id.setdefault(str(pnid), []).append(row)
    return {
        "masters_by_id": masters_by_id,
        "pnid_rows_by_id": pnid_rows_by_id,
    }


def production_types(rows: list[dict[str, Any]]) -> set[str]:
    return {str(row.get("position_master_type_id") or "") for row in rows}


def matches_structural_master(position_name: object, master_rows: list[dict[str, Any]]) -> bool:
    return any(
        str(row.get("position_master_type_id") or "") == "5"
        and token_match(position_name, row.get("position_name"), min_tokens=3)
        for row in master_rows
    )


def matches_non_structural_pnid(position_name: object, pnid_rows: list[dict[str, Any]]) -> bool:
    for row in pnid_rows:
        if str(row.get("position_master_type_id") or "") == "5":
            continue
        if token_match(position_name, row.get("cluster_label"), min_tokens=3):
            return True
        if token_match(position_name, row.get("position_name"), min_tokens=3):
            return True
    return False


def workbook_paths(roots: list[Path]) -> list[Path]:
    paths: list[Path] = []
    for root in roots:
        if not root.exists():
            continue
        for path in root.rglob("*.xlsx"):
            if path.name.startswith("~$"):
                continue
            paths.append(path)
    return sorted(paths)


def read_upload_headers(path: Path) -> list[Any] | None:
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except Exception:
        return None
    if "KPI Template" not in workbook.sheetnames:
        return None
    sheet = workbook["KPI Template"]
    return [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1, max_col=len(EXPECTED_HEADERS)))]


def classify_workbook(path: Path) -> str:
    parts = path.parts
    if "upload-ready" in parts:
        return "upload-ready"
    if "KPI_Upload_Final_20260618" in parts:
        return "zip-expanded-copy"
    if "outputs" in parts:
        return "auxiliary-output"
    return "generated-workbook"


def audit_workbook(path: Path, reference: dict[str, Any]) -> tuple[dict[str, Any], list[dict[str, Any]]]:
    issues: list[dict[str, Any]] = []
    masters_by_id = reference["masters_by_id"]
    pnid_rows_by_id = reference["pnid_rows_by_id"]
    workbook = load_workbook(path, read_only=True, data_only=True)
    if "KPI Template" not in workbook.sheetnames:
        return (
            {
                "path": str(path),
                "category": "not-upload-workbook",
                "rows": 0,
                "status": "SKIPPED",
            },
            issues,
        )
    sheet = workbook["KPI Template"]
    headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1, max_col=len(EXPECTED_HEADERS)))]
    if headers != EXPECTED_HEADERS:
        issues.append(
            {
                "severity": "error",
                "issue_type": "HEADER_SCHEMA_MISMATCH",
                "workbook": str(path),
                "row": "",
                "position": "",
                "pmid": "",
                "pnid": "",
                "message": "KPI Template headers do not match the expected 24-column upload schema.",
            }
        )

    counts = Counter()
    for row_idx, row in enumerate(
        sheet.iter_rows(min_row=2, max_col=len(EXPECTED_HEADERS), values_only=True),
        start=2,
    ):
        title = row[10]
        if title in (None, ""):
            continue
        counts["rows"] += 1
        position_name = row[3]
        pmid = str(row[4] or "").strip()
        pnid = str(row[22] or "").strip()

        def add_issue(issue_type: str, message: str, severity: str = "error") -> None:
            issues.append(
                {
                    "severity": severity,
                    "issue_type": issue_type,
                    "workbook": str(path),
                    "row": row_idx,
                    "position": position_name,
                    "kpi_type": row[7],
                    "kpi_title": title,
                    "pmid": pmid,
                    "pnid": pnid,
                    "message": message,
                }
            )

        if pmid and pnid:
            counts["double_identity_rows"] += 1
            add_issue("DOUBLE_IDENTITY", "Row has both PMID and PNID populated.")
            continue
        if not pmid and not pnid:
            counts["blank_identity_rows"] += 1
            add_issue("BLANK_IDENTITY", "Row has neither PMID nor PNID populated.")
            continue

        if pmid:
            counts["pmid_rows"] += 1
            master_rows = masters_by_id.get(pmid, [])
            if not master_rows:
                add_issue("INVALID_PMID", "PMID does not exist in reference position_master_rows.")
                continue
            types = production_types(master_rows)
            if types != {"5"}:
                add_issue(
                    "NON_STRUCTURAL_AS_PMID",
                    f"PMID has non-structural production types {sorted(types)}; expected PNID ownership.",
                )
            pnid_rows = pnid_rows_by_id.get(pmid, [])
            if pnid_rows and matches_non_structural_pnid(position_name, pnid_rows) and not matches_structural_master(position_name, master_rows):
                add_issue(
                    "NON_STRUCTURAL_AS_PMID_BY_IDENTITY",
                    "PMID value also exists as PNID and row position identity matches non-structural PNID.",
                )
        else:
            counts["pnid_rows"] += 1
            pnid_rows = pnid_rows_by_id.get(pnid, [])
            if not pnid_rows:
                add_issue("INVALID_PNID", "PNID does not exist as rows[].cluster_id in reference.")
                continue
            types = production_types(pnid_rows)
            if not types or "5" in types:
                add_issue(
                    "STRUCTURAL_TYPE_IN_PNID",
                    f"PNID maps to structural production type(s) {sorted(types)}.",
                )
            master_rows = masters_by_id.get(pnid, [])
            if master_rows and matches_structural_master(position_name, master_rows):
                add_issue(
                    "STRUCTURAL_AS_PNID_BY_IDENTITY",
                    "PNID value also exists as structural PMID and row position identity matches the structural position.",
                )

    summary = {
        "path": str(path),
        "category": classify_workbook(path),
        "rows": counts["rows"],
        "pmid_rows": counts["pmid_rows"],
        "pnid_rows": counts["pnid_rows"],
        "blank_identity_rows": counts["blank_identity_rows"],
        "double_identity_rows": counts["double_identity_rows"],
        "issues": len(issues),
        "status": "FAIL" if issues else "PASS",
    }
    return summary, issues


def write_csv(path: Path, rows: list[dict[str, Any]], fields: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields)
        writer.writeheader()
        writer.writerows(rows)


def main() -> int:
    args = parse_args()
    reference = load_reference(args.reference)
    summaries: list[dict[str, Any]] = []
    issues: list[dict[str, Any]] = []
    skipped = 0
    for path in workbook_paths(args.roots):
        headers = read_upload_headers(path)
        if headers is None:
            skipped += 1
            continue
        summary, workbook_issues = audit_workbook(path, reference)
        summaries.append(summary)
        issues.extend(workbook_issues)

    issue_counts = Counter(issue["issue_type"] for issue in issues)
    category_counts = Counter(summary["category"] for summary in summaries)
    status_counts = Counter(summary["status"] for summary in summaries)
    total_rows = sum(int(summary["rows"]) for summary in summaries)
    total_pmid_rows = sum(int(summary["pmid_rows"]) for summary in summaries)
    total_pnid_rows = sum(int(summary["pnid_rows"]) for summary in summaries)

    args.output_dir.mkdir(parents=True, exist_ok=True)
    write_csv(
        args.output_dir / "workbook_identity_scope_summary.csv",
        summaries,
        [
            "status",
            "category",
            "rows",
            "pmid_rows",
            "pnid_rows",
            "blank_identity_rows",
            "double_identity_rows",
            "issues",
            "path",
        ],
    )
    write_csv(
        args.output_dir / "identity_scope_issues.csv",
        issues,
        [
            "severity",
            "issue_type",
            "workbook",
            "row",
            "position",
            "kpi_type",
            "kpi_title",
            "pmid",
            "pnid",
            "message",
        ],
    )
    summary_json = {
        "generated_at": datetime.now(timezone.utc).isoformat(timespec="seconds"),
        "roots": [str(root) for root in args.roots],
        "reference": str(args.reference),
        "scanned_upload_workbooks": len(summaries),
        "skipped_non_upload_workbooks": skipped,
        "status_counts": dict(status_counts),
        "category_counts": dict(category_counts),
        "total_rows": total_rows,
        "total_pmid_rows": total_pmid_rows,
        "total_pnid_rows": total_pnid_rows,
        "total_issues": len(issues),
        "issue_counts": dict(issue_counts),
    }
    (args.output_dir / "summary.json").write_text(
        json.dumps(summary_json, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )

    print(f"scanned_upload_workbooks={len(summaries)}")
    print(f"skipped_non_upload_workbooks={skipped}")
    print(f"total_rows={total_rows}")
    print(f"total_pmid_rows={total_pmid_rows}")
    print(f"total_pnid_rows={total_pnid_rows}")
    print(f"total_issues={len(issues)}")
    print(f"issue_counts={dict(issue_counts)}")
    print(f"output_dir={args.output_dir}")
    return 1 if issues else 0


if __name__ == "__main__":
    raise SystemExit(main())
