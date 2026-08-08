#!/usr/bin/env python3
"""Build a workbook-level alias review artifact for Subholding roster mapping."""

from __future__ import annotations

import json
import re
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo

OUT = Path("outputs/kamus-group2-subholding-roster-mapping-20260806")
DRAFT = OUT / "subholding_roster_position_first_mapping_20260806.json"
INVENTORY = Path("configs/kamus_kpi_group2_visible_20260807.json")

NAVY, TEAL, PALE = "173651", "138074", "E9F1F8"
BODY = "Aptos"

HQ_ALIASES = {
    (
        "KAMUS KPI SPMT.xlsx",
        "KAMUS KPI SUBHOLDING/KAMUS KPI SPMT/KAMUS KPI HO SPMT/KAMUS KPI SPMT.xlsx",
    ): (
        "KAMUS KPI SUBHOLDING/KAMUS KPI SPMT/KAMUS KPI SUBHOLDING SPMT/"
        "Kamus KPI SPMT - Mapping dengan Kontrak Manajemen.xlsx"
    ),
    (
        "KAMUS KPI SPTP.xlsx",
        "KAMUS KPI SUBHOLDING/KAMUS KPI SPTP_/KAMUS KPI HO SPTP/KAMUS KPI SPTP.xlsx",
    ): (
        "KAMUS KPI SUBHOLDING/KAMUS KPI SPTP /KAMUS KPI SUBHOLDING SPTP/"
        "Kamus KPI SPTP - Mapping dengan Kontrak Manajemen.xlsx"
    ),
    (
        "KAMUS KPI SPJM.xlsx",
        "KAMUS KPI SUBHOLDING/KAMUS KPI SPJM_/KAMUS KPI HO SPJM/KAMUS KPI SPJM.xlsx",
    ): (
        "KAMUS KPI SUBHOLDING/KAMUS KPI SPJM /KAMUS KPI SUBHOLDING SPJM/"
        "Kamus KPI SPJM - Mapping dengan Kontrak Manajemen.xlsx"
    ),
    (
        "KAMUS KPI SPSL.xlsx",
        "KAMUS KPI SUBHOLDING/KAMUS KPI SPSL/KAMUS KPI HO SPSL/KAMUS KPI SPSL.xlsx",
    ): (
        "KAMUS KPI SUBHOLDING/KAMUS KPI SPSL/KAMUS KPI SUBHOLDING SPSL/"
        "Kamus KPI SPSL - Mapping dengan Kontrak Manajemen.xlsx"
    ),
}


def norm(value: object) -> str:
    return "" if value is None else str(value).strip()


def norm_path(path: object) -> str:
    text = norm(path).replace("\\", "/")
    for prefix in (
        "KAMUS KPI PELINDO GROUP 2 (REGIONAL, CABANG DAN SUBHOLDING)/",
        "KAMUS KPI PELINDO GROUP 3 (AFILIASI, NON CLUSTER, DANA PENSIUN)/",
    ):
        if text.startswith(prefix):
            text = text[len(prefix) :]
    return text


def suggest_paths(
    *,
    workbook_title: str,
    folder: str,
    candidate_top: str | None,
    candidate_share: int,
    row_count: int,
    inv_paths: list[str],
    inv_by_base: dict[str, list[str]],
) -> list[tuple[str, str, int]]:
    suggestions: list[tuple[str, str, int]] = []
    if (workbook_title, folder) in HQ_ALIASES:
        suggestions.append((HQ_ALIASES[(workbook_title, folder)], "hq_alias_prior", 100))

    # Exact basename matches from R&W title / folder beat weak resolver candidates.
    base = (workbook_title or (folder.split("/")[-1] if folder else "")).casefold()
    for path in inv_by_base.get(base, []):
        suggestions.append((path, "basename_exact", 95))
    folder_base = folder.split("/")[-1].casefold() if folder else ""
    if folder_base and folder_base != base:
        for path in inv_by_base.get(folder_base, []):
            suggestions.append((path, "folder_basename_exact", 92))

    # Parent-folder company cue, e.g. .../KAMUS KPI PT TERMINAL PETIKEMAS SURABAYA/<file>.xlsx
    folder_parts = [part for part in folder.split("/") if part]
    if len(folder_parts) >= 2:
        parent = folder_parts[-2]
        parent_hits = [
            path
            for path in inv_paths
            if parent.casefold() in path.casefold() and path.split("/")[-1].casefold().startswith("kamus")
        ]
        # Prefer inventory file whose directory contains the same parent folder name.
        for path in parent_hits:
            if f"/{parent}/" in path or path.casefold().find(parent.casefold()) >= 0:
                suggestions.append((path, "folder_parent_match", 88))

    # Dominant candidate only if it covers most rows for this alias key.
    if candidate_top and candidate_top in inv_paths and row_count > 0 and (candidate_share / row_count) >= 0.5:
        suggestions.append((candidate_top, "dominant_candidate_in_inventory", 85))

    segments = [part for part in re.split(r"[/ _-]+", f"{folder} {workbook_title}") if len(part) > 2]
    stop = {
        "kamus",
        "kpi",
        "xlsx",
        "xlsm",
        "dengan",
        "kontrak",
        "manajemen",
        "mapping",
        "pelindo",
        "group",
        "dan",
        "ho",
    }
    segments = [part for part in segments if part.casefold() not in stop]
    scored: list[tuple[int, str]] = []
    for path in inv_paths:
        score = 0
        path_cf = path.casefold()
        for segment in segments:
            if segment.casefold() in path_cf:
                score += 1
        if folder_base and path.split("/")[-1].casefold() == folder_base:
            score += 5
        if workbook_title and path.split("/")[-1].casefold() == workbook_title.casefold():
            score += 4
        if score:
            scored.append((score, path))
    scored.sort(reverse=True)
    for score, path in scored[:5]:
        suggestions.append((path, f"path_token_score_{score}", min(70 + score, 84)))

    dedup: dict[str, tuple[str, int]] = {}
    for path, reason, score in suggestions:
        if path not in dedup or score > dedup[path][1]:
            dedup[path] = (reason, score)
    return sorted(
        ((path, reason, score) for path, (reason, score) in dedup.items()),
        key=lambda item: -item[2],
    )[:5]


def main() -> None:
    draft = json.loads(DRAFT.read_text(encoding="utf-8"))
    inventory = json.loads(INVENTORY.read_text(encoding="utf-8"))

    inv_paths = sorted(
        {
            norm(row.get("source_workbook"))
            for row in inventory.get("kamus_kpi_v2", [])
            if row.get("include_in_position_config") and norm(row.get("source_workbook"))
        }
    )
    inv_by_base: dict[str, list[str]] = defaultdict(list)
    for path in inv_paths:
        inv_by_base[path.split("/")[-1].casefold()].append(path)

    unresolved = [
        row
        for row in draft.get("rows", [])
        if "INVENTORY_UNRESOLVED" in norm(row.get("Roster Review Tag"))
    ]
    groups: dict[tuple[str, str], dict] = defaultdict(
        lambda: {
            "rows": 0,
            "employees": 0,
            "nipps": set(),
            "companies": Counter(),
            "candidate_wb": Counter(),
            "sheets": Counter(),
        }
    )
    for row in unresolved:
        workbook_title = norm(row.get("Reviewed Workbook Title"))
        folder = norm_path(row.get("Reviewed Folder"))
        bucket = groups[(workbook_title, folder)]
        bucket["rows"] += 1
        nipps = [part.strip() for part in norm(row.get("Active Employee NIPPs")).split(";") if part.strip()]
        bucket["nipps"].update(nipps)
        bucket["employees"] += int(row.get("Active Employees") or len(nipps) or 0)
        if row.get("Company"):
            bucket["companies"][norm(row.get("Company"))] += 1
        if row.get("Candidate Source Workbook"):
            bucket["candidate_wb"][norm(row.get("Candidate Source Workbook"))] += 1
        if row.get("Reviewed Worksheet Title"):
            bucket["sheets"][norm(row.get("Reviewed Worksheet Title"))] += 1

    rows_out: list[dict[str, object]] = []
    for (workbook_title, folder), bucket in sorted(
        groups.items(), key=lambda item: (-item[1]["rows"], item[0][0], item[0][1])
    ):
        candidate_top = bucket["candidate_wb"].most_common(1)[0][0] if bucket["candidate_wb"] else ""
        candidate_share = bucket["candidate_wb"].most_common(1)[0][1] if bucket["candidate_wb"] else 0
        ranked = suggest_paths(
            workbook_title=workbook_title,
            folder=folder,
            candidate_top=candidate_top or None,
            candidate_share=candidate_share,
            row_count=bucket["rows"],
            inv_paths=inv_paths,
            inv_by_base=inv_by_base,
        )
        proposed = ranked[0][0] if ranked else ""
        reason = ranked[0][1] if ranked else "no_suggestion"
        ambiguous = ""
        if (
            workbook_title == "KAMUS KPI WILAYAH TERMINAL PETIKEMAS.xlsx"
            and folder.endswith("KAMUS KPI WILAYAH TERMINAL PETIKEMAS/KAMUS KPI WILAYAH TERMINAL PETIKEMAS.xlsx")
        ):
            ambiguous = "YES — folder R&W mengarah ke induk wilayah; perlu pecah per terminal/wilayah"
            proposed = ""
            reason = "needs_split_by_wilayah"
        elif len(ranked) >= 2 and ranked[0][2] < 90 and ranked[0][2] == ranked[1][2]:
            ambiguous = "YES — beberapa kandidat skor sama"

        if reason in {
            "hq_alias_prior",
            "dominant_candidate_in_inventory",
            "basename_exact",
            "folder_basename_exact",
            "folder_parent_match",
        } and not ambiguous:
            confidence = "high"
        elif proposed and not ambiguous:
            confidence = "medium"
        else:
            confidence = "low"

        rows_out.append(
            {
                "No.": len(rows_out) + 1,
                "Reviewed Workbook Title": workbook_title,
                "Reviewed Folder (normalized)": folder,
                "Position Rows Affected": bucket["rows"],
                "Unique NIPPs Affected": len(bucket["nipps"]),
                "Sum Active Employees": bucket["employees"],
                "Top Companies": "; ".join(f"{name} ({count})" for name, count in bucket["companies"].most_common(5)),
                "Dominant Candidate Workbook": candidate_top,
                "Candidate Workbook Share": bucket["candidate_wb"].most_common(1)[0][1] if bucket["candidate_wb"] else 0,
                "Proposed Inventory Workbook": proposed,
                "Proposal Reason": reason,
                "Proposal Confidence": confidence,
                "Ambiguous / Needs Split": ambiguous,
                "Alt Inventory 2": ranked[1][0] if len(ranked) > 1 else "",
                "Alt Inventory 3": ranked[2][0] if len(ranked) > 2 else "",
                "Sample Reviewed Worksheets": "; ".join(
                    f"{name} ({count})" for name, count in bucket["sheets"].most_common(5)
                ),
                "Reviewer Approve Alias": "",
                "Reviewer Final Inventory Workbook": "",
                "Reviewer Notes": "",
            }
        )

    book = Workbook()
    summary = book.active
    summary.title = "Ringkasan"
    summary["A1"] = "Workbook Alias Review — Subholding Roster 2026-08-06"
    summary["A1"].font = Font(name=BODY, size=16, bold=True, color="FFFFFF")
    summary["A1"].fill = PatternFill("solid", fgColor=NAVY)
    summary.merge_cells("A1:F1")
    summary["A2"] = (
        "Fokus tahap 1: samakan path workbook R&W ke path inventory resmi. "
        "Isi kolom Reviewer Approve Alias (YES/NO/NEEDS_SPLIT). Worksheet alias menyusul setelah ini dikunci."
    )
    summary["A2"].fill = PatternFill("solid", fgColor=PALE)
    summary.merge_cells("A2:F2")
    unique_nipps = set().union(*(bucket["nipps"] for bucket in groups.values())) if groups else set()
    meta = [
        ("Generated at", datetime.now().astimezone().isoformat(timespec="seconds")),
        ("Source mapping", str(OUT / "Position_First_Mapping_Subholding_20260806.xlsx")),
        ("Inventory", str(INVENTORY)),
        ("Unresolved position rows", len(unresolved)),
        ("Distinct workbook+folder aliases", len(rows_out)),
        ("Unique NIPPs touched", len(unique_nipps)),
        (
            "How to review",
            "Sort by Position Rows Affected desc. YES=terima Proposed. NEEDS_SPLIT=pecah. NO=isi Final path sendiri.",
        ),
    ]
    summary["A4"] = "Item"
    summary["B4"] = "Value"
    for cell in (summary["A4"], summary["B4"]):
        cell.fill = PatternFill("solid", fgColor=TEAL)
        cell.font = Font(name=BODY, bold=True, color="FFFFFF")
    for offset, (label, value) in enumerate(meta, start=5):
        summary.cell(offset, 1, label)
        summary.cell(offset, 2, value)
    summary.column_dimensions["A"].width = 36
    summary.column_dimensions["B"].width = 110

    alias = book.create_sheet("Workbook Alias")
    headers = list(rows_out[0].keys())
    alias["A1"] = "Workbook Alias — isi keputusan reviewer di kolom kanan"
    alias["A1"].fill = PatternFill("solid", fgColor=NAVY)
    alias["A1"].font = Font(name=BODY, size=14, bold=True, color="FFFFFF")
    alias.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    for col, header in enumerate(headers, start=1):
        cell = alias.cell(3, col, header)
        cell.fill = PatternFill("solid", fgColor=TEAL)
        cell.font = Font(name=BODY, bold=True, color="FFFFFF")
        cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
    for row_idx, row in enumerate(rows_out, start=4):
        for col_idx, header in enumerate(headers, start=1):
            cell = alias.cell(row_idx, col_idx, row[header])
            cell.font = Font(name=BODY, size=9)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if header == "Proposal Confidence":
                fills = {"high": "D9EAD3", "medium": "FFF2CC", "low": "F4CCCC"}
                if row[header] in fills:
                    cell.fill = PatternFill("solid", fgColor=fills[str(row[header])])
    end_row = 3 + len(rows_out)
    table = Table(displayName="WorkbookAliasTable", ref=f"A3:{get_column_letter(len(headers))}{end_row}")
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    alias.add_table(table)
    approve_col = get_column_letter(headers.index("Reviewer Approve Alias") + 1)
    validation = DataValidation(type="list", formula1='"YES,NO,NEEDS_SPLIT"', allow_blank=True)
    alias.add_data_validation(validation)
    validation.add(f"{approve_col}4:{approve_col}{end_row}")
    widths = [6, 34, 55, 12, 12, 12, 36, 55, 10, 55, 22, 12, 28, 45, 45, 40, 14, 55, 28]
    for idx, width in enumerate(widths[: len(headers)], start=1):
        alias.column_dimensions[get_column_letter(idx)].width = width
    alias.freeze_panes = "D4"
    alias.row_dimensions[3].height = 32

    instructions = book.create_sheet("Cara Review")
    instructions["A1"] = "Langkah review workbook alias"
    instructions["A1"].font = Font(name=BODY, size=14, bold=True, color="FFFFFF")
    instructions["A1"].fill = PatternFill("solid", fgColor=NAVY)
    steps = [
        "1. Buka sheet Workbook Alias; urutkan Position Rows Affected terbesar dulu.",
        "2. Proposal Confidence=high: biasanya Approve YES (SPMT/SPTP/SPJM/SPSL HQ sudah di-prior).",
        "3. Jika Ambiguous/Needs Split terisi (induk Wilayah Terminal Petikemas): pilih NEEDS_SPLIT.",
        "4. Jika usulan salah: Approve NO, isi Reviewer Final Inventory Workbook dengan path exact inventory.",
        "5. Path valid relatif seperti: KAMUS KPI SUBHOLDING/.../Nama File.xlsx",
        "6. Setelah alias workbook dikunci, baru generate Worksheet Alias.",
        (
            "Contoh benar: .../KAMUS KPI HO SPMT/KAMUS KPI SPMT.xlsx → "
            "KAMUS KPI SUBHOLDING/KAMUS KPI SPMT/KAMUS KPI SUBHOLDING SPMT/"
            "Kamus KPI SPMT - Mapping dengan Kontrak Manajemen.xlsx"
        ),
    ]
    for offset, line in enumerate(steps, start=3):
        instructions.cell(offset, 1, line)
    instructions.column_dimensions["A"].width = 140

    out_path = OUT / "Workbook_Alias_Review_Subholding_20260806.xlsx"
    book.save(out_path)

    summary_lines = [
        "# Workbook Alias Review Summary",
        "",
        f"- Unresolved position rows: **{len(unresolved)}**",
        f"- Distinct workbook+folder keys: **{len(rows_out)}**",
        f"- Artifact: `{out_path}`",
        "",
        "| Rows | Reviewed workbook | Proposed inventory | Confidence |",
        "| ---: | --- | --- | --- |",
    ]
    for row in rows_out[:35]:
        proposed = row["Proposed Inventory Workbook"] or "(split/manual)"
        summary_lines.append(
            f"| {row['Position Rows Affected']} | `{row['Reviewed Workbook Title']}` | `{proposed}` | {row['Proposal Confidence']} |"
        )
    (OUT / "WORKBOOK_ALIAS_SUMMARY.md").write_text("\n".join(summary_lines) + "\n", encoding="utf-8")

    print(
        json.dumps(
            {
                "alias_keys": len(rows_out),
                "unresolved_rows": len(unresolved),
                "high": sum(1 for row in rows_out if row["Proposal Confidence"] == "high"),
                "medium": sum(1 for row in rows_out if row["Proposal Confidence"] == "medium"),
                "low": sum(1 for row in rows_out if row["Proposal Confidence"] == "low"),
                "xlsx": str(out_path),
            },
            ensure_ascii=False,
            indent=2,
        )
    )


if __name__ == "__main__":
    main()
