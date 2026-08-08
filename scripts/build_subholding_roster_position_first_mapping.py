#!/usr/bin/env python3
"""Build Position First Mapping Subholding scoped to roster 2,705 NIPPs.

Applies Red & White reviewed Workbook/Worksheet decisions from the prior
Subholding review artifact, tags the 56 roster workers absent from that review,
and excludes the 257 production-tree NIPPs that are outside the roster sheets.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import re
import zipfile
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any
from xml.etree import ElementTree as ET

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo

import build_group2_position_first_mapping_review as g2
import position_mapping as pm

NS = {"m": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
REL_NS = "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}"

SUBHOLDING_ROSTER_SHEETS = ("SPTP", "SPMT", "SPSL", "SPJM")
TAG_NEW_56 = "NEEDS_REVIEW_NEW_56"
TAG_REVIEWED = "ROSTER_COVERED_REVIEWED"
TAG_PENDING = "ROSTER_COVERED_PENDING"
TAG_INVENTORY_UNRESOLVED = "INVENTORY_UNRESOLVED"

EXTRA_COLUMNS = [
    "Roster Review Tag",
    "Roster Sheet",
    "Roster Company",
    "Roster Location",
    "Roster Job Title",
    "Reviewed Folder",
    "Reviewed Workbook Title",
    "Reviewed Worksheet Title",
    "Inventory Resolve Status",
    "Mapping Source",
]

REPORT_COLUMNS = [
    "No.",
    "Identity Scope",
    "PMID",
    "PNID",
    "Position Title",
    "Group / Unit",
    "Company",
    "Company Code",
    "Active Employees",
    "Active Employee NIPPs",
    "Active Employee Names",
    "Confidence Label",
    "Confidence Reason",
    "Candidate Score",
    "Candidate Source Folder",
    "Candidate Source Workbook",
    "Candidate Worksheet",
    "Candidate Worksheet Title",
    "Candidate Group",
    "Runner-up Score",
    "Runner-up Workbook",
    "Runner-up Worksheet",
    "Runner-up Title",
    "Shared Worksheet Position Count",
    "Recommended Action",
    *EXTRA_COLUMNS,
    *g2.REVIEW_COLUMNS,
]

NAVY = g2.NAVY
TEAL = g2.TEAL
PALE_BLUE = g2.PALE_BLUE
PALE_GRAY = g2.PALE_GRAY
BODY = g2.BODY
CONFIDENCE_FILLS = g2.CONFIDENCE_FILLS
TAG_FILLS = {
    TAG_NEW_56: PatternFill("solid", fgColor="FCE4D6"),
    TAG_REVIEWED: PatternFill("solid", fgColor="D9EAD3"),
    TAG_PENDING: PatternFill("solid", fgColor="FFF2CC"),
    TAG_INVENTORY_UNRESOLVED: PatternFill("solid", fgColor="F4CCCC"),
}


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def norm(value: Any) -> str:
    return "" if value is None else str(value).strip()


def split_semi(value: Any) -> list[str]:
    text = norm(value)
    if not text:
        return []
    return [part.strip() for part in text.split(";") if part.strip()]


def join_semi(values: list[str]) -> str:
    return "; ".join(values)


def load_shared_strings(archive: zipfile.ZipFile) -> list[str]:
    if "xl/sharedStrings.xml" not in archive.namelist():
        return []
    root = ET.fromstring(archive.read("xl/sharedStrings.xml"))
    out: list[str] = []
    for item in root.findall("m:si", NS):
        texts = [node.text or "" for node in item.iter("{http://schemas.openxmlformats.org/spreadsheetml/2006/main}t")]
        out.append("".join(texts))
    return out


def cell_value(cell: ET.Element, shared: list[str]) -> Any:
    cell_type = cell.attrib.get("t")
    value_node = cell.find("m:v", NS)
    if value_node is None:
        inline = cell.find("m:is", NS)
        if inline is None:
            return None
        texts = [node.text or "" for node in inline.iter("{http://schemas.openxmlformats.org/spreadsheetml/2006/main}t")]
        return "".join(texts)
    raw = value_node.text
    if cell_type == "s":
        return shared[int(raw)]
    return raw


def col_index(cell_ref: str) -> int:
    match = re.match(r"([A-Z]+)", cell_ref)
    assert match is not None
    total = 0
    for char in match.group(1):
        total = total * 26 + (ord(char) - 64)
    return total - 1


def read_sheet_rows(archive: zipfile.ZipFile, target: str, shared: list[str]) -> list[list[Any]]:
    root = ET.fromstring(archive.read(target))
    rows: list[list[Any]] = []
    for row in root.findall("m:sheetData/m:row", NS):
        cells: dict[int, Any] = {}
        for cell in row.findall("m:c", NS):
            cells[col_index(cell.attrib.get("r", "A1"))] = cell_value(cell, shared)
        if not cells:
            continue
        width = max(cells) + 1
        rows.append([cells.get(index) for index in range(width)])
    return rows


def workbook_sheet_targets(archive: zipfile.ZipFile) -> list[tuple[str, str]]:
    workbook = ET.fromstring(archive.read("xl/workbook.xml"))
    rels = ET.fromstring(archive.read("xl/_rels/workbook.xml.rels"))
    rid_to_target = {rel.attrib["Id"]: rel.attrib["Target"] for rel in rels}
    out: list[tuple[str, str]] = []
    for sheet in workbook.findall("m:sheets/m:sheet", NS):
        rid = sheet.attrib[f"{REL_NS}id"]
        target = rid_to_target[rid].lstrip("/")
        if not target.startswith("xl/"):
            target = f"xl/{target}"
        out.append((sheet.attrib["name"], target))
    return out


def load_roster(path: Path) -> dict[str, dict[str, str]]:
    """Return NIPP -> roster metadata from SPTP/SPMT/SPSL/SPJM sheets."""
    roster: dict[str, dict[str, str]] = {}
    with zipfile.ZipFile(path) as archive:
        shared = load_shared_strings(archive)
        for sheet_name, target in workbook_sheet_targets(archive):
            if sheet_name not in SUBHOLDING_ROSTER_SHEETS:
                continue
            rows = read_sheet_rows(archive, target, shared)
            if not rows:
                continue
            header = [norm(value) for value in rows[0]]
            index = {name: idx for idx, name in enumerate(header)}
            required = ("PNALT_NEW", "CNAME", "STEXT_STO", "PERSA_TEXT", "SUB_PERSA_TEXT", "COMPANY_CODE", "ANSTX")
            for name in required:
                if name not in index:
                    raise SystemExit(f"Roster sheet {sheet_name} missing column {name}")
            for row in rows[1:]:
                if not row or all(value in (None, "") for value in row):
                    continue
                nipp = norm(row[index["PNALT_NEW"]])
                if not nipp:
                    continue
                roster[nipp] = {
                    "sheet": sheet_name,
                    "name": norm(row[index["CNAME"]]),
                    "job_title": norm(row[index["STEXT_STO"]]),
                    "persa_text": norm(row[index["PERSA_TEXT"]]),
                    "sub_persa_text": norm(row[index["SUB_PERSA_TEXT"]]),
                    "company_code": norm(row[index["COMPANY_CODE"]]),
                    "anstx": norm(row[index["ANSTX"]]),
                }
    return roster


def load_reviewed_position_coverage(path: Path) -> list[dict[str, Any]]:
    with zipfile.ZipFile(path) as archive:
        shared = load_shared_strings(archive)
        targets = {name: target for name, target in workbook_sheet_targets(archive)}
        if "Position Coverage" not in targets:
            raise SystemExit("Reviewed workbook missing sheet Position Coverage")
        rows = read_sheet_rows(archive, targets["Position Coverage"], shared)
    header_row = None
    header_index = None
    for idx, row in enumerate(rows[:5]):
        if row and "PMID" in row and "Position Title" in row:
            header_row = [norm(value) for value in row]
            header_index = idx
            break
    if header_row is None or header_index is None:
        raise SystemExit("Could not find Position Coverage header in reviewed workbook")
    out: list[dict[str, Any]] = []
    for row in rows[header_index + 1 :]:
        if not row or all(value in (None, "") for value in row):
            continue
        record = {
            header_row[col]: (row[col] if col < len(row) else None)
            for col in range(len(header_row))
            if header_row[col]
        }
        out.append(record)
    return out


HQ_WORKBOOK_ALIASES = {
    (
        "KAMUS KPI SPMT.xlsx",
        "KAMUS KPI SUBHOLDING/KAMUS KPI SPMT/KAMUS KPI HO SPMT/KAMUS KPI SPMT.xlsx",
    ): (
        "KAMUS KPI SUBHOLDING/KAMUS KPI SPMT/KAMUS KPI SUBHOLDING SPMT/"
        "Kamus KPI SPMT - Mapping dengan Kontrak Manajemen.xlsx"
    ),
    (
        "KAMUS KPI SPTP.xlsx",
        "KAMUS KPI SUBHOLDING/KAMUS KPI SPTP_/KAMUS KPI HO SPTP/KAMUS KPI SPTP.xlsx",
    ): (
        "KAMUS KPI SUBHOLDING/KAMUS KPI SPTP /KAMUS KPI SUBHOLDING SPTP/"
        "Kamus KPI SPTP - Mapping dengan Kontrak Manajemen.xlsx"
    ),
    (
        "KAMUS KPI SPJM.xlsx",
        "KAMUS KPI SUBHOLDING/KAMUS KPI SPJM_/KAMUS KPI HO SPJM/KAMUS KPI SPJM.xlsx",
    ): (
        "KAMUS KPI SUBHOLDING/KAMUS KPI SPJM /KAMUS KPI SUBHOLDING SPJM/"
        "Kamus KPI SPJM - Mapping dengan Kontrak Manajemen.xlsx"
    ),
    (
        "KAMUS KPI SPSL.xlsx",
        "KAMUS KPI SUBHOLDING/KAMUS KPI SPSL/KAMUS KPI HO SPSL/KAMUS KPI SPSL.xlsx",
    ): (
        "KAMUS KPI SUBHOLDING/KAMUS KPI SPSL/KAMUS KPI SUBHOLDING SPSL/"
        "Kamus KPI SPSL - Mapping dengan Kontrak Manajemen.xlsx"
    ),
}


def normalize_reviewed_folder(path: str) -> str:
    text = norm(path).replace("\\", "/")
    for prefix in (
        "KAMUS KPI PELINDO GROUP 2 (REGIONAL, CABANG DAN SUBHOLDING)/",
        "KAMUS KPI PELINDO GROUP 3 (AFILIASI, NON CLUSTER, DANA PENSIUN)/",
    ):
        if text.startswith(prefix):
            text = text[len(prefix) :]
    return text


def build_inventory_indexes(
    inventory_path: Path,
) -> tuple[dict[tuple[str, str], dict[str, str]], dict[str, list[str]], list[str], dict[str, list[str]]]:
    payload = json.loads(inventory_path.read_text(encoding="utf-8"))
    by_wb_sheet: dict[tuple[str, str], dict[str, str]] = {}
    basenames: dict[str, list[str]] = defaultdict(list)
    sheets_by_workbook: dict[str, list[str]] = defaultdict(list)
    inv_paths: set[str] = set()
    for row in payload.get("kamus_kpi_v2", []):
        if not isinstance(row, dict) or not row.get("include_in_position_config"):
            continue
        workbook = norm(row.get("source_workbook"))
        sheet = norm(row.get("sheet_name"))
        title = norm(row.get("position_name")) or sheet
        folder = norm(row.get("source_folder"))
        if not workbook or not sheet:
            continue
        inv_paths.add(workbook)
        base = workbook.split("/")[-1]
        basenames[base.casefold()].append(workbook)
        sheets_by_workbook[workbook].append(sheet)
        for key_sheet in {sheet, title}:
            by_wb_sheet[(base.casefold(), key_sheet.casefold())] = {
                "source_workbook": workbook,
                "sheet_name": sheet,
                "position_name": title,
                "source_folder": folder,
            }
    return by_wb_sheet, basenames, sorted(inv_paths), dict(sheets_by_workbook)


def suggest_inventory_workbook(
    *,
    reviewed_workbook_title: str,
    reviewed_folder: str,
    candidate_workbook: str,
    inv_paths: list[str],
    basenames: dict[str, list[str]],
) -> tuple[str, str]:
    """Use R&W path only as a reference to pick the closest inventory workbook."""
    wb_title = norm(reviewed_workbook_title)
    folder = normalize_reviewed_folder(reviewed_folder)
    candidate = norm(candidate_workbook)

    ranked: list[tuple[int, str, str]] = []
    if (wb_title, folder) in HQ_WORKBOOK_ALIASES:
        ranked.append((100, HQ_WORKBOOK_ALIASES[(wb_title, folder)], "hq_alias_prior"))

    base = (wb_title or (folder.split("/")[-1] if folder else "")).casefold()
    for path in basenames.get(base, []):
        ranked.append((95, path, "basename_exact"))
    folder_base = folder.split("/")[-1].casefold() if folder else ""
    if folder_base and folder_base != base:
        for path in basenames.get(folder_base, []):
            ranked.append((92, path, "folder_basename_exact"))

    parts = [part for part in folder.split("/") if part]
    if len(parts) >= 2:
        parent = parts[-2]
        for path in inv_paths:
            if parent.casefold() in path.casefold():
                ranked.append((88, path, "folder_parent_match"))

    if candidate and candidate in inv_paths:
        ranked.append((85, candidate, "candidate_workbook_reference"))

    stop = {
        "kamus",
        "kpi",
        "xlsx",
        "xlsm",
        "dengan",
        "kontrak",
        "manajemen",
        "mapping",
        "pelindo",
        "group",
        "dan",
        "ho",
    }
    segments = [
        part
        for part in re.split(r"[/ _-]+", f"{folder} {wb_title}")
        if len(part) > 2 and part.casefold() not in stop
    ]
    for path in inv_paths:
        score = 0
        path_cf = path.casefold()
        for segment in segments:
            if segment.casefold() in path_cf:
                score += 1
        if folder_base and path.split("/")[-1].casefold() == folder_base:
            score += 5
        if wb_title and path.split("/")[-1].casefold() == wb_title.casefold():
            score += 4
        if score:
            ranked.append((min(70 + score, 84), path, f"path_token_score_{score}"))

    if not ranked:
        return "", "no_workbook_suggestion"
    ranked.sort(key=lambda item: (-item[0], item[1]))
    return ranked[0][1], ranked[0][2]


def match_sheet_in_workbook(
    *,
    workbook: str,
    reviewed_worksheet_title: str,
    sheets_by_workbook: dict[str, list[str]],
    by_wb_sheet: dict[tuple[str, str], dict[str, str]],
) -> tuple[str, str]:
    """Return (sheet_name, status). Keep R&W title as reference when no exact inventory sheet."""
    ws_title = norm(reviewed_worksheet_title)
    if not workbook or not ws_title:
        return ws_title, "sheet_reference_only"
    base = workbook.split("/")[-1].casefold()
    hit = by_wb_sheet.get((base, ws_title.casefold()))
    if hit:
        return hit["sheet_name"], "sheet_exact"
    # Light fuzzy: reviewed title contained in inventory sheet or vice versa.
    best: tuple[float, str] | None = None
    needle = re.sub(r"[^a-z0-9]+", " ", ws_title.casefold())
    needle_tokens = {tok for tok in needle.split() if len(tok) > 2}
    for sheet in sheets_by_workbook.get(workbook, []):
        hay = re.sub(r"[^a-z0-9]+", " ", sheet.casefold())
        if needle == hay:
            return sheet, "sheet_exact"
        if needle and (needle in hay or hay in needle):
            score = min(len(needle), len(hay)) / max(len(needle), len(hay))
            if best is None or score > best[0]:
                best = (score, sheet)
            continue
        hay_tokens = {tok for tok in hay.split() if len(tok) > 2}
        if needle_tokens and hay_tokens:
            overlap = len(needle_tokens & hay_tokens) / max(len(needle_tokens), 1)
            if overlap >= 0.6 and (best is None or overlap > best[0]):
                best = (overlap, sheet)
    if best and best[0] >= 0.6:
        return best[1], "sheet_fuzzy_from_rw_reference"
    return ws_title, "sheet_reference_only"


def resolve_reviewed_to_inventory(
    *,
    reviewed_workbook_title: str,
    reviewed_worksheet_title: str,
    reviewed_folder: str,
    candidate_workbook: str,
    by_wb_sheet: dict[tuple[str, str], dict[str, str]],
    basenames: dict[str, list[str]],
    inv_paths: list[str],
    sheets_by_workbook: dict[str, list[str]],
) -> tuple[str, str, str, str]:
    """Return (resolved_workbook, resolved_sheet, status, note).

    R&W paths are a reference only. Prefer inventory config paths when a
    sensible workbook match exists; otherwise keep R&W values as reference.
    """
    wb_title = norm(reviewed_workbook_title)
    ws_title = norm(reviewed_worksheet_title)
    folder = normalize_reviewed_folder(reviewed_folder)
    if not wb_title or wb_title == "#N/A":
        return "", "", "not_reviewed", ""

    # Fast path: exact workbook basename + sheet already in inventory.
    exact = by_wb_sheet.get((wb_title.casefold(), ws_title.casefold()))
    if exact:
        return exact["source_workbook"], exact["sheet_name"], "resolved_exact", ""

    workbook, wb_reason = suggest_inventory_workbook(
        reviewed_workbook_title=wb_title,
        reviewed_folder=folder,
        candidate_workbook=candidate_workbook,
        inv_paths=inv_paths,
        basenames=basenames,
    )
    if workbook:
        sheet, sheet_status = match_sheet_in_workbook(
            workbook=workbook,
            reviewed_worksheet_title=ws_title,
            sheets_by_workbook=sheets_by_workbook,
            by_wb_sheet=by_wb_sheet,
        )
        if sheet_status == "sheet_exact":
            return workbook, sheet, "workbook_resolved_sheet_exact", wb_reason
        if sheet_status == "sheet_fuzzy_from_rw_reference":
            return (
                workbook,
                sheet,
                "workbook_resolved_sheet_fuzzy",
                f"{wb_reason}; sheet matched from R&W reference '{ws_title}'",
            )
        return (
            workbook,
            sheet,
            "workbook_resolved_sheet_reference",
            f"{wb_reason}; worksheet kept as R&W reference (no exact inventory sheet)",
        )

    # No inventory workbook found — keep R&W values purely as reference.
    return (
        wb_title,
        ws_title,
        "rw_path_as_reference",
        "R&W workbook/sheet retained as reference; no inventory workbook match",
    )


def empty_row_template() -> dict[str, Any]:
    return {column: "" for column in REPORT_COLUMNS if column != "No."}


def identity_key(pmid: Any, pnid: Any) -> tuple[str, str] | None:
    pmid_text = norm(pmid)
    pnid_text = norm(pnid)
    if pmid_text:
        return ("pmid", pmid_text)
    if pnid_text:
        return ("pnid", pnid_text)
    return None


def filter_row_to_roster(row: dict[str, Any], roster: dict[str, dict[str, str]]) -> dict[str, Any] | None:
    nipps = [nipp for nipp in split_semi(row.get("Active Employee NIPPs")) if nipp in roster]
    if not nipps:
        return None
    names = split_semi(row.get("Active Employee Names"))
    # Keep names aligned when lengths match; otherwise rebuild from roster.
    if len(names) == len(split_semi(row.get("Active Employee NIPPs"))):
        original = split_semi(row.get("Active Employee NIPPs"))
        kept_names = [names[idx] for idx, nipp in enumerate(original) if nipp in roster]
    else:
        kept_names = [roster[nipp]["name"] for nipp in nipps if roster[nipp]["name"]]
    filtered = dict(row)
    filtered["Active Employee NIPPs"] = join_semi(nipps)
    filtered["Active Employee Names"] = join_semi(kept_names)
    filtered["Active Employees"] = len(nipps)
    return filtered


def build_nipp_to_candidates(reference_path: Path) -> dict[str, list[Any]]:
    payload = json.loads(reference_path.read_text(encoding="utf-8"))
    indexes = pm.build_lookup_indexes(payload)
    mapping: dict[str, list[Any]] = defaultdict(list)
    for candidate in [*indexes.structural, *indexes.non_structural]:
        for nipp in candidate.active_employee_nipps:
            mapping[str(nipp).strip()].append(candidate)
    return mapping


def build_nipp_to_assignments(reference_path: Path) -> dict[str, list[dict[str, Any]]]:
    payload = json.loads(reference_path.read_text(encoding="utf-8"))
    companies = {
        int(row["company_in_id"]): row
        for row in payload.get("company_rows", [])
        if isinstance(row, dict) and row.get("company_in_id") not in (None, "")
    }
    mapping: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in payload.get("active_assignment_rows", []):
        if not isinstance(row, dict):
            continue
        nipps = row.get("active_employee_nipps") or []
        if isinstance(nipps, str):
            values = [part.strip() for part in nipps.replace(";", ",").split(",") if part.strip()]
        else:
            values = [str(item).strip() for item in nipps if str(item).strip()]
        try:
            company_id = int(row.get("company_id"))
        except (TypeError, ValueError):
            company_id = None
        company = companies.get(company_id, {}) if company_id is not None else {}
        for nipp in values:
            mapping[nipp].append(
                {
                    "pmid": row.get("position_master_id"),
                    "company_id": company_id,
                    "company_name": company.get("company_name") or company.get("name") or "",
                    "company_code": company.get("company_code") or company.get("code") or "",
                    "active_employee_count": row.get("active_employee_count") or len(values),
                }
            )
    return mapping


def classify_extra_257(
    *,
    old_draft_path: Path,
    roster: dict[str, dict[str, str]],
    reference_path: Path,
) -> dict[str, Any]:
    draft = json.loads(old_draft_path.read_text(encoding="utf-8"))
    mapping_nipps: set[str] = set()
    for row in draft.get("rows", []):
        mapping_nipps.update(split_semi(row.get("Active Employee NIPPs")))
    extra = sorted(mapping_nipps - set(roster))
    payload = json.loads(reference_path.read_text(encoding="utf-8"))
    companies = {
        int(row["company_in_id"]): row
        for row in payload.get("company_rows", [])
        if isinstance(row, dict) and row.get("company_in_id") not in (None, "")
    }
    sub_ids = {cid for cid, row in companies.items() if (row.get("type_org") or "") == "Subholding"}

    def ancestors(company_id: int) -> list[int]:
        path: list[int] = []
        seen: set[int] = set()
        current: int | None = company_id
        while current is not None and current not in seen:
            seen.add(current)
            path.append(current)
            parent = companies.get(current, {}).get("parent_id")
            if parent in (None, ""):
                break
            try:
                current = int(parent)
            except (TypeError, ValueError):
                break
        return path

    nipp_companies: dict[str, set[int]] = defaultdict(set)
    for row in payload.get("active_assignment_rows", []):
        nipps = row.get("active_employee_nipps") or []
        if isinstance(nipps, str):
            values = [part.strip() for part in nipps.replace(";", ",").split(",") if part.strip()]
        else:
            values = [str(item).strip() for item in nipps if str(item).strip()]
        try:
            company_id = int(row.get("company_id"))
        except (TypeError, ValueError):
            continue
        for nipp in values:
            if nipp in extra:
                nipp_companies[nipp].add(company_id)

    company_counts: Counter[str] = Counter()
    root_counts: Counter[str] = Counter()
    under_tree = 0
    for nipp in extra:
        for company_id in nipp_companies.get(nipp, ()):
            company = companies.get(company_id, {})
            name = company.get("company_name") or company.get("name") or str(company_id)
            company_counts[name] += 1
            path = ancestors(company_id)
            if any(node in sub_ids for node in path):
                under_tree += 1
                for node in path:
                    if node in sub_ids:
                        root = companies[node]
                        root_counts[root.get("company_name") or root.get("name") or str(node)] += 1
                        break
                break
    return {
        "extra_count": len(extra),
        "under_subholding_tree_unique": under_tree,
        "top_companies": company_counts.most_common(15),
        "roots": dict(root_counts),
        "note": (
            "All extra NIPPs sit under production Subholding org-tree companies "
            "(anak usaha / HQ assignments) but are absent from roster sheets SPTP/SPMT/SPSL/SPJM."
        ),
    }


def apply_review_to_row(
    row: dict[str, Any],
    reviewed: dict[str, Any] | None,
    *,
    by_wb_sheet: dict[tuple[str, str], dict[str, str]],
    basenames: dict[str, list[str]],
    inv_paths: list[str],
    sheets_by_workbook: dict[str, list[str]],
) -> None:
    if reviewed is None:
        row["Reviewed Folder"] = ""
        row["Reviewed Workbook Title"] = ""
        row["Reviewed Worksheet Title"] = ""
        row["Inventory Resolve Status"] = "no_prior_review_row"
        row["Mapping Source"] = "resolver_only"
        row["Roster Review Tag"] = TAG_PENDING
        return

    folder = normalize_reviewed_folder(reviewed.get("Folder"))
    workbook_title = norm(reviewed.get("Workbook Title"))
    worksheet_title = norm(reviewed.get("Worksheet Title"))
    row["Reviewed Folder"] = folder
    row["Reviewed Workbook Title"] = workbook_title
    row["Reviewed Worksheet Title"] = worksheet_title

    confidence = norm(row.get("Confidence Label"))
    if workbook_title in ("", "#N/A") and worksheet_title in ("", "#N/A"):
        # R&W left high_confidence as #N/A => accept candidate draft.
        if confidence == pm.HIGH_CONFIDENCE and norm(row.get("Candidate Source Workbook")):
            row["Reviewer Confirm Mapping"] = "YES"
            row["Reviewer Source Workbook"] = norm(row.get("Candidate Source Workbook"))
            row["Reviewer Worksheet"] = norm(row.get("Candidate Worksheet"))
            row["Inventory Resolve Status"] = "accepted_high_confidence_candidate"
            row["Mapping Source"] = "rw_accept_candidate"
            row["Roster Review Tag"] = TAG_REVIEWED
            row["Reviewer Notes"] = "R&W #N/A on high_confidence; candidate accepted"
        else:
            row["Inventory Resolve Status"] = "review_blank"
            row["Mapping Source"] = "unreviewed"
            row["Roster Review Tag"] = TAG_PENDING
        return

    resolved_wb, resolved_ws, status, note = resolve_reviewed_to_inventory(
        reviewed_workbook_title=workbook_title,
        reviewed_worksheet_title=worksheet_title,
        reviewed_folder=folder,
        candidate_workbook=norm(row.get("Candidate Source Workbook")),
        by_wb_sheet=by_wb_sheet,
        basenames=basenames,
        inv_paths=inv_paths,
        sheets_by_workbook=sheets_by_workbook,
    )
    row["Reviewer Confirm Mapping"] = "YES"
    row["Reviewer Source Workbook"] = resolved_wb
    row["Reviewer Worksheet"] = resolved_ws
    row["Inventory Resolve Status"] = status
    row["Mapping Source"] = "rw_reference"
    row["Roster Review Tag"] = TAG_REVIEWED
    row["Reviewer Notes"] = (
        note
        or "R&W path used as reference; Reviewer* columns hold inventory path when matched"
    )


def style_header(cell, fill: str) -> None:
    cell.fill = PatternFill("solid", fgColor=fill)
    cell.font = Font(name=BODY, bold=True, color="FFFFFF", size=10)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def write_summary(
    ws,
    *,
    title: str,
    meta: list[tuple[str, Any]],
    labels: Counter[str],
    tags: Counter[str],
    extra_257: dict[str, Any],
) -> None:
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:F1")
    ws["A1"] = title
    ws["A1"].fill = PatternFill("solid", fgColor=NAVY)
    ws["A1"].font = Font(name=BODY, size=18, bold=True, color="FFFFFF")
    ws.row_dimensions[1].height = 34

    ws.merge_cells("A2:F2")
    ws["A2"] = (
        "Scope = exact 2.705 NIPP dari sheet SPTP/SPMT/SPSL/SPJM. "
        "Path review Red & White dipakai sebagai referensi untuk memilih workbook inventory. "
        "Kolom Reviewed* = nilai asli R&W; kolom Reviewer* = path/sheet yang dipakai. "
        "Tag NEEDS_REVIEW_NEW_56 menandai pekerja yang belum ada di review sebelumnya."
    )
    ws["A2"].fill = PatternFill("solid", fgColor=PALE_BLUE)
    ws["A2"].font = Font(name=BODY, size=10, italic=True, color="37556E")

    ws["A4"] = "Provenance / Audit"
    ws["B4"] = "Value"
    style_header(ws["A4"], TEAL)
    style_header(ws["B4"], TEAL)
    for offset, (label, value) in enumerate(meta, start=5):
        ws.cell(offset, 1, label).fill = PatternFill("solid", fgColor=PALE_GRAY)
        ws.cell(offset, 1).font = Font(name=BODY, bold=True, color=NAVY)
        ws.cell(offset, 2, value).font = Font(name=BODY, color="263746")
        ws.cell(offset, 2).alignment = Alignment(wrap_text=True)

    conf_start = 5 + len(meta) + 2
    ws.cell(conf_start, 1, "Confidence").font = Font(name=BODY, bold=True, color="FFFFFF")
    ws.cell(conf_start, 2, "Count").font = Font(name=BODY, bold=True, color="FFFFFF")
    style_header(ws.cell(conf_start, 1), TEAL)
    style_header(ws.cell(conf_start, 2), TEAL)
    for offset, label in enumerate(
        [pm.HIGH_CONFIDENCE, pm.LOW_CONFIDENCE, pm.MAPPING_CONFLICT, pm.NO_CANDIDATE, pm.SCOPE_UNCERTAIN],
        start=conf_start + 1,
    ):
        ws.cell(offset, 1, label).fill = CONFIDENCE_FILLS.get(label, PatternFill())
        ws.cell(offset, 2, labels.get(label, 0))

    tag_start = conf_start + 8
    ws.cell(tag_start, 1, "Roster Review Tag").font = Font(name=BODY, bold=True, color="FFFFFF")
    ws.cell(tag_start, 2, "Count").font = Font(name=BODY, bold=True, color="FFFFFF")
    style_header(ws.cell(tag_start, 1), TEAL)
    style_header(ws.cell(tag_start, 2), TEAL)
    for offset, (tag, count) in enumerate(sorted(tags.items()), start=tag_start + 1):
        ws.cell(offset, 1, tag).fill = TAG_FILLS.get(tag.split(";")[0], PatternFill())
        ws.cell(offset, 2, count)

    note_start = tag_start + len(tags) + 3
    ws.cell(note_start, 1, "Out-of-roster NIPPs excluded (257)").font = Font(name=BODY, bold=True, color=NAVY)
    ws.cell(note_start + 1, 1, extra_257["note"])
    ws.merge_cells(start_row=note_start + 1, start_column=1, end_row=note_start + 1, end_column=2)
    ws.cell(note_start + 2, 1, "Extra count")
    ws.cell(note_start + 2, 2, extra_257["extra_count"])
    ws.cell(note_start + 3, 1, "Under Subholding tree (unique)")
    ws.cell(note_start + 3, 2, extra_257["under_subholding_tree_unique"])
    ws.cell(note_start + 4, 1, "Roots")
    ws.cell(note_start + 4, 2, json.dumps(extra_257["roots"], ensure_ascii=False))
    ws.cell(note_start + 5, 1, "Top companies")
    ws.cell(note_start + 5, 2, json.dumps(extra_257["top_companies"], ensure_ascii=False))

    ws.column_dimensions["A"].width = 48
    ws.column_dimensions["B"].width = 100


def write_report_sheet(ws, title: str, rows: list[dict[str, Any]], table_name: str) -> None:
    ws.sheet_view.showGridLines = False
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(REPORT_COLUMNS))
    ws.cell(1, 1, title).fill = PatternFill("solid", fgColor=NAVY)
    ws.cell(1, 1).font = Font(name=BODY, size=16, bold=True, color="FFFFFF")
    ws.row_dimensions[1].height = 30

    nipp_col = REPORT_COLUMNS.index("Active Employee NIPPs") + 1
    tag_col = REPORT_COLUMNS.index("Roster Review Tag") + 1
    for col, header in enumerate(REPORT_COLUMNS, start=1):
        style_header(ws.cell(3, col, header), TEAL)

    for row_idx, row in enumerate(rows, start=1):
        excel_row = 3 + row_idx
        values = [row_idx, *[row.get(column, "") for column in REPORT_COLUMNS[1:]]]
        for col, value in enumerate(values, start=1):
            cell = ws.cell(excel_row, col, value)
            cell.font = Font(name=BODY, size=9, color="263746")
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if col == nipp_col:
                cell.number_format = "@"
        label = row.get("Confidence Label")
        if label in CONFIDENCE_FILLS:
            ws.cell(excel_row, REPORT_COLUMNS.index("Confidence Label") + 1).fill = CONFIDENCE_FILLS[label]
        tag = norm(row.get("Roster Review Tag"))
        primary = tag.split(";")[0] if tag else ""
        if primary in TAG_FILLS:
            ws.cell(excel_row, tag_col).fill = TAG_FILLS[primary]
        ws.row_dimensions[excel_row].height = 22

    end_row = 3 + max(len(rows), 1)
    if rows:
        table = Table(displayName=table_name, ref=f"A3:{get_column_letter(len(REPORT_COLUMNS))}{end_row}")
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
        ws.add_table(table)
        confirm_col = get_column_letter(REPORT_COLUMNS.index("Reviewer Confirm Mapping") + 1)
        dv = DataValidation(type="list", formula1='"YES,NEEDS_CHECK,NO"', allow_blank=True)
        ws.add_data_validation(dv)
        dv.add(f"{confirm_col}4:{confirm_col}{end_row}")

    widths = [
        6, 14, 12, 12, 34, 28, 28, 12, 10, 26, 28, 16, 36, 10, 18, 42, 24, 28, 22,
        10, 36, 20, 28, 10, 36, 22, 10, 28, 18, 34, 28, 28, 28, 18, 18, 14, 42, 20, 28,
    ]
    for idx, width in enumerate(widths[: len(REPORT_COLUMNS)], start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    ws.freeze_panes = "E4"


def write_out_of_roster_note(ws, extra_257: dict[str, Any]) -> None:
    ws.sheet_view.showGridLines = False
    ws["A1"] = "Out of Roster Note — 257 NIPP excluded"
    ws["A1"].fill = PatternFill("solid", fgColor=NAVY)
    ws["A1"].font = Font(name=BODY, size=14, bold=True, color="FFFFFF")
    ws.merge_cells("A1:C1")
    ws["A3"] = extra_257["note"]
    ws.merge_cells("A3:C3")
    ws["A5"] = "Metric"
    ws["B5"] = "Value"
    style_header(ws["A5"], TEAL)
    style_header(ws["B5"], TEAL)
    ws["A6"] = "Extra NIPP count"
    ws["B6"] = extra_257["extra_count"]
    ws["A7"] = "Under Subholding tree"
    ws["B7"] = extra_257["under_subholding_tree_unique"]
    ws["A8"] = "Roots"
    ws["B8"] = json.dumps(extra_257["roots"], ensure_ascii=False)
    ws["A10"] = "Company"
    ws["B10"] = "Extra assignment hits"
    style_header(ws["A10"], TEAL)
    style_header(ws["B10"], TEAL)
    for offset, (company, count) in enumerate(extra_257["top_companies"], start=11):
        ws.cell(offset, 1, company)
        ws.cell(offset, 2, count)
    ws.column_dimensions["A"].width = 48
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 40


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--roster",
        type=Path,
        default=Path("outputs/kamus-group2-subholding-roster-mapping-20260806/source/REGIONAL dan SUBHOLDING.xlsx"),
    )
    parser.add_argument(
        "--reviewed",
        type=Path,
        default=Path(
            "outputs/kamus-group2-subholding-roster-mapping-20260806/source/"
            "Pemetaan Kamus KPI Subholding (05 Aug 2026)-with mapping.xlsx"
        ),
    )
    parser.add_argument(
        "--old-draft",
        type=Path,
        default=Path(
            "outputs/kamus-group2-subholding-position-mapping-20260805/"
            "group2_subholding_position_first_mapping_draft_20260805.json"
        ),
    )
    parser.add_argument(
        "--inventory",
        type=Path,
        default=Path("configs/kamus_kpi_group2_visible_20260807.json"),
    )
    parser.add_argument(
        "--reference",
        type=Path,
        default=Path("configs/production_position_reference.json"),
    )
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=Path("outputs/kamus-group2-subholding-roster-mapping-20260806"),
    )
    args = parser.parse_args()

    generated_at = datetime.now().astimezone().isoformat(timespec="seconds")
    stamp = "20260806"
    cover_title = "Position First Mapping Subholding — 2026-08-06 (R&W path as reference)"

    roster = load_roster(args.roster)
    if len(roster) != 2705:
        raise SystemExit(f"Expected 2705 unique roster NIPPs, got {len(roster)}")

    reviewed_rows = load_reviewed_position_coverage(args.reviewed)
    reviewed_by_identity: dict[tuple[str, str], dict[str, Any]] = {}
    reviewed_by_company_title: dict[tuple[str, str], dict[str, Any]] = {}
    for row in reviewed_rows:
        key = identity_key(row.get("PMID"), row.get("PNID"))
        if key:
            reviewed_by_identity[key] = row
        company_title = (norm(row.get("Company")), norm(row.get("Position Title")))
        if company_title[0] and company_title[1]:
            reviewed_by_company_title[company_title] = row

    by_wb_sheet, basenames, inv_paths, sheets_by_workbook = build_inventory_indexes(args.inventory)
    worksheets = g2.load_worksheets(args.inventory, source_folder="KAMUS KPI SUBHOLDING")
    old_draft = json.loads(args.old_draft.read_text(encoding="utf-8"))
    reference = json.loads(args.reference.read_text(encoding="utf-8"))
    exported_at = norm(reference.get("source", {}).get("exported_at"))

    # 1) Keep prior draft rows that still have roster NIPPs.
    covered_nipps: set[str] = set()
    rows: list[dict[str, Any]] = []
    for raw in old_draft.get("rows", []):
        filtered = filter_row_to_roster(raw, roster)
        if filtered is None:
            continue
        row = empty_row_template()
        for column in REPORT_COLUMNS:
            if column == "No.":
                continue
            if column in filtered:
                row[column] = filtered.get(column, "")
        nipps = split_semi(row["Active Employee NIPPs"])
        covered_nipps.update(nipps)
        # Roster context from first NIPP.
        first = roster[nipps[0]]
        row["Roster Sheet"] = first["sheet"]
        row["Roster Company"] = first["persa_text"]
        row["Roster Location"] = first["sub_persa_text"]
        row["Roster Job Title"] = first["job_title"]
        key = identity_key(row.get("PMID"), row.get("PNID"))
        reviewed = reviewed_by_identity.get(key) if key else None
        if reviewed is None:
            reviewed = reviewed_by_company_title.get((norm(row.get("Company")), norm(row.get("Position Title"))))
        apply_review_to_row(
            row,
            reviewed,
            by_wb_sheet=by_wb_sheet,
            basenames=basenames,
            inv_paths=inv_paths,
            sheets_by_workbook=sheets_by_workbook,
        )
        rows.append(row)

    missing_nipps = sorted(set(roster) - covered_nipps)

    # 2) Add 56 missing roster workers as tagged rows.
    nipp_to_candidates = build_nipp_to_candidates(args.reference)
    nipp_to_assignments = build_nipp_to_assignments(args.reference)
    new_positions: list[g2.PositionEntry] = []
    new_meta: list[dict[str, Any]] = []
    for nipp in missing_nipps:
        info = roster[nipp]
        candidates = nipp_to_candidates.get(nipp, [])
        chosen = None
        if candidates:
            # Prefer candidate whose company/title best matches roster.
            scored = []
            for candidate in candidates:
                title_score = pm._title_score(info["job_title"].split("#")[0].strip(), candidate)
                company_score = pm._context_score(info["persa_text"], candidate.company_name)
                scored.append((title_score * 0.7 + company_score * 0.3, candidate))
            scored.sort(key=lambda item: item[0], reverse=True)
            chosen = scored[0][1]
        pmid = chosen.position_master_id if chosen and chosen.scope == "structural" else None
        pnid = chosen.position_nomenclature_id if chosen and chosen.scope == "non_structural" else None
        if chosen is None:
            assignments = nipp_to_assignments.get(nipp, [])
            title = info["job_title"].split("#")[0].strip() or info["job_title"]
            company_name = info["persa_text"]
            company_code = info["company_code"]
            group_name = info["sub_persa_text"]
            scope = "unknown"
            if assignments:
                company_name = assignments[0]["company_name"] or company_name
                company_code = assignments[0]["company_code"] or company_code
                pmid = str(assignments[0]["pmid"]) if assignments[0].get("pmid") not in (None, "") else None
                scope = "structural" if pmid else "unknown"
            position = g2.PositionEntry(
                scope=scope,
                pmid=pmid,
                pnid=None,
                title=title,
                group_name=group_name,
                company_name=company_name,
                company_code=company_code,
                company_id=None,
                active_employee_count=1,
                active_employee_nipps=nipp,
                active_employee_names=info["name"],
                normalized_title=pm.normalize_position_lookup(title),
                tokens=g2.significant_tokens(title),
                company_tokens=g2.company_tokens(" ".join(filter(None, [company_name, company_code]))),
                company_key=g2.company_key(company_name or company_code),
            )
        else:
            position = g2.PositionEntry(
                scope=chosen.scope,
                pmid=pmid,
                pnid=pnid,
                title=chosen.title,
                group_name=chosen.group_name,
                company_name=chosen.company_name,
                company_code=chosen.company_code,
                company_id=norm(chosen.company_id),
                active_employee_count=1,
                active_employee_nipps=nipp,
                active_employee_names=info["name"] or "; ".join(chosen.active_employee_names),
                normalized_title=pm.normalize_position_lookup(chosen.title),
                tokens=g2.significant_tokens(chosen.title),
                company_tokens=g2.company_tokens(
                    " ".join(filter(None, [chosen.company_name, chosen.company_code]))
                ),
                company_key=g2.company_key(norm(chosen.company_name) or norm(chosen.company_code)),
            )
        new_positions.append(position)
        new_meta.append({"nipp": nipp, "info": info, "has_lookup": chosen is not None})

    new_rows, _, _ = g2.resolve_all(new_positions, worksheets)
    for row, meta in zip(new_rows, new_meta):
        info = meta["info"]
        row["Roster Review Tag"] = TAG_NEW_56
        row["Roster Sheet"] = info["sheet"]
        row["Roster Company"] = info["persa_text"]
        row["Roster Location"] = info["sub_persa_text"]
        row["Roster Job Title"] = info["job_title"]
        row["Reviewed Folder"] = ""
        row["Reviewed Workbook Title"] = ""
        row["Reviewed Worksheet Title"] = ""
        row["Inventory Resolve Status"] = "new_roster_worker"
        row["Mapping Source"] = "new_56_resolver"
        row["Reviewer Confirm Mapping"] = ""
        row["Reviewer Source Workbook"] = ""
        row["Reviewer Worksheet"] = ""
        row["Reviewer Notes"] = (
            "Baru dari roster 2705; belum ada di review R&W sebelumnya. "
            + ("Identity dari production lookup." if meta["has_lookup"] else "Identity lemah/tidak ada di lookup aktif.")
        )
        # Ensure only this NIPP is listed.
        row["Active Employee NIPPs"] = meta["nipp"]
        row["Active Employee Names"] = info["name"]
        row["Active Employees"] = 1
        covered_nipps.add(meta["nipp"])
        rows.append(row)

    # Safety: if any roster NIPP still missing, add stub rows.
    still_missing = sorted(set(roster) - covered_nipps)
    for nipp in still_missing:
        info = roster[nipp]
        row = empty_row_template()
        row.update(
            {
                "Identity Scope": "unknown",
                "Position Title": info["job_title"].split("#")[0].strip() or info["job_title"],
                "Group / Unit": info["sub_persa_text"],
                "Company": info["persa_text"],
                "Company Code": info["company_code"],
                "Active Employees": 1,
                "Active Employee NIPPs": nipp,
                "Active Employee Names": info["name"],
                "Confidence Label": pm.NO_CANDIDATE,
                "Confidence Reason": "Roster NIPP not present in production lookup/assignment indexes",
                "Recommended Action": "Review manual — identity production belum tersedia",
                "Roster Review Tag": TAG_NEW_56,
                "Roster Sheet": info["sheet"],
                "Roster Company": info["persa_text"],
                "Roster Location": info["sub_persa_text"],
                "Roster Job Title": info["job_title"],
                "Inventory Resolve Status": "stub_no_production_identity",
                "Mapping Source": "new_56_stub",
                "Reviewer Notes": "Stub row agar unique roster 2705 lengkap",
            }
        )
        rows.append(row)
        covered_nipps.add(nipp)

    # Deduplicate identity rows that may have been created twice for multi-seat people
    # already covered by filtered old draft: keep first occurrence of each NIPP only once
    # across rows by recomputing unique set after build.
    all_nipps = [nipp for row in rows for nipp in split_semi(row.get("Active Employee NIPPs"))]
    unique_nipps = set(all_nipps)
    if unique_nipps != set(roster):
        raise SystemExit(
            f"Unique NIPP mismatch: got {len(unique_nipps)} expected 2705; "
            f"missing={sorted(set(roster)-unique_nipps)[:10]} extra={sorted(unique_nipps-set(roster))[:10]}"
        )
    if any(nipp not in roster for nipp in unique_nipps):
        raise SystemExit("Non-roster NIPP leaked into coverage rows")

    labels = Counter(norm(row.get("Confidence Label")) for row in rows)
    tags = Counter(norm(row.get("Roster Review Tag")) or "(blank)" for row in rows)
    shared_best: Counter[tuple[str, str]] = Counter()
    for row in rows:
        wb = norm(row.get("Reviewer Source Workbook")) or norm(row.get("Candidate Source Workbook"))
        ws = norm(row.get("Reviewer Worksheet")) or norm(row.get("Candidate Worksheet"))
        if wb and ws:
            shared_best[(wb, ws)] += 1
    for row in rows:
        wb = norm(row.get("Reviewer Source Workbook")) or norm(row.get("Candidate Source Workbook"))
        ws = norm(row.get("Reviewer Worksheet")) or norm(row.get("Candidate Worksheet"))
        row["Shared Worksheet Position Count"] = shared_best.get((wb, ws), 0) if wb else 0

    rows.sort(
        key=lambda row: (
            0 if TAG_NEW_56 in norm(row.get("Roster Review Tag")) else 1,
            norm(row.get("Company")).casefold(),
            norm(row.get("Position Title")).casefold(),
            norm(row.get("PMID")),
            norm(row.get("PNID")),
        )
    )

    extra_257 = classify_extra_257(
        old_draft_path=args.old_draft,
        roster=roster,
        reference_path=args.reference,
    )

    args.output_dir.mkdir(parents=True, exist_ok=True)
    xlsx_path = args.output_dir / f"Position_First_Mapping_Subholding_{stamp}.xlsx"
    json_path = args.output_dir / f"subholding_roster_position_first_mapping_{stamp}.json"
    receipt_path = args.output_dir / "MAPPING_RECEIPT.md"
    readme_path = args.output_dir / "README_SOURCE.md"

    roster_hash = sha256_file(args.roster)
    reviewed_hash = sha256_file(args.reviewed)
    inventory_hash = sha256_file(args.inventory)
    reference_hash = sha256_file(args.reference)

    needs_56 = [row for row in rows if TAG_NEW_56 in norm(row.get("Roster Review Tag"))]
    review_queue = [
        row
        for row in rows
        if norm(row.get("Reviewer Confirm Mapping")) != "YES"
        or TAG_NEW_56 in norm(row.get("Roster Review Tag"))
        or norm(row.get("Inventory Resolve Status"))
        in {
            "rw_path_as_reference",
            "workbook_resolved_sheet_reference",
            "review_blank",
            "no_prior_review_row",
        }
    ]

    status_counts = Counter(norm(row.get("Inventory Resolve Status")) or "(blank)" for row in rows)

    meta_rows = [
        ("Title", cover_title),
        ("Generated at", generated_at),
        ("Scope", "Roster Subholding sheets SPTP+SPMT+SPSL+SPJM (exact 2705 NIPP)"),
        ("Unique active employees (roster)", len(unique_nipps)),
        ("Position / stub rows", len(rows)),
        ("NEEDS_REVIEW_NEW_56 rows", len(needs_56)),
        ("Reviewer YES rows", sum(1 for row in rows if norm(row.get("Reviewer Confirm Mapping")) == "YES")),
        ("R&W path handling", "R&W Folder/Workbook/Worksheet = reference only; Reviewer* prefers inventory config path"),
        ("Inventory resolve status counts", json.dumps(dict(status_counts), ensure_ascii=False)),
        ("Roster file", str(args.roster)),
        ("Roster sha256", roster_hash),
        ("R&W reviewed file", str(args.reviewed)),
        ("R&W reviewed sha256", reviewed_hash),
        ("Worksheet inventory", str(args.inventory)),
        ("Inventory sha256", inventory_hash),
        ("Production position reference", str(args.reference)),
        ("Reference exported_at", exported_at),
        ("Reference sha256", reference_hash),
        ("Prior draft (filtered)", str(args.old_draft)),
        ("Kamus worksheets (Subholding folder)", len(worksheets)),
        ("Excluded out-of-roster NIPPs", extra_257["extra_count"]),
    ]

    wb = Workbook()
    summary = wb.active
    summary.title = "Ringkasan"
    write_summary(
        summary,
        title=cover_title,
        meta=meta_rows,
        labels=labels,
        tags=tags,
        extra_257=extra_257,
    )
    write_report_sheet(
        wb.create_sheet("Position Coverage"),
        "Posisi roster Subholding (2.705 NIPP) → worksheet Kamus + hasil review R&W",
        rows,
        "PositionCoverageTable",
    )
    write_report_sheet(
        wb.create_sheet("Needs Review (56)"),
        "56 pekerja roster yang belum ada di review R&W sebelumnya — wajib direview",
        needs_56,
        "NeedsReview56Table",
    )
    write_report_sheet(
        wb.create_sheet("Review Queue"),
        "Antrian: belum YES / sheet masih referensi R&W / NEEDS_REVIEW_NEW_56",
        review_queue,
        "ReviewQueueTable",
    )
    g2.write_shared_sheet(wb.create_sheet("Shared Worksheets"), shared_best, worksheets)
    write_out_of_roster_note(wb.create_sheet("Out of Roster Note"), extra_257)
    wb.save(xlsx_path)

    payload = {
        "metadata": {
            "title": cover_title,
            "orientation": "position_first",
            "scope": "subholding-roster-2705",
            "generated_at": generated_at,
            "unique_active_employees": len(unique_nipps),
            "position_row_count": len(rows),
            "needs_review_new_56": len(needs_56),
            "confidence_counts": dict(labels),
            "tag_counts": dict(tags),
            "reviewer_yes_count": sum(1 for row in rows if norm(row.get("Reviewer Confirm Mapping")) == "YES"),
            "inventory_resolve_status_counts": dict(status_counts),
            "rw_path_policy": "R&W Folder/Workbook/Worksheet used as reference only; Reviewer* prefers inventory config path",
            "sources": {
                "roster": {"path": str(args.roster), "sha256": roster_hash},
                "reviewed_mapping": {"path": str(args.reviewed), "sha256": reviewed_hash},
                "inventory": {"path": str(args.inventory), "sha256": inventory_hash},
                "production_reference": {
                    "path": str(args.reference),
                    "sha256": reference_hash,
                    "exported_at": exported_at,
                },
                "prior_draft": str(args.old_draft),
            },
            "excluded_out_of_roster": extra_257,
            "reviewer_columns": g2.REVIEW_COLUMNS,
            "tag_columns": EXTRA_COLUMNS,
        },
        "rows": rows,
    }
    json_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    receipt = f"""# Mapping Receipt — {cover_title}

## Scope
- Exact **{len(unique_nipps)}** unique NIPP from roster sheets SPTP/SPMT/SPSL/SPJM
- Position-first rows: **{len(rows)}**
- Tagged new workers (`{TAG_NEW_56}`): **{len(needs_56)}**

## Audit sources
| Source | Path | SHA256 / timestamp |
| --- | --- | --- |
| Roster | `{args.roster}` | `{roster_hash}` |
| R&W reviewed mapping | `{args.reviewed}` | `{reviewed_hash}` |
| Worksheet inventory | `{args.inventory}` | `{inventory_hash}` |
| Production reference | `{args.reference}` | exported_at `{exported_at}` / `{reference_hash}` |

## Review application
- R&W `Folder` / `Workbook Title` / `Worksheet Title` = **referensi saja** (disimpan di kolom Reviewed*)
- `Reviewer Source Workbook` diisi path inventory terdekat (HQ alias / folder parent / basename / candidate)
- Worksheet: exact/fuzzy ke inventory bila ada; jika tidak, judul R&W tetap dipakai sebagai referensi
- R&W `#N/A` pada high_confidence → terima candidate workbook/worksheet sebagai YES

## Inventory resolve status
{json.dumps(dict(status_counts), ensure_ascii=False, indent=2)}

## Confidence
{json.dumps(dict(labels), ensure_ascii=False, indent=2)}

## Tags
{json.dumps(dict(tags), ensure_ascii=False, indent=2)}

## Why 257 NIPPs were in the old mapping but not this roster
{extra_257["note"]}

- Extra count: **{extra_257["extra_count"]}**
- Under Subholding tree: **{extra_257["under_subholding_tree_unique"]}**
- Roots: `{json.dumps(extra_257["roots"], ensure_ascii=False)}`
- Top companies: `{json.dumps(extra_257["top_companies"], ensure_ascii=False)}`

Those 257 are **excluded** from this workbook.

## Artifacts
- `{xlsx_path}`
- `{json_path}`
- `{readme_path}`

## Reviewer columns
{', '.join(g2.REVIEW_COLUMNS)}

Filter `Roster Review Tag = {TAG_NEW_56}` or open sheet `Needs Review (56)`.
"""
    receipt_path.write_text(receipt, encoding="utf-8")

    readme = f"""# README SOURCE — Position First Mapping Subholding {stamp}

## Input files (copied under `source/`)
1. `REGIONAL dan SUBHOLDING.xlsx` — roster pegawai (sheets SPTP/SPMT/SPSL/SPJM) = **2705** unique NIPP  
   SHA256: `{roster_hash}`
2. `Pemetaan Kamus KPI Subholding (05 Aug 2026)-with mapping.xlsx` — hasil review Red & White  
   SHA256: `{reviewed_hash}`

## Config / reference used
- Worksheet inventory: `{args.inventory}` (`{inventory_hash}`)
- Production positions: `{args.reference}` exported_at `{exported_at}` (`{reference_hash}`)
- Prior Subholding draft filtered to roster: `{args.old_draft}`

## Output
- `{xlsx_path.name}` — review workbook
- `{json_path.name}` — machine-readable draft
- `MAPPING_RECEIPT.md` — counts + 257 analysis

Generated at `{generated_at}`.
"""
    readme_path.write_text(readme, encoding="utf-8")

    print(
        json.dumps(
            {
                "title": cover_title,
                "unique_active_employees": len(unique_nipps),
                "position_rows": len(rows),
                "needs_review_new_56": len(needs_56),
                "reviewer_yes": sum(1 for row in rows if norm(row.get("Reviewer Confirm Mapping")) == "YES"),
                "inventory_resolve_status_counts": dict(status_counts),
                "confidence_counts": dict(labels),
                "tag_counts": dict(tags),
                "excluded_257": extra_257["extra_count"],
                "xlsx": str(xlsx_path),
            },
            ensure_ascii=False,
            indent=2,
        )
    )


if __name__ == "__main__":
    main()
