#!/usr/bin/env python3
"""Apply explicitly approved mapping override candidates to a discovered config."""

from __future__ import annotations

import argparse
import json
from collections import Counter
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


APPROVED_VALUES = {"1", "true", "yes", "y", "approved", "approve", "ok"}
CONFIG_KEY_COLUMNS = ("Source Workbook", "Sheet")


def norm(value: Any) -> str:
    return "" if value is None else str(value).strip()


def is_approved(value: Any) -> bool:
    return norm(value).lower() in APPROVED_VALUES


def load_override_rows(path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    if "Override Candidates" not in workbook.sheetnames:
        raise ValueError(f"{path} does not contain Override Candidates")
    worksheet = workbook["Override Candidates"]
    rows = worksheet.iter_rows(values_only=True)
    headers = [norm(value) for value in next(rows)]
    return [dict(zip(headers, row)) for row in rows]


def config_key(row: dict[str, Any]) -> tuple[str, str]:
    return norm(row.get("source_workbook")), norm(row.get("sheet_name"))


def override_key(row: dict[str, Any]) -> tuple[str, str]:
    return norm(row.get("Source Workbook")), norm(row.get("Sheet"))


def approved_override_map(rows: list[dict[str, Any]]) -> tuple[dict[tuple[str, str], dict[str, Any]], list[str]]:
    errors: list[str] = []
    approved: dict[tuple[str, str], dict[str, Any]] = {}
    seen = Counter()
    for row in rows:
        if not is_approved(row.get("Approved")):
            continue
        key = override_key(row)
        seen[key] += 1
        scope = norm(row.get("Suggested Position Scope"))
        pmid = norm(row.get("Suggested Position Master ID"))
        pnid = norm(row.get("Suggested Position Nomenklatur ID"))
        reviewer_scope = norm(row.get("Reviewer Selected Scope"))
        reviewer_pmid = norm(row.get("Reviewer Selected Position Master ID"))
        reviewer_pnid = norm(row.get("Reviewer Selected Position Nomenklatur ID"))
        if reviewer_scope or reviewer_pmid or reviewer_pnid:
            scope = reviewer_scope
            pmid = reviewer_pmid
            pnid = reviewer_pnid
        if not key[0] or not key[1]:
            errors.append("approved row is missing Source Workbook or Sheet")
            continue
        if scope not in {"structural", "non_structural"}:
            errors.append(f"approved row has unsupported scope {scope or '<blank>'}: {key}")
            continue
        if scope == "structural" and (not pmid or pnid):
            errors.append(f"approved structural row must have PMID only: {key}")
            continue
        if scope == "non_structural" and (not pnid or pmid):
            errors.append(f"approved non_structural row must have PNID only: {key}")
            continue
        approved[key] = row
    for key, count in seen.items():
        if count > 1:
            errors.append(f"multiple approved rows for {key}")
    return approved, errors


def apply_overrides(config: dict[str, Any], overrides: dict[tuple[str, str], dict[str, Any]]) -> tuple[int, list[str]]:
    errors: list[str] = []
    positions = config.get("positions", [])
    config_counts = Counter(config_key(row) for row in positions)
    applied = 0
    for row in positions:
        key = config_key(row)
        override = overrides.get(key)
        if not override:
            continue
        if config_counts[key] > 1:
            errors.append(f"config has duplicate position rows for {key}")
            continue
        scope = norm(override.get("Suggested Position Scope"))
        pmid = norm(override.get("Suggested Position Master ID"))
        pnid = norm(override.get("Suggested Position Nomenklatur ID"))
        row["position_scope"] = scope
        row["position_master_id"] = pmid or None
        row["position_nomenclature_id"] = pnid or None
        row["mapping_override_approved"] = True
        row["mapping_review_status"] = "approved"
        row["portaverse_position_title"] = norm(override.get("Suggested Position Title")) or None
        row["portaverse_group_name"] = norm(override.get("Suggested Group")) or None
        row["portaverse_company_name"] = norm(override.get("Suggested Company")) or None
        row["portaverse_company_code"] = norm(override.get("Suggested Company Code")) or None
        row["candidate_position_master_id"] = pmid or None
        row["candidate_position_nomenclature_id"] = pnid or None
        row["active_employee_name"] = norm(override.get("Active Employee Name")) or None
        row["active_employee_nipp"] = norm(override.get("Active Employee NIPP")) or None
        row["cluster_label"] = norm(override.get("Suggested Position Title")) if scope == "non_structural" else None
        applied += 1
    missing = sorted(set(overrides) - {config_key(row) for row in positions})
    errors.extend(f"approved override does not match config row: {key}" for key in missing)
    return applied, errors


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--config", required=True, type=Path)
    parser.add_argument("--overrides", required=True, type=Path)
    parser.add_argument("--output", required=True, type=Path)
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()

    config = json.loads(args.config.read_text(encoding="utf-8"))
    override_rows = load_override_rows(args.overrides)
    overrides, errors = approved_override_map(override_rows)
    applied, apply_errors = apply_overrides(config, overrides)
    errors.extend(apply_errors)
    if errors:
        for error in errors:
            print(f"ERROR: {error}")
        return 1
    print(f"approved_rows={len(overrides)}")
    print(f"applied_rows={applied}")
    if args.dry_run:
        print("dry_run=true")
        return 0
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(json.dumps(config, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(f"output={args.output}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
