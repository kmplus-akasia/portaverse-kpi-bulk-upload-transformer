#!/usr/bin/env python3
"""Reconcile Red & White Subholding mapping reviews vs inventory + merge to 2,705.

Inputs:
  1) Main R&W review (Position Coverage + Mapping)
  2) Fresh mapping R&W review for the 56 new workers (Pemetaan + Belum 56 + Mapping)
  3) Group 2 kamus inventory
  4) Roster SPTP/SPMT/SPSL/SPJM (exact 2,705 NIPPs)
  5) Production position reference (resolve PMID/PNID → NIPPs for File1)
  6) Automated fresh mapping (high_confidence fallback when R&W path empty)

Outputs:
  - Gap report: R&W Folder/Workbook/Worksheet vs inventory status + nearest candidates
  - Confirmed mapping workbook scoped to 2,705 roster NIPPs with inventory availability
  - Conversion readiness by position identity (PMID/PNID)
"""

from __future__ import annotations

import argparse
import json
import re
import zipfile
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any
from xml.etree import ElementTree as ET

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

import position_mapping as pm

NS = {"m": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
REL = "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}"
NAVY, TEAL, PALE = "173651", "138074", "E9F1F8"
GREEN, YELLOW, ORANGE, GRAY, RED = "D9EAD3", "FFF2CC", "FCE4D6", "E7E6E6", "F4CCCC"
BODY = "Aptos"

STATUS_FOUND = "found"
STATUS_FOLDER_MISMATCH = "found_folder_mismatch"
STATUS_SHEET_MISSING = "workbook_ok_sheet_missing"
STATUS_WB_MISSING = "workbook_not_in_inventory"
STATUS_EMPTY = "empty"

READY_STATUSES = {STATUS_FOUND, STATUS_FOLDER_MISMATCH}

CONV_RW_FOUND = "siap__red_white_inventory_found"
CONV_RW_FOLDER = "siap__red_white_folder_mismatch"
CONV_AUTO_HIGH = "siap__automated_high_confidence"
CONV_RW_SHEET = "belum__red_white_sheet_unresolved"
CONV_RW_WB = "belum__red_white_workbook_missing"
CONV_AUTO_LOW = "belum__automated_low_confidence"
CONV_AUTO_NONE = "belum__automated_no_candidate"
CONV_AUTO_CONFLICT = "belum__automated_mapping_conflict"
CONV_MISSING_ID = "belum__missing_production_identity"
CONV_EMPTY = "belum__empty_no_high_confidence"

CONV_LABELS = {
    CONV_RW_FOUND: (
        "Siap konversi — path Red & White resolve ke workbook+sheet inventory "
        "(Folder/Workbook/Worksheet review cocok)."
    ),
    CONV_RW_FOLDER: (
        "Siap konversi — sheet+workbook R&W ketemu di inventory; folder R&W tidak selaras "
        "(folder diabaikan untuk konversi formulir)."
    ),
    CONV_AUTO_HIGH: (
        "Siap konversi — path R&W kosong; memakai automated mapping high_confidence "
        "dari Position First Mapping (workbook+sheet dari inventory)."
    ),
    CONV_RW_SHEET: (
        "Belum konversi — R&W punya workbook yang ada di inventory, tetapi Worksheet Title "
        "tidak resolve (sering judul panjang vs tab ≤31 karakter / beda nama)."
    ),
    CONV_RW_WB: (
        "Belum konversi — Workbook Title R&W tidak ada di inventory "
        "(contoh: Pandu, Branch Kalimas, nama file beda)."
    ),
    CONV_AUTO_LOW: (
        "Belum konversi — tidak ada path R&W yang resolve; automated mapping low_confidence "
        "(ada kandidat tetapi tidak lolos strict high-confidence)."
    ),
    CONV_AUTO_NONE: (
        "Belum konversi — tidak ada path R&W yang resolve; automated mapping no_candidate "
        "(tidak ada sheet kamus yang lolos threshold)."
    ),
    CONV_AUTO_CONFLICT: (
        "Belum konversi — tidak ada path R&W yang resolve; automated mapping mapping_conflict "
        "(lebih dari satu kandidat / konflik skor)."
    ),
    CONV_MISSING_ID: (
        "Belum konversi — NIPP roster belum punya identitas produksi PMID/PNID "
        "(stub / missing production identity)."
    ),
    CONV_EMPTY: (
        "Belum konversi — path R&W kosong dan tidak ada automated high_confidence "
        "yang bisa di-overlay."
    ),
}


def norm(value: Any) -> str:
    return "" if value is None else str(value).strip()


def load_shared_strings(archive: zipfile.ZipFile) -> list[str]:
    if "xl/sharedStrings.xml" not in archive.namelist():
        return []
    root = ET.fromstring(archive.read("xl/sharedStrings.xml"))
    out: list[str] = []
    for item in root.findall("m:si", NS):
        texts = [
            node.text or ""
            for node in item.iter("{http://schemas.openxmlformats.org/spreadsheetml/2006/main}t")
        ]
        out.append("".join(texts))
    return out


def cell_value(cell: ET.Element, shared: list[str]) -> Any:
    cell_type = cell.attrib.get("t")
    inline = cell.find("m:is", NS)
    if inline is not None:
        return "".join(
            node.text or ""
            for node in inline.iter("{http://schemas.openxmlformats.org/spreadsheetml/2006/main}t")
        )
    value_node = cell.find("m:v", NS)
    if value_node is None:
        return None
    raw = value_node.text
    return shared[int(raw)] if cell_type == "s" else raw


def col_index(cell_ref: str) -> int:
    match = re.match(r"([A-Z]+)", cell_ref or "A1")
    assert match is not None
    total = 0
    for char in match.group(1):
        total = total * 26 + (ord(char) - 64)
    return total - 1


def read_sheet_rows(archive: zipfile.ZipFile, target: str, shared: list[str]) -> list[list[Any]]:
    root = ET.fromstring(archive.read(target))
    rows: list[list[Any]] = []
    for row in root.findall("m:sheetData/m:row", NS):
        cells: dict[int, Any] = {}
        for cell in row.findall("m:c", NS):
            cells[col_index(cell.attrib.get("r", "A1"))] = cell_value(cell, shared)
        if not cells:
            continue
        width = max(cells) + 1
        rows.append([cells.get(index) for index in range(width)])
    return rows


def workbook_sheet_map(archive: zipfile.ZipFile) -> dict[str, str]:
    workbook = ET.fromstring(archive.read("xl/workbook.xml"))
    rels = ET.fromstring(archive.read("xl/_rels/workbook.xml.rels"))
    rid = {rel.attrib["Id"]: rel.attrib["Target"] for rel in rels}
    out: dict[str, str] = {}
    for sheet in workbook.findall("m:sheets/m:sheet", NS):
        target = rid[sheet.attrib[f"{REL}id"]].lstrip("/")
        if not target.startswith("xl/"):
            target = f"xl/{target}"
        out[sheet.attrib["name"]] = target
    return out


def sheet_rows(path: Path, sheet_name: str) -> list[list[Any]]:
    with zipfile.ZipFile(path) as archive:
        shared = load_shared_strings(archive)
        sheets = workbook_sheet_map(archive)
        if sheet_name not in sheets:
            raise KeyError(f"Sheet {sheet_name!r} not in {path.name}: {sorted(sheets)}")
        return read_sheet_rows(archive, sheets[sheet_name], shared)


def parse_table(rows: list[list[Any]], required_keys: list[str]) -> list[dict[str, str]]:
    header_idx = None
    header: list[str] = []
    for index, row in enumerate(rows[:10]):
        values = [norm(cell) for cell in row]
        if sum(1 for key in required_keys if key in values) >= min(2, len(required_keys)):
            header_idx = index
            header = values
            break
    if header_idx is None:
        raise RuntimeError(f"Header not found for keys {required_keys}")
    index_map = {name: idx for idx, name in enumerate(header) if name}
    out: list[dict[str, str]] = []
    for row in rows[header_idx + 1 :]:
        if not row or all(cell in (None, "") for cell in row):
            continue
        out.append(
            {
                name: (norm(row[idx]) if idx < len(row) else "")
                for name, idx in index_map.items()
            }
        )
    return out


def extract_nipps_from_pegawai(text: str) -> list[str]:
    values = re.findall(r"\(([^()\n]+)\)", norm(text))
    return [value.strip() for value in values if value.strip()]


def load_roster_nipps(path: Path) -> set[str]:
    roster: set[str] = set()
    with zipfile.ZipFile(path) as archive:
        shared = load_shared_strings(archive)
        sheets = workbook_sheet_map(archive)
        for sheet_name in ("SPTP", "SPMT", "SPSL", "SPJM"):
            if sheet_name not in sheets:
                continue
            rows = read_sheet_rows(archive, sheets[sheet_name], shared)
            header = [norm(cell) for cell in rows[0]]
            index = {name: idx for idx, name in enumerate(header)}
            for row in rows[1:]:
                if not row:
                    continue
                nipp = norm(row[index["PNALT_NEW"]]) if "PNALT_NEW" in index and index["PNALT_NEW"] < len(row) else ""
                if nipp:
                    roster.add(nipp)
    return roster


@dataclass(frozen=True)
class InventoryEntry:
    source_folder: str
    source_workbook: str
    sheet_name: str
    position_name: str
    basename: str
    stem: str


def load_inventory(path: Path) -> list[InventoryEntry]:
    payload = json.loads(path.read_text(encoding="utf-8"))
    entries: list[InventoryEntry] = []
    for row in payload.get("kamus_kpi_v2", []):
        if not isinstance(row, dict) or not row.get("include_in_position_config"):
            continue
        workbook = norm(row.get("source_workbook"))
        entries.append(
            InventoryEntry(
                source_folder=norm(row.get("source_folder")),
                source_workbook=workbook,
                sheet_name=norm(row.get("sheet_name")),
                position_name=norm(row.get("position_name")),
                basename=Path(workbook).name,
                stem=Path(workbook).stem,
            )
        )
    return entries


def sheet_compatible(rw_title: str, inv_sheet: str, inv_position: str) -> bool:
    left = rw_title.casefold()
    sheet = inv_sheet.casefold()
    position = inv_position.casefold()
    if not left:
        return False
    if left == sheet or (position and left == position):
        return True
    if len(sheet) == 31 and left.startswith(sheet.rstrip()):
        return True
    if len(left) == 31 and sheet.startswith(left.rstrip()):
        return True
    if position and len(position) == 31 and left.startswith(position.rstrip()):
        return True
    if left[:31] == sheet or (position and left[:31] == position):
        return True
    if sheet.startswith(left) or left.startswith(sheet):
        return True
    if position and (position.startswith(left) or left.startswith(position)):
        return True
    return False


def token_overlap_score(left: str, right: str) -> float:
    a = set(pm.normalize_position_lookup(left).split())
    b = set(pm.normalize_position_lookup(right).split())
    if not a or not b:
        return 0.0
    return len(a & b) / max(len(a), len(b))


class InventoryIndex:
    def __init__(self, entries: list[InventoryEntry]) -> None:
        self.entries = entries
        self.by_base: dict[str, list[InventoryEntry]] = defaultdict(list)
        self.bases: set[str] = set()
        for entry in entries:
            key = entry.basename.casefold()
            stem = entry.stem.casefold()
            self.by_base[key].append(entry)
            self.by_base[stem].append(entry)
            self.bases.add(key)
            self.bases.add(stem)

    def workbook_candidates(self, workbook_title: str) -> list[InventoryEntry]:
        title = norm(workbook_title)
        if not title:
            return []
        keys = {
            title.casefold(),
            Path(title).name.casefold(),
            Path(title).stem.casefold(),
        }
        if not title.casefold().endswith(".xlsx"):
            keys.add(f"{title.casefold()}.xlsx")
        hits: list[InventoryEntry] = []
        for key in keys:
            hits.extend(self.by_base.get(key, []))
        if hits:
            return hits
        # partial contains
        needle = title.casefold()
        for base, rows in self.by_base.items():
            if needle in base or base in needle:
                hits.extend(rows)
                if hits:
                    break
        return hits

    def match(
        self, folder: str, workbook_title: str, worksheet_title: str
    ) -> tuple[str, InventoryEntry | None, list[InventoryEntry]]:
        folder_n = norm(folder)
        workbook_n = norm(workbook_title)
        sheet_n = norm(worksheet_title)
        if not workbook_n and not sheet_n:
            return STATUS_EMPTY, None, []
        candidates = self.workbook_candidates(workbook_n)
        if not candidates:
            return STATUS_WB_MISSING, None, self.nearest_workbooks(workbook_n, limit=5)
        sheet_hits = [
            entry
            for entry in candidates
            if sheet_compatible(sheet_n, entry.sheet_name, entry.position_name)
        ]
        if sheet_hits:
            if folder_n:
                folder_cf = folder_n.casefold()
                folder_hits = [
                    entry
                    for entry in sheet_hits
                    if folder_cf in entry.source_folder.casefold()
                    or folder_cf in entry.source_workbook.casefold()
                ]
                if folder_hits:
                    return STATUS_FOUND, folder_hits[0], folder_hits[:5]
                return STATUS_FOLDER_MISMATCH, sheet_hits[0], sheet_hits[:5]
            return STATUS_FOUND, sheet_hits[0], sheet_hits[:5]
        nearest = self.nearest_sheets(candidates, sheet_n, limit=5)
        return STATUS_SHEET_MISSING, None, nearest

    def nearest_workbooks(self, workbook_title: str, limit: int = 5) -> list[InventoryEntry]:
        title = pm.normalize_title(workbook_title)
        scored: list[tuple[float, InventoryEntry]] = []
        seen: set[str] = set()
        for entry in self.entries:
            key = entry.basename.casefold()
            if key in seen:
                continue
            seen.add(key)
            score = token_overlap_score(title, entry.basename)
            if score <= 0:
                continue
            scored.append((score, entry))
        scored.sort(key=lambda item: (-item[0], item[1].basename))
        return [entry for _, entry in scored[:limit]]

    def nearest_sheets(
        self, workbook_entries: list[InventoryEntry], worksheet_title: str, limit: int = 5
    ) -> list[InventoryEntry]:
        scored: list[tuple[float, InventoryEntry]] = []
        seen: set[tuple[str, str]] = set()
        for entry in workbook_entries:
            key = (entry.source_workbook, entry.sheet_name)
            if key in seen:
                continue
            seen.add(key)
            score = max(
                token_overlap_score(worksheet_title, entry.sheet_name),
                token_overlap_score(worksheet_title, entry.position_name),
            )
            if score <= 0:
                continue
            scored.append((score, entry))
        scored.sort(key=lambda item: (-item[0], item[1].sheet_name))
        return [entry for _, entry in scored[:limit]]


def style_header(cell, fill: str = TEAL) -> None:
    cell.fill = PatternFill("solid", fgColor=fill)
    cell.font = Font(name=BODY, bold=True, color="FFFFFF", size=10)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def write_table(ws, title: str, headers: list[str], rows: list[dict[str, Any]], table_name: str) -> None:
    ws.sheet_view.showGridLines = False
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws.cell(1, 1, title).fill = PatternFill("solid", fgColor=NAVY)
    ws.cell(1, 1).font = Font(name=BODY, size=14, bold=True, color="FFFFFF")
    for col, header in enumerate(headers, start=1):
        style_header(ws.cell(3, col, header))
    status_col = headers.index("Status Inventory") + 1 if "Status Inventory" in headers else None
    for r_idx, row in enumerate(rows, start=4):
        for c_idx, header in enumerate(headers, start=1):
            cell = ws.cell(r_idx, c_idx, row.get(header, ""))
            cell.font = Font(name=BODY, size=9)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        if status_col is not None:
            status = norm(row.get("Status Inventory"))
            fill = {
                STATUS_FOUND: GREEN,
                STATUS_FOLDER_MISMATCH: YELLOW,
                STATUS_SHEET_MISSING: ORANGE,
                STATUS_WB_MISSING: RED,
                STATUS_EMPTY: GRAY,
            }.get(status)
            if fill:
                ws.cell(r_idx, status_col).fill = PatternFill("solid", fgColor=fill)
    end_row = 3 + max(len(rows), 1)
    if rows:
        table = Table(
            displayName=table_name,
            ref=f"A3:{get_column_letter(len(headers))}{end_row}",
        )
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
        ws.add_table(table)
    for idx, header in enumerate(headers, start=1):
        width = 18
        if "Path" in header or "Workbook" in header or "Folder" in header or "Kandidat" in header:
            width = 42
        if header in {"Pegawai", "Alasan / Catatan"}:
            width = 36
        if header in {"PMID", "PNID", "No.", "Status Inventory"}:
            width = 14 if header != "Status Inventory" else 26
        ws.column_dimensions[get_column_letter(idx)].width = width
    ws.freeze_panes = "A4"


def format_candidate(entry: InventoryEntry | None) -> str:
    if entry is None:
        return ""
    return f"{entry.source_workbook} :: {entry.sheet_name}"


def format_candidates(entries: list[InventoryEntry]) -> str:
    lines = []
    for entry in entries[:5]:
        score_hint = entry.position_name or entry.sheet_name
        lines.append(f"- {entry.basename} / {entry.sheet_name} ({score_hint})")
    return "\n".join(lines)


def path_row(
    *,
    source: str,
    folder: str,
    workbook_title: str,
    worksheet_title: str,
    index: InventoryIndex,
    extra: dict[str, Any] | None = None,
) -> dict[str, Any]:
    status, best, nearest = index.match(folder, workbook_title, worksheet_title)
    row = {
        "Sumber Review": source,
        "Folder R&W": folder,
        "Workbook Title R&W": workbook_title,
        "Worksheet Title R&W": worksheet_title,
        "Status Inventory": status,
        "Inventory Workbook Path": best.source_workbook if best else "",
        "Inventory Sheet": best.sheet_name if best else "",
        "Inventory Position Name": best.position_name if best else "",
        "Inventory Folder": best.source_folder if best else "",
        "Kandidat Terdekat": format_candidates(nearest if nearest else ([best] if best else [])),
    }
    if extra:
        row.update(extra)
    return row


def load_file1_coverage(path: Path) -> list[dict[str, str]]:
    return parse_table(
        sheet_rows(path, "Position Coverage"),
        ["PMID", "Position Title", "Folder", "Workbook Title", "Worksheet Title"],
    )


def load_file1_mapping(path: Path) -> list[dict[str, str]]:
    return parse_table(
        sheet_rows(path, "Mapping"),
        ["Folder", "Workbook Title", "Worksheet Title"],
    )


def load_file2_pemetaan(path: Path) -> list[dict[str, str]]:
    return parse_table(sheet_rows(path, "Pemetaan"), ["PMID", "Pegawai", "Judul Posisi"])


def load_file2_56(path: Path) -> list[dict[str, str]]:
    return parse_table(
        sheet_rows(path, "Belum di Mapping Awal (56)"),
        ["PMID", "Pegawai", "Folder", "Workbook Title", "Worksheet Title"],
    )


def load_file2_mapping(path: Path) -> list[dict[str, str]]:
    return parse_table(
        sheet_rows(path, "Mapping"),
        ["Folder", "Workbook Title", "Worksheet Title"],
    )


def build_pmid_pnid_nipp_maps(
    reference_path: Path,
) -> tuple[dict[str, pm.LookupCandidate], dict[str, pm.LookupCandidate]]:
    payload = json.loads(reference_path.read_text(encoding="utf-8"))
    indexes = pm.build_lookup_indexes(payload)
    by_pmid = {
        str(candidate.position_master_id): candidate
        for candidate in indexes.structural
        if candidate.position_master_id
    }
    by_pnid = {
        str(candidate.position_nomenclature_id): candidate
        for candidate in indexes.non_structural
        if candidate.position_nomenclature_id
    }
    return by_pmid, by_pnid


def parse_auto_nipps(row: dict[str, Any]) -> list[str]:
    raw = row.get("Active Employee NIPPs")
    if isinstance(raw, list):
        return [norm(item) for item in raw if norm(item)]
    text = norm(raw)
    if not text:
        return []
    parts = re.split(r"[;\n,]+", text)
    return [part.strip() for part in parts if part.strip()]


def load_automated_by_nipp(path: Path) -> dict[str, dict[str, Any]]:
    payload = json.loads(path.read_text(encoding="utf-8"))
    by_nipp: dict[str, dict[str, Any]] = {}
    for row in payload.get("rows", []):
        if not isinstance(row, dict):
            continue
        for nipp in parse_auto_nipps(row):
            by_nipp[nipp] = row
    return by_nipp


def workbook_basename(path_or_title: str) -> str:
    text = norm(path_or_title)
    if not text:
        return ""
    return Path(text).name


def classify_conversion(
    *,
    pmid: str,
    pnid: str,
    path_source: str,
    inventory_status: str,
    auto_confidence: str,
) -> tuple[str, str, str]:
    """Return (siap|belum, kategori, penjelasan)."""
    if not pmid and not pnid:
        return "belum", CONV_MISSING_ID, CONV_LABELS[CONV_MISSING_ID]

    if inventory_status == STATUS_FOUND:
        if path_source.startswith("Automated high_confidence"):
            return "siap", CONV_AUTO_HIGH, CONV_LABELS[CONV_AUTO_HIGH]
        return "siap", CONV_RW_FOUND, CONV_LABELS[CONV_RW_FOUND]

    if inventory_status == STATUS_FOLDER_MISMATCH:
        if path_source.startswith("Automated high_confidence"):
            return "siap", CONV_AUTO_HIGH, CONV_LABELS[CONV_AUTO_HIGH]
        return "siap", CONV_RW_FOLDER, CONV_LABELS[CONV_RW_FOLDER]

    if inventory_status == STATUS_SHEET_MISSING:
        return "belum", CONV_RW_SHEET, CONV_LABELS[CONV_RW_SHEET]
    if inventory_status == STATUS_WB_MISSING:
        return "belum", CONV_RW_WB, CONV_LABELS[CONV_RW_WB]

    # empty / unresolved without usable path
    if auto_confidence == pm.HIGH_CONFIDENCE:
        # High confidence should have been overlaid; if still empty, treat as empty.
        return "belum", CONV_EMPTY, CONV_LABELS[CONV_EMPTY]
    if auto_confidence == pm.LOW_CONFIDENCE:
        return "belum", CONV_AUTO_LOW, CONV_LABELS[CONV_AUTO_LOW]
    if auto_confidence == "mapping_conflict":
        return "belum", CONV_AUTO_CONFLICT, CONV_LABELS[CONV_AUTO_CONFLICT]
    if auto_confidence == pm.NO_CANDIDATE:
        return "belum", CONV_AUTO_NONE, CONV_LABELS[CONV_AUTO_NONE]
    return "belum", CONV_EMPTY, CONV_LABELS[CONV_EMPTY]


def identity_key(pmid: str, pnid: str) -> tuple[str, str]:
    return (norm(pmid), norm(pnid))


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--file1",
        type=Path,
        default=Path(
            "/Users/alfredoteja/Downloads/"
            "Pemetaan Kamus KPI Subholding (05 Aug 2026)-with mapping (1).xlsx"
        ),
    )
    parser.add_argument(
        "--file2",
        type=Path,
        default=Path(
            "/Users/alfredoteja/Downloads/"
            "Position_First_Mapping_Subholding_Fresh_20260806_182856-with mapping.xlsx"
        ),
    )
    parser.add_argument(
        "--inventory",
        type=Path,
        default=Path("configs/kamus_kpi_group2_visible_20260807.json"),
    )
    parser.add_argument(
        "--roster",
        type=Path,
        default=Path(
            "outputs/kamus-group2-subholding-roster-mapping-20260806/source/"
            "REGIONAL dan SUBHOLDING.xlsx"
        ),
    )
    parser.add_argument(
        "--reference",
        type=Path,
        default=Path("configs/production_position_reference.json"),
    )
    parser.add_argument(
        "--automated-mapping",
        type=Path,
        default=Path(
            "outputs/kamus-group2-subholding-roster-fresh-20260806/"
            "subholding_roster_fresh_mapping_LATEST.json"
        ),
    )
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=Path("outputs/kamus-group2-subholding-rw-reconcile-20260807"),
    )
    args = parser.parse_args()

    stamp = datetime.now().astimezone().strftime("%Y%m%d_%H%M%S")
    generated_at = datetime.now().astimezone().isoformat(timespec="seconds")
    args.output_dir.mkdir(parents=True, exist_ok=True)

    roster = load_roster_nipps(args.roster)
    if len(roster) != 2705:
        raise SystemExit(f"Expected 2705 roster NIPPs, got {len(roster)}")

    inventory = load_inventory(args.inventory)
    index = InventoryIndex(inventory)
    by_pmid, by_pnid = build_pmid_pnid_nipp_maps(args.reference)
    automated_by_nipp = load_automated_by_nipp(args.automated_mapping)

    file1_coverage = load_file1_coverage(args.file1)
    file1_mapping = load_file1_mapping(args.file1)
    file2_pemetaan = load_file2_pemetaan(args.file2)
    file2_56 = load_file2_56(args.file2)
    file2_mapping = load_file2_mapping(args.file2)

    # --- Gap report rows (unique paths + per-row coverage/56) ---
    gap_unique: dict[tuple[str, str, str, str], dict[str, Any]] = {}
    for source, rows in (
        ("File1 Mapping", file1_mapping),
        ("File2 Mapping", file2_mapping),
    ):
        for row in rows:
            key = (
                source,
                norm(row.get("Folder")),
                norm(row.get("Workbook Title")),
                norm(row.get("Worksheet Title")),
            )
            if key in gap_unique:
                continue
            gap_unique[key] = path_row(
                source=source,
                folder=key[1],
                workbook_title=key[2],
                worksheet_title=key[3],
                index=index,
            )

    gap_coverage_rows = [
        path_row(
            source="File1 Position Coverage",
            folder=norm(row.get("Folder")),
            workbook_title=norm(row.get("Workbook Title")),
            worksheet_title=norm(row.get("Worksheet Title")),
            index=index,
            extra={
                "PMID": norm(row.get("PMID")),
                "PNID": norm(row.get("PNID")),
                "Position Title": norm(row.get("Position Title")),
                "Company": norm(row.get("Company")),
            },
        )
        for row in file1_coverage
    ]

    gap_56_rows = [
        path_row(
            source="File2 Belum di Mapping Awal (56)",
            folder=norm(row.get("Folder")),
            workbook_title=norm(row.get("Workbook Title")),
            worksheet_title=norm(row.get("Worksheet Title")),
            index=index,
            extra={
                "PMID": norm(row.get("PMID")),
                "PNID": norm(row.get("PNID")),
                "Judul Posisi": norm(row.get("Judul Posisi")),
                "Pegawai": norm(row.get("Pegawai")),
            },
        )
        for row in file2_56
    ]

    unique_gap_rows = list(gap_unique.values())
    status_unique = Counter(row["Status Inventory"] for row in unique_gap_rows)
    status_cov = Counter(row["Status Inventory"] for row in gap_coverage_rows)
    status_56 = Counter(row["Status Inventory"] for row in gap_56_rows)

    # --- Merge confirmed 2,705 ---
    # Prefer File2 Pemetaan as base (exact roster). Overlay R&W path from:
    # 1) File2 Belum-56 by NIPP
    # 2) File1 Coverage by PMID/PNID for other NIPPs

    path_by_nipp: dict[str, dict[str, str]] = {}
    for row in file2_56:
        folder = norm(row.get("Folder"))
        workbook = norm(row.get("Workbook Title"))
        worksheet = norm(row.get("Worksheet Title"))
        for nipp in extract_nipps_from_pegawai(row.get("Pegawai", "")):
            if nipp in roster:
                path_by_nipp[nipp] = {
                    "source": "File2 Belum-56",
                    "folder": folder,
                    "workbook": workbook,
                    "worksheet": worksheet,
                    "pmid": norm(row.get("PMID")),
                    "pnid": norm(row.get("PNID")),
                    "title": norm(row.get("Judul Posisi")),
                    "company": norm(row.get("Perusahaan")),
                    "company_in_id": norm(row.get("company_in_id")),
                    "unit": norm(row.get("Unit / Group")),
                    "pegawai_line": next(
                        (
                            line.strip()
                            for line in norm(row.get("Pegawai")).splitlines()
                            if f"({nipp})" in line
                        ),
                        f"- ({nipp})",
                    ),
                }

    for row in file1_coverage:
        pmid = norm(row.get("PMID"))
        pnid = norm(row.get("PNID"))
        candidate = by_pmid.get(pmid) if pmid else None
        if candidate is None and pnid:
            candidate = by_pnid.get(pnid)
        if candidate is None:
            continue
        folder = norm(row.get("Folder"))
        workbook = norm(row.get("Workbook Title"))
        worksheet = norm(row.get("Worksheet Title"))
        names = list(candidate.active_employee_names)
        nipps = [str(nipp).strip() for nipp in candidate.active_employee_nipps]
        for idx, nipp in enumerate(nipps):
            if nipp not in roster or nipp in path_by_nipp:
                continue
            name = names[idx] if idx < len(names) else ""
            path_by_nipp[nipp] = {
                "source": "File1 Position Coverage",
                "folder": folder,
                "workbook": workbook,
                "worksheet": worksheet,
                "pmid": pmid,
                "pnid": pnid,
                "title": norm(row.get("Position Title")) or (candidate.title or ""),
                "company": norm(row.get("Company")) or (candidate.company_name or ""),
                "company_in_id": norm(candidate.company_id),
                "unit": norm(row.get("Group / Unit")) or (candidate.group_name or ""),
                "pegawai_line": f"- {name} ({nipp})" if name else f"- ({nipp})",
            }

    # Base rows from File2 Pemetaan (guarantees 2,705)
    empty_before_auto = 0
    auto_high_overlay_count = 0
    confirmed_by_nipp: dict[str, dict[str, Any]] = {}
    for row in file2_pemetaan:
        nipps = extract_nipps_from_pegawai(row.get("Pegawai", ""))
        for nipp in nipps:
            if nipp not in roster:
                continue
            overlay = path_by_nipp.get(nipp, {})
            folder = overlay.get("folder") or ""
            workbook = overlay.get("workbook") or ""
            worksheet = overlay.get("worksheet") or ""
            path_source = overlay.get("source") or "File2 Pemetaan (path kosong)"
            auto_row = automated_by_nipp.get(nipp, {})
            auto_confidence = norm(auto_row.get("Confidence Label"))

            if not folder and not workbook and not worksheet:
                empty_before_auto += 1
                if (
                    auto_confidence == pm.HIGH_CONFIDENCE
                    and norm(auto_row.get("Candidate Source Workbook"))
                    and norm(auto_row.get("Candidate Worksheet"))
                ):
                    folder = norm(auto_row.get("Candidate Source Folder"))
                    workbook = workbook_basename(norm(auto_row.get("Candidate Source Workbook")))
                    worksheet = norm(auto_row.get("Candidate Worksheet"))
                    path_source = "Automated high_confidence"
                    auto_high_overlay_count += 1

            status, best, nearest = index.match(folder, workbook, worksheet)
            if not folder and not workbook and not worksheet:
                status = STATUS_EMPTY

            pmid = overlay.get("pmid") or norm(row.get("PMID"))
            pnid = overlay.get("pnid") or norm(row.get("PNID"))
            if not pmid and not pnid and auto_row:
                pmid = norm(auto_row.get("PMID"))
                pnid = norm(auto_row.get("PNID"))

            siap, kategori, penjelasan = classify_conversion(
                pmid=pmid,
                pnid=pnid,
                path_source=path_source,
                inventory_status=status,
                auto_confidence=auto_confidence,
            )

            confirmed_by_nipp[nipp] = {
                "NIPP": nipp,
                "Nama": (
                    overlay.get("pegawai_line", "").removeprefix("- ").rsplit(" (", 1)[0]
                    if overlay.get("pegawai_line")
                    else ""
                ),
                "PMID": pmid,
                "PNID": pnid,
                "Judul Posisi": overlay.get("title")
                or norm(row.get("Judul Posisi"))
                or norm(auto_row.get("Position Title")),
                "Perusahaan": overlay.get("company")
                or norm(row.get("Perusahaan"))
                or norm(auto_row.get("Company")),
                "company_in_id": overlay.get("company_in_id")
                or norm(row.get("company_in_id"))
                or norm(auto_row.get("Company ID")),
                "Unit / Group": overlay.get("unit")
                or norm(row.get("Unit / Group"))
                or norm(auto_row.get("Group / Unit")),
                "Pegawai": overlay.get("pegawai_line")
                or next(
                    (
                        line.strip()
                        for line in norm(row.get("Pegawai")).splitlines()
                        if f"({nipp})" in line
                    ),
                    f"- ({nipp})",
                ),
                "Sumber Path Mapping": path_source,
                "Folder Mapping": folder,
                "Workbook Title Mapping": workbook,
                "Worksheet Title Mapping": worksheet,
                "Status Inventory": status,
                "Inventory Workbook Path": best.source_workbook if best else "",
                "Inventory Sheet": best.sheet_name if best else "",
                "Inventory Position Name": best.position_name if best else "",
                "Kandidat Terdekat": format_candidates(
                    nearest if nearest else ([best] if best else [])
                ),
                "Confidence Otomatis": auto_confidence,
                "Status Konversi": siap,
                "Kategori Konversi": kategori,
                "Penjelasan Konversi": penjelasan,
                "Flag Mapping Awal": norm(row.get("Flag Mapping Awal")),
                "Status Mapping File2": norm(row.get("Status Mapping")),
            }

    missing_from_pemetaan = sorted(roster - set(confirmed_by_nipp))
    if missing_from_pemetaan:
        raise SystemExit(
            f"File2 Pemetaan missing roster NIPPs: {missing_from_pemetaan[:10]} "
            f"(count={len(missing_from_pemetaan)})"
        )

    confirmed_rows = [confirmed_by_nipp[nipp] for nipp in sorted(confirmed_by_nipp)]
    confirmed_status = Counter(row["Status Inventory"] for row in confirmed_rows)
    with_path = sum(
        1
        for row in confirmed_rows
        if row["Folder Mapping"] or row["Workbook Title Mapping"] or row["Worksheet Title Mapping"]
    )
    nipp_siap = sum(1 for row in confirmed_rows if row["Status Konversi"] == "siap")
    nipp_belum = len(confirmed_rows) - nipp_siap
    nipp_kategori = Counter(row["Kategori Konversi"] for row in confirmed_rows)

    # Identity-level (PMID/PNID) conversion readiness
    by_identity: dict[tuple[str, str], list[dict[str, Any]]] = defaultdict(list)
    for row in confirmed_rows:
        by_identity[identity_key(row["PMID"], row["PNID"])].append(row)

    identity_rows: list[dict[str, Any]] = []
    for (pmid, pnid), rows in sorted(
        by_identity.items(),
        key=lambda item: (item[0][0] or "zzz", item[0][1] or "zzz"),
    ):
        # Prefer a "siap" row; else first row (shared path per identity).
        preferred = next((row for row in rows if row["Status Konversi"] == "siap"), rows[0])
        mixed = len({row["Kategori Konversi"] for row in rows}) > 1
        identity_rows.append(
            {
                "PMID": pmid,
                "PNID": pnid,
                "Judul Posisi": preferred["Judul Posisi"],
                "Perusahaan": preferred["Perusahaan"],
                "company_in_id": preferred["company_in_id"],
                "Unit / Group": preferred["Unit / Group"],
                "Jumlah NIPP Roster": len(rows),
                "NIPP Sample": "; ".join(row["NIPP"] for row in rows[:8]),
                "Sumber Path Mapping": preferred["Sumber Path Mapping"],
                "Folder Mapping": preferred["Folder Mapping"],
                "Workbook Title Mapping": preferred["Workbook Title Mapping"],
                "Worksheet Title Mapping": preferred["Worksheet Title Mapping"],
                "Status Inventory": preferred["Status Inventory"],
                "Inventory Workbook Path": preferred["Inventory Workbook Path"],
                "Inventory Sheet": preferred["Inventory Sheet"],
                "Confidence Otomatis": preferred["Confidence Otomatis"],
                "Status Konversi": preferred["Status Konversi"],
                "Kategori Konversi": preferred["Kategori Konversi"],
                "Penjelasan Konversi": preferred["Penjelasan Konversi"],
                "Kategori Campuran antar NIPP": "ya" if mixed else "tidak",
            }
        )

    identity_siap = [row for row in identity_rows if row["Status Konversi"] == "siap"]
    identity_belum = [row for row in identity_rows if row["Status Konversi"] == "belum"]
    identity_kategori = Counter(row["Kategori Konversi"] for row in identity_rows)
    identity_siap_kategori = Counter(row["Kategori Konversi"] for row in identity_siap)
    identity_belum_kategori = Counter(row["Kategori Konversi"] for row in identity_belum)

    # --- Workbooks ---
    gap_path = args.output_dir / f"RW_Inventory_Gap_Report_Subholding_{stamp}.xlsx"
    confirmed_path = args.output_dir / f"Confirmed_Mapping_Subholding_2705_RW_{stamp}.xlsx"
    receipt_path = args.output_dir / f"RECONCILE_RECEIPT_{stamp}.md"
    latest_gap = args.output_dir / "RW_Inventory_Gap_Report_Subholding_LATEST.xlsx"
    latest_confirmed = args.output_dir / "Confirmed_Mapping_Subholding_2705_RW_LATEST.xlsx"

    gap_book = Workbook()
    guide = gap_book.active
    guide.title = "Baca Dulu"
    guide["A1"] = "Gap Report — R&W Path vs Inventory Kamus Group 2"
    guide["A1"].font = Font(name=BODY, size=16, bold=True, color="FFFFFF")
    guide["A1"].fill = PatternFill("solid", fgColor=NAVY)
    guide.merge_cells("A1:B1")
    guide["A3"] = (
        "Status Inventory:\n"
        f"- {STATUS_FOUND}: workbook + sheet ketemu (folder cocok bila diisi)\n"
        f"- {STATUS_FOLDER_MISMATCH}: sheet+workbook ketemu, folder R&W tidak selaras\n"
        f"- {STATUS_SHEET_MISSING}: workbook ada di inventory, worksheet title R&W tidak match "
        "(sering karena judul panjang vs tab 31 karakter)\n"
        f"- {STATUS_WB_MISSING}: workbook title R&W tidak ada di inventory\n"
        f"- {STATUS_EMPTY}: path R&W kosong\n\n"
        "Kandidat Terdekat = suggestion dari token overlap (bukan konfirmasi)."
    )
    guide["A3"].alignment = Alignment(wrap_text=True, vertical="top")
    guide.merge_cells("A3:B3")
    guide.row_dimensions[3].height = 140
    guide.column_dimensions["A"].width = 110

    summary = gap_book.create_sheet("Ringkasan")
    summary["A1"] = "Ringkasan Gap R&W vs Inventory"
    summary["A1"].fill = PatternFill("solid", fgColor=NAVY)
    summary["A1"].font = Font(name=BODY, size=16, bold=True, color="FFFFFF")
    meta = [
        ("Generated at", generated_at),
        ("File1", str(args.file1)),
        ("File2", str(args.file2)),
        ("Inventory", str(args.inventory)),
        ("Roster unique NIPP", len(roster)),
        ("Unique Mapping paths (File1+File2)", len(unique_gap_rows)),
        ("File1 Coverage rows", len(gap_coverage_rows)),
        ("File2 Belum-56 rows", len(gap_56_rows)),
    ]
    for offset, (label, value) in enumerate(meta, start=3):
        summary.cell(offset, 1, label).font = Font(name=BODY, bold=True, color=NAVY)
        summary.cell(offset, 2, value)
    start = 3 + len(meta) + 2
    summary.cell(start, 1, "Status (unique Mapping paths)").font = Font(name=BODY, bold=True)
    for offset, (status, count) in enumerate(sorted(status_unique.items()), start=start + 1):
        summary.cell(offset, 1, status)
        summary.cell(offset, 2, count)
    summary.column_dimensions["A"].width = 48
    summary.column_dimensions["B"].width = 100

    write_table(
        gap_book.create_sheet("Unique Paths"),
        "Unique Folder/Workbook/Worksheet dari sheet Mapping (File1+File2)",
        [
            "Sumber Review",
            "Folder R&W",
            "Workbook Title R&W",
            "Worksheet Title R&W",
            "Status Inventory",
            "Inventory Workbook Path",
            "Inventory Sheet",
            "Inventory Position Name",
            "Inventory Folder",
            "Kandidat Terdekat",
        ],
        unique_gap_rows,
        "UniquePathsTable",
    )
    write_table(
        gap_book.create_sheet("File1 Coverage"),
        "Setiap baris Position Coverage File1 + status inventory",
        [
            "Sumber Review",
            "PMID",
            "PNID",
            "Position Title",
            "Company",
            "Folder R&W",
            "Workbook Title R&W",
            "Worksheet Title R&W",
            "Status Inventory",
            "Inventory Workbook Path",
            "Inventory Sheet",
            "Kandidat Terdekat",
        ],
        gap_coverage_rows,
        "File1CoverageGapTable",
    )
    write_table(
        gap_book.create_sheet("File2 Belum56"),
        "Path R&W untuk 56 pekerja tambahan + status inventory",
        [
            "Sumber Review",
            "PMID",
            "PNID",
            "Judul Posisi",
            "Pegawai",
            "Folder R&W",
            "Workbook Title R&W",
            "Worksheet Title R&W",
            "Status Inventory",
            "Inventory Workbook Path",
            "Inventory Sheet",
            "Kandidat Terdekat",
        ],
        gap_56_rows,
        "File2Belum56GapTable",
    )
    for sheet_name, predicate in [
        ("WB Missing", lambda row: row["Status Inventory"] == STATUS_WB_MISSING),
        ("Sheet Missing", lambda row: row["Status Inventory"] == STATUS_SHEET_MISSING),
        ("Folder Mismatch", lambda row: row["Status Inventory"] == STATUS_FOLDER_MISMATCH),
    ]:
        subset = [row for row in unique_gap_rows if predicate(row)]
        write_table(
            gap_book.create_sheet(sheet_name),
            f"{sheet_name} — unique Mapping paths ({len(subset)})",
            [
                "Sumber Review",
                "Folder R&W",
                "Workbook Title R&W",
                "Worksheet Title R&W",
                "Status Inventory",
                "Kandidat Terdekat",
            ],
            subset,
            re.sub(r"[^A-Za-z0-9]", "", sheet_name) + "Tbl",
        )
    gap_book.save(gap_path)
    latest_gap.write_bytes(gap_path.read_bytes())

    conf_book = Workbook()
    cguide = conf_book.active
    cguide.title = "Baca Dulu"
    cguide["A1"] = "Confirmed Mapping Subholding 2.705 — R&W + Auto High + Konversi"
    cguide["A1"].font = Font(name=BODY, size=16, bold=True, color="FFFFFF")
    cguide["A1"].fill = PatternFill("solid", fgColor=NAVY)
    cguide.merge_cells("A1:B1")
    cguide["A3"] = (
        "Base pekerja: File2 Pemetaan (exact 2.705 NIPP).\n"
        "Path mapping: File2 Belum-56 → File1 Position Coverage → "
        "fallback Automated high_confidence bila path R&W kosong.\n"
        "Status Inventory = apakah Folder/Workbook/Worksheet resolve ke inventory.\n"
        "Status Konversi (per NIPP & per PMID/PNID) = siap dipakai untuk konversi kamus → formulir upload."
    )
    cguide["A3"].alignment = Alignment(wrap_text=True, vertical="top")
    cguide.merge_cells("A3:B3")
    cguide.row_dimensions[3].height = 120
    cguide.column_dimensions["A"].width = 110

    csum = conf_book.create_sheet("Ringkasan")
    csum["A1"] = "Ringkasan Confirmed 2.705 + Kesiapan Konversi"
    csum["A1"].fill = PatternFill("solid", fgColor=NAVY)
    csum["A1"].font = Font(name=BODY, size=16, bold=True, color="FFFFFF")
    cmeta = [
        ("Generated at", generated_at),
        ("Unique NIPP", len(confirmed_rows)),
        ("NIPP empty R&W sebelum auto overlay", empty_before_auto),
        ("NIPP diisi Automated high_confidence", auto_high_overlay_count),
        ("NIPP dengan path mapping (R&W atau auto)", with_path),
        ("NIPP masih empty setelah overlay", confirmed_status.get(STATUS_EMPTY, 0)),
        ("Identitas PMID/PNID unik", len(identity_rows)),
        ("Identitas SIAP konversi", len(identity_siap)),
        ("Identitas BELUM konversi", len(identity_belum)),
        ("NIPP SIAP konversi", nipp_siap),
        ("NIPP BELUM konversi", nipp_belum),
    ]
    for offset, (label, value) in enumerate(cmeta, start=3):
        csum.cell(offset, 1, label).font = Font(name=BODY, bold=True, color=NAVY)
        csum.cell(offset, 2, value)

    start = 3 + len(cmeta) + 2
    csum.cell(start, 1, "Status Inventory (NIPP)").font = Font(name=BODY, bold=True)
    for offset, (status, count) in enumerate(sorted(confirmed_status.items()), start=start + 1):
        csum.cell(offset, 1, status)
        csum.cell(offset, 2, count)

    start2 = start + len(confirmed_status) + 3
    csum.cell(start2, 1, "Kategori Konversi (PMID/PNID)").font = Font(name=BODY, bold=True)
    csum.cell(start2, 2, "Jumlah identitas")
    csum.cell(start2, 3, "Penjelasan")
    for offset, (kategori, count) in enumerate(sorted(identity_kategori.items()), start=start2 + 1):
        csum.cell(offset, 1, kategori)
        csum.cell(offset, 2, count)
        csum.cell(offset, 3, CONV_LABELS.get(kategori, ""))
        csum.cell(offset, 3).alignment = Alignment(wrap_text=True)
    csum.column_dimensions["A"].width = 44
    csum.column_dimensions["B"].width = 28
    csum.column_dimensions["C"].width = 90

    confirmed_headers = [
        "NIPP",
        "Nama",
        "PMID",
        "PNID",
        "Judul Posisi",
        "Perusahaan",
        "company_in_id",
        "Unit / Group",
        "Pegawai",
        "Sumber Path Mapping",
        "Folder Mapping",
        "Workbook Title Mapping",
        "Worksheet Title Mapping",
        "Status Inventory",
        "Inventory Workbook Path",
        "Inventory Sheet",
        "Inventory Position Name",
        "Kandidat Terdekat",
        "Confidence Otomatis",
        "Status Konversi",
        "Kategori Konversi",
        "Penjelasan Konversi",
        "Flag Mapping Awal",
        "Status Mapping File2",
    ]
    write_table(
        conf_book.create_sheet("Confirmed 2705"),
        "Mapping confirmed scoped roster Subholding 2.705 + status inventory + kesiapan konversi",
        confirmed_headers,
        confirmed_rows,
        "Confirmed2705Table",
    )

    identity_headers = [
        "PMID",
        "PNID",
        "Judul Posisi",
        "Perusahaan",
        "company_in_id",
        "Unit / Group",
        "Jumlah NIPP Roster",
        "NIPP Sample",
        "Sumber Path Mapping",
        "Folder Mapping",
        "Workbook Title Mapping",
        "Worksheet Title Mapping",
        "Status Inventory",
        "Inventory Workbook Path",
        "Inventory Sheet",
        "Confidence Otomatis",
        "Status Konversi",
        "Kategori Konversi",
        "Penjelasan Konversi",
        "Kategori Campuran antar NIPP",
    ]
    write_table(
        conf_book.create_sheet("Identitas Konversi"),
        f"Kesiapan konversi per identitas PMID/PNID ({len(identity_rows)} unik)",
        identity_headers,
        identity_rows,
        "IdentityConversionTable",
    )
    write_table(
        conf_book.create_sheet("Siap Konversi"),
        f"Identitas SIAP dikonversi ke formulir upload ({len(identity_siap)})",
        identity_headers,
        identity_siap,
        "ReadyIdentityTable",
    )
    write_table(
        conf_book.create_sheet("Belum Konversi"),
        f"Identitas BELUM bisa dikonversi ({len(identity_belum)})",
        identity_headers,
        identity_belum,
        "BlockedIdentityTable",
    )

    for sheet_name, predicate, title in [
        (
            "Inventory Found",
            lambda row: row["Status Inventory"] == STATUS_FOUND,
            "Path mapping resolve penuh ke inventory",
        ),
        (
            "Perlu Resolve Sheet",
            lambda row: row["Status Inventory"] == STATUS_SHEET_MISSING,
            "Workbook ada; worksheet title belum match",
        ),
        (
            "Workbook Missing",
            lambda row: row["Status Inventory"] == STATUS_WB_MISSING,
            "Workbook title tidak ada di inventory",
        ),
        (
            "Masih Empty",
            lambda row: row["Status Inventory"] == STATUS_EMPTY,
            "Path kosong setelah fallback automated high_confidence",
        ),
        (
            "Auto High Overlay",
            lambda row: str(row.get("Sumber Path Mapping", "")).startswith("Automated high_confidence"),
            "NIPP yang diisi dari automated high_confidence",
        ),
    ]:
        subset = [row for row in confirmed_rows if predicate(row)]
        write_table(
            conf_book.create_sheet(sheet_name),
            f"{title} ({len(subset)} NIPP)",
            confirmed_headers,
            subset,
            re.sub(r"[^A-Za-z0-9]", "", sheet_name)[:20] + "Tbl",
        )
    conf_book.save(confirmed_path)
    latest_confirmed.write_bytes(confirmed_path.read_bytes())

    readiness_json = {
        "generated_at": generated_at,
        "identity_total": len(identity_rows),
        "identity_siap": len(identity_siap),
        "identity_belum": len(identity_belum),
        "identity_siap_by_kategori": dict(identity_siap_kategori),
        "identity_belum_by_kategori": dict(identity_belum_kategori),
        "identity_all_by_kategori": dict(identity_kategori),
        "nipp_siap": nipp_siap,
        "nipp_belum": nipp_belum,
        "nipp_by_kategori": dict(nipp_kategori),
        "empty_before_auto": empty_before_auto,
        "auto_high_overlay_count": auto_high_overlay_count,
        "labels": CONV_LABELS,
    }
    readiness_path = args.output_dir / f"CONVERSION_READINESS_{stamp}.json"
    readiness_path.write_text(json.dumps(readiness_json, ensure_ascii=False, indent=2), encoding="utf-8")
    (args.output_dir / "CONVERSION_READINESS_LATEST.json").write_text(
        json.dumps(readiness_json, ensure_ascii=False, indent=2), encoding="utf-8"
    )

    def _fmt_cat(counter: Counter[str]) -> str:
        lines = []
        for key, count in sorted(counter.items(), key=lambda item: (-item[1], item[0])):
            lines.append(f"- `{key}`: **{count}** — {CONV_LABELS.get(key, '')}")
        return "\n".join(lines) if lines else "- (none)"

    receipt = f"""# R&W Subholding Reconcile Receipt

Generated: `{generated_at}`

## Inputs
- File1: `{args.file1}`
- File2: `{args.file2}`
- Inventory: `{args.inventory}`
- Roster: `{args.roster}` ({len(roster)} NIPP)
- Automated mapping: `{args.automated_mapping}`

## Gap report (unique Mapping paths File1+File2)
{json.dumps(dict(status_unique), ensure_ascii=False, indent=2)}

## Confirmed 2.705
- Rows: **{len(confirmed_rows)}**
- Empty R&W sebelum auto: **{empty_before_auto}**
- Overlay Automated high_confidence: **{auto_high_overlay_count}**
- With mapping path after overlay: **{with_path}**
- Inventory status: `{json.dumps(dict(confirmed_status), ensure_ascii=False)}`

## Conversion readiness — position identity (PMID/PNID)

Total identitas: **{len(identity_rows)}**

### SIAP dikonversi: **{len(identity_siap)}**
{_fmt_cat(identity_siap_kategori)}

### BELUM dikonversi: **{len(identity_belum)}**
{_fmt_cat(identity_belum_kategori)}

### NIPP view
- SIAP: **{nipp_siap}** / BELUM: **{nipp_belum}**

## Artifacts
- `{gap_path}`
- `{confirmed_path}`
- `{readiness_path}`
- Latest: `{latest_gap}`
- Latest: `{latest_confirmed}`
- Latest readiness: `{args.output_dir / "CONVERSION_READINESS_LATEST.json"}`
"""
    receipt_path.write_text(receipt, encoding="utf-8")

    print(
        json.dumps(
            {
                "generated_at": generated_at,
                "artifact_stamp": stamp,
                "roster_nipps": len(roster),
                "unique_mapping_paths": len(unique_gap_rows),
                "unique_path_status": dict(status_unique),
                "file1_coverage_status": dict(status_cov),
                "file2_56_status": dict(status_56),
                "confirmed_rows": len(confirmed_rows),
                "empty_before_auto": empty_before_auto,
                "auto_high_overlay_count": auto_high_overlay_count,
                "confirmed_with_mapping_path": with_path,
                "confirmed_status": dict(confirmed_status),
                "identity_total": len(identity_rows),
                "identity_siap": len(identity_siap),
                "identity_belum": len(identity_belum),
                "identity_siap_by_kategori": dict(identity_siap_kategori),
                "identity_belum_by_kategori": dict(identity_belum_kategori),
                "nipp_siap": nipp_siap,
                "nipp_belum": nipp_belum,
                "gap_report": str(gap_path),
                "confirmed_mapping": str(confirmed_path),
                "conversion_readiness": str(readiness_path),
            },
            ensure_ascii=False,
            indent=2,
        )
    )


if __name__ == "__main__":
    main()
