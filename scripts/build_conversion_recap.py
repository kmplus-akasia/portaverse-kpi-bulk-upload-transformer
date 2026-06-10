#!/usr/bin/env python3
"""Build a workbook that monitors KPI bulk conversion outputs."""

from __future__ import annotations

import argparse
import csv
import json
import re
from collections import Counter, defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter


UPLOAD_SHEET = "KPI Template"
TARGET_COMPANY_ID = "1"
TARGET_COMPANY_NAME = "PT Pelabuhan Indonesia (Persero)"
REPORT_SCOPE_DEFAULT = "Group 1 HO"


def norm(value: Any) -> str:
    return "" if value is None else str(value).strip()


def numeric(value: Any) -> float:
    text = norm(value).replace("%", "").replace(",", ".")
    if not text:
        return 0.0
    try:
        return float(text)
    except ValueError:
        return 0.0


def is_close(left: float, right: float, tolerance: float = 0.01) -> bool:
    return abs(left - right) <= tolerance


def yes_no(value: bool) -> str:
    return "YES" if value else "NO"


def report_title(scope: str, generated_at: datetime) -> str:
    return f"Kamus KPI Conversion Report - {scope} - {generated_at.strftime('%m/%d/%Y at %H.%M.%S')}"


def safe_sheet_title(value: str) -> str:
    title = "".join(ch for ch in value if ch not in r"[]:*?/\\").strip()
    return (title or "Sheet")[:31]


def find_workbooks(output_dirs: list[Path]) -> list[Path]:
    workbooks: dict[str, Path] = {}
    for output_dir in output_dirs:
        if not output_dir.exists():
            continue
        for path in output_dir.rglob("*.xlsx"):
            if path.name.startswith("~$"):
                continue
            lower_name = path.name.lower()
            if "conversion report" in lower_name or "conversion recap" in lower_name or "monitoring_recap" in lower_name:
                continue
            # Prefer the larger batch output when the same source workbook exists
            # in multiple run folders.
            key = path.name
            current = workbooks.get(key)
            if current is None or "pre_restructure_batch" in path.parts:
                workbooks[key] = path
    return sorted(workbooks.values(), key=lambda item: item.name.lower())


def report_path_for(workbook_path: Path) -> Path:
    return workbook_path.with_suffix(".report.csv")


def load_report_counts(report_path: Path) -> tuple[Counter, list[dict[str, str]]]:
    counts: Counter = Counter()
    rows: list[dict[str, str]] = []
    if not report_path.exists():
        counts["missing_report"] += 1
        return counts, rows
    with report_path.open(newline="", encoding="utf-8") as handle:
        for row in csv.DictReader(handle):
            rows.append(row)
            severity = norm(row.get("severity")).lower()
            message = norm(row.get("message"))
            counts[f"{severity}_count"] += 1
            if severity == "error" and "Polarity" in message:
                counts["missing_polarity"] += 1
            if "KAI formula may not be percentage-based" in message:
                counts["kai_formula_warning"] += 1
            if "Merged duplicate OUTPUT row" in message:
                counts["merged_duplicate_output"] += 1
            if "enum_issue category=" in message:
                counts["enum_issue"] += 1
            if "category=cross_column" in message:
                counts["cross_column_enum"] += 1
            if "mapping_corrected" in message:
                counts["mapping_corrected"] += 1
            if "mapping_conflict" in message:
                counts["mapping_conflict"] += 1
    return counts, rows


def read_upload_rows(workbook_path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    worksheet = workbook[UPLOAD_SHEET] if UPLOAD_SHEET in workbook.sheetnames else workbook.active
    headers = [norm(cell.value) for cell in next(worksheet.iter_rows(min_row=1, max_row=1))]
    rows: list[dict[str, Any]] = []
    for excel_row in worksheet.iter_rows(min_row=2, values_only=True):
        if not any(norm(value) for value in excel_row):
            continue
        rows.append(dict(zip(headers, excel_row)))
    return rows


def reference_status(reference_path: Path | None) -> str:
    if not reference_path:
        return "Not provided"
    if not reference_path.exists():
        return f"Missing: {reference_path}"
    payload = json.loads(reference_path.read_text(encoding="utf-8"))
    source = payload.get("source", {})
    return (
        f"{source.get('profile', 'unknown')} reference; "
        f"nomenclature={len(payload.get('rows', []))}; "
        f"position_master_org={len(payload.get('position_master_rows', []))}; "
        f"organizations={len(payload.get('organization_rows', []))}; "
        f"companies={len(payload.get('company_rows', []))}"
    )


def load_config_positions(config_path: Path | None) -> dict[str, dict[str, Any]]:
    if not config_path or not config_path.exists():
        return {}
    payload = json.loads(config_path.read_text(encoding="utf-8"))
    lookup: dict[str, dict[str, Any]] = {}
    for row in payload.get("positions", []):
        position_name = norm(row.get("position_name"))
        if position_name and position_name not in lookup:
            lookup[position_name] = row
    return lookup


def load_reference_indexes(reference_path: Path | None) -> tuple[dict[str, dict[str, Any]], dict[str, dict[str, Any]]]:
    if not reference_path or not reference_path.exists():
        return {}, {}
    payload = json.loads(reference_path.read_text(encoding="utf-8"))
    by_pnid: dict[str, dict[str, Any]] = {}
    for row in payload.get("rows", []):
        if norm(row.get("company_id")) != TARGET_COMPANY_ID:
            continue
        pnid = norm(row.get("cluster_id"))
        if pnid and pnid not in by_pnid:
            by_pnid[pnid] = row
    by_pmid: dict[str, dict[str, Any]] = {}
    for row in payload.get("position_master_rows", []):
        if norm(row.get("company_id")) != TARGET_COMPANY_ID:
            continue
        pmid = norm(row.get("position_master_id"))
        if not pmid:
            continue
        current = by_pmid.get(pmid)
        if current is None or (
            not current.get("is_position_organization_active") and row.get("is_position_organization_active")
        ):
            by_pmid[pmid] = row
    return by_pnid, by_pmid


def enrich_position_context(
    raw_position_title: str,
    pmid: str,
    pnid: str,
    config_lookup: dict[str, dict[str, Any]],
    ref_by_pnid: dict[str, dict[str, Any]],
    ref_by_pmid: dict[str, dict[str, Any]],
) -> dict[str, str]:
    config = config_lookup.get(raw_position_title, {})
    scope = norm(config.get("position_scope")) or ("non_structural" if pnid else "structural" if pmid else "")
    if scope == "non_structural":
        ref = ref_by_pnid.get(pnid, {})
        cluster_label = norm(config.get("cluster_label")) or norm(ref.get("cluster_label"))
        return {
            "Position Scope": "Non-structural",
            "Raw Kamus KPI Position Title": raw_position_title,
            "Portaverse Position Title": cluster_label or norm(config.get("portaverse_position_title")) or norm(ref.get("position_name")),
            "Cluster Label": cluster_label,
            "Group Master Name": norm(config.get("portaverse_group_name"))
            or norm(ref.get("active_group_name"))
            or norm(ref.get("group_name")),
            "Company Name": norm(config.get("portaverse_company_name"))
            or norm(ref.get("active_company_name"))
            or norm(ref.get("company_name")),
        }
    ref = ref_by_pmid.get(pmid, {})
    return {
        "Position Scope": "Structural" if scope == "structural" else "",
        "Raw Kamus KPI Position Title": raw_position_title,
        "Portaverse Position Title": norm(config.get("portaverse_position_title")) or norm(ref.get("position_name")),
        "Cluster Label": "",
        "Group Master Name": norm(config.get("portaverse_group_name")) or norm(ref.get("group_name")),
        "Company Name": norm(config.get("portaverse_company_name")) or norm(ref.get("company_name")),
    }


def parse_weight_audit_row(issue_row: dict[str, Any]) -> dict[str, Any]:
    message = norm(issue_row.get("Message"))
    parsed = dict(re.findall(r"([a-z_]+)=([^;.]*)", message))
    level = parsed.get("level", "")
    kpi_type = parsed.get("kpi_type", "")
    parent = parsed.get("parent", "")
    raw_total = numeric(parsed.get("raw_total", ""))
    output_total = numeric(parsed.get("output_total", ""))
    difference = round(output_total - 100, 4)
    if level == "position":
        check_scope = "Position total"
        what_to_check = f"Total {kpi_type} weight for this position"
    elif level == "impact":
        check_scope = "Output under Impact"
        what_to_check = "Total OUTPUT weight under this IMPACT parent"
    elif level == "output":
        check_scope = "KAI under Output"
        what_to_check = "Total KAI weight under this OUTPUT parent"
    else:
        check_scope = level
        what_to_check = f"Total {kpi_type} weight"
    cause = parsed.get("cause", "")
    action = "Adjust during allocation; not an upload blocker."
    if cause == "Converter Issue":
        action = "Review converter logic before upload."
    elif cause == "Dedupe Adjustment":
        action = "Review duplicate OUTPUT merge result, then adjust during allocation if needed."
    return {
        "Source Workbook": issue_row.get("Source Workbook", ""),
        "Sheet / Posisi": issue_row.get("Sheet / Posisi", ""),
        "Check Scope": check_scope,
        "What To Check": what_to_check,
        "Parent KPI": parent,
        "Expected Total": 100,
        "Raw Total": raw_total,
        "Converted Total": output_total,
        "Difference vs 100": difference,
        "Finding": "Total weight is not 100",
        "Cause": cause,
        "Recommended Action": action,
        "Raw Technical Message": message,
    }


def config_sheet_for_position(config_lookup: dict[str, dict[str, Any]], raw_position_title: str) -> str:
    return norm(config_lookup.get(raw_position_title, {}).get("sheet_name"))


def matching_position_keys(
    position_rows: dict[tuple[str, str, str], dict[str, Any]],
    config_lookup: dict[str, dict[str, Any]],
    sheet_name: str,
) -> list[tuple[str, str, str]]:
    matches = []
    for key in position_rows:
        raw_position_title = key[2]
        if raw_position_title == sheet_name or config_sheet_for_position(config_lookup, raw_position_title) == sheet_name:
            matches.append(key)
    return matches


def token_set(value: str) -> set[str]:
    return {token for token in re.sub(r"[^a-z0-9]+", " ", norm(value).lower()).split() if len(token) > 1}


def find_config_for_issue(
    config_rows: list[dict[str, Any]],
    source_name: str,
    sheet_name: str,
) -> dict[str, Any]:
    candidates = [row for row in config_rows if norm(row.get("Sheet / Posisi")) == sheet_name]
    if not candidates:
        return {}
    source_tokens = token_set(source_name)
    return max(
        candidates,
        key=lambda row: len(source_tokens & token_set(norm(row.get("Source Workbook")))),
    )


def is_kpi_item_finding(report_row: dict[str, str]) -> bool:
    record_type = norm(report_row.get("record_type")).upper()
    message = norm(report_row.get("message")).lower()
    if record_type not in {"IMPACT", "OUTPUT", "KAI"}:
        return False
    return any(
        token in message
        for token in [
            "missing required upload field",
            "missing polarity defaulted",
            "missing kai nature",
            "missing unit",
            "satuan",
            "formula may not be percentage-based",
        ]
    )


def kpi_item_finding_row(
    source_name: str,
    report_row: dict[str, str],
    config_rows: list[dict[str, Any]],
) -> dict[str, Any]:
    sheet_name = norm(report_row.get("sheet_name"))
    config = find_config_for_issue(config_rows, source_name, sheet_name)
    return {
        "KPI Owner": norm(config.get("Raw Kamus KPI Position Title")) or sheet_name,
        "KPI Workbook Source": norm(config.get("Source Workbook")) or source_name,
        "Converted Workbook": source_name,
        "Sheet / Posisi": sheet_name,
        "Raw Kamus KPI Group Name": norm(config.get("Raw Kamus KPI Group Name")),
        "KPI Title": norm(report_row.get("title")),
        "KPI Type": norm(report_row.get("record_type")).upper(),
        "Severity": norm(report_row.get("severity")).upper(),
        "Finding": norm(report_row.get("message")),
        "Source Row": norm(report_row.get("source_row")),
        "PMID": norm(config.get("Position Master ID")),
        "PNID": norm(config.get("Position Nomenklatur ID")),
        "Position Scope": norm(config.get("Position Scope")),
    }


def build_weight_summary_rows(weight_analysis_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    definitions = {
        "Position total": (
            "Total weight for one KPI type across one position.",
            "Check during allocation only.",
        ),
        "Output under Impact": (
            "Total OUTPUT weight attached to one IMPACT parent.",
            "Check during allocation only.",
        ),
        "KAI under Output": (
            "Total KAI weight attached to one OUTPUT parent.",
            "Check during allocation only.",
        ),
    }
    rows: list[dict[str, Any]] = []
    for scope, (meaning, action) in definitions.items():
        scoped = [row for row in weight_analysis_rows if row.get("Check Scope") == scope]
        rows.append(
            {
                "Check Scope": scope,
                "Meaning": meaning,
                "Finding Count": len(scoped),
                "Raw Data Issue": sum(1 for row in scoped if row.get("Cause") == "Raw Data Issue"),
                "Dedupe Adjustment": sum(1 for row in scoped if row.get("Cause") == "Dedupe Adjustment"),
                "Converter Issue": sum(1 for row in scoped if row.get("Cause") == "Converter Issue"),
                "Upload Blocker": "NO",
                "Recommended Action": action,
            }
        )
    rows.append(
        {
            "Check Scope": "How to use",
            "Meaning": "Use Weight Analysis only to see where managers need allocation adjustment.",
            "Finding Count": len(weight_analysis_rows),
            "Raw Data Issue": sum(1 for row in weight_analysis_rows if row.get("Cause") == "Raw Data Issue"),
            "Dedupe Adjustment": sum(1 for row in weight_analysis_rows if row.get("Cause") == "Dedupe Adjustment"),
            "Converter Issue": sum(1 for row in weight_analysis_rows if row.get("Cause") == "Converter Issue"),
            "Upload Blocker": "NO",
            "Recommended Action": "Do not block upload only because weight is not 100.",
        }
    )
    return rows


def finalize_position_checks(position_rows: dict[tuple[str, str, str], dict[str, Any]]) -> None:
    for record in position_rows.values():
        scope = norm(record.get("Position Scope"))
        pmid = norm(record.get("PMID"))
        pnid = norm(record.get("PNID"))
        company_ok = norm(record.get("Company Name")) in {"", TARGET_COMPANY_NAME}
        if scope == "Non-structural":
            id_ok = bool(pnid) and not pmid
        elif scope == "Structural":
            id_ok = bool(pmid) and not pnid
        else:
            id_ok = bool(pmid) ^ bool(pnid)
        total_ok = all(
            is_close(float(record.get(f"{kpi_type} Weight") or 0), 100.0)
            for kpi_type in ["IMPACT", "OUTPUT", "KAI"]
        )
        no_error = int(record.get("Errors") or 0) == 0
        record["PMID/PNID Valid"] = yes_no(id_ok and company_ok)
        record["Total Weight Valid"] = yes_no(total_ok)
        record["No Error"] = yes_no(no_error)
        record["Ready To Upload"] = yes_no(id_ok and company_ok and no_error)


def build_recap(
    output_dirs: list[Path],
    config_path: Path | None,
    toc_path: Path | None,
    reference_path: Path | None,
    report_scope: str = REPORT_SCOPE_DEFAULT,
) -> Workbook:
    generated_at_dt = datetime.now()
    generated_at = generated_at_dt.astimezone(timezone.utc).isoformat(timespec="seconds")
    workbook_title = report_title(report_scope, generated_at_dt)
    workbook_paths = find_workbooks(output_dirs)
    config_lookup = load_config_positions(config_path)
    ref_by_pnid, ref_by_pmid = load_reference_indexes(reference_path)

    position_rows: dict[tuple[str, str, str], dict[str, Any]] = {}
    workbook_rows: list[dict[str, Any]] = []
    issue_rows: list[dict[str, Any]] = []
    formula_warning_rows: list[dict[str, Any]] = []
    weight_analysis_rows: list[dict[str, Any]] = []
    kpi_item_finding_rows: list[dict[str, Any]] = []
    config_rows: list[dict[str, Any]] = []
    unresolved_position_rows: list[dict[str, Any]] = []
    if config_path and config_path.exists():
        payload = json.loads(config_path.read_text(encoding="utf-8"))
        for row in payload.get("positions", []):
            config_row = {
                "Source Workbook": norm(row.get("source_workbook")),
                "Sheet / Posisi": norm(row.get("sheet_name")),
                "Raw Kamus KPI Group Name": norm(row.get("group_name")),
                "Raw Kamus KPI Position Title": norm(row.get("position_name")),
                "Portaverse Position Title": norm(row.get("portaverse_position_title")),
                "Cluster Label": norm(row.get("cluster_label")),
                "Position Master ID": norm(row.get("position_master_id")),
                "Position Nomenklatur ID": norm(row.get("position_nomenclature_id")),
                "Position Scope": norm(row.get("position_scope")),
                "Group Master Name": norm(row.get("portaverse_group_name")),
                "Company Name": norm(row.get("portaverse_company_name")),
                "Direktorat": norm(row.get("directorate_name")),
            }
            config_rows.append(config_row)
            if not config_row["Position Master ID"] and not config_row["Position Nomenklatur ID"]:
                unresolved_position_rows.append(
                    {
                        "Source Workbook": config_row["Source Workbook"],
                        "Sheet / Posisi": config_row["Sheet / Posisi"],
                        "Raw Kamus KPI Group Name": config_row["Raw Kamus KPI Group Name"],
                        "Raw Kamus KPI Position Title": config_row["Raw Kamus KPI Position Title"],
                        "Direktorat": config_row["Direktorat"],
                        "Resolution Status": "Unresolved",
                        "Recommended Action": "Confirm PMID/PNID against production company_id=1, then add manual override mapping.",
                    }
                )

    for workbook_path in workbook_paths:
        upload_rows = read_upload_rows(workbook_path)
        report_counts, report_rows = load_report_counts(report_path_for(workbook_path))
        source_name = workbook_path.name
        workbook_counter = Counter()

        for row in upload_rows:
            raw_position_title = norm(row.get("Posisi"))
            pmid = norm(row.get("Position Master ID (Required)"))
            pnid = norm(row.get("Position Nomenklatur ID"))
            context = enrich_position_context(raw_position_title, pmid, pnid, config_lookup, ref_by_pnid, ref_by_pmid)
            position_key = (
                norm(row.get("Direktorat")),
                norm(row.get("Group")),
                raw_position_title,
            )
            record = position_rows.setdefault(
                position_key,
                {
                    "Direktorat": position_key[0],
                    "Raw Kamus KPI Group Name": position_key[1],
                    "Group Master Name": context["Group Master Name"],
                    "Raw Kamus KPI Position Title": context["Raw Kamus KPI Position Title"],
                    "Portaverse Position Title": context["Portaverse Position Title"],
                    "Cluster Label": context["Cluster Label"],
                    "Company Name": context["Company Name"],
                    "Position Scope": context["Position Scope"],
                    "Source Workbook": source_name,
                    "PMID": pmid,
                    "PNID": pnid,
                    "RKM Code ID": norm(row.get("RKM Code ID")),
                    "IMPACT Count": 0,
                    "OUTPUT Count": 0,
                    "KAI Count": 0,
                    "IMPACT Weight": 0.0,
                    "OUTPUT Weight": 0.0,
                    "KAI Weight": 0.0,
                    "Missing Polarity": 0,
                    "KAI Formula Warning": 0,
                    "Merged Duplicate OUTPUT": 0,
                    "Errors": 0,
                    "Warnings": 0,
                    "Infos": 0,
                    "PMID/PNID Valid": "",
                    "Total Weight Valid": "",
                    "No Error": "",
                    "Ready To Upload": "",
                    "Upload Status": "Not Uploaded",
                    "Upload Notes": "",
                },
            )
            kpi_type = norm(row.get("KPI Type")).upper()
            if kpi_type in {"IMPACT", "OUTPUT", "KAI"}:
                record[f"{kpi_type} Count"] += 1
                record[f"{kpi_type} Weight"] += numeric(row.get("Weight (%)"))
                workbook_counter[f"{kpi_type.lower()}_count"] += 1
                workbook_counter[f"{kpi_type.lower()}_weight"] += numeric(row.get("Weight (%)"))

        for report_row in report_rows:
            severity = norm(report_row.get("severity")).lower()
            sheet_name = norm(report_row.get("sheet_name"))
            message = norm(report_row.get("message"))
            kpi_title = norm(report_row.get("title"))
            matching_keys = matching_position_keys(position_rows, config_lookup, sheet_name)
            for key in matching_keys:
                record = position_rows[key]
                if severity == "error":
                    record["Errors"] += 1
                elif severity == "warning":
                    record["Warnings"] += 1
                elif severity == "info":
                    record["Infos"] += 1
                if severity == "error" and "Polarity" in message:
                    record["Missing Polarity"] += 1
                if "KAI formula may not be percentage-based" in message:
                    record["KAI Formula Warning"] += 1
                if "Merged duplicate OUTPUT row" in message:
                    record["Merged Duplicate OUTPUT"] += 1

            issue_row = {
                "Source Workbook": source_name,
                "Sheet / Posisi": sheet_name,
                "Severity": severity,
                "Record Type": norm(report_row.get("record_type")),
                "Source Row": norm(report_row.get("source_row")),
                "Title": kpi_title,
                "Message": message,
            }
            if severity == "error":
                issue_rows.append(issue_row)
            if "KAI formula may not be percentage-based" in message:
                formula_warning_rows.append(issue_row)
            if norm(report_row.get("record_type")) == "weight_audit":
                weight_analysis_rows.append(parse_weight_audit_row(issue_row))
            if is_kpi_item_finding(report_row):
                kpi_item_finding_rows.append(kpi_item_finding_row(source_name, report_row, config_rows))

        workbook_rows.append(
            {
                "Source Workbook": source_name,
                "Positions": len({norm(row.get("Posisi")) for row in upload_rows if norm(row.get("Posisi"))}),
                "Rows": len(upload_rows),
                "IMPACT Count": workbook_counter["impact_count"],
                "OUTPUT Count": workbook_counter["output_count"],
                "KAI Count": workbook_counter["kai_count"],
                "IMPACT Weight": workbook_counter["impact_weight"],
                "OUTPUT Weight": workbook_counter["output_weight"],
                "KAI Weight": workbook_counter["kai_weight"],
                "Errors": report_counts["error_count"],
                "Warnings": report_counts["warning_count"],
                "Infos": report_counts["info_count"],
                "Missing Polarity": report_counts["missing_polarity"],
                "KAI Formula Warning": report_counts["kai_formula_warning"],
                "Merged Duplicate OUTPUT": report_counts["merged_duplicate_output"],
                "Enum Issues": report_counts["enum_issue"],
                "Cross Column Enum": report_counts["cross_column_enum"],
                "Mapping Corrected": report_counts["mapping_corrected"],
                "Mapping Conflict": report_counts["mapping_conflict"],
                "Report Path": str(report_path_for(workbook_path)),
            }
        )

    finalize_position_checks(position_rows)
    configured_source_sheet_count = len(config_rows)
    resolved_source_sheet_count = sum(
        1 for row in config_rows if norm(row.get("Position Master ID")) or norm(row.get("Position Nomenklatur ID"))
    )
    unresolved_source_sheet_count = configured_source_sheet_count - resolved_source_sheet_count
    duplicate_position_key_count = resolved_source_sheet_count - len(position_rows)

    wb = Workbook()
    wb.remove(wb.active)
    write_sheet(
        wb,
        "Summary",
        [
            {"Metric": "Report Title", "Value": workbook_title},
            {"Metric": "Converted Kamus KPI Scope", "Value": report_scope},
            {"Metric": "Generated At UTC", "Value": generated_at},
            {"Metric": "Workbook Count", "Value": len(workbook_paths)},
            {"Metric": "Position Count", "Value": len(position_rows)},
            {"Metric": "Configured Source Sheet Count", "Value": configured_source_sheet_count},
            {"Metric": "Resolved Source Sheet Count", "Value": resolved_source_sheet_count},
            {"Metric": "Unresolved Source Sheet Count", "Value": unresolved_source_sheet_count},
            {"Metric": "Duplicate Position Key Count", "Value": duplicate_position_key_count},
            {"Metric": "Total IMPACT", "Value": sum(row["IMPACT Count"] for row in position_rows.values())},
            {"Metric": "Total OUTPUT", "Value": sum(row["OUTPUT Count"] for row in position_rows.values())},
            {"Metric": "Total KAI", "Value": sum(row["KAI Count"] for row in position_rows.values())},
            {"Metric": "Total Missing Polarity", "Value": sum(row["Missing Polarity"] for row in position_rows.values())},
            {"Metric": "Total KAI Formula Warning", "Value": sum(row["KAI Formula Warning"] for row in position_rows.values())},
            {"Metric": "Recommended Update Mode", "Value": "Automated via scripts/build_conversion_recap.py after every conversion run"},
            {"Metric": "Position Reference", "Value": reference_status(reference_path)},
            {"Metric": "Target Company", "Value": f"{TARGET_COMPANY_ID} - {TARGET_COMPANY_NAME}"},
        ],
    )
    write_sheet(wb, "Workbook Recap", workbook_rows)
    write_sheet(
        wb,
        "Position Recap",
        sorted(
            position_rows.values(),
            key=lambda row: (
                row["Direktorat"],
                row["Raw Kamus KPI Group Name"],
                row["Raw Kamus KPI Position Title"],
            ),
        ),
    )
    write_sheet(wb, "Weight Summary", build_weight_summary_rows(weight_analysis_rows))
    write_sheet(wb, "Weight Analysis", weight_analysis_rows)
    write_sheet(wb, "KPI Item Findings", kpi_item_finding_rows)
    write_sheet(
        wb,
        "Unresolved Positions",
        sorted(
            unresolved_position_rows,
            key=lambda row: (
                row["Source Workbook"],
                row["Sheet / Posisi"],
                row["Raw Kamus KPI Position Title"],
            ),
        ),
    )
    write_sheet(wb, "PNID Config", config_rows)
    write_sheet(
        wb,
        "DB Discrepancy",
        [
            {
                "Check": "Production reference export",
                "Status": "Available" if reference_path and reference_path.exists() else "Missing",
                "Detail": str(reference_path) if reference_path else "No reference path provided.",
            },
            {
                "Check": "Current converter mapping source",
                "Status": "Use offline JSON",
                "Detail": "Run converter with --mapping configs/production_position_reference.json so other devices do not need DB access.",
            },
        ],
    )
    if toc_path and toc_path.exists():
        append_toc_sheet(wb, toc_path)
    return wb


def write_sheet(wb: Workbook, title: str, rows: list[dict[str, Any]]) -> None:
    ws = wb.create_sheet(safe_sheet_title(title))
    ws.sheet_properties.tabColor = "1F4E78"
    if not rows:
        ws.append(["No data"])
        return
    headers = list(rows[0].keys())
    ws.append(headers)
    for row in rows:
        ws.append([row.get(header, "") for header in headers])
    style_table(ws)


def append_toc_sheet(wb: Workbook, toc_path: Path) -> None:
    try:
        source_wb = load_workbook(toc_path, read_only=True, data_only=True)
        source_ws = source_wb.active
        ws = wb.create_sheet("TOC Source")
        ws.sheet_properties.tabColor = "1F4E78"
        for row in source_ws.iter_rows(values_only=True):
            ws.append(list(row))
        style_table(ws)
    except Exception as exc:  # pragma: no cover - defensive for user-provided files
        write_sheet(wb, "TOC Source", [{"Status": "Could not read TOC", "Detail": str(exc)}])


def style_table(ws: Any) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    subheader_fill = PatternFill("solid", fgColor="D9EAF7")
    thin_gray = Side(style="thin", color="D9E2EC")
    border = Border(left=thin_gray, right=thin_gray, top=thin_gray, bottom=thin_gray)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 32
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if cell.row % 2 == 0:
                cell.fill = PatternFill("solid", fgColor="F8FAFC")
            if isinstance(cell.value, (int, float)):
                cell.number_format = "#,##0.00" if "Weight" in norm(ws.cell(1, cell.column).value) else "#,##0"
    if ws.title == "Summary":
        for cell in ws["A"]:
            cell.font = Font(bold=True, color="1F2937")
            if cell.row > 1:
                cell.fill = subheader_fill
    if ws.title == "Position Recap" and ws.max_row > 1:
        headers = {norm(cell.value): cell.column for cell in ws[1]}
        status_col = headers.get("Upload Status")
        if status_col:
            validation = DataValidation(
                type="list",
                formula1='"Not Uploaded,Uploaded,Failed,Hold"',
                allow_blank=True,
            )
            ws.add_data_validation(validation)
            validation.add(f"{get_column_letter(status_col)}2:{get_column_letter(status_col)}{ws.max_row}")
        for header in ["PMID/PNID Valid", "Total Weight Valid", "No Error", "Ready To Upload"]:
            col_index = headers.get(header)
            if not col_index:
                continue
            letter = get_column_letter(col_index)
            cell_range = f"{letter}2:{letter}{ws.max_row}"
            ws.conditional_formatting.add(
                cell_range,
                CellIsRule(operator="equal", formula=['"YES"'], fill=PatternFill("solid", fgColor="E2F0D9")),
            )
            ws.conditional_formatting.add(
                cell_range,
                CellIsRule(operator="equal", formula=['"NO"'], fill=PatternFill("solid", fgColor="FCE4D6")),
            )
    for column_index in range(1, ws.max_column + 1):
        header = norm(ws.cell(1, column_index).value)
        max_len = 10
        for cell in ws.iter_rows(min_col=column_index, max_col=column_index, values_only=True):
            max_len = max(max_len, len(norm(cell[0])) if cell else 0)
        preferred_widths = {
            "Source Workbook": 46,
            "KPI Workbook Source": 54,
            "Converted Workbook": 46,
            "Report Path": 54,
            "Message": 72,
            "Meaning": 60,
            "Finding": 72,
            "Raw Technical Message": 72,
            "What To Check": 44,
            "Parent KPI": 52,
            "Recommended Action": 48,
            "Title": 52,
            "KPI Title": 52,
            "Parent": 52,
            "Formula": 72,
            "Raw Kamus KPI Group Name": 34,
            "Group Master Name": 34,
            "Raw Kamus KPI Position Title": 34,
            "Portaverse Position Title": 38,
            "Upload Notes": 36,
        }
        ws.column_dimensions[get_column_letter(column_index)].width = preferred_widths.get(
            header,
            min(max_len + 2, 52),
        )


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--output-dir", action="append", type=Path, default=[])
    parser.add_argument("--config", type=Path, default=Path("configs/pre_restructure_positions.json"))
    parser.add_argument("--reference", type=Path, default=Path("configs/production_position_reference.json"))
    parser.add_argument("--toc", type=Path)
    parser.add_argument("--output", type=Path, required=True)
    parser.add_argument("--report-scope", default=REPORT_SCOPE_DEFAULT)
    args = parser.parse_args()

    output_dirs = args.output_dir or [Path("output/final_conversion")]
    wb = build_recap(output_dirs, args.config, args.toc, args.reference, args.report_scope)
    args.output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(args.output)
    print(f"Wrote {args.output}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
