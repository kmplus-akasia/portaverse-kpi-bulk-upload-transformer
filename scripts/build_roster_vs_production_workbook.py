#!/usr/bin/env python3
"""Build Roster vs Production workbook for Subholding 2,705 NIPPs.

Compares:
- position title
- organization title
- hierarchy / atasan org (STEXT_STO mid-segment vs group + ancestors)
- NIPP availability in Portaverse lookup / assignment
- production reference snapshot sync date (exported_at)
"""

from __future__ import annotations

import json
import re
import zipfile
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any
from xml.etree import ElementTree as ET

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

import position_mapping as pm

NS = {"m": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
REL = "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}"
OUT = Path("outputs/kamus-group2-subholding-roster-fresh-20260806")
ROSTER_PATH = Path(
    "outputs/kamus-group2-subholding-roster-mapping-20260806/source/REGIONAL dan SUBHOLDING.xlsx"
)
REFERENCE_PATH = Path("configs/production_position_reference.json")
NAVY, TEAL, PALE = "173651", "138074", "E9F1F8"
BODY = "Aptos"
FILLS = {
    "exact": "D9EAD3",
    "contains": "D9EAD3",
    "strong_overlap": "FFF2CC",
    "partial_overlap": "FCE4D6",
    "mismatch": "F4CCCC",
    "roster_org_empty": "E7E6E6",
    "production_org_empty": "E7E6E6",
    "n/a": "E7E6E6",
    "n/a_missing_prod": "E7E6E6",
    "missing": "E7E6E6",
    "in_lookup": "D9EAD3",
    "in_assignment_only": "FFF2CC",
    "absent": "F4CCCC",
    "nipp_mismatch_suspect": "FCE4D6",
}
STRONG = {"exact", "contains", "strong_overlap"}

HEADERS = [
    "NIPP",
    "Nama Roster",
    "Sheet Roster",
    "Judul Posisi Roster",
    "Judul Posisi Production",
    "Kesesuaian Judul",
    "Unit Org Roster (STEXT_STO)",
    "Group Production",
    "Kesesuaian Title Organisasi",
    "Path Ancestor Production",
    "Kesesuaian Hierarki/Atasan Org",
    "Status NIPP Portaverse",
    "PMID",
    "PNID",
    "Company Production",
    "Scope Production",
    "Tanggal Terakhir Sync Snapshot",
    "Catatan",
]


def norm(value: Any) -> str:
    return "" if value is None else str(value).strip()


def parse_stext(stext: str) -> tuple[str, str, str, str]:
    raw = re.sub(r"\s+", " ", (stext or "").replace("\n", " ").replace("\r", " ")).strip()
    parts = [part.strip() for part in raw.split("#")]
    title = parts[0] if parts else ""
    org = parts[1] if len(parts) >= 2 else ""
    flag = parts[2] if len(parts) >= 3 else ""
    return title, org, flag, raw


def load_shared_strings(archive: zipfile.ZipFile) -> list[str]:
    root = ET.fromstring(archive.read("xl/sharedStrings.xml"))
    out: list[str] = []
    for item in root.findall("m:si", NS):
        texts = [
            node.text or ""
            for node in item.iter("{http://schemas.openxmlformats.org/spreadsheetml/2006/main}t")
        ]
        out.append("".join(texts))
    return out


def cell_value(cell: ET.Element, shared: list[str]) -> Any:
    cell_type = cell.attrib.get("t")
    value_node = cell.find("m:v", NS)
    if value_node is None:
        return None
    raw = value_node.text
    return shared[int(raw)] if cell_type == "s" else raw


def col_index(cell_ref: str) -> int:
    match = re.match(r"([A-Z]+)", cell_ref)
    assert match is not None
    total = 0
    for char in match.group(1):
        total = total * 26 + (ord(char) - 64)
    return total - 1


def load_roster(path: Path) -> dict[str, dict[str, str]]:
    roster: dict[str, dict[str, str]] = {}
    with zipfile.ZipFile(path) as archive:
        shared = load_shared_strings(archive)
        workbook = ET.fromstring(archive.read("xl/workbook.xml"))
        rels = ET.fromstring(archive.read("xl/_rels/workbook.xml.rels"))
        rid = {rel.attrib["Id"]: rel.attrib["Target"] for rel in rels}
        for sheet in workbook.findall("m:sheets/m:sheet", NS):
            name = sheet.attrib["name"]
            if name not in {"SPTP", "SPMT", "SPSL", "SPJM"}:
                continue
            target = "xl/" + rid[sheet.attrib[f"{REL}id"]].lstrip("/")
            root = ET.fromstring(archive.read(target))
            rows: list[list[Any]] = []
            for row in root.findall("m:sheetData/m:row", NS):
                cells: dict[int, Any] = {}
                for cell in row.findall("m:c", NS):
                    cells[col_index(cell.attrib.get("r", "A1"))] = cell_value(cell, shared)
                if cells:
                    width = max(cells) + 1
                    rows.append([cells.get(index) for index in range(width)])
            header = [norm(value) for value in rows[0]]
            index = {name: idx for idx, name in enumerate(header)}
            for row in rows[1:]:
                if not row or all(value in (None, "") for value in row):
                    continue
                nipp = norm(row[index["PNALT_NEW"]])
                if not nipp:
                    continue
                title, org, flag, raw = parse_stext(str(row[index["STEXT_STO"]] or ""))
                roster[nipp] = {
                    "sheet": name,
                    "name": norm(row[index["CNAME"]]),
                    "roster_title": title,
                    "roster_org_unit": org,
                    "roster_flag": flag,
                    "stext_raw": raw,
                    "persa_text": norm(row[index["PERSA_TEXT"]]),
                }
    return roster


def title_match(left: str, right: str) -> str:
    a = pm.normalize_position_lookup(left)
    b = pm.normalize_position_lookup(right)
    if not a or not b:
        return "missing"
    if a == b:
        return "exact"
    if a in b or b in a:
        return "contains"
    ta, tb = set(a.split()), set(b.split())
    overlap = len(ta & tb) / max(len(ta), 1)
    if overlap >= 0.8:
        return "strong_overlap"
    if overlap >= 0.5:
        return "partial_overlap"
    return "mismatch"


def org_match(roster_org: str, group_name: str | None, ancestor_names: list[str]) -> str:
    roster_key = pm.normalize_position_lookup(roster_org)
    if not roster_key:
        return "roster_org_empty"
    candidates = [pm.normalize_position_lookup(group_name or "")]
    candidates.extend(pm.normalize_position_lookup(name) for name in ancestor_names)
    candidates = [item for item in candidates if item]
    if not candidates:
        return "production_org_empty"
    if any(roster_key == item for item in candidates):
        return "exact"
    if any(roster_key in item or item in roster_key for item in candidates):
        return "contains"
    roster_tokens = set(roster_key.split())
    best = 0.0
    for item in candidates:
        tokens = set(item.split())
        if roster_tokens and tokens:
            best = max(best, len(roster_tokens & tokens) / max(len(roster_tokens), 1))
    if best >= 0.8:
        return "strong_overlap"
    if best >= 0.5:
        return "partial_overlap"
    return "mismatch"


def expand_nipps(raw: Any) -> list[str]:
    if raw is None:
        return []
    values = raw if isinstance(raw, list) else [raw]
    out: list[str] = []
    for value in values:
        for part in re.split(r"\s*;\s*|\s*,\s*", str(value)):
            part = part.strip()
            if part:
                out.append(part)
    return out


def main() -> None:
    OUT.mkdir(parents=True, exist_ok=True)
    roster = load_roster(ROSTER_PATH)
    if len(roster) != 2705:
        raise SystemExit(f"Expected 2705 roster NIPPs, got {len(roster)}")

    reference = json.loads(REFERENCE_PATH.read_text(encoding="utf-8"))
    exported_at = norm((reference.get("source") or {}).get("exported_at"))
    indexes = pm.build_lookup_indexes(reference)
    org_by_id = {
        int(row["group_master_id"]): row
        for row in reference.get("organization_rows", [])
        if row.get("group_master_id") not in (None, "")
    }

    lookup_by_nipp: dict[str, list[pm.LookupCandidate]] = defaultdict(list)
    for candidate in [*indexes.structural, *indexes.non_structural]:
        for nipp in candidate.active_employee_nipps:
            lookup_by_nipp[str(nipp).strip()].append(candidate)

    assign_by_nipp: dict[str, list[dict[str, Any]]] = defaultdict(list)
    assign_by_name: dict[str, list[tuple[str, dict[str, Any]]]] = defaultdict(list)
    for row in reference.get("active_assignment_rows") or []:
        names = expand_nipps(row.get("active_employee_names"))
        nipps = expand_nipps(row.get("active_employee_nipps"))
        for nipp in nipps:
            assign_by_nipp[nipp].append(row)
        for idx, name in enumerate(names):
            key = name.upper()
            nipp = nipps[idx] if idx < len(nipps) else ""
            if key:
                assign_by_name[key].append((nipp, row))

    def org_ancestors(group_id: int) -> list[dict[str, Any]]:
        out: list[dict[str, Any]] = []
        seen: set[int] = set()
        current: int | None = group_id
        while current is not None and current not in seen:
            seen.add(current)
            row = org_by_id.get(current)
            if not row:
                break
            out.append(row)
            parent = row.get("parent_id")
            if parent in (None, ""):
                break
            try:
                current = int(parent)
            except (TypeError, ValueError):
                break
        return out

    rows_out: list[dict[str, Any]] = []
    title_stats: Counter[str] = Counter()
    org_stats: Counter[str] = Counter()
    hier_stats: Counter[str] = Counter()
    presence_stats: Counter[str] = Counter()

    for nipp, info in sorted(roster.items(), key=lambda item: item[0]):
        candidates = lookup_by_nipp.get(nipp) or []
        notes: list[str] = []
        if candidates:
            scored = []
            for candidate in candidates:
                title_score = pm._title_score(info["roster_title"], candidate)
                company_score = pm._context_score(info["persa_text"], candidate.company_name)
                scored.append((title_score * 0.7 + company_score * 0.3, candidate))
            scored.sort(key=lambda item: -item[0])
            chosen = scored[0][1]
            source_row = chosen.source_row or {}
            raw_group_id = source_row.get("group_master_id")
            try:
                group_id = int(raw_group_id) if raw_group_id not in (None, "") else None
            except (TypeError, ValueError):
                group_id = None
            ancestors = org_ancestors(group_id) if group_id is not None else []
            ancestor_names = [norm(row.get("group_name")) for row in ancestors]
            if not ancestor_names and chosen.group_ancestor_names:
                ancestor_names = [norm(name) for name in chosen.group_ancestor_names]
            title_label = title_match(info["roster_title"], chosen.title or "")
            org_label = org_match(info["roster_org_unit"], chosen.group_name, ancestor_names)
            hier_label = org_label
            presence = "in_lookup"
            prod_title = chosen.title or ""
            prod_group = chosen.group_name or ""
            prod_company = chosen.company_name or ""
            scope = chosen.scope
            pmid = chosen.position_master_id or ""
            pnid = chosen.position_nomenclature_id or ""
            ancestor_path = " > ".join([name for name in ancestor_names if name][:8])
        elif nipp in assign_by_nipp:
            presence = "in_assignment_only"
            assign = assign_by_nipp[nipp][0]
            pmid = str(assign.get("position_master_id") or "")
            pnid = ""
            group_id_raw = assign.get("group_master_id")
            try:
                group_id = int(group_id_raw) if group_id_raw not in (None, "") else None
            except (TypeError, ValueError):
                group_id = None
            ancestors = org_ancestors(group_id) if group_id is not None else []
            ancestor_names = [norm(row.get("group_name")) for row in ancestors]
            prod_group = norm((org_by_id.get(group_id) or {}).get("group_name")) if group_id else ""
            prod_title = ""
            prod_company = ""
            scope = "assignment_only"
            ancestor_path = " > ".join([name for name in ancestor_names if name][:8])
            title_label = "n/a_missing_prod"
            org_label = org_match(info["roster_org_unit"], prod_group, ancestor_names) if prod_group or ancestor_names else "n/a_missing_prod"
            hier_label = org_label
            notes.append("Ada di active_assignment_rows tetapi tidak di structural/non-structural lookup")
        else:
            presence = "absent"
            name_key = info["name"].upper()
            suspects = assign_by_name.get(name_key) or []
            exact_name = [(other, row) for other, row in suspects if other and other != nipp]
            if exact_name:
                presence = "nipp_mismatch_suspect"
                other_nipp, row = exact_name[0]
                notes.append(f"Nama sama di production dengan NIPP berbeda: {other_nipp}")
                pmid = str(row.get("position_master_id") or "")
            else:
                pmid = ""
            pnid = ""
            prod_title = ""
            prod_group = ""
            prod_company = ""
            scope = ""
            ancestor_path = ""
            title_label = "n/a_missing_prod"
            org_label = "n/a_missing_prod"
            hier_label = "n/a_missing_prod"

        title_stats[title_label] += 1
        org_stats[org_label] += 1
        hier_stats[hier_label] += 1
        presence_stats[presence] += 1
        rows_out.append(
            {
                "NIPP": nipp,
                "Nama Roster": info["name"],
                "Sheet Roster": info["sheet"],
                "Judul Posisi Roster": info["roster_title"],
                "Judul Posisi Production": prod_title,
                "Kesesuaian Judul": title_label,
                "Unit Org Roster (STEXT_STO)": info["roster_org_unit"],
                "Group Production": prod_group,
                "Kesesuaian Title Organisasi": org_label,
                "Path Ancestor Production": ancestor_path,
                "Kesesuaian Hierarki/Atasan Org": hier_label,
                "Status NIPP Portaverse": presence,
                "PMID": pmid,
                "PNID": pnid,
                "Company Production": prod_company,
                "Scope Production": scope,
                "Tanggal Terakhir Sync Snapshot": exported_at,
                "Catatan": "; ".join(notes),
            }
        )

    title_good = sum(1 for row in rows_out if row["Kesesuaian Judul"] in STRONG)
    org_good = sum(1 for row in rows_out if row["Kesesuaian Title Organisasi"] in STRONG)
    hier_good = sum(1 for row in rows_out if row["Kesesuaian Hierarki/Atasan Org"] in STRONG)
    both_good = sum(
        1
        for row in rows_out
        if row["Kesesuaian Judul"] in STRONG and row["Kesesuaian Title Organisasi"] in STRONG
    )

    book = Workbook()
    guide = book.active
    guide.title = "Baca Dulu"
    guide["A1"] = "Roster vs Production — Subholding 2.705"
    guide["A1"].font = Font(name=BODY, size=16, bold=True, color="FFFFFF")
    guide["A1"].fill = PatternFill("solid", fgColor=NAVY)
    guide.merge_cells("A1:B1")
    blocks = [
        (
            "Bagaimana atasan/hierarki dibaca?",
            "Roster tidak punya NIPP/nama atasan orang. "
            "Atasan org = segmen tengah STEXT_STO: <Judul> # <Unit Org> # <Flag>. "
            "Dibandingkan ke group_name production + rantai parent_id (Path Ancestor).",
        ),
        (
            "Status NIPP Portaverse",
            "in_lookup = ada di structural/non-structural lookup (siap mapping).\n"
            "in_assignment_only = ada assignment aktif tetapi posisi belum masuk lookup mapping.\n"
            "absent = tidak ada di snapshot reference.\n"
            "nipp_mismatch_suspect = nama sama, NIPP production berbeda.",
        ),
        (
            "Tanggal sync",
            f"Tanggal Terakhir Sync Snapshot = source.exported_at production reference: {exported_at}. "
            "Ini tanggal export snapshot, bukan last-sync per pegawai.",
        ),
    ]
    row_idx = 3
    for title, body in blocks:
        guide.cell(row_idx, 1, title).font = Font(name=BODY, bold=True, color=NAVY, size=12)
        guide.cell(row_idx, 1).fill = PatternFill("solid", fgColor=PALE)
        guide.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=2)
        row_idx += 1
        guide.cell(row_idx, 1, body).alignment = Alignment(wrap_text=True, vertical="top")
        guide.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=2)
        guide.row_dimensions[row_idx].height = 90
        row_idx += 2
    guide.column_dimensions["A"].width = 110

    summary = book.create_sheet("Ringkasan")
    summary["A1"] = "Ringkasan Roster vs Production (2.705)"
    summary["A1"].font = Font(name=BODY, size=16, bold=True, color="FFFFFF")
    summary["A1"].fill = PatternFill("solid", fgColor=NAVY)
    summary.merge_cells("A1:B1")
    metrics = [
        ("Unique NIPP roster", len(rows_out)),
        ("Tanggal sync snapshot", exported_at),
        ("NIPP in_lookup", presence_stats.get("in_lookup", 0)),
        ("NIPP in_assignment_only", presence_stats.get("in_assignment_only", 0)),
        ("NIPP absent", presence_stats.get("absent", 0)),
        ("NIPP mismatch suspect", presence_stats.get("nipp_mismatch_suspect", 0)),
        ("Judul cocok kuat", title_good),
        ("Title organisasi cocok kuat", org_good),
        ("Hierarki/atasan org cocok kuat", hier_good),
        ("Judul + org cocok kuat", both_good),
        ("Judul mismatch", title_stats.get("mismatch", 0)),
        ("Org mismatch", org_stats.get("mismatch", 0)),
        ("Unit roster kosong", org_stats.get("roster_org_empty", 0)),
    ]
    summary["A3"] = "Metric"
    summary["B3"] = "Nilai"
    for col in (1, 2):
        summary.cell(3, col).fill = PatternFill("solid", fgColor=TEAL)
        summary.cell(3, col).font = Font(name=BODY, bold=True, color="FFFFFF")
    for offset, (label, value) in enumerate(metrics, start=4):
        summary.cell(offset, 1, label)
        summary.cell(offset, 2, value)
    summary.column_dimensions["A"].width = 40
    summary.column_dimensions["B"].width = 50

    def write_detail(sheet_name: str, subset: list[dict[str, Any]], banner: str) -> None:
        sheet = book.create_sheet(sheet_name)
        sheet["A1"] = banner
        sheet["A1"].font = Font(name=BODY, size=14, bold=True, color="FFFFFF")
        sheet["A1"].fill = PatternFill("solid", fgColor=NAVY)
        sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(HEADERS))
        for col, header in enumerate(HEADERS, start=1):
            cell = sheet.cell(3, col, header)
            cell.fill = PatternFill("solid", fgColor=TEAL)
            cell.font = Font(name=BODY, bold=True, color="FFFFFF", size=10)
            cell.alignment = Alignment(wrap_text=True, horizontal="center")
        for r_idx, row in enumerate(subset, start=4):
            for c_idx, header in enumerate(HEADERS, start=1):
                cell = sheet.cell(r_idx, c_idx, row.get(header, ""))
                cell.font = Font(name=BODY, size=9)
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                label = norm(row.get(header))
                if header in {
                    "Kesesuaian Judul",
                    "Kesesuaian Title Organisasi",
                    "Kesesuaian Hierarki/Atasan Org",
                    "Status NIPP Portaverse",
                }:
                    color = FILLS.get(label)
                    if color:
                        cell.fill = PatternFill("solid", fgColor=color)
        end_row = 3 + max(len(subset), 1)
        if subset:
            table = Table(
                displayName=re.sub(r"[^A-Za-z0-9]", "", sheet_name)[:20] + "Tbl",
                ref=f"A3:{get_column_letter(len(HEADERS))}{end_row}",
            )
            table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
            sheet.add_table(table)
        widths = {
            "NIPP": 12,
            "Nama Roster": 28,
            "Sheet Roster": 10,
            "Judul Posisi Roster": 30,
            "Judul Posisi Production": 30,
            "Kesesuaian Judul": 14,
            "Unit Org Roster (STEXT_STO)": 28,
            "Group Production": 28,
            "Kesesuaian Title Organisasi": 16,
            "Path Ancestor Production": 40,
            "Kesesuaian Hierarki/Atasan Org": 16,
            "Status NIPP Portaverse": 18,
            "PMID": 10,
            "PNID": 10,
            "Company Production": 28,
            "Scope Production": 14,
            "Tanggal Terakhir Sync Snapshot": 24,
            "Catatan": 40,
        }
        for idx, header in enumerate(HEADERS, start=1):
            sheet.column_dimensions[get_column_letter(idx)].width = widths.get(header, 14)
        sheet.freeze_panes = "D4"

    write_detail("Detail per NIPP", rows_out, f"Detail kesesuaian roster vs production ({len(rows_out)} NIPP)")
    write_detail(
        "Judul Mismatch",
        [row for row in rows_out if row["Kesesuaian Judul"] == "mismatch"],
        "Judul mismatch",
    )
    write_detail(
        "Org Mismatch",
        [row for row in rows_out if row["Kesesuaian Title Organisasi"] == "mismatch"],
        "Title organisasi mismatch",
    )
    write_detail(
        "Hierarki Mismatch",
        [row for row in rows_out if row["Kesesuaian Hierarki/Atasan Org"] == "mismatch"],
        "Hierarki/atasan org mismatch",
    )
    write_detail(
        "NIPP Tidak di Lookup",
        [row for row in rows_out if row["Status NIPP Portaverse"] != "in_lookup"],
        "NIPP tidak di production lookup",
    )
    write_detail(
        "NIPP Suspect Mismatch",
        [row for row in rows_out if row["Status NIPP Portaverse"] == "nipp_mismatch_suspect"],
        "Suspect NIPP beda / nama sama",
    )

    output = OUT / "Roster_vs_Production_Subholding_2705_20260806.xlsx"
    book.save(output)
    md = OUT / "ROSTER_VS_PRODUCTION.md"
    md.write_text(
        f"""# Roster vs Production — Subholding 2.705

## Sync
- Tanggal snapshot: `{exported_at}`

## Ringkas
| Metric | Count |
| --- | ---: |
| in_lookup | {presence_stats.get('in_lookup', 0)} |
| in_assignment_only | {presence_stats.get('in_assignment_only', 0)} |
| absent | {presence_stats.get('absent', 0)} |
| nipp_mismatch_suspect | {presence_stats.get('nipp_mismatch_suspect', 0)} |
| Judul cocok kuat | {title_good} |
| Org cocok kuat | {org_good} |
| Hierarki cocok kuat | {hier_good} |

## Artifact
`{output}`
""",
        encoding="utf-8",
    )
    print(
        json.dumps(
            {
                "roster": len(rows_out),
                "exported_at": exported_at,
                "presence": dict(presence_stats),
                "title_good": title_good,
                "org_good": org_good,
                "hier_good": hier_good,
                "both_good": both_good,
                "output": str(output),
            },
            ensure_ascii=False,
            indent=2,
        )
    )


if __name__ == "__main__":
    main()
