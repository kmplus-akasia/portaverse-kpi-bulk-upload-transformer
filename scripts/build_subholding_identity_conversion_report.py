#!/usr/bin/env python3
"""Position-First Subholding workbook: kamus mapping + KPI upload tracker.

Sheets (1 row = 1 PMID/PNID identity):
  - Baca Dulu / Ringkasan
  - Pemetaan Kamus KPI
  - Detail Kamus KPI
  - Tracker Upload Kamus KPI

Status Kesiapan rules:
  - R&W inventory found / folder mismatch → siap
  - Automated high_confidence (empty R&W) → siap
  - Automated high_confidence when R&W is #N/A or fails but auto exists → siap
  - D subtype alias_atau_judul_mirip → siap (accepted)
"""

from __future__ import annotations

import argparse
import json
import re
import zipfile
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any
from xml.etree import ElementTree as ET

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo

import position_mapping as pm

NS = {"m": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
REL = "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}"
NAVY, TEAL = "173651", "138074"
GREEN, YELLOW, ORANGE, GRAY, RED = "D9EAD3", "FFF2CC", "FCE4D6", "E7E6E6", "F4CCCC"
BODY = "Aptos"
SCHEMA_VERSION = "subholding-kpi-tracker-v3"

SHARED_PREFIX = [
    "PMID",
    "PNID",
    "Judul Posisi",
    "Perusahaan",
    "Unit / Group",
    "Status Kesiapan",
    "Jumlah Pegawai",
    "Pegawai",
]

PEMETAAN_COLUMNS = SHARED_PREFIX + [
    "Alasan Status",
    "Treatment",
    "Detail Breakdown",
    "No.",
    "company_in_id",
    "Folder R&W (raw)",
    "Nama File R&W (raw)",
    "Worksheet R&W (raw)",
    "Sheet Inventory (resolved)",
    "File Kamus (usulan otomatis)",
    "Sheet Kamus (usulan otomatis)",
    "Keputusan Reviewer",
    "Catatan Reviewer",
]

DETAIL_COLUMNS = SHARED_PREFIX + [
    "File Kamus (resolved)",
    "Sheet Kamus (resolved)",
    "Nama File Formulir Upload KPI",
    "Judul Formulir Upload Versi Sebelumnya",
    "Link Google Sheet Formulir Upload KPI",
    "Jumlah IMPACT",
    "Jumlah OUTPUT",
    "Jumlah KAI",
    "Total Bobot IMPACT (%)",
    "Total Bobot OUTPUT (%)",
    "Total Bobot KAI (%)",
    "Bobot IMPACT OK?",
    "Bobot OUTPUT OK?",
    "Bobot KAI OK?",
    "Status Formulir",
]

TRACKER_COLUMNS = SHARED_PREFIX + [
    "Nama File Formulir Upload KPI",
    "Judul Formulir Upload Versi Sebelumnya",
    "Link Google Sheet Formulir Upload KPI",
    "Link Evidence Upload KPI",
    "Status Upload KPI",
    "Timestamp Upload KPI",
    "PIC Upload KPI",
    "Catatan Upload",
]

FORMULIR_COLUMNS = [
    "No.",
    "Nama File Formulir Upload KPI",
    "Judul Formulir Upload Versi Sebelumnya",
    "Link Google Sheet Formulir Upload KPI",
    "Jumlah Identity",
    "Jumlah Pegawai",
    "Daftar Identity (PMID/PNID — Judul Posisi)",
    "Daftar Pegawai (lengkap)",
    "Status Upload KPI",
    "Timestamp Upload KPI",
    "PIC Upload KPI",
    "Catatan Upload",
]

UPLOAD_STATUS_CHOICES = "Belum Upload,Sedang Proses,Sudah Upload,Revisi,Ditahan,Belum ada formulir"
REVIEWER_CHOICES = "OK,PERLU REVISI R&W,TOLAK,LAINNYA"


def norm(value: Any) -> str:
    return "" if value is None else str(value).strip()


def is_na(value: Any) -> bool:
    return norm(value).upper() in {"", "#N/A", "N/A", "NA", "-", "NULL"}


def numeric(value: Any) -> float:
    text = norm(value).replace("%", "").replace(",", ".")
    if not text:
        return 0.0
    try:
        return float(text)
    except ValueError:
        return 0.0


def is_close(left: float, right: float, tolerance: float = 0.01) -> bool:
    return abs(left - right) <= tolerance


def yes_no(value: bool) -> str:
    return "YES" if value else "NO"


def load_shared_strings(archive: zipfile.ZipFile) -> list[str]:
    if "xl/sharedStrings.xml" not in archive.namelist():
        return []
    root = ET.fromstring(archive.read("xl/sharedStrings.xml"))
    out: list[str] = []
    for item in root.findall("m:si", NS):
        texts = [
            node.text or ""
            for node in item.iter("{http://schemas.openxmlformats.org/spreadsheetml/2006/main}t")
        ]
        out.append("".join(texts))
    return out


def cell_value(cell: ET.Element, shared: list[str]) -> Any:
    cell_type = cell.attrib.get("t")
    inline = cell.find("m:is", NS)
    if inline is not None:
        return "".join(
            node.text or ""
            for node in inline.iter("{http://schemas.openxmlformats.org/spreadsheetml/2006/main}t")
        )
    value_node = cell.find("m:v", NS)
    if value_node is None:
        return None
    raw = value_node.text
    return shared[int(raw)] if cell_type == "s" else raw


def col_index(cell_ref: str) -> int:
    match = re.match(r"([A-Z]+)", cell_ref or "A1")
    assert match is not None
    total = 0
    for char in match.group(1):
        total = total * 26 + (ord(char) - 64)
    return total - 1


def read_sheet_rows(archive: zipfile.ZipFile, target: str, shared: list[str]) -> list[list[Any]]:
    root = ET.fromstring(archive.read(target))
    rows: list[list[Any]] = []
    for row in root.findall("m:sheetData/m:row", NS):
        cells: dict[int, Any] = {}
        for cell in row.findall("m:c", NS):
            cells[col_index(cell.attrib.get("r", "A1"))] = cell_value(cell, shared)
        if not cells:
            continue
        width = max(cells) + 1
        rows.append([cells.get(index) for index in range(width)])
    return rows


def workbook_sheet_map(archive: zipfile.ZipFile) -> dict[str, str]:
    workbook = ET.fromstring(archive.read("xl/workbook.xml"))
    rels = ET.fromstring(archive.read("xl/_rels/workbook.xml.rels"))
    rid = {rel.attrib["Id"]: rel.attrib["Target"] for rel in rels}
    out: dict[str, str] = {}
    for sheet in workbook.findall("m:sheets/m:sheet", NS):
        target = rid[sheet.attrib[f"{REL}id"]].lstrip("/")
        if not target.startswith("xl/"):
            target = f"xl/{target}"
        out[sheet.attrib["name"]] = target
    return out


def sheet_table(path: Path, sheet_name: str) -> list[dict[str, str]]:
    with zipfile.ZipFile(path) as archive:
        shared = load_shared_strings(archive)
        sheets = workbook_sheet_map(archive)
        if sheet_name not in sheets:
            raise KeyError(f"Sheet {sheet_name!r} not in {path.name}")
        rows = read_sheet_rows(archive, sheets[sheet_name], shared)
    header_idx = None
    header: list[str] = []
    for index, row in enumerate(rows[:12]):
        values = [norm(cell) for cell in row]
        if "PMID" in values and (
            "Kategori Konversi" in values or "Status Inventory" in values or "NIPP" in values
        ):
            header_idx = index
            header = values
            break
    if header_idx is None:
        raise RuntimeError(f"Header not found in {sheet_name}")
    out: list[dict[str, str]] = []
    for row in rows[header_idx + 1 :]:
        if not row or all(cell in (None, "") for cell in row):
            continue
        out.append(
            {
                name: (norm(row[idx]) if idx < len(row) else "")
                for idx, name in enumerate(header)
                if name
            }
        )
    return out


def token_score(left: str, right: str) -> float:
    a = set(pm.normalize_position_lookup(left).split())
    b = set(pm.normalize_position_lookup(right).split())
    if not a or not b:
        return 0.0
    return len(a & b) / max(len(a), len(b))


def parse_kandidat(text: str) -> list[tuple[str, str, str]]:
    out: list[tuple[str, str, str]] = []
    for line in norm(text).splitlines():
        line = line.strip().lstrip("- ").strip()
        if " / " not in line:
            continue
        workbook, rest = line.split(" / ", 1)
        sheet = rest.split(" (", 1)[0].strip()
        position = ""
        if "(" in rest and rest.endswith(")"):
            position = rest[rest.find("(") + 1 : -1]
        out.append((workbook.strip(), sheet, position))
    return out


def mismatch_subtype(rw_sheet: str, kandidat_list: list[tuple[str, str, str]]) -> tuple[str, str, float]:
    rw = norm(rw_sheet)
    if not kandidat_list:
        return "tidak_ada_kandidat", "", 0.0
    best_workbook, best_sheet, best_position, best_score = "", "", "", -1.0
    for workbook, sheet, position in kandidat_list:
        score = max(token_score(rw, sheet), token_score(rw, position))
        if score > best_score:
            best_workbook, best_sheet, best_position, best_score = workbook, sheet, position, score
    label = best_sheet
    if best_position and best_position != best_sheet:
        label = f"{best_sheet} ({best_position})"
    if len(best_sheet) == 31 and rw.casefold().startswith(best_sheet.casefold().rstrip()):
        return "truncasi_31_char", label, best_score
    if rw[:31].casefold() == best_sheet.casefold():
        return "truncasi_31_char", label, best_score
    if best_score >= 0.6:
        return "alias_atau_judul_mirip", label, best_score
    if best_score >= 0.35:
        return "kandidat_lemah_perlu_konfirmasi", label, best_score
    return "judul_berbeda_jauh", label, best_score


def workbook_basename(value: str) -> str:
    text = norm(value)
    if not text or is_na(text):
        return ""
    return Path(text).name


def style_header(cell) -> None:
    cell.fill = PatternFill("solid", fgColor=TEAL)
    cell.font = Font(name=BODY, bold=True, color="FFFFFF", size=10)
    cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")


def empty_metric() -> dict[str, Any]:
    return {
        "formulir_filename": "",
        "impact_count": 0,
        "output_count": 0,
        "kai_count": 0,
        "impact_weight": 0.0,
        "output_weight": 0.0,
        "kai_weight": 0.0,
    }


def identity_metric_key(pmid: str, pnid: str) -> tuple[str, str] | None:
    if pmid:
        return ("pmid", pmid)
    if pnid:
        return ("pnid", pnid)
    return None


def aggregate_upload_metrics(upload_ready_dir: Path) -> dict[tuple[str, str], dict[str, Any]]:
    """Scan upload-ready formulir; key by ('pmid'| 'pnid', id)."""
    metrics: dict[tuple[str, str], dict[str, Any]] = {}
    if not upload_ready_dir.exists():
        return metrics
    for path in sorted(upload_ready_dir.glob("*.xlsx")):
        if path.name.startswith("~$"):
            continue
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            worksheet = workbook[workbook.sheetnames[0]]
            rows_iter = worksheet.iter_rows(values_only=True)
            header_row = next(rows_iter, None)
            if not header_row:
                continue
            headers = [norm(cell) for cell in header_row]
            try:
                pmid_idx = headers.index("Position Master ID (Required)")
                pnid_idx = headers.index("Position Nomenklatur ID")
                type_idx = headers.index("KPI Type")
                weight_idx = headers.index("Weight (%)")
            except ValueError:
                continue
            for row in rows_iter:
                if not row or all(cell is None or str(cell).strip() == "" for cell in row[:5]):
                    continue
                pmid = norm(row[pmid_idx] if pmid_idx < len(row) else "")
                pnid = norm(row[pnid_idx] if pnid_idx < len(row) else "")
                key = identity_metric_key(pmid, pnid)
                if key is None:
                    continue
                record = metrics.setdefault(key, empty_metric())
                if not record["formulir_filename"]:
                    record["formulir_filename"] = path.name
                kpi_type = norm(row[type_idx] if type_idx < len(row) else "").upper()
                weight = numeric(row[weight_idx] if weight_idx < len(row) else None)
                if kpi_type == "IMPACT":
                    record["impact_count"] += 1
                    record["impact_weight"] += weight
                elif kpi_type == "OUTPUT":
                    record["output_count"] += 1
                    record["output_weight"] += weight
                elif kpi_type == "KAI":
                    record["kai_count"] += 1
                    record["kai_weight"] += weight
        finally:
            workbook.close()
    return metrics


def scan_formulir_identity_index(
    upload_ready_dir: Path,
) -> dict[str, list[tuple[str, str]]]:
    """Map formulir filename -> list of identity keys present in the file."""
    index: dict[str, list[tuple[str, str]]] = {}
    if not upload_ready_dir.exists():
        return index
    for path in sorted(upload_ready_dir.glob("*.xlsx")):
        if path.name.startswith("~$"):
            continue
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            worksheet = workbook[workbook.sheetnames[0]]
            rows_iter = worksheet.iter_rows(values_only=True)
            header_row = next(rows_iter, None)
            if not header_row:
                continue
            headers = [norm(cell) for cell in header_row]
            try:
                pmid_idx = headers.index("Position Master ID (Required)")
                pnid_idx = headers.index("Position Nomenklatur ID")
            except ValueError:
                continue
            seen: set[tuple[str, str]] = set()
            ordered: list[tuple[str, str]] = []
            for row in rows_iter:
                pmid = norm(row[pmid_idx] if pmid_idx < len(row) else "")
                pnid = norm(row[pnid_idx] if pnid_idx < len(row) else "")
                key = identity_metric_key(pmid, pnid)
                if key is None or key in seen:
                    continue
                seen.add(key)
                ordered.append(key)
            index[path.name] = ordered
        finally:
            workbook.close()
    return index


def load_unresolved_keys(allowlist_path: Path | None) -> set[tuple[str, str]]:
    if allowlist_path is None or not allowlist_path.exists():
        return set()
    payload = json.loads(allowlist_path.read_text(encoding="utf-8"))
    keys: set[tuple[str, str]] = set()
    for row in payload.get("held_inventory_unresolved", []):
        key = identity_metric_key(norm(row.get("pmid")), norm(row.get("pnid")))
        if key is not None:
            keys.add(key)
    return keys


def write_data_sheet(
    ws,
    title: str,
    columns: list[str],
    rows: list[dict[str, Any]],
    table_name: str,
    stamp: str,
    *,
    freeze_at: str = "G4",
    dropdowns: dict[str, str] | None = None,
    widths: dict[str, int] | None = None,
) -> None:
    ws.sheet_view.showGridLines = False
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columns))
    titled = f"{title}  ·  {stamp}"
    ws.cell(1, 1, titled).fill = PatternFill("solid", fgColor=NAVY)
    ws.cell(1, 1).font = Font(name=BODY, size=14, bold=True, color="FFFFFF")
    for col, header in enumerate(columns, start=1):
        style_header(ws.cell(3, col, header))

    status_col = columns.index("Status Kesiapan") + 1 if "Status Kesiapan" in columns else None
    treatment_col = columns.index("Treatment") + 1 if "Treatment" in columns else None
    formulir_col = columns.index("Status Formulir") + 1 if "Status Formulir" in columns else None
    upload_col = columns.index("Status Upload KPI") + 1 if "Status Upload KPI" in columns else None

    for r_idx, row in enumerate(rows, start=4):
        for c_idx, header in enumerate(columns, start=1):
            value = row.get(header, "")
            cell = ws.cell(r_idx, c_idx, value)
            cell.font = Font(name=BODY, size=9)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if header.startswith("Total Bobot") and isinstance(value, (int, float)):
                cell.number_format = "0.00"
        if status_col is not None:
            status = norm(row.get("Status Kesiapan"))
            if status == "Siap konversi":
                ws.cell(r_idx, status_col).fill = PatternFill("solid", fgColor=GREEN)
            elif status.startswith("Belum"):
                ws.cell(r_idx, status_col).fill = PatternFill("solid", fgColor=ORANGE)
        if treatment_col is not None:
            treatment = norm(row.get("Treatment"))
            if treatment.startswith("D."):
                ws.cell(r_idx, treatment_col).fill = PatternFill("solid", fgColor=YELLOW)
            elif treatment.startswith("C."):
                ws.cell(r_idx, treatment_col).fill = PatternFill("solid", fgColor=RED)
        if formulir_col is not None:
            formulir = norm(row.get("Status Formulir"))
            if formulir == "Ada di batch upload":
                ws.cell(r_idx, formulir_col).fill = PatternFill("solid", fgColor=GREEN)
            elif formulir == "Ditahan unresolved":
                ws.cell(r_idx, formulir_col).fill = PatternFill("solid", fgColor=RED)
            elif formulir.startswith("Belum"):
                ws.cell(r_idx, formulir_col).fill = PatternFill("solid", fgColor=GRAY)
        if upload_col is not None:
            upload_status = norm(row.get("Status Upload KPI"))
            if upload_status == "Sudah Upload":
                ws.cell(r_idx, upload_col).fill = PatternFill("solid", fgColor=GREEN)
            elif upload_status == "Belum Upload":
                ws.cell(r_idx, upload_col).fill = PatternFill("solid", fgColor=YELLOW)
            elif upload_status in {"Revisi", "Ditahan"}:
                ws.cell(r_idx, upload_col).fill = PatternFill("solid", fgColor=ORANGE)
            elif upload_status == "Belum ada formulir":
                ws.cell(r_idx, upload_col).fill = PatternFill("solid", fgColor=GRAY)

    end_row = 3 + max(len(rows), 1)
    if rows:
        table = Table(
            displayName=table_name,
            ref=f"A3:{get_column_letter(len(columns))}{end_row}",
        )
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
        ws.add_table(table)

    default_widths = {
        "PMID": 10,
        "PNID": 10,
        "Judul Posisi": 28,
        "Perusahaan": 28,
        "Unit / Group": 24,
        "Status Kesiapan": 16,
        "Jumlah Pegawai": 12,
        "Pegawai": 32,
        "Alasan Status": 34,
        "Treatment": 40,
        "Detail Breakdown": 42,
        "No.": 6,
        "company_in_id": 12,
        "Jumlah NIPP": 10,
        "Pegawai (sample)": 28,
        "Pegawai": 32,
        "Jumlah Pegawai": 12,
        "Folder R&W (raw)": 36,
        "Nama File R&W (raw)": 34,
        "Worksheet R&W (raw)": 30,
        "Sheet Inventory (resolved)": 28,
        "File Kamus (usulan otomatis)": 34,
        "Sheet Kamus (usulan otomatis)": 28,
        "File Kamus (resolved)": 36,
        "Sheet Kamus (resolved)": 28,
        "Keputusan Reviewer": 18,
        "Catatan Reviewer": 24,
        "Nama File Formulir Upload KPI": 42,
        "Link Google Sheet Formulir Upload KPI": 36,
        "Link Evidence Upload KPI": 36,
        "Jumlah IMPACT": 12,
        "Jumlah OUTPUT": 12,
        "Jumlah KAI": 12,
        "Total Bobot IMPACT (%)": 14,
        "Total Bobot OUTPUT (%)": 14,
        "Total Bobot KAI (%)": 14,
        "Bobot IMPACT OK?": 12,
        "Bobot OUTPUT OK?": 12,
        "Bobot KAI OK?": 12,
        "Judul Formulir Upload Versi Sebelumnya": 42,
        "Daftar Identity (PMID/PNID — Judul Posisi)": 48,
        "Daftar Pegawai (lengkap)": 48,
        "Jumlah Identity": 12,
        "Status Formulir": 20,
        "Status Upload KPI": 16,
        "Timestamp Upload KPI": 20,
        "PIC Upload KPI": 18,
        "Catatan Upload": 28,
    }
    merged_widths = {**default_widths, **(widths or {})}
    for idx, header in enumerate(columns, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = merged_widths.get(header, 16)
    ws.row_dimensions[3].height = 34
    ws.freeze_panes = freeze_at

    if rows and dropdowns:
        for header, formula in dropdowns.items():
            if header not in columns:
                continue
            validation = DataValidation(type="list", formula1=f'"{formula}"', allow_blank=True)
            validation.error = "Pilih dari daftar"
            validation.errorTitle = header
            ws.add_data_validation(validation)
            col_letter = get_column_letter(columns.index(header) + 1)
            validation.add(f"{col_letter}4:{col_letter}{end_row}")


def classify_row(
    *,
    artifact_status: str,
    artifact_kategori: str,
    pmid: str,
    pnid: str,
    rw_folder: str,
    rw_workbook: str,
    rw_worksheet: str,
    inventory_sheet: str,
    auto_confidence: str,
    auto_workbook: str,
    auto_sheet: str,
    kandidat_text: str,
    high_keys: set[tuple[str, str]],
) -> dict[str, str]:
    key = (pmid, pnid)
    subtype, best_candidate, best_score = mismatch_subtype(rw_worksheet, parse_kandidat(kandidat_text))

    if artifact_kategori == "siap__red_white_inventory_found":
        return {
            "Status Kesiapan": "Siap konversi",
            "Alasan Status": "R&W resolve ke inventory",
            "Treatment": "(siap) Tidak perlu treatment",
            "Detail Breakdown": f"Inventory sheet: {inventory_sheet}",
            "Sheet Inventory (resolved)": inventory_sheet,
            "Subtype": "",
            "Best Candidate": "",
            "Best Score": "",
        }
    if artifact_kategori == "siap__red_white_folder_mismatch":
        return {
            "Status Kesiapan": "Siap konversi",
            "Alasan Status": "R&W sheet ketemu (folder mismatch diabaikan)",
            "Treatment": "(siap) Tidak perlu treatment",
            "Detail Breakdown": f"Inventory sheet: {inventory_sheet}",
            "Sheet Inventory (resolved)": inventory_sheet,
            "Subtype": "",
            "Best Candidate": "",
            "Best Score": "",
        }
    if artifact_kategori == "siap__automated_high_confidence":
        return {
            "Status Kesiapan": "Siap konversi",
            "Alasan Status": "Automated high_confidence (path R&W kosong)",
            "Treatment": "(siap) Tidak perlu treatment",
            "Detail Breakdown": f"Auto: {auto_workbook} / {auto_sheet}",
            "Sheet Inventory (resolved)": auto_sheet,
            "Subtype": "",
            "Best Candidate": "",
            "Best Score": "",
        }

    if key in high_keys and artifact_kategori in {
        "belum__red_white_workbook_missing",
        "belum__red_white_sheet_unresolved",
    }:
        if (is_na(rw_workbook) and is_na(rw_worksheet)) or (
            artifact_kategori == "belum__red_white_sheet_unresolved"
            and auto_confidence == pm.HIGH_CONFIDENCE
        ):
            return {
                "Status Kesiapan": "Siap konversi",
                "Alasan Status": "Automated high_confidence (R&W #N/A/gagal → auto)",
                "Treatment": "(siap) Tidak perlu treatment",
                "Detail Breakdown": f"Auto: {auto_workbook} / {auto_sheet}",
                "Sheet Inventory (resolved)": auto_sheet,
                "Subtype": "",
                "Best Candidate": "",
                "Best Score": "",
            }

    if (
        artifact_kategori == "belum__red_white_sheet_unresolved"
        and subtype == "alias_atau_judul_mirip"
        and best_candidate
    ):
        return {
            "Status Kesiapan": "Siap konversi",
            "Alasan Status": "R&W sheet alias/urutan kata mirip — accepted",
            "Treatment": "(siap) Alias/urutan kata diterima",
            "Detail Breakdown": (
                f"R&W «{rw_worksheet}» ≈ inventory «{best_candidate}» "
                f"(score {best_score:.2f})"
            ),
            "Sheet Inventory (resolved)": best_candidate.split(" (", 1)[0],
            "Subtype": subtype,
            "Best Candidate": best_candidate,
            "Best Score": f"{best_score:.3f}",
        }

    if artifact_kategori == "belum__missing_production_identity":
        return {
            "Status Kesiapan": "Belum konversi",
            "Alasan Status": "Tanpa PMID/PNID produksi",
            "Treatment": "A. Lengkapi identitas produksi (PMID/PNID)",
            "Detail Breakdown": "Stub roster — belum punya identity master/nomenclature",
            "Sheet Inventory (resolved)": "",
            "Subtype": "",
            "Best Candidate": best_candidate,
            "Best Score": f"{best_score:.3f}" if best_score else "",
        }
    if artifact_kategori == "belum__red_white_sheet_unresolved":
        subtype_label = {
            "truncasi_31_char": "truncasi 31 karakter",
            "kandidat_lemah_perlu_konfirmasi": "kandidat lemah — perlu konfirmasi",
            "judul_berbeda_jauh": "judul berbeda jauh",
            "tidak_ada_kandidat": "tidak ada kandidat di workbook",
        }.get(subtype, subtype)
        return {
            "Status Kesiapan": "Belum konversi",
            "Alasan Status": "Review R&W: worksheet belum resolve",
            "Treatment": "D. Resolve judul sheet R&W ↔ tab inventory",
            "Detail Breakdown": (
                f"Subtype: {subtype_label}. "
                f"R&W «{rw_worksheet}» → kandidat «{best_candidate or '-'}»"
                + (f" (score {best_score:.2f})" if best_score else "")
            ),
            "Sheet Inventory (resolved)": "",
            "Subtype": subtype,
            "Best Candidate": best_candidate,
            "Best Score": f"{best_score:.3f}" if best_score else "",
        }
    if artifact_kategori == "belum__red_white_workbook_missing":
        detail = (
            "R&W mengisi #N/A"
            if is_na(rw_workbook)
            else f"File R&W «{rw_workbook}» tidak ada di inventory"
        )
        return {
            "Status Kesiapan": "Belum konversi",
            "Alasan Status": "Review R&W: workbook tidak di inventory",
            "Treatment": "C. Alias / tambah workbook ke inventory",
            "Detail Breakdown": detail,
            "Sheet Inventory (resolved)": "",
            "Subtype": "",
            "Best Candidate": "",
            "Best Score": "",
        }
    if artifact_kategori == "belum__automated_mapping_conflict":
        return {
            "Status Kesiapan": "Belum konversi",
            "Alasan Status": "Tanpa R&W resolve; auto konflik kandidat",
            "Treatment": "E. Putuskan konflik kandidat automated",
            "Detail Breakdown": f"Auto conflict — usulan: {auto_workbook} / {auto_sheet}",
            "Sheet Inventory (resolved)": "",
            "Subtype": "",
            "Best Candidate": "",
            "Best Score": "",
        }
    if artifact_kategori == "belum__automated_no_candidate":
        return {
            "Status Kesiapan": "Belum konversi",
            "Alasan Status": "Tanpa R&W resolve; auto no_candidate",
            "Treatment": "F. Cari/buat sheet kamus",
            "Detail Breakdown": "Tidak ada sheet kamus yang lolos threshold",
            "Sheet Inventory (resolved)": "",
            "Subtype": "",
            "Best Candidate": "",
            "Best Score": "",
        }
    if artifact_kategori == "belum__automated_low_confidence":
        return {
            "Status Kesiapan": "Belum konversi",
            "Alasan Status": "Tanpa R&W resolve; auto low_confidence",
            "Treatment": "G. Review manual kandidat lemah",
            "Detail Breakdown": f"Usulan lemah: {auto_workbook} / {auto_sheet}",
            "Sheet Inventory (resolved)": "",
            "Subtype": "",
            "Best Candidate": "",
            "Best Score": "",
        }
    return {
        "Status Kesiapan": "Belum konversi",
        "Alasan Status": artifact_status or "unknown",
        "Treatment": "H. Lainnya",
        "Detail Breakdown": artifact_kategori,
        "Sheet Inventory (resolved)": "",
        "Subtype": "",
        "Best Candidate": "",
        "Best Score": "",
    }


def resolve_kamus_paths(row: dict[str, Any]) -> tuple[str, str]:
    sheet = norm(row.get("Sheet Inventory (resolved)")) or norm(row.get("Sheet Kamus (usulan otomatis)"))
    file_name = norm(row.get("Nama File R&W (raw)"))
    if is_na(file_name) or not file_name:
        file_name = norm(row.get("File Kamus (usulan otomatis)"))
    if row["Status Kesiapan"] != "Siap konversi":
        # Prefer auto usulan when unresolved R&W
        if norm(row.get("File Kamus (usulan otomatis)")):
            file_name = norm(row.get("File Kamus (usulan otomatis)"))
        if norm(row.get("Sheet Kamus (usulan otomatis)")) and not sheet:
            sheet = norm(row.get("Sheet Kamus (usulan otomatis)"))
    return file_name, sheet


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--confirmed",
        type=Path,
        default=Path(
            "outputs/kamus-group2-subholding-rw-reconcile-20260807/"
            "Confirmed_Mapping_Subholding_2705_RW_LATEST.xlsx"
        ),
    )
    parser.add_argument(
        "--automated-mapping",
        type=Path,
        default=Path(
            "outputs/kamus-group2-subholding-roster-fresh-20260806/"
            "subholding_roster_fresh_mapping_LATEST.json"
        ),
    )
    parser.add_argument(
        "--upload-ready-dir",
        type=Path,
        default=Path(
            "outputs/kamus-group2-subholding-upload-ready-siap-20260807_impact10/upload-ready"
        ),
    )
    parser.add_argument(
        "--upload-config",
        type=Path,
        default=Path(
            "outputs/kamus-group2-subholding-upload-ready-siap-20260807_impact10/"
            "subholding_siap_upload_config_20260807_143506.json"
        ),
    )
    parser.add_argument(
        "--upload-allowlist",
        type=Path,
        default=Path(
            "outputs/kamus-group2-subholding-upload-ready-siap-20260807_impact10/"
            "siap_allowlist_20260807_143506.json"
        ),
    )
    parser.add_argument(
        "--previous-upload-ready-dir",
        type=Path,
        default=Path(
            "outputs/kamus-group2-subholding-upload-ready-siap-20260807_sheetreuse/upload-ready"
        ),
        help="Previously uploaded formulir batch (sheetreuse) for prior filename mapping",
    )
    parser.add_argument(
        "--formulir-upload-ready-dir",
        type=Path,
        default=Path(
            "outputs/kamus-group2-subholding-upload-ready-siap-20260807_impact10/upload-ready"
        ),
        help="Formulir set featured on sheet Formulir Upload KPI (full upload-ready batch)",
    )
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=Path("outputs/kamus-group2-subholding-rw-reconcile-20260807"),
    )
    args = parser.parse_args()

    stamp = datetime.now().astimezone().strftime("%Y%m%d_%H%M%S")
    generated_at = datetime.now().astimezone().isoformat(timespec="seconds")
    args.output_dir.mkdir(parents=True, exist_ok=True)

    identity_src = sheet_table(args.confirmed, "Identitas Konversi")
    nipp_src = sheet_table(args.confirmed, "Confirmed 2705")
    kandidat_by_identity: dict[tuple[str, str], str] = {}
    pegawai_by_identity: dict[tuple[str, str], list[str]] = defaultdict(list)
    for row in nipp_src:
        key = (norm(row.get("PMID")), norm(row.get("PNID")))
        if row.get("Kandidat Terdekat") and key not in kandidat_by_identity:
            kandidat_by_identity[key] = row["Kandidat Terdekat"]
        pegawai = norm(row.get("Pegawai")) or norm(row.get("Nama"))
        nipp = norm(row.get("NIPP"))
        line = pegawai if pegawai else (f"- ({nipp})" if nipp else "")
        if line and line not in pegawai_by_identity[key]:
            pegawai_by_identity[key].append(line)

    fresh = json.loads(args.automated_mapping.read_text(encoding="utf-8"))
    high_keys: set[tuple[str, str]] = set()
    auto_by_key: dict[tuple[str, str], dict[str, Any]] = {}
    for row in fresh.get("rows", []):
        key = (norm(row.get("PMID")), norm(row.get("PNID")))
        auto_by_key[key] = row
        if row.get("Confidence Label") == pm.HIGH_CONFIDENCE:
            high_keys.add(key)

    upload_metrics = aggregate_upload_metrics(args.upload_ready_dir)
    previous_metrics = aggregate_upload_metrics(args.previous_upload_ready_dir)
    formulir_metrics = aggregate_upload_metrics(args.formulir_upload_ready_dir)
    formulir_index = scan_formulir_identity_index(args.formulir_upload_ready_dir)
    unresolved_keys = load_unresolved_keys(args.upload_allowlist)

    rows_out: list[dict[str, Any]] = []
    for index, src in enumerate(identity_src, start=1):
        pmid = norm(src.get("PMID"))
        pnid = norm(src.get("PNID"))
        key = (pmid, pnid)
        auto = auto_by_key.get(key, {})
        kandidat = kandidat_by_identity.get(key, "")
        rw_folder = norm(src.get("Folder Mapping"))
        rw_workbook = workbook_basename(src.get("Workbook Title Mapping", ""))
        raw_wb_title = norm(src.get("Workbook Title Mapping"))
        if not rw_workbook and raw_wb_title:
            rw_workbook = raw_wb_title
        rw_worksheet = norm(src.get("Worksheet Title Mapping"))
        classified = classify_row(
            artifact_status=norm(src.get("Status Konversi")),
            artifact_kategori=norm(src.get("Kategori Konversi")),
            pmid=pmid,
            pnid=pnid,
            rw_folder=rw_folder,
            rw_workbook=rw_workbook,
            rw_worksheet=rw_worksheet,
            inventory_sheet=norm(src.get("Inventory Sheet")),
            auto_confidence=norm(src.get("Confidence Otomatis"))
            or norm(auto.get("Confidence Label")),
            auto_workbook=workbook_basename(norm(auto.get("Candidate Source Workbook"))),
            auto_sheet=norm(auto.get("Candidate Worksheet")),
            kandidat_text=kandidat,
            high_keys=high_keys,
        )
        pegawai_lines = pegawai_by_identity.get(key, [])
        sample = "\n".join(pegawai_lines)  # full list — do not truncate

        metric_key = identity_metric_key(pmid, pnid)
        # Prefer remediate formulir when present; else current upload-ready batch.
        metric = empty_metric()
        previous_filename = ""
        if metric_key and metric_key in formulir_metrics:
            metric = formulir_metrics[metric_key]
        elif metric_key and metric_key in upload_metrics:
            metric = upload_metrics[metric_key]
        if metric_key and metric_key in previous_metrics:
            previous_filename = previous_metrics[metric_key].get("formulir_filename", "")
        has_formulir = bool(metric.get("formulir_filename"))
        if has_formulir:
            status_formulir = "Ada di batch upload"
            status_upload = "Belum Upload"
        elif metric_key is not None and metric_key in unresolved_keys:
            status_formulir = "Ditahan unresolved"
            status_upload = "Belum ada formulir"
        else:
            status_formulir = "Belum ada formulir"
            status_upload = "Belum ada formulir"

        rows_out.append(
            {
                "PMID": pmid,
                "PNID": pnid,
                "Judul Posisi": norm(src.get("Judul Posisi")),
                "Perusahaan": norm(src.get("Perusahaan")),
                "Status Kesiapan": classified["Status Kesiapan"],
                "Alasan Status": classified["Alasan Status"],
                "Treatment": classified["Treatment"],
                "Detail Breakdown": classified["Detail Breakdown"],
                "No.": index,
                "company_in_id": norm(src.get("company_in_id")),
                "Unit / Group": norm(src.get("Unit / Group")),
                "Jumlah Pegawai": int(float(src["Jumlah NIPP Roster"]))
                if norm(src.get("Jumlah NIPP Roster"))
                else len(pegawai_lines),
                "Pegawai": sample or norm(src.get("NIPP Sample")),
                "Jumlah NIPP": int(float(src["Jumlah NIPP Roster"]))
                if norm(src.get("Jumlah NIPP Roster"))
                else len(pegawai_lines),
                "Pegawai (sample)": sample or norm(src.get("NIPP Sample")),
                "Folder R&W (raw)": rw_folder,
                "Nama File R&W (raw)": rw_workbook,
                "Worksheet R&W (raw)": rw_worksheet,
                "Sheet Inventory (resolved)": classified["Sheet Inventory (resolved)"],
                "File Kamus (usulan otomatis)": workbook_basename(
                    norm(auto.get("Candidate Source Workbook"))
                ),
                "Sheet Kamus (usulan otomatis)": norm(auto.get("Candidate Worksheet")),
                "Keputusan Reviewer": "",
                "Catatan Reviewer": "",
                "_subtype": classified["Subtype"],
                "_best_candidate": classified["Best Candidate"],
                "_best_score": classified["Best Score"],
                "_priority": 0
                if classified["Status Kesiapan"] == "Siap konversi"
                else {
                    "A.": 1,
                    "D.": 2,
                    "C.": 3,
                    "E.": 4,
                    "F.": 5,
                    "G.": 6,
                }.get(classified["Treatment"][:2], 9),
                "_metric": metric,
                "_previous_formulir": previous_filename,
                "_status_formulir": status_formulir,
                "_status_upload": status_upload,
            }
        )

    rows_out.sort(
        key=lambda row: (
            row["_priority"],
            row["Perusahaan"],
            row["Judul Posisi"],
            row["PMID"],
            row["PNID"],
        )
    )
    for index, row in enumerate(rows_out, start=1):
        row["No."] = index

    siap = [row for row in rows_out if row["Status Kesiapan"] == "Siap konversi"]
    belum = [row for row in rows_out if row["Status Kesiapan"] != "Siap konversi"]
    bermasalah_rw = [
        row
        for row in belum
        if row["Alasan Status"].startswith("Review R&W")
        or row["Treatment"].startswith("D.")
        or row["Treatment"].startswith("C.")
    ]
    alias_accepted = [
        row for row in siap if "alias/urutan kata" in row["Alasan Status"].casefold()
    ]
    d_remaining = [row for row in belum if row["Treatment"].startswith("D.")]
    formulir_ada = [row for row in rows_out if row["_status_formulir"] == "Ada di batch upload"]
    formulir_unresolved = [
        row for row in rows_out if row["_status_formulir"] == "Ditahan unresolved"
    ]
    formulir_belum = [row for row in rows_out if row["_status_formulir"] == "Belum ada formulir"]

    pemetaan_rows: list[dict[str, Any]] = []
    detail_rows: list[dict[str, Any]] = []
    tracker_rows: list[dict[str, Any]] = []
    for row in rows_out:
        metric = row["_metric"]
        file_kamus, sheet_kamus = resolve_kamus_paths(row)
        has_counts = bool(metric.get("formulir_filename"))
        pemetaan_rows.append({col: row.get(col, "") for col in PEMETAAN_COLUMNS})
        detail_rows.append(
            {
                "PMID": row["PMID"],
                "PNID": row["PNID"],
                "Judul Posisi": row["Judul Posisi"],
                "Perusahaan": row["Perusahaan"],
                "Unit / Group": row["Unit / Group"],
                "Status Kesiapan": row["Status Kesiapan"],
                "Jumlah Pegawai": row["Jumlah Pegawai"],
                "Pegawai": row["Pegawai"],
                "File Kamus (resolved)": file_kamus,
                "Sheet Kamus (resolved)": sheet_kamus,
                "Nama File Formulir Upload KPI": metric.get("formulir_filename", ""),
                "Judul Formulir Upload Versi Sebelumnya": row["_previous_formulir"],
                "Link Google Sheet Formulir Upload KPI": "",
                "Jumlah IMPACT": metric["impact_count"] if has_counts else "",
                "Jumlah OUTPUT": metric["output_count"] if has_counts else "",
                "Jumlah KAI": metric["kai_count"] if has_counts else "",
                "Total Bobot IMPACT (%)": round(metric["impact_weight"], 2) if has_counts else "",
                "Total Bobot OUTPUT (%)": round(metric["output_weight"], 2) if has_counts else "",
                "Total Bobot KAI (%)": round(metric["kai_weight"], 2) if has_counts else "",
                "Bobot IMPACT OK?": yes_no(is_close(metric["impact_weight"], 100.0))
                if has_counts
                else "",
                "Bobot OUTPUT OK?": yes_no(is_close(metric["output_weight"], 100.0))
                if has_counts
                else "",
                "Bobot KAI OK?": yes_no(is_close(metric["kai_weight"], 100.0)) if has_counts else "",
                "Status Formulir": row["_status_formulir"],
            }
        )
        tracker_rows.append(
            {
                "PMID": row["PMID"],
                "PNID": row["PNID"],
                "Judul Posisi": row["Judul Posisi"],
                "Perusahaan": row["Perusahaan"],
                "Unit / Group": row["Unit / Group"],
                "Status Kesiapan": row["Status Kesiapan"],
                "Jumlah Pegawai": row["Jumlah Pegawai"],
                "Pegawai": row["Pegawai"],
                "Nama File Formulir Upload KPI": metric.get("formulir_filename", ""),
                "Judul Formulir Upload Versi Sebelumnya": row["_previous_formulir"],
                "Link Google Sheet Formulir Upload KPI": "",
                "Link Evidence Upload KPI": "",
                "Status Upload KPI": row["_status_upload"],
                "Timestamp Upload KPI": "",
                "PIC Upload KPI": "",
                "Catatan Upload": "",
            }
        )

    # One row per formulir featured in --formulir-upload-ready-dir (remediate package)
    by_metric_key = {
        identity_metric_key(norm(r["PMID"]), norm(r["PNID"])): r
        for r in rows_out
        if identity_metric_key(norm(r["PMID"]), norm(r["PNID"]))
    }
    formulir_rows: list[dict[str, Any]] = []
    for index, (filename, keys) in enumerate(sorted(formulir_index.items()), start=1):
        identity_lines: list[str] = []
        pegawai_lines: list[str] = []
        previous_names: set[str] = set()
        pegawai_count = 0
        for key in keys:
            row = by_metric_key.get(key)
            if row is None:
                scope, ident = key
                identity_lines.append(f"{scope.upper()} {ident}")
                continue
            label_id = row["PMID"] or row["PNID"]
            identity_lines.append(f"{label_id} — {row['Judul Posisi']} ({row['Perusahaan']})")
            if row.get("_previous_formulir"):
                previous_names.add(row["_previous_formulir"])
            pegawai_text = norm(row.get("Pegawai"))
            if pegawai_text:
                for line in pegawai_text.splitlines():
                    line = line.strip()
                    if line and line not in pegawai_lines:
                        pegawai_lines.append(line)
            try:
                pegawai_count += int(row.get("Jumlah Pegawai") or 0)
            except (TypeError, ValueError):
                pass
        formulir_rows.append(
            {
                "No.": index,
                "Nama File Formulir Upload KPI": filename,
                "Judul Formulir Upload Versi Sebelumnya": "\n".join(sorted(previous_names)),
                "Link Google Sheet Formulir Upload KPI": "",
                "Jumlah Identity": len(keys),
                "Jumlah Pegawai": pegawai_count,
                "Daftar Identity (PMID/PNID — Judul Posisi)": "\n".join(identity_lines),
                "Daftar Pegawai (lengkap)": "\n".join(pegawai_lines),
                "Status Upload KPI": "Belum Upload",
                "Timestamp Upload KPI": "",
                "PIC Upload KPI": "",
                "Catatan Upload": "",
            }
        )

    out_path = args.output_dir / f"Position_First_Identity_Conversion_Subholding_{stamp}.xlsx"
    latest = args.output_dir / "Position_First_Identity_Conversion_Subholding_LATEST.xlsx"

    book = Workbook()
    guide = book.active
    guide.title = "Baca Dulu"
    guide.sheet_view.showGridLines = False
    guide["A1"] = f"Position First — Kamus KPI + Upload Tracker Subholding  ·  {stamp}"
    guide["A1"].font = Font(name=BODY, size=16, bold=True, color="FFFFFF")
    guide["A1"].fill = PatternFill("solid", fgColor=NAVY)
    guide.merge_cells("A1:B1")
    guide["A3"] = (
        f"Generated: {generated_at}\n"
        f"Schema: {SCHEMA_VERSION}\n\n"
        "1 baris = 1 identitas posisi (PMID/PNID) di ketiga sheet utama.\n"
        "Urutan baris sama di Pemetaan / Detail / Tracker (row N = identity yang sama).\n\n"
        "Sheet utama:\n"
        "1. Pemetaan Kamus KPI — mapping R&W ↔ inventory + treatment reviewer\n"
        "2. Detail Kamus KPI — file/sheet kamus, formulir upload (+ versi sebelumnya), bobot IMPACT/OUTPUT/KAI\n"
        "3. Tracker Upload Kamus KPI — link Google Sheet, evidence, status upload, timestamp, PIC\n"
        "4. Formulir Upload KPI — 1 baris per file formulir upload-ready (seluruh identity terkonversi)\n\n"
        "Cara pakai tracker:\n"
        "• Kolom Pegawai menampilkan daftar pekerja lengkap (tidak dipotong).\n"
        "• Sheet Formulir Upload KPI: daftar semua file upload-ready; kolom versi sebelumnya = file sheetreuse bila ada.\n"
        "• Setelah formulir di-upload ke Google Drive, isi «Link Google Sheet Formulir Upload KPI».\n"
        "• Isi «Link Evidence Upload KPI» + ubah «Status Upload KPI» + Timestamp + PIC.\n"
        "• Filter Status Kesiapan di Pemetaan untuk Siap vs Belum (sheet filter terpisah dihapus).\n"
        "• Kolom Pegawai / Jumlah Pegawai ada di ketiga sheet.\n\n"
        "Nama formulir upload (kolom Nama File Formulir):\n"
        "  «Judul file kamus mentah» - «YYYYMMDD_HHMM» («N pekerja», «M identity»).xlsx\n"
        "  Contoh HO: cari «Kamus KPI SPTP - Mapping…», «Kamus KPI SPJM - Mapping…», "
        "«Kamus KPI SPSL - Mapping…», «Kamus KPI SPMT - Mapping…».\n\n"
        "Status Formulir (Detail):\n"
        "• Ada di batch upload — identity ada di formulir hasil konversi terbaru\n"
        "• Ditahan unresolved — siap tapi sheet inventory belum resolve\n"
        "• Belum ada formulir — belum dikonversi\n\n"
        "Dropdown Status Upload: Belum Upload | Sedang Proses | Sudah Upload | Revisi | Ditahan | Belum ada formulir\n"
        "Dropdown Keputusan Reviewer: OK | PERLU REVISI R&W | TOLAK | LAINNYA"
    )
    guide["A3"].alignment = Alignment(wrap_text=True, vertical="top")
    guide.merge_cells("A3:B3")
    guide.row_dimensions[3].height = 280
    guide.column_dimensions["A"].width = 110

    summary = book.create_sheet("Ringkasan")
    summary["A1"] = f"Ringkasan kesiapan + coverage formulir  ·  {stamp}"
    summary["A1"].fill = PatternFill("solid", fgColor=NAVY)
    summary["A1"].font = Font(name=BODY, size=16, bold=True, color="FFFFFF")
    meta = [
        ("Generated at", generated_at),
        ("Schema version", SCHEMA_VERSION),
        ("Upload-ready dir", str(args.upload_ready_dir)),
        ("Unique identity (PMID/PNID)", len(rows_out)),
        ("Siap konversi", len(siap)),
        ("  ↳ termasuk alias/urutan kata accepted", len(alias_accepted)),
        ("Belum konversi", len(belum)),
        ("Bermasalah review R&W (C+D sisa)", len(bermasalah_rw)),
        ("D tersisa (bukan alias)", len(d_remaining)),
        ("Formulir: Ada di batch upload", len(formulir_ada)),
        ("Formulir: Ditahan unresolved", len(formulir_unresolved)),
        ("Formulir: Belum ada formulir", len(formulir_belum)),
        ("Identities dengan metrik KPI ter-scan", len(upload_metrics)),
    ]
    for offset, (label, value) in enumerate(meta, start=3):
        summary.cell(offset, 1, label).font = Font(name=BODY, bold=True, color=NAVY)
        summary.cell(offset, 2, value)

    next_row = 3 + len(meta) + 2
    summary.cell(next_row, 1, "Breakdown Alasan Status (Siap)").font = Font(name=BODY, bold=True)
    for offset, (label, count) in enumerate(
        Counter(row["Alasan Status"] for row in siap).most_common(), start=next_row + 1
    ):
        summary.cell(offset, 1, label)
        summary.cell(offset, 2, count)

    start_belum = next_row + 1 + len(Counter(row["Alasan Status"] for row in siap)) + 2
    summary.cell(start_belum, 1, "Breakdown Treatment (Belum)").font = Font(name=BODY, bold=True)
    for offset, (label, count) in enumerate(
        Counter(row["Treatment"] for row in belum).most_common(), start=start_belum + 1
    ):
        summary.cell(offset, 1, label)
        summary.cell(offset, 2, count)

    d_start = start_belum + len(Counter(row["Treatment"] for row in belum)) + 3
    summary.cell(d_start, 1, "Subtype D tersisa").font = Font(name=BODY, bold=True)
    for offset, (label, count) in enumerate(
        Counter(row["_subtype"] for row in d_remaining).most_common(), start=d_start + 1
    ):
        summary.cell(offset, 1, label or "(kosong)")
        summary.cell(offset, 2, count)
    summary.column_dimensions["A"].width = 70
    summary.column_dimensions["B"].width = 48

    write_data_sheet(
        book.create_sheet("Pemetaan Kamus KPI"),
        "Pemetaan Kamus KPI — seluruh PMID/PNID Subholding",
        PEMETAAN_COLUMNS,
        pemetaan_rows,
        "PemetaanKamusKPI",
        stamp,
        freeze_at="I4",
        dropdowns={"Keputusan Reviewer": REVIEWER_CHOICES},
    )
    write_data_sheet(
        book.create_sheet("Detail Kamus KPI"),
        "Detail Kamus KPI — formulir + bobot IMPACT/OUTPUT/KAI",
        DETAIL_COLUMNS,
        detail_rows,
        "DetailKamusKPI",
        stamp,
        freeze_at="I4",
    )
    write_data_sheet(
        book.create_sheet("Tracker Upload Kamus KPI"),
        "Tracker Upload Kamus KPI — status operasional upload",
        TRACKER_COLUMNS,
        tracker_rows,
        "TrackerUploadKamusKPI",
        stamp,
        freeze_at="I4",
        dropdowns={"Status Upload KPI": UPLOAD_STATUS_CHOICES},
    )
    write_data_sheet(
        book.create_sheet("Formulir Upload KPI"),
        (
            f"Formulir Upload KPI — remedi Impact10 ({len(formulir_rows)} file, "
            f"{sum(int(r['Jumlah Identity']) for r in formulir_rows)} identity)"
        ),
        FORMULIR_COLUMNS,
        formulir_rows,
        "FormulirUploadKPI",
        stamp,
        freeze_at="C4",
        dropdowns={"Status Upload KPI": UPLOAD_STATUS_CHOICES},
    )

    book.save(out_path)
    latest.write_bytes(out_path.read_bytes())

    receipt = {
        "generated_at": generated_at,
        "schema_version": SCHEMA_VERSION,
        "artifact": str(out_path),
        "latest": str(latest),
        "upload_ready_dir": str(args.upload_ready_dir),
        "previous_upload_ready_dir": str(args.previous_upload_ready_dir),
        "formulir_upload_ready_dir": str(args.formulir_upload_ready_dir),
        "upload_config": str(args.upload_config),
        "upload_allowlist": str(args.upload_allowlist),
        "identity_total": len(rows_out),
        "siap": len(siap),
        "belum": len(belum),
        "alias_accepted": len(alias_accepted),
        "bermasalah_review_rw": len(bermasalah_rw),
        "d_remaining": len(d_remaining),
        "formulir_ada_di_batch": len(formulir_ada),
        "formulir_ditahan_unresolved": len(formulir_unresolved),
        "formulir_belum_ada": len(formulir_belum),
        "upload_metrics_identities": len(upload_metrics),
        "formulir_sheet_files": len(formulir_rows),
        "formulir_sheet_identities": sum(int(r["Jumlah Identity"]) for r in formulir_rows),
        "sheets": [
            "Baca Dulu",
            "Ringkasan",
            "Pemetaan Kamus KPI",
            "Detail Kamus KPI",
            "Tracker Upload Kamus KPI",
            "Formulir Upload KPI",
        ],
        "siap_by_alasan": dict(Counter(row["Alasan Status"] for row in siap)),
        "belum_by_treatment": dict(Counter(row["Treatment"] for row in belum)),
        "d_remaining_subtypes": dict(Counter(row["_subtype"] for row in d_remaining)),
    }
    (args.output_dir / f"IDENTITY_CONVERSION_SLIM_RECEIPT_{stamp}.json").write_text(
        json.dumps(receipt, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    (args.output_dir / "IDENTITY_CONVERSION_SLIM_RECEIPT_LATEST.json").write_text(
        json.dumps(receipt, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    print(json.dumps(receipt, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
