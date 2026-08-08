#!/usr/bin/env python3
"""Fresh Position First Mapping for Subholding roster (2,705 NIPPs).

No Red & White review is applied. Mapping comes only from:
- roster sheets SPTP/SPMT/SPSL/SPJM
- production position reference
- automatic resolver against KAMUS KPI SUBHOLDING worksheets
"""

from __future__ import annotations

import argparse
import hashlib
import json
import re
import zipfile
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any
from xml.etree import ElementTree as ET

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo

import build_group2_position_first_mapping_review as g2
import position_mapping as pm

NS = {"m": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
REL_NS = "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}"
SUB_SHEETS = ("SPTP", "SPMT", "SPSL", "SPJM")

NAVY, TEAL, PALE = "173651", "138074", "E9F1F8"
GREEN, YELLOW, ORANGE, GRAY, RED = "D9EAD3", "FFF2CC", "FCE4D6", "E7E6E6", "F4CCCC"
BODY = "Aptos"

MAIN_COLUMNS = [
    "PMID",
    "PNID",
    "Judul Posisi",
    "Perusahaan",
    "Status Mapping",
    "Flag Mapping Awal",
    "No.",
    "company_in_id",
    "Unit / Group",
    "Pegawai",
    "File Kamus (usulan otomatis)",
    "Sheet Kamus (usulan otomatis)",
    "Judul Posisi Kamus",
    "Sumber Judul Kamus",
    "Folder Kamus",
    "Skor Usulan",
    "Runner-up File",
    "Runner-up Sheet",
    "Runner-up Skor",
    "Alasan Confidence",
    "Rekomendasi",
    "Catatan Reviewer",
    "Keputusan Reviewer",
]

NEW_56_SOURCE = Path(
    "outputs/kamus-group2-subholding-roster-mapping-20260806/"
    "subholding_roster_position_first_mapping_20260806.json"
)
FLAG_NEW_56 = "BARU — belum di mapping initial"
FLAG_COVERED = "Sudah di mapping initial"
ROSTER_SHEET_TO_COMPANY_CODE = {
    "SPTP": "SPTP",
    "SPMT": "SPMT",
    "SPSL": "SPSL",
    "SPJM": "SPJM",
}


def norm(value: Any) -> str:
    return "" if value is None else str(value).strip()


def split_semi(value: Any) -> list[str]:
    text = norm(value)
    if not text:
        return []
    return [part.strip() for part in text.split(";") if part.strip()]


def join_semi(values: list[str]) -> str:
    return "; ".join(values)


def format_company(name: Any, code: Any) -> str:
    company_name = norm(name)
    company_code = norm(code)
    if company_name and company_code:
        return f"{company_name} ({company_code})"
    return company_name or company_code


def build_company_index(
    payload: dict[str, Any],
) -> tuple[dict[str, tuple[str, str, str]], dict[str, tuple[str, str, str]]]:
    """Map company code / normalized name → (company_in_id, company_name, company_code)."""
    by_code: dict[str, tuple[str, str, str]] = {}
    by_name: dict[str, tuple[str, str, str]] = {}
    for row in payload.get("company_rows") or []:
        if not isinstance(row, dict):
            continue
        company_in_id = row.get("company_in_id")
        if company_in_id in (None, ""):
            company_in_id = row.get("company_id")
        if company_in_id in (None, ""):
            continue
        company_name = norm(row.get("company_name"))
        company_code = norm(row.get("company_code")).upper()
        entry = (str(company_in_id), company_name, company_code)
        if company_code:
            by_code[company_code] = entry
        name_key = pm.normalize_title(company_name)
        if name_key:
            by_name[name_key] = entry
    return by_code, by_name


def resolve_company_identity(
    *,
    roster_sheet: str,
    persa_text: str,
    company_code: str,
    by_code: dict[str, tuple[str, str, str]],
    by_name: dict[str, tuple[str, str, str]],
) -> tuple[str, str, str]:
    """Return (company_in_id, company_name, company_code)."""
    sheet_code = ROSTER_SHEET_TO_COMPANY_CODE.get(norm(roster_sheet).upper(), "")
    raw_code = norm(company_code).upper()
    # Prefer Subholding sheet code (SPTP/…) over roster SAP numeric codes like 6000.
    for code in (sheet_code, raw_code if raw_code and not raw_code.isdigit() else ""):
        if code and code in by_code:
            return by_code[code]
    name_key = pm.normalize_title(persa_text)
    if name_key and name_key in by_name:
        return by_name[name_key]
    if name_key:
        for key, entry in by_name.items():
            if name_key in key or key in name_key:
                return entry
    return ("", norm(persa_text), sheet_code or raw_code)


def format_pegawai(names: Any, nipps: Any) -> str:
    name_list = split_semi(names)
    nipp_list = split_semi(nipps)
    width = max(len(name_list), len(nipp_list))
    lines: list[str] = []
    for idx in range(width):
        name = name_list[idx] if idx < len(name_list) else ""
        nipp = nipp_list[idx] if idx < len(nipp_list) else ""
        if name and nipp:
            lines.append(f"- {name} ({nipp})")
        elif nipp:
            lines.append(f"- ({nipp})")
        elif name:
            lines.append(f"- {name}")
    return "\n".join(lines)


def extract_nipps_from_pegawai(value: Any) -> list[str]:
    text = norm(value)
    if not text:
        return []
    return re.findall(r"\(([^()\n]+)\)", text)


def load_new_56_nipps(path: Path = NEW_56_SOURCE) -> set[str]:
    if not path.exists():
        raise SystemExit(f"Missing new-56 source mapping: {path}")
    payload = json.loads(path.read_text(encoding="utf-8"))
    nipps: set[str] = set()
    for row in payload.get("rows") or []:
        if not str(row.get("Mapping Source") or "").startswith("new_56"):
            continue
        for nipp in split_semi(row.get("Active Employee NIPPs")):
            nipps.add(nipp)
    if len(nipps) != 56:
        raise SystemExit(f"Expected 56 new-initial NIPPs, got {len(nipps)} from {path}")
    return nipps


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def load_shared_strings(archive: zipfile.ZipFile) -> list[str]:
    if "xl/sharedStrings.xml" not in archive.namelist():
        return []
    root = ET.fromstring(archive.read("xl/sharedStrings.xml"))
    out: list[str] = []
    for item in root.findall("m:si", NS):
        texts = [node.text or "" for node in item.iter("{http://schemas.openxmlformats.org/spreadsheetml/2006/main}t")]
        out.append("".join(texts))
    return out


def cell_value(cell: ET.Element, shared: list[str]) -> Any:
    cell_type = cell.attrib.get("t")
    value_node = cell.find("m:v", NS)
    if value_node is None:
        return None
    raw = value_node.text
    if cell_type == "s":
        return shared[int(raw)]
    return raw


def col_index(cell_ref: str) -> int:
    match = re.match(r"([A-Z]+)", cell_ref)
    assert match is not None
    total = 0
    for char in match.group(1):
        total = total * 26 + (ord(char) - 64)
    return total - 1


def read_sheet_rows(archive: zipfile.ZipFile, target: str, shared: list[str]) -> list[list[Any]]:
    root = ET.fromstring(archive.read(target))
    rows: list[list[Any]] = []
    for row in root.findall("m:sheetData/m:row", NS):
        cells: dict[int, Any] = {}
        for cell in row.findall("m:c", NS):
            cells[col_index(cell.attrib.get("r", "A1"))] = cell_value(cell, shared)
        if not cells:
            continue
        width = max(cells) + 1
        rows.append([cells.get(index) for index in range(width)])
    return rows


def workbook_sheet_targets(archive: zipfile.ZipFile) -> list[tuple[str, str]]:
    workbook = ET.fromstring(archive.read("xl/workbook.xml"))
    rels = ET.fromstring(archive.read("xl/_rels/workbook.xml.rels"))
    rid_to_target = {rel.attrib["Id"]: rel.attrib["Target"] for rel in rels}
    out: list[tuple[str, str]] = []
    for sheet in workbook.findall("m:sheets/m:sheet", NS):
        rid = sheet.attrib[f"{REL_NS}id"]
        target = rid_to_target[rid].lstrip("/")
        if not target.startswith("xl/"):
            target = f"xl/{target}"
        out.append((sheet.attrib["name"], target))
    return out


def load_roster(path: Path) -> dict[str, dict[str, str]]:
    roster: dict[str, dict[str, str]] = {}
    with zipfile.ZipFile(path) as archive:
        shared = load_shared_strings(archive)
        for sheet_name, target in workbook_sheet_targets(archive):
            if sheet_name not in SUB_SHEETS:
                continue
            rows = read_sheet_rows(archive, target, shared)
            if not rows:
                continue
            header = [norm(value) for value in rows[0]]
            index = {name: idx for idx, name in enumerate(header)}
            for row in rows[1:]:
                if not row or all(value in (None, "") for value in row):
                    continue
                nipp = norm(row[index["PNALT_NEW"]])
                if not nipp:
                    continue
                roster[nipp] = {
                    "sheet": sheet_name,
                    "name": norm(row[index["CNAME"]]),
                    "job_title": norm(row[index["STEXT_STO"]]),
                    "persa_text": norm(row[index["PERSA_TEXT"]]),
                    "sub_persa_text": norm(row[index["SUB_PERSA_TEXT"]]),
                    "company_code": norm(row[index["COMPANY_CODE"]]),
                }
    return roster


def status_mapping(confidence: str) -> str:
    return {
        pm.HIGH_CONFIDENCE: "Usulan kuat (high)",
        pm.LOW_CONFIDENCE: "Usulan lemah (low) — perlu cek",
        pm.MAPPING_CONFLICT: "Konflik kandidat — perlu pilih",
        pm.NO_CANDIDATE: "Belum ada kandidat Kamus",
        pm.SCOPE_UNCERTAIN: "Scope identitas belum jelas",
    }.get(confidence, confidence or "(kosong)")


def style_header(cell, fill: str = TEAL) -> None:
    cell.fill = PatternFill("solid", fgColor=fill)
    cell.font = Font(name=BODY, bold=True, color="FFFFFF", size=10)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def write_guide(ws) -> None:
    ws.sheet_view.showGridLines = False
    ws["A1"] = "Baca Dulu — Position First Mapping Subholding (Fresh, tanpa R&W)"
    ws["A1"].font = Font(name=BODY, size=16, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor=NAVY)
    ws.merge_cells("A1:B1")
    blocks = [
        (
            "Apa ini?",
            "Pemetaan posisi untuk exact 2.705 NIPP Subholding (sheet SPTP/SPMT/SPSL/SPJM). "
            "Dibangun fresh: tanpa memakai hasil review Red & White.",
        ),
        (
            "Sumber data",
            "1) Roster REGIONAL dan SUBHOLDING.xlsx\n"
            "2) Production position reference\n"
            "3) Inventory Kamus KPI Group 2 — hanya folder KAMUS KPI SUBHOLDING",
        ),
        (
            "Cara baca kolom usulan",
            "Kolom kiri (freeze): PMID, PNID, Judul Posisi, Perusahaan, Status Mapping, Flag Mapping Awal.\n"
            "Keputusan Reviewer ada di kolom paling kanan.\n"
            "Sheet Kamus = nama tab Excel (maks. 31 karakter). "
            "Judul Posisi Kamus = position_name inventory (kosong jika tidak ada).\n"
            "company_in_id diisi dari production company_rows (termasuk stub tanpa PMID, via sheet SPTP/SPMT/SPSL/SPJM).",
        ),
        (
            "Sheet mana yang dilihat?",
            "Ringkasan → Pemetaan (semua) → Belum di Mapping Awal (56) → Perlu Dicek → Tanpa Kandidat.",
        ),
    ]
    row = 3
    for title, body in blocks:
        ws.cell(row, 1, title).font = Font(name=BODY, bold=True, color=NAVY, size=12)
        ws.cell(row, 1).fill = PatternFill("solid", fgColor=PALE)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        row += 1
        ws.cell(row, 1, body).alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        ws.row_dimensions[row].height = 64
        row += 2
    ws.column_dimensions["A"].width = 110


def write_summary(ws, meta: list[tuple[str, Any]], labels: Counter[str]) -> None:
    ws.sheet_view.showGridLines = False
    ws["A1"] = "Position First Mapping Subholding — Fresh (tanpa R&W)"
    ws["A1"].font = Font(name=BODY, size=18, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor=NAVY)
    ws.merge_cells("A1:B1")
    ws["A2"] = "Scope exact 2.705 NIPP roster Subholding. Usulan Kamus hanya dari folder Subholding."
    ws["A2"].fill = PatternFill("solid", fgColor=PALE)
    ws.merge_cells("A2:B2")
    style_header(ws.cell(4, 1, "Item"))
    style_header(ws.cell(4, 2, "Nilai"))
    for offset, (label, value) in enumerate(meta, start=5):
        ws.cell(offset, 1, label).fill = PatternFill("solid", fgColor="F3F6F8")
        ws.cell(offset, 1).font = Font(name=BODY, bold=True, color=NAVY)
        ws.cell(offset, 2, value).alignment = Alignment(wrap_text=True)
    start = 5 + len(meta) + 2
    style_header(ws.cell(start, 1, "Confidence"))
    style_header(ws.cell(start, 2, "Jumlah"))
    for offset, label in enumerate(
        [pm.HIGH_CONFIDENCE, pm.LOW_CONFIDENCE, pm.MAPPING_CONFLICT, pm.NO_CANDIDATE, pm.SCOPE_UNCERTAIN],
        start=start + 1,
    ):
        ws.cell(offset, 1, label).fill = {
            pm.HIGH_CONFIDENCE: PatternFill("solid", fgColor=GREEN),
            pm.LOW_CONFIDENCE: PatternFill("solid", fgColor=YELLOW),
            pm.MAPPING_CONFLICT: PatternFill("solid", fgColor=RED),
            pm.NO_CANDIDATE: PatternFill("solid", fgColor=GRAY),
        }.get(label, PatternFill())
        ws.cell(offset, 2, labels.get(label, 0))
    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 100


def write_table(ws, title: str, rows: list[dict[str, Any]], table_name: str) -> None:
    ws.sheet_view.showGridLines = False
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(MAIN_COLUMNS))
    ws.cell(1, 1, title).fill = PatternFill("solid", fgColor=NAVY)
    ws.cell(1, 1).font = Font(name=BODY, size=14, bold=True, color="FFFFFF")
    for col, header in enumerate(MAIN_COLUMNS, start=1):
        style_header(ws.cell(3, col, header))
    pegawai_col = MAIN_COLUMNS.index("Pegawai") + 1
    status_col = MAIN_COLUMNS.index("Status Mapping") + 1
    flag_col = MAIN_COLUMNS.index("Flag Mapping Awal") + 1
    for r_idx, row in enumerate(rows, start=4):
        pegawai_lines = norm(row.get("Pegawai")).count("\n") + (1 if norm(row.get("Pegawai")) else 0)
        if pegawai_lines > 1:
            ws.row_dimensions[r_idx].height = min(15 * pegawai_lines + 6, 90)
        for c_idx, header in enumerate(MAIN_COLUMNS, start=1):
            cell = ws.cell(r_idx, c_idx, row.get(header, ""))
            cell.font = Font(name=BODY, size=9)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if c_idx == pegawai_col:
                cell.number_format = "@"
        conf = norm(row.get("_confidence"))
        fills = {
            pm.HIGH_CONFIDENCE: GREEN,
            pm.LOW_CONFIDENCE: YELLOW,
            pm.MAPPING_CONFLICT: RED,
            pm.NO_CANDIDATE: GRAY,
        }
        if conf in fills:
            ws.cell(r_idx, status_col).fill = PatternFill("solid", fgColor=fills[conf])
        if norm(row.get("Status Mapping")).startswith("Belum ada identitas"):
            ws.cell(r_idx, status_col).fill = PatternFill("solid", fgColor=ORANGE)
        if norm(row.get("Flag Mapping Awal")) == FLAG_NEW_56:
            ws.cell(r_idx, flag_col).fill = PatternFill("solid", fgColor=ORANGE)
    end_row = 3 + max(len(rows), 1)
    if rows:
        table = Table(displayName=table_name, ref=f"A3:{get_column_letter(len(MAIN_COLUMNS))}{end_row}")
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
        ws.add_table(table)
        decide_col = get_column_letter(MAIN_COLUMNS.index("Keputusan Reviewer") + 1)
        dv = DataValidation(type="list", formula1='"YES,NEEDS_CHECK,NO"', allow_blank=True)
        ws.add_data_validation(dv)
        dv.add(f"{decide_col}4:{decide_col}{end_row}")
    widths = {
        "No.": 6,
        "Status Mapping": 32,
        "Keputusan Reviewer": 14,
        "Flag Mapping Awal": 28,
        "PMID": 10,
        "PNID": 10,
        "Judul Posisi": 36,
        "Perusahaan": 36,
        "company_in_id": 12,
        "Unit / Group": 24,
        "Pegawai": 36,
        "File Kamus (usulan otomatis)": 52,
        "Sheet Kamus (usulan otomatis)": 34,
        "Judul Posisi Kamus": 40,
        "Sumber Judul Kamus": 14,
        "Folder Kamus": 18,
        "Skor Usulan": 10,
        "Runner-up File": 40,
        "Runner-up Sheet": 24,
        "Runner-up Skor": 10,
        "Alasan Confidence": 36,
        "Rekomendasi": 28,
        "Catatan Reviewer": 24,
    }
    for idx, header in enumerate(MAIN_COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = widths.get(header, 16)
    # Freeze after PMID | PNID | Judul Posisi | Perusahaan | Status Mapping | Flag Mapping Awal
    ws.freeze_panes = "G4"


def build_positions_from_roster(
    *,
    roster: dict[str, dict[str, str]],
    reference_path: Path,
) -> tuple[list[g2.PositionEntry], list[dict[str, str]], set[str]]:
    payload = json.loads(reference_path.read_text(encoding="utf-8"))
    indexes = pm.build_lookup_indexes(payload)
    by_code, by_name = build_company_index(payload)
    nipp_to_candidates: dict[str, list[Any]] = defaultdict(list)
    for candidate in [*indexes.structural, *indexes.non_structural]:
        for nipp in candidate.active_employee_nipps:
            key = str(nipp).strip()
            if key in roster:
                nipp_to_candidates[key].append(candidate)

    # Group by identity so one PMID/PNID becomes one row with roster-filtered people.
    identity_people: dict[tuple[str, str], dict[str, Any]] = {}
    covered: set[str] = set()
    for nipp, candidates in nipp_to_candidates.items():
        info = roster[nipp]
        scored: list[tuple[float, Any]] = []
        for candidate in candidates:
            title_score = pm._title_score(info["job_title"].split("#")[0].strip(), candidate)
            company_score = pm._context_score(info["persa_text"], candidate.company_name)
            scored.append((title_score * 0.7 + company_score * 0.3, candidate))
        scored.sort(key=lambda item: item[0], reverse=True)
        chosen = scored[0][1]
        if chosen.scope == "structural" and chosen.position_master_id:
            identity = ("pmid", str(chosen.position_master_id))
        elif chosen.position_nomenclature_id:
            identity = ("pnid", str(chosen.position_nomenclature_id))
        else:
            continue
        bucket = identity_people.setdefault(
            identity,
            {
                "candidate": chosen,
                "nipps": [],
                "names": [],
                "roster_sheets": Counter(),
                "roster_jobs": [],
            },
        )
        if nipp not in bucket["nipps"]:
            bucket["nipps"].append(nipp)
            bucket["names"].append(info["name"] or "")
            bucket["roster_sheets"][info["sheet"]] += 1
            bucket["roster_jobs"].append(info["job_title"])
            covered.add(nipp)

    positions: list[g2.PositionEntry] = []
    roster_meta: list[dict[str, str]] = []
    for identity, bucket in identity_people.items():
        chosen = bucket["candidate"]
        nipps = bucket["nipps"]
        names = bucket["names"]
        sheet = bucket["roster_sheets"].most_common(1)[0][0]
        sample = roster[nipps[0]]
        company_id = norm(chosen.company_id)
        company_name = norm(chosen.company_name)
        company_code = norm(chosen.company_code)
        if not company_id:
            company_id, resolved_name, resolved_code = resolve_company_identity(
                roster_sheet=sheet,
                persa_text=company_name or sample["persa_text"],
                company_code=company_code,
                by_code=by_code,
                by_name=by_name,
            )
            company_name = company_name or resolved_name
            company_code = company_code or resolved_code
        positions.append(
            g2.PositionEntry(
                scope=chosen.scope,
                pmid=chosen.position_master_id if identity[0] == "pmid" else None,
                pnid=chosen.position_nomenclature_id if identity[0] == "pnid" else None,
                title=chosen.title,
                group_name=chosen.group_name,
                company_name=company_name or chosen.company_name,
                company_code=company_code or chosen.company_code,
                company_id=company_id,
                active_employee_count=len(nipps),
                active_employee_nipps=join_semi(nipps),
                active_employee_names=join_semi(names),
                normalized_title=pm.normalize_position_lookup(chosen.title),
                tokens=g2.significant_tokens(chosen.title),
                company_tokens=g2.company_tokens(
                    " ".join(filter(None, [company_name or chosen.company_name, company_code or chosen.company_code]))
                ),
                company_key=g2.company_key(
                    norm(company_name or chosen.company_name) or norm(company_code or chosen.company_code)
                ),
            )
        )
        roster_meta.append(
            {
                "sheet": sheet,
                "persa_text": sample["persa_text"],
                "sub_persa_text": sample["sub_persa_text"],
                "job_title": sample["job_title"],
            }
        )

    missing = sorted(set(roster) - covered)
    stubs: list[g2.PositionEntry] = []
    stub_meta: list[dict[str, str]] = []
    for nipp in missing:
        info = roster[nipp]
        title = info["job_title"].split("#")[0].strip() or info["job_title"] or f"Roster NIPP {nipp}"
        company_id, company_name, company_code = resolve_company_identity(
            roster_sheet=info["sheet"],
            persa_text=info["persa_text"],
            company_code=info["company_code"],
            by_code=by_code,
            by_name=by_name,
        )
        stubs.append(
            g2.PositionEntry(
                scope="unknown",
                pmid=None,
                pnid=None,
                title=title,
                group_name=info["sub_persa_text"],
                company_name=company_name or info["persa_text"],
                company_code=company_code,
                company_id=company_id or None,
                active_employee_count=1,
                active_employee_nipps=nipp,
                active_employee_names=info["name"],
                normalized_title=pm.normalize_position_lookup(title),
                tokens=g2.significant_tokens(title),
                company_tokens=g2.company_tokens(
                    " ".join(filter(None, [company_name or info["persa_text"], company_code]))
                ),
                company_key=g2.company_key(company_name or info["persa_text"] or company_code),
            )
        )
        stub_meta.append(
            {
                "sheet": info["sheet"],
                "persa_text": info["persa_text"],
                "sub_persa_text": info["sub_persa_text"],
                "job_title": info["job_title"],
                "stub": "1",
            }
        )
    return positions + stubs, roster_meta + stub_meta, set(missing)


def to_readable(row: dict[str, Any], meta: dict[str, str], index: int, new_56: set[str]) -> dict[str, Any]:
    confidence = norm(row.get("Confidence Label"))
    is_stub = meta.get("stub") == "1"
    nipps = split_semi(row.get("Active Employee NIPPs"))
    has_new = any(nipp in new_56 for nipp in nipps)
    title_source = norm(row.get("Title Match Source"))
    if title_source == "position_name":
        sumber = "position_name"
    elif title_source == "sheet_name":
        sumber = "sheet_name"
    else:
        sumber = title_source
    # Position title: inventory value only — never fall back to sheet tab name.
    position_title = norm(row.get("Candidate Worksheet Title"))
    sheet_title = norm(row.get("Candidate Worksheet"))
    if position_title and sheet_title and position_title == sheet_title:
        # Same string usually means extractor copied the (possibly truncated) tab name.
        # Keep both visible so reviewers see the raw inventory fields as stored.
        pass
    return {
        "PMID": norm(row.get("PMID")),
        "PNID": norm(row.get("PNID")),
        "Judul Posisi": norm(row.get("Position Title")),
        "Perusahaan": format_company(row.get("Company"), row.get("Company Code")),
        "Status Mapping": (
            "Belum ada identitas production — perlu review manual"
            if is_stub
            else status_mapping(confidence)
        ),
        "Flag Mapping Awal": FLAG_NEW_56 if has_new else FLAG_COVERED,
        "No.": index,
        "company_in_id": norm(row.get("Company ID")),
        "Unit / Group": norm(row.get("Group / Unit")),
        "Pegawai": format_pegawai(row.get("Active Employee Names"), row.get("Active Employee NIPPs")),
        "File Kamus (usulan otomatis)": norm(row.get("Candidate Source Workbook")),
        "Sheet Kamus (usulan otomatis)": sheet_title,
        "Judul Posisi Kamus": position_title,
        "Sumber Judul Kamus": sumber,
        "Folder Kamus": norm(row.get("Candidate Source Folder")),
        "Skor Usulan": row.get("Candidate Score") if row.get("Candidate Score") != "" else "",
        "Runner-up File": norm(row.get("Runner-up Workbook")),
        "Runner-up Sheet": norm(row.get("Runner-up Worksheet")),
        "Runner-up Skor": row.get("Runner-up Score") if row.get("Runner-up Score") != "" else "",
        "Alasan Confidence": norm(row.get("Confidence Reason")),
        "Rekomendasi": norm(row.get("Recommended Action")),
        "Catatan Reviewer": "",
        "Keputusan Reviewer": "",
        "_confidence": confidence,
    }


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--roster",
        type=Path,
        default=Path(
            "outputs/kamus-group2-subholding-roster-mapping-20260806/source/"
            "REGIONAL dan SUBHOLDING.xlsx"
        ),
    )
    parser.add_argument(
        "--inventory",
        type=Path,
        default=Path("configs/kamus_kpi_group2_visible_20260807.json"),
    )
    parser.add_argument(
        "--reference",
        type=Path,
        default=Path("configs/production_position_reference.json"),
    )
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=Path("outputs/kamus-group2-subholding-roster-fresh-20260806"),
    )
    args = parser.parse_args()

    generated_at = datetime.now().astimezone().isoformat(timespec="seconds")
    roster = load_roster(args.roster)
    if len(roster) != 2705:
        raise SystemExit(f"Expected 2705 roster NIPPs, got {len(roster)}")

    worksheets = g2.load_worksheets(args.inventory, source_folder="KAMUS KPI SUBHOLDING")
    new_56 = load_new_56_nipps()
    positions, roster_meta, missing_identity = build_positions_from_roster(
        roster=roster,
        reference_path=args.reference,
    )
    resolved_rows, labels, shared_best = g2.resolve_all(positions, worksheets)

    # Attach roster meta in the same order as positions/resolved_rows
    if len(resolved_rows) != len(roster_meta):
        raise SystemExit("Internal error: resolved rows and roster meta length mismatch")

    readable = [
        to_readable(row, meta, idx, new_56)
        for idx, (row, meta) in enumerate(zip(resolved_rows, roster_meta), start=1)
    ]

    all_nipps = [nipp for row in readable for nipp in extract_nipps_from_pegawai(row["Pegawai"])]
    unique_nipps = set(all_nipps)
    if unique_nipps != set(roster):
        raise SystemExit(
            f"Unique NIPP mismatch: got {len(unique_nipps)}; "
            f"missing={sorted(set(roster) - unique_nipps)[:10]}; "
            f"extra={sorted(unique_nipps - set(roster))[:10]}"
        )

    flagged_new = [row for row in readable if row["Flag Mapping Awal"] == FLAG_NEW_56]
    flagged_nipp_count = len(
        {nipp for row in flagged_new for nipp in extract_nipps_from_pegawai(row["Pegawai"]) if nipp in new_56}
    )
    if flagged_nipp_count != 56:
        raise SystemExit(f"Expected 56 flagged new-initial NIPPs in rows, got {flagged_nipp_count}")

    # Ensure usulan file stays in Subholding folder when present
    non_sub = [
        row
        for row in readable
        if norm(row["File Kamus (usulan otomatis)"])
        and "SUBHOLDING" not in norm(row["File Kamus (usulan otomatis)"]).upper()
    ]

    args.output_dir.mkdir(parents=True, exist_ok=True)
    # Copy provenance pointer for roster
    source_dir = args.output_dir / "source"
    source_dir.mkdir(exist_ok=True)
    target_roster = source_dir / args.roster.name
    if not target_roster.exists():
        target_roster.write_bytes(args.roster.read_bytes())

    reference = json.loads(args.reference.read_text(encoding="utf-8"))
    exported_at = norm(reference.get("source", {}).get("exported_at"))
    roster_hash = sha256_file(args.roster)
    inventory_hash = sha256_file(args.inventory)
    reference_hash = sha256_file(args.reference)

    xlsx_stamp = datetime.now().astimezone().strftime("%Y%m%d_%H%M%S")
    xlsx_path = args.output_dir / f"Position_First_Mapping_Subholding_Fresh_{xlsx_stamp}.xlsx"
    json_path = args.output_dir / f"subholding_roster_fresh_mapping_{xlsx_stamp}.json"
    receipt_path = args.output_dir / f"MAPPING_RECEIPT_{xlsx_stamp}.md"
    # Stable latest pointers for convenience
    latest_xlsx = args.output_dir / "Position_First_Mapping_Subholding_Fresh_LATEST.xlsx"
    latest_json = args.output_dir / "subholding_roster_fresh_mapping_LATEST.json"

    needs_check = [row for row in readable if norm(row.get("_confidence")) != pm.HIGH_CONFIDENCE]
    no_candidate = [row for row in readable if norm(row.get("_confidence")) == pm.NO_CANDIDATE]
    stub_rows = [
        row for row in readable if str(row["Status Mapping"]).startswith("Belum ada identitas")
    ]

    meta_rows = [
        ("Judul", "Position First Mapping Subholding — Fresh (tanpa R&W)"),
        ("Generated at", generated_at),
        ("Artifact stamp", xlsx_stamp),
        ("Unique NIPP", len(unique_nipps)),
        ("Baris posisi", len(readable)),
        ("Tanpa identity production", len(missing_identity)),
        ("Belum di mapping initial (56 NIPP)", flagged_nipp_count),
        ("Worksheet Kamus Subholding", len(worksheets)),
        ("Usulan non-Subholding (harusnya 0)", len(non_sub)),
        ("Roster", str(args.roster)),
        ("Roster sha256", roster_hash),
        ("Inventory", str(args.inventory)),
        ("Inventory sha256", inventory_hash),
        ("Production reference", str(args.reference)),
        ("Reference exported_at", exported_at),
        ("Reference sha256", reference_hash),
        ("Kebijakan", "Fresh start — tidak memakai review Red & White"),
    ]

    wb = Workbook()
    guide = wb.active
    guide.title = "Baca Dulu"
    write_guide(guide)
    write_summary(wb.create_sheet("Ringkasan"), meta_rows, labels)
    write_table(
        wb.create_sheet("Pemetaan"),
        "Pemetaan fresh Subholding (2.705 NIPP) — usulan otomatis Kamus Subholding",
        readable,
        "PemetaanFreshTable",
    )
    write_table(
        wb.create_sheet("Belum di Mapping Awal (56)"),
        "56 NIPP roster yang belum terinput di position-first mapping initial",
        flagged_new,
        "BelumMappingAwal56Table",
    )
    write_table(
        wb.create_sheet("Perlu Dicek"),
        "Antrian: bukan high_confidence",
        needs_check,
        "PerluDicekTable",
    )
    write_table(
        wb.create_sheet("Tanpa Kandidat"),
        "Belum ada kandidat worksheet Kamus Subholding",
        no_candidate,
        "TanpaKandidatTable",
    )
    if stub_rows:
        write_table(
            wb.create_sheet("Tanpa Identity Production"),
            "NIPP roster yang belum ketemu identity PMID/PNID aktif",
            stub_rows,
            "TanpaIdentityTable",
        )
    wb.save(xlsx_path)
    # Copy as LATEST for easy discovery without overwriting timestamped artifact.
    latest_xlsx.write_bytes(xlsx_path.read_bytes())

    # Drop internal-only keys from persisted readable rows.
    readable_public = [{k: v for k, v in row.items() if not str(k).startswith("_")} for row in readable]

    payload = {
        "metadata": {
            "title": "Position First Mapping Subholding — Fresh (tanpa R&W)",
            "orientation": "position_first",
            "scope": "subholding-roster-2705-fresh",
            "uses_red_and_white_review": False,
            "generated_at": generated_at,
            "artifact_stamp": xlsx_stamp,
            "unique_active_employees": len(unique_nipps),
            "position_row_count": len(readable),
            "missing_production_identity_nipps": len(missing_identity),
            "new_56_initial_mapping_nipps": flagged_nipp_count,
            "confidence_counts": dict(labels),
            "non_subholding_proposal_rows": len(non_sub),
            "shared_worksheet_count": sum(1 for count in shared_best.values() if count >= 2),
            "sources": {
                "roster": {"path": str(args.roster), "sha256": roster_hash},
                "inventory": {"path": str(args.inventory), "sha256": inventory_hash},
                "production_reference": {
                    "path": str(args.reference),
                    "sha256": reference_hash,
                    "exported_at": exported_at,
                },
            },
        },
        "rows": resolved_rows,
        "readable_rows": readable_public,
        "missing_production_identity_nipps": sorted(missing_identity),
        "new_56_initial_mapping_nipps": sorted(new_56),
    }
    json_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    latest_json.write_text(json_path.read_text(encoding="utf-8"), encoding="utf-8")

    receipt = f"""# Mapping Receipt — Subholding Fresh (tanpa R&W)

## Scope
- Exact **{len(unique_nipps)}** unique NIPP from roster SPTP/SPMT/SPSL/SPJM
- Position rows: **{len(readable)}**
- Red & White review: **not used**

## Sources
- Roster: `{args.roster}` (`{roster_hash}`)
- Inventory: `{args.inventory}` (`{inventory_hash}`)
- Production: `{args.reference}` exported_at `{exported_at}` (`{reference_hash}`)

## Confidence
{json.dumps(dict(labels), ensure_ascii=False, indent=2)}

## Gaps
- NIPP without production identity: **{len(missing_identity)}**
- NIPP belum di mapping initial: **{flagged_nipp_count}**
- Non-Subholding proposals: **{len(non_sub)}**

## Artifacts
- Timestamped: `{xlsx_path}`
- Timestamped JSON: `{json_path}`
- Latest pointer: `{latest_xlsx}`
"""
    receipt_path.write_text(receipt, encoding="utf-8")

    print(
        json.dumps(
            {
                "unique_active_employees": len(unique_nipps),
                "position_rows": len(readable),
                "missing_production_identity": len(missing_identity),
                "new_56_initial_mapping_nipps": flagged_nipp_count,
                "confidence_counts": dict(labels),
                "non_subholding_proposals": len(non_sub),
                "artifact_stamp": xlsx_stamp,
                "xlsx": str(xlsx_path),
                "json": str(json_path),
                "latest_xlsx": str(latest_xlsx),
            },
            ensure_ascii=False,
            indent=2,
        )
    )


if __name__ == "__main__":
    main()
