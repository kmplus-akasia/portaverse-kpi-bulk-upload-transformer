#!/usr/bin/env python3
"""Split a generated project-position KPI upload workbook into project workbooks."""

from __future__ import annotations

import argparse
import csv
import re
import shutil
import zipfile
from pathlib import Path

import openpyxl


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser()
    parser.add_argument("--source-workbook", required=True, type=Path)
    parser.add_argument("--mapping-audit", required=True, type=Path)
    parser.add_argument("--output-dir", required=True, type=Path)
    parser.add_argument("--zip-output", required=True, type=Path)
    parser.add_argument("--manifest-output", required=True, type=Path)
    return parser.parse_args()


def safe_name(value: str) -> str:
    return re.sub(r"[^A-Za-z0-9._ -]+", "", value).strip().replace("  ", " ")


def project_bucket(group_name: str) -> str:
    text = group_name.lower()
    if "bali maritime tourism hub" in text or "bmth" in text:
        return "BMTH"
    if "jict" in text or "koja" in text:
        return "JICT KOJA"
    if "kijing" in text:
        return "Kijing"
    if "kalibaru" in text and "npea" in text:
        return "Terminal Kalibaru dan NPEA"
    if "npea" in text:
        return "NPEA"
    if "kalibaru" in text:
        return "Terminal Kalibaru"
    return "Other Project"


def main() -> int:
    args = parse_args()
    pmid_to_project: dict[str, str] = {}
    with args.mapping_audit.open(newline="") as handle:
        for row in csv.DictReader(handle):
            pmid_to_project[str(row["position_master_id"])] = project_bucket(row["group_name"])

    source = openpyxl.load_workbook(args.source_workbook, read_only=True, data_only=False)
    source_sheet = source["KPI Template"]
    rows_by_project: dict[str, list[tuple]] = {}
    for row in source_sheet.iter_rows(min_row=2, values_only=True):
        title = row[10] if len(row) > 10 else None
        if title in (None, ""):
            continue
        pmid = str(row[4] or "").strip()
        project = pmid_to_project.get(pmid, "Other Project")
        rows_by_project.setdefault(project, []).append(row)

    if args.output_dir.exists():
        shutil.rmtree(args.output_dir)
    args.output_dir.mkdir(parents=True, exist_ok=True)

    manifest_rows: list[dict[str, object]] = []
    for project, rows in sorted(rows_by_project.items()):
        workbook = openpyxl.load_workbook(args.source_workbook)
        sheet = workbook["KPI Template"]
        if sheet.max_row > 1:
            sheet.delete_rows(2, sheet.max_row - 1)
        for row in rows:
            sheet.append(list(row))
        output_path = args.output_dir / f"KPI Upload - Project Positions - {safe_name(project)} - 20260615.xlsx"
        workbook.save(output_path)
        positions = sorted({str(row[4] or "") for row in rows})
        manifest_rows.append(
            {
                "project": project,
                "workbook": str(output_path),
                "rows": len(rows),
                "positions": len(positions),
            }
        )

    if args.zip_output.exists():
        args.zip_output.unlink()
    with zipfile.ZipFile(args.zip_output, "w", zipfile.ZIP_DEFLATED) as archive:
        for row in manifest_rows:
            path = Path(str(row["workbook"]))
            archive.write(path, arcname=path.name)
    with zipfile.ZipFile(args.zip_output) as archive:
        bad_file = archive.testzip()
    if bad_file:
        raise RuntimeError(f"bad zip member: {bad_file}")

    with args.manifest_output.open("w", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=["project", "workbook", "rows", "positions"])
        writer.writeheader()
        writer.writerows(manifest_rows)

    print(f"project_workbooks={len(manifest_rows)}")
    print(f"total_rows={sum(int(row['rows']) for row in manifest_rows)}")
    print(f"zip_output={args.zip_output}")
    print(f"manifest_output={args.manifest_output}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
