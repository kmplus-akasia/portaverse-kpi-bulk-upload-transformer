#!/usr/bin/env python3
"""Build position/worker crosscheck list from File Backlog Kamus KPI.xlsx."""

from __future__ import annotations

import csv
import json
import re
from collections import OrderedDict
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
BACKLOG = Path("/Users/alfredoteja/Downloads/File Backlog Kamus KPI.xlsx")
KAMUS_PATH = ROOT / "configs/kamus_kpi_ho_visible_20260727.json"
REF_PATH = ROOT / "configs/production_position_reference.json"
OLD_CONFIG_PATH = ROOT / "output/group1_ho_v2_20260709_latest_prod/group1_ho_v2_20260709_refreshed.config.json"
KEBERLANJUTAN_CONFIG = ROOT / "output/group_keberlanjutan_5_positions_20260727/support/group_keberlanjutan_5_positions.config.json"
OUT_DIR = ROOT / "output/backlog_kamus_kpi_20260727"

MANUAL_OVERRIDES: dict[tuple[str, str], dict[str, str | None]] = {}


def norm_name(value: object) -> str:
    text = str(value or "").lower().strip()
    text = re.sub(r"\bdh\b", "department head", text)
    text = text.replace("&", " dan ")
    text = re.sub(r"[-_/(),.]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def load_manual_overrides() -> dict[tuple[str, str], dict[str, str | None]]:
    overrides: dict[tuple[str, str], dict[str, str | None]] = {}
    if KEBERLANJUTAN_CONFIG.exists():
        payload = json.loads(KEBERLANJUTAN_CONFIG.read_text(encoding="utf-8"))
        for pos in payload.get("positions", []):
            overrides[(pos["sheet_name"], "Group Keberlanjutan Korporasi")] = {
                "shape": "PMID" if pos.get("position_master_id") else "PNID",
                "id": str(pos.get("position_master_id") or pos.get("position_nomenclature_id")),
                "production_title": pos.get("portaverse_position_title"),
                "mapping_source": "approved_run_20260727",
            }
    return overrides


def split_positions(raw: str) -> list[str]:
    if not raw or raw.lower().strip() == "semua posisi":
        return ["__ALL__"]
    parts = re.split(r",\s*", raw.strip().rstrip(","))
    aliases = {
        "dh pengelolaan keberlanjutan": "department head pengelolaan keberlanjutan",
        "officer pengelolaan keberlanjutan": "officer pengelolaan keberlanjutan",
        "manager implementasi dan pelaporan keberlanjutan": "manager implementasi dan pelaporan keberlanjutan korporasi",
        "officer implementasi dan pelaporan keberlanjutan": "officer implementasi dan pelaporan keberlanjutan korporasi",
        "dh perencanaan dan evaluasi pemasaran": "department head perencanaan dan evaluasi pemasaran",
        "dh pengelolaan pelanggan": "department head pengelolaan pelanggan",
        "department head dan officer aset tetap": ["department head aset tetap", "officer aset tetap"],
        "dh komunikasi korporasi": "department head komunikasi korporasi",
        "group head": "group head",
    }
    out: list[str] = []
    for part in parts:
        part = part.strip().rstrip(",")
        if not part:
            continue
        key = norm_name(part)
        alias = aliases.get(key, key)
        if isinstance(alias, list):
            out.extend(alias)
        else:
            out.append(alias)
    return out


def main() -> None:
    kamus_payload = json.loads(KAMUS_PATH.read_text(encoding="utf-8"))
    kamus_items = kamus_payload.get("kamus_kpi_v2", [])
    ref = json.loads(REF_PATH.read_text(encoding="utf-8"))
    old_config = json.loads(OLD_CONFIG_PATH.read_text(encoding="utf-8"))
    overrides = load_manual_overrides()

    struct_emp: dict[str, dict[str, object]] = {}
    for row in ref.get("structural_lookup_rows", []):
        struct_emp[str(row.get("position_master_id"))] = row
    for row in ref.get("position_master_employee_summary_rows", []):
        pmid = str(row.get("position_master_id"))
        struct_emp.setdefault(pmid, row)

    pnid_rows: dict[str, list[dict[str, object]]] = {}
    for row in ref.get("non_structural_lookup_rows", []):
        pnid_rows.setdefault(str(row.get("cluster_id")), []).append(row)

    kamus_by_group: dict[str, list[dict[str, object]]] = {}
    for item in kamus_items:
        if not item.get("include_in_position_config"):
            continue
        folder = str(item.get("source_folder") or item.get("group_name") or "")
        kamus_by_group.setdefault(folder, []).append(item)

    wb = load_workbook(BACKLOG, read_only=True, data_only=True)
    ws = wb["Sheet1"]
    rows = list(ws.iter_rows(values_only=True))
    headers = [h for h in rows[0] if h]
    backlog_entries: dict[tuple[str, str], dict[str, str]] = {}
    for row in rows[1:]:
        if not any(v is not None and str(v).strip() for v in row):
            continue
        data = {headers[i]: row[i] for i in range(len(headers))}
        group = str(data.get("Group") or "").strip()
        nama_file = str(data.get("Nama File") or "").strip()
        backlog_entries[(group, nama_file)] = {
            "direktorat": str(data.get("Direktorat") or "").strip(),
            "group": group,
            "posisi_raw": str(data.get("Posisi ") or data.get("Posisi") or "").strip().rstrip(","),
            "nama_file": nama_file,
            "status": str(data.get("Status") or "").strip(),
            "tanggal_update": str(data.get("Tanggal Update Dokumen") or data.get("Tanggal Pencatatan Backlog") or "").strip(),
        }

    def group_items(group_name: str) -> list[dict[str, object]]:
        if group_name in kamus_by_group:
            return kamus_by_group[group_name]
        for folder, items in kamus_by_group.items():
            if norm_name(folder) == norm_name(group_name):
                return items
        return []

    def resolve_from_old(item: dict[str, object]) -> dict[str, str | None]:
        sheet = item.get("sheet_name")
        folder = item.get("source_folder") or item.get("group_name")
        override = overrides.get((str(sheet), str(folder)))
        if override:
            return override
        matches = [
            p
            for p in old_config.get("positions", [])
            if p.get("sheet_name") == sheet
            and str(folder) in str(p.get("source_workbook", ""))
        ]
        if not matches:
            matches = [
                p
                for p in old_config.get("positions", [])
                if p.get("sheet_name") == sheet
                and norm_name(folder or "") in norm_name(str(p.get("source_workbook", "")))
            ]
        if not matches:
            return {"shape": None, "id": None, "production_title": None, "mapping_source": "unmapped"}
        pos = matches[0]
        pmid = pos.get("position_master_id")
        pnid = pos.get("position_nomenclature_id")
        return {
            "shape": "PMID" if pmid else "PNID",
            "id": str(pmid or pnid),
            "production_title": pos.get("portaverse_position_title") or pos.get("position_name"),
            "mapping_source": "approved_config_20260709",
        }

    def employees(shape: str | None, ident: str | None) -> tuple[list[dict[str, str | None]], str | None]:
        if not shape or not ident:
            return [], "identity belum ter-resolve"
        if shape == "PMID":
            row = struct_emp.get(ident)
            if row and row.get("active_employee_names"):
                names = [x.strip() for x in str(row["active_employee_names"]).split(";")]
                nipps = [x.strip() for x in str(row.get("active_employee_nipps") or "").split(";")]
                while len(nipps) < len(names):
                    nipps.append("")
                return [{"name": n, "nipp": nip or None} for n, nip in zip(names, nipps)], None
            pm_rows = [r for r in ref.get("position_master_rows", []) if str(r.get("position_master_id")) == ident]
            if pm_rows:
                return [], "PMID ada di reference; pekerja aktif perlu verifikasi manual"
            return [], f"PMID {ident} tidak ditemukan di reference 20260727"
        cluster_rows = pnid_rows.get(ident, [])
        if not cluster_rows:
            ref_rows = [r for r in ref.get("rows", []) if str(r.get("cluster_id")) == ident]
            if ref_rows:
                return [], f"PNID {ident} ({ref_rows[0].get('cluster_label')}) — pekerja cluster perlu verifikasi"
            return [], f"PNID {ident} tidak ditemukan"
        workers: list[dict[str, str | None]] = []
        for row in cluster_rows:
            names = [x.strip() for x in str(row.get("active_employee_names") or "").split(";") if x.strip()]
            nipps = [x.strip() for x in str(row.get("active_employee_nipps") or "").split(";") if x.strip()]
            if not names:
                workers.append({"name": row.get("position_name"), "nipp": None})
            else:
                while len(nipps) < len(names):
                    nipps.append("")
                workers.extend({"name": n, "nipp": nip or None} for n, nip in zip(names, nipps))
        return workers, None

    def match_items(group_name: str, target_norm: str) -> list[dict[str, object]]:
        items = group_items(group_name)
        matched: list[dict[str, object]] = []
        for item in items:
            names = [norm_name(item.get("position_name")), norm_name(item.get("sheet_name"))]
            for name in names:
                if not name:
                    continue
                if name == target_norm or target_norm in name or name in target_norm:
                    matched.append(item)
                    break
                if target_norm.replace(" korporasi", "") in name or name.replace(" korporasi", "") in target_norm:
                    matched.append(item)
                    break
        if group_name == "Group Pengendalian Proyek" and target_norm in {
            "pimpinan proyek",
            "deputi pimpinan proyek",
            "manager proyek",
            "officer proyek",
        }:
            role_prefix = target_norm
            matched = [
                item
                for item in items
                if norm_name(item.get("sheet_name", "")).startswith(role_prefix)
                or norm_name(item.get("position_name", "")) == role_prefix
            ]
        return matched

    report: list[dict[str, object]] = []
    for entry in backlog_entries.values():
        group = entry["group"]
        targets = split_positions(entry["posisi_raw"])
        if targets == ["__ALL__"]:
            selected_items = group_items(group)
        else:
            selected_items = []
            for target in targets:
                selected_items.extend(match_items(group, target))
        if not selected_items and targets != ["__ALL__"]:
            for target in targets:
                report.append(
                    {
                        "backlog_direktorat": entry["direktorat"],
                        "backlog_group": group,
                        "backlog_nama_file": entry["nama_file"],
                        "backlog_posisi_diminta": entry["posisi_raw"],
                        "worksheet": None,
                        "kamus_position": target,
                        "identity_shape": None,
                        "identity_id": None,
                        "production_title": None,
                        "mapping_source": "worksheet_not_found",
                        "employee_names": "",
                        "employee_nipps": "",
                        "status_crosscheck": "BLOCKED — worksheet tidak ditemukan",
                    }
                )
            continue
        for item in selected_items:
            ident = resolve_from_old(item)
            workers, note = employees(ident["shape"], ident["id"])
            names = "; ".join(w["name"] or "?" for w in workers)
            nipps = "; ".join(w["nipp"] or "?" for w in workers)
            status = "OK — siap generate" if ident["id"] and workers else (
                "PERLU REVIEW — identity/pekerja belum lengkap" if ident["id"] else "BLOCKED — belum mapped PMID/PNID"
            )
            if note and not workers:
                status = f"PERLU REVIEW — {note}"
            report.append(
                {
                    "backlog_direktorat": entry["direktorat"],
                    "backlog_group": group,
                    "backlog_nama_file": entry["nama_file"],
                    "backlog_posisi_diminta": entry["posisi_raw"],
                    "worksheet": item.get("sheet_name"),
                    "kamus_position": item.get("position_name"),
                    "identity_shape": ident["shape"],
                    "identity_id": ident["id"],
                    "production_title": ident["production_title"],
                    "mapping_source": ident["mapping_source"],
                    "employee_names": names,
                    "employee_nipps": nipps,
                    "status_crosscheck": status,
                }
            )

    unique: OrderedDict[tuple[object, ...], dict[str, object]] = OrderedDict()
    for row in report:
        key = (row["backlog_group"], row["worksheet"], row["identity_id"], row["kamus_position"])
        unique[key] = row
    report = list(unique.values())

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    json_path = OUT_DIR / "crosscheck_posisi_pekerja.json"
    csv_path = OUT_DIR / "crosscheck_posisi_pekerja.csv"
    json_path.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    if report:
        with csv_path.open("w", newline="", encoding="utf-8-sig") as handle:
            writer = csv.DictWriter(handle, fieldnames=list(report[0].keys()))
            writer.writeheader()
            writer.writerows(report)

    groups = sorted({r["backlog_group"] for r in report})
    print(f"Backlog groups: {len(backlog_entries)}")
    print(f"Crosscheck rows: {len(report)}")
    print(f"Mapped identities: {sum(1 for r in report if r['identity_id'])}")
    print(f"With workers: {sum(1 for r in report if r['employee_names'])}")
    print(f"Saved JSON: {json_path}")
    print(f"Saved CSV: {csv_path}")
    for group in groups:
        items = [r for r in report if r["backlog_group"] == group]
        print(f"\n## {group} ({len(items)} baris)")
        for row in items:
            ident = f"{row['identity_shape']} {row['identity_id']}" if row["identity_id"] else "BELUM TER-RESOLVE"
            emp = row["employee_names"] or row["status_crosscheck"]
            print(f"  - {row['kamus_position']} | {row['worksheet']} | {ident} | {row['production_title'] or '—'} | {emp}")


if __name__ == "__main__":
    main()
