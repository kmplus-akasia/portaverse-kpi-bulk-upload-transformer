#!/usr/bin/env python3
"""Reconcile roster organization fields (title + org superior unit) vs production."""

from __future__ import annotations

import json
import re
import zipfile
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any
from xml.etree import ElementTree as ET

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

import position_mapping as pm

NS = {"m": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
REL = "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}"
OUT = Path("outputs/kamus-group2-subholding-roster-fresh-20260806")
ROSTER_PATH = Path(
    "outputs/kamus-group2-subholding-roster-mapping-20260806/source/REGIONAL dan SUBHOLDING.xlsx"
)
REFERENCE_PATH = Path("configs/production_position_reference.json")
NAVY, TEAL, PALE = "173651", "138074", "E9F1F8"
BODY = "Aptos"
FILLS = {
    "exact": "D9EAD3",
    "contains": "D9EAD3",
    "strong_overlap": "FFF2CC",
    "partial_overlap": "FCE4D6",
    "mismatch": "F4CCCC",
    "roster_org_empty": "E7E6E6",
    "production_org_empty": "E7E6E6",
    "n/a": "E7E6E6",
    "n/a_missing_prod": "E7E6E6",
    "missing": "E7E6E6",
}


def norm(value: Any) -> str:
    return "" if value is None else str(value).strip()


def parse_stext(stext: str) -> tuple[str, str, str, str]:
    raw = re.sub(r"\s+", " ", (stext or "").replace("\n", " ").replace("\r", " ")).strip()
    parts = [part.strip() for part in raw.split("#")]
    title = parts[0] if parts else ""
    org = parts[1] if len(parts) >= 2 else ""
    flag = parts[2] if len(parts) >= 3 else ""
    return title, org, flag, raw


def load_shared_strings(archive: zipfile.ZipFile) -> list[str]:
    root = ET.fromstring(archive.read("xl/sharedStrings.xml"))
    out: list[str] = []
    for item in root.findall("m:si", NS):
        texts = [node.text or "" for node in item.iter("{http://schemas.openxmlformats.org/spreadsheetml/2006/main}t")]
        out.append("".join(texts))
    return out


def cell_value(cell: ET.Element, shared: list[str]) -> Any:
    cell_type = cell.attrib.get("t")
    value_node = cell.find("m:v", NS)
    if value_node is None:
        return None
    raw = value_node.text
    return shared[int(raw)] if cell_type == "s" else raw


def col_index(cell_ref: str) -> int:
    match = re.match(r"([A-Z]+)", cell_ref)
    assert match is not None
    total = 0
    for char in match.group(1):
        total = total * 26 + (ord(char) - 64)
    return total - 1


def load_roster(path: Path) -> dict[str, dict[str, str]]:
    roster: dict[str, dict[str, str]] = {}
    with zipfile.ZipFile(path) as archive:
        shared = load_shared_strings(archive)
        workbook = ET.fromstring(archive.read("xl/workbook.xml"))
        rels = ET.fromstring(archive.read("xl/_rels/workbook.xml.rels"))
        rid = {rel.attrib["Id"]: rel.attrib["Target"] for rel in rels}
        for sheet in workbook.findall("m:sheets/m:sheet", NS):
            name = sheet.attrib["name"]
            if name not in {"SPTP", "SPMT", "SPSL", "SPJM"}:
                continue
            target = "xl/" + rid[sheet.attrib[f"{REL}id"]].lstrip("/")
            root = ET.fromstring(archive.read(target))
            rows: list[list[Any]] = []
            for row in root.findall("m:sheetData/m:row", NS):
                cells: dict[int, Any] = {}
                for cell in row.findall("m:c", NS):
                    cells[col_index(cell.attrib.get("r", "A1"))] = cell_value(cell, shared)
                if cells:
                    width = max(cells) + 1
                    rows.append([cells.get(index) for index in range(width)])
            header = [norm(value) for value in rows[0]]
            index = {name: idx for idx, name in enumerate(header)}
            for row in rows[1:]:
                if not row or all(value in (None, "") for value in row):
                    continue
                nipp = norm(row[index["PNALT_NEW"]])
                if not nipp:
                    continue
                title, org, flag, raw = parse_stext(str(row[index["STEXT_STO"]] or ""))
                roster[nipp] = {
                    "sheet": name,
                    "name": norm(row[index["CNAME"]]),
                    "jobid": norm(row[index["JOBID"]]),
                    "stext_raw": raw,
                    "roster_title": title,
                    "roster_org_unit": org,
                    "roster_flag": flag,
                    "subdi": norm(row[index["SUBDI"]]),
                    "company_code": norm(row[index["COMPANY_CODE"]]),
                    "persa_text": norm(row[index["PERSA_TEXT"]]),
                    "sub_persa_text": norm(row[index["SUB_PERSA_TEXT"]]),
                    "anstx": norm(row[index["ANSTX"]]),
                }
    return roster


def title_match(left: str, right: str) -> str:
    a = pm.normalize_position_lookup(left)
    b = pm.normalize_position_lookup(right)
    if not a or not b:
        return "missing"
    if a == b:
        return "exact"
    if a in b or b in a:
        return "contains"
    ta, tb = set(a.split()), set(b.split())
    overlap = len(ta & tb) / max(len(ta), 1)
    if overlap >= 0.8:
        return "strong_overlap"
    if overlap >= 0.5:
        return "partial_overlap"
    return "mismatch"


def org_match(roster_org: str, group_name: str | None, ancestor_names: list[str]) -> str:
    roster_key = pm.normalize_position_lookup(roster_org)
    if not roster_key:
        return "roster_org_empty"
    candidates = [pm.normalize_position_lookup(group_name or "")]
    candidates.extend(pm.normalize_position_lookup(name) for name in ancestor_names)
    candidates = [item for item in candidates if item]
    if not candidates:
        return "production_org_empty"
    if any(roster_key == item for item in candidates):
        return "exact"
    if any(roster_key in item or item in roster_key for item in candidates):
        return "contains"
    roster_tokens = set(roster_key.split())
    best = 0.0
    for item in candidates:
        tokens = set(item.split())
        if roster_tokens and tokens:
            best = max(best, len(roster_tokens & tokens) / max(len(roster_tokens), 1))
    if best >= 0.8:
        return "strong_overlap"
    if best >= 0.5:
        return "partial_overlap"
    return "mismatch"


def main() -> None:
    OUT.mkdir(parents=True, exist_ok=True)
    roster = load_roster(ROSTER_PATH)
    reference = json.loads(REFERENCE_PATH.read_text(encoding="utf-8"))
    indexes = pm.build_lookup_indexes(reference)
    org_by_id = {
        int(row["group_master_id"]): row
        for row in reference.get("organization_rows", [])
        if row.get("group_master_id") not in (None, "")
    }

    def org_ancestors(group_id: int) -> list[dict[str, Any]]:
        out: list[dict[str, Any]] = []
        seen: set[int] = set()
        current: int | None = group_id
        while current is not None and current not in seen:
            seen.add(current)
            row = org_by_id.get(current)
            if not row:
                break
            out.append(row)
            parent = row.get("parent_id")
            if parent in (None, ""):
                break
            try:
                current = int(parent)
            except (TypeError, ValueError):
                break
        return out

    nipp_candidates: dict[str, list[Any]] = defaultdict(list)
    for candidate in [*indexes.structural, *indexes.non_structural]:
        for nipp in candidate.active_employee_nipps:
            key = str(nipp).strip()
            if key in roster:
                nipp_candidates[key].append(candidate)

    rows_out: list[dict[str, Any]] = []
    title_stats: Counter[str] = Counter()
    org_stats: Counter[str] = Counter()
    presence: Counter[str] = Counter()

    for nipp, info in roster.items():
        candidates = nipp_candidates.get(nipp, [])
        if not candidates:
            presence["not_in_production_lookup"] += 1
            title_stats["n/a_missing_prod"] += 1
            org_stats["n/a_missing_prod"] += 1
            rows_out.append(
                {
                    "NIPP": nipp,
                    "Nama Roster": info["name"],
                    "Roster Sheet": info["sheet"],
                    "Status Presence": "Tidak ada di production lookup aktif",
                    "Judul Posisi Roster": info["roster_title"],
                    "Unit/Atasan Org Roster (dari STEXT_STO)": info["roster_org_unit"],
                    "STEXT_STO mentah": info["stext_raw"],
                    "SUBDI Roster": info["subdi"],
                    "Perusahaan Roster": info["persa_text"],
                    "Lokasi Roster": info["sub_persa_text"],
                    "Judul Posisi Production": "",
                    "Group Production": "",
                    "Ancestor Org Production": "",
                    "Company Production": "",
                    "Scope Production": "",
                    "PMID": "",
                    "PNID": "",
                    "Group Master ID": "",
                    "Kesesuaian Judul": "n/a_missing_prod",
                    "Kesesuaian Unit/Atasan Org": "n/a_missing_prod",
                    "Catatan": "NIPP roster tidak ketemu di structural/non-structural lookup aktif",
                }
            )
            continue

        presence["in_production_lookup"] += 1
        scored: list[tuple[float, Any]] = []
        for candidate in candidates:
            title_score = pm._title_score(info["roster_title"], candidate)
            company_score = pm._context_score(info["persa_text"], candidate.company_name)
            scored.append((title_score * 0.7 + company_score * 0.3, candidate))
        scored.sort(key=lambda item: -item[0])
        chosen = scored[0][1]
        source_row = chosen.source_row or {}
        raw_group_id = source_row.get("group_master_id")
        try:
            group_id = int(raw_group_id) if raw_group_id not in (None, "") else None
        except (TypeError, ValueError):
            group_id = None
        ancestors = org_ancestors(group_id) if group_id is not None else []
        ancestor_names = [norm(row.get("group_name")) for row in ancestors]
        if not ancestor_names and chosen.group_ancestor_names:
            ancestor_names = [norm(name) for name in chosen.group_ancestor_names]
        title_label = title_match(info["roster_title"], chosen.title or "")
        org_label = org_match(info["roster_org_unit"], chosen.group_name, ancestor_names)
        title_stats[title_label] += 1
        org_stats[org_label] += 1

        note = ""
        group_master_id = str(raw_group_id or "")
        if info["subdi"] and group_master_id and group_master_id == info["subdi"]:
            note = "SUBDI = group_master_id"
        elif info["subdi"]:
            ancestor_ids = [str(row.get("group_master_id")) for row in ancestors]
            if info["subdi"] in ancestor_ids:
                note = "SUBDI cocok ancestor group_master_id"
            else:
                note = "SUBDI tidak cocok group_master_id"

        rows_out.append(
            {
                "NIPP": nipp,
                "Nama Roster": info["name"],
                "Roster Sheet": info["sheet"],
                "Status Presence": "Ada di production lookup",
                "Judul Posisi Roster": info["roster_title"],
                "Unit/Atasan Org Roster (dari STEXT_STO)": info["roster_org_unit"],
                "STEXT_STO mentah": info["stext_raw"],
                "SUBDI Roster": info["subdi"],
                "Perusahaan Roster": info["persa_text"],
                "Lokasi Roster": info["sub_persa_text"],
                "Judul Posisi Production": chosen.title or "",
                "Group Production": chosen.group_name or "",
                "Ancestor Org Production": " > ".join([name for name in ancestor_names if name][:6]),
                "Company Production": chosen.company_name or "",
                "Scope Production": chosen.scope,
                "PMID": chosen.position_master_id or "",
                "PNID": chosen.position_nomenclature_id or "",
                "Group Master ID": group_master_id,
                "Kesesuaian Judul": title_label,
                "Kesesuaian Unit/Atasan Org": org_label,
                "Catatan": note,
            }
        )

    strong = {"exact", "contains", "strong_overlap"}
    title_good = sum(1 for row in rows_out if row["Kesesuaian Judul"] in strong)
    org_good = sum(1 for row in rows_out if row["Kesesuaian Unit/Atasan Org"] in strong)
    both_good = sum(
        1
        for row in rows_out
        if row["Kesesuaian Judul"] in strong and row["Kesesuaian Unit/Atasan Org"] in strong
    )

    book = Workbook()
    guide = book.active
    guide.title = "Baca Dulu"
    guide["A1"] = "Rekap Kesesuaian Organisasi: Roster vs Production"
    guide["A1"].font = Font(name=BODY, size=16, bold=True, color="FFFFFF")
    guide["A1"].fill = PatternFill("solid", fgColor=NAVY)
    guide.merge_cells("A1:B1")
    blocks = [
        (
            "Bagaimana atasan dibaca dari workbook roster?",
            "Workbook roster (REGIONAL dan SUBHOLDING.xlsx) TIDAK punya kolom nama/NIPP atasan orang.\n\n"
            "Atasan organisasi dibaca dari kolom STEXT_STO dengan pemisah '#':\n\n"
            "  <Judul Posisi> # <Unit Organisasi / Atasan Org> # <Flag>\n\n"
            "Contoh:\n"
            "  Senior Officer HSSE # Dinas Perencanaan dan Pengendalian Operasi # Penugasan\n"
            "  → Judul posisi = Senior Officer HSSE\n"
            "  → Unit/Atasan org = Dinas Perencanaan dan Pengendalian Operasi\n"
            "  → Flag = Penugasan\n\n"
            "Kolom pendukung: SUBDI (ID unit org roster), PERSA_TEXT/SUB_PERSA_TEXT, JOBID.\n"
            "Jadi 'atasan' di rekap ini = unit organisasi atasan, bukan nama pejabat atasan.",
        ),
        (
            "Apa yang dibandingkan dengan production?",
            "1) Judul posisi: STEXT_STO kiri '#' vs position_name production tempat NIPP menjabat\n"
            "2) Unit/Atasan org: STEXT_STO tengah '#' vs group_name + rantai parent org production\n"
            "3) Presence: apakah NIPP roster ada di lookup production aktif\n"
            "4) Catatan: apakah SUBDI cocok group_master_id",
        ),
        (
            "Label kesesuaian",
            "exact/contains = cocok kuat; strong_overlap = mirip kuat; partial_overlap = mirip sebagian;\n"
            "mismatch = tidak cocok; roster_org_empty = unit di STEXT_STO kosong;\n"
            "n/a_missing_prod = NIPP tidak ketemu di production lookup.",
        ),
    ]
    row_idx = 3
    for title, body in blocks:
        guide.cell(row_idx, 1, title).font = Font(name=BODY, bold=True, color=NAVY, size=12)
        guide.cell(row_idx, 1).fill = PatternFill("solid", fgColor=PALE)
        guide.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=2)
        row_idx += 1
        guide.cell(row_idx, 1, body).alignment = Alignment(wrap_text=True, vertical="top")
        guide.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=2)
        guide.row_dimensions[row_idx].height = 130
        row_idx += 2
    guide.column_dimensions["A"].width = 110

    summary = book.create_sheet("Ringkasan")
    summary["A1"] = "Ringkasan Kesesuaian Roster Subholding (2.705) vs Production"
    summary["A1"].font = Font(name=BODY, size=16, bold=True, color="FFFFFF")
    summary["A1"].fill = PatternFill("solid", fgColor=NAVY)
    summary.merge_cells("A1:B1")
    meta = [
        ("Unique NIPP roster", len(roster)),
        ("Ada di production lookup", presence["in_production_lookup"]),
        ("Tidak ada di production lookup", presence["not_in_production_lookup"]),
        ("Judul cocok kuat (exact/contains/strong)", title_good),
        ("Unit/Atasan org cocok kuat", org_good),
        ("Judul + Unit cocok kuat", both_good),
        ("Judul mismatch", title_stats.get("mismatch", 0)),
        ("Unit/Atasan mismatch", org_stats.get("mismatch", 0)),
        ("Unit roster kosong", org_stats.get("roster_org_empty", 0)),
    ]
    summary["A3"] = "Item"
    summary["B3"] = "Nilai"
    for cell in (summary["A3"], summary["B3"]):
        cell.fill = PatternFill("solid", fgColor=TEAL)
        cell.font = Font(name=BODY, bold=True, color="FFFFFF")
    for offset, (label, value) in enumerate(meta, start=4):
        summary.cell(offset, 1, label)
        summary.cell(offset, 2, value)
    summary["A15"] = "Distribusi Kesesuaian Judul"
    summary["A15"].font = Font(name=BODY, bold=True, color=NAVY)
    cursor = 16
    for label, count in title_stats.most_common():
        summary.cell(cursor, 1, label)
        summary.cell(cursor, 2, count)
        cursor += 1
    cursor += 1
    summary.cell(cursor, 1, "Distribusi Kesesuaian Unit/Atasan Org").font = Font(
        name=BODY, bold=True, color=NAVY
    )
    cursor += 1
    for label, count in org_stats.most_common():
        summary.cell(cursor, 1, label)
        summary.cell(cursor, 2, count)
        cursor += 1
    summary.column_dimensions["A"].width = 48
    summary.column_dimensions["B"].width = 18

    headers = list(rows_out[0].keys())
    detail = book.create_sheet("Detail per NIPP")
    detail["A1"] = "Detail kesesuaian per NIPP"
    detail["A1"].fill = PatternFill("solid", fgColor=NAVY)
    detail["A1"].font = Font(name=BODY, bold=True, color="FFFFFF", size=13)
    detail.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    for col, header in enumerate(headers, start=1):
        cell = detail.cell(3, col, header)
        cell.fill = PatternFill("solid", fgColor=TEAL)
        cell.font = Font(name=BODY, bold=True, color="FFFFFF")
        cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
    title_col = headers.index("Kesesuaian Judul") + 1
    org_col = headers.index("Kesesuaian Unit/Atasan Org") + 1
    nipp_col = headers.index("NIPP") + 1
    for row_no, row in enumerate(rows_out, start=4):
        for col, header in enumerate(headers, start=1):
            cell = detail.cell(row_no, col, row[header])
            cell.font = Font(name=BODY, size=8)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if col == nipp_col:
                cell.number_format = "@"
        for col, key in ((title_col, "Kesesuaian Judul"), (org_col, "Kesesuaian Unit/Atasan Org")):
            label = str(row[key])
            if label in FILLS:
                detail.cell(row_no, col).fill = PatternFill("solid", fgColor=FILLS[label])
    end_row = 3 + len(rows_out)
    table = Table(displayName="OrgReconcileTable", ref=f"A3:{get_column_letter(len(headers))}{end_row}")
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    detail.add_table(table)
    widths = [12, 22, 10, 22, 28, 28, 40, 12, 24, 18, 28, 24, 40, 24, 12, 10, 10, 12, 14, 16, 28]
    for idx, width in enumerate(widths[: len(headers)], start=1):
        detail.column_dimensions[get_column_letter(idx)].width = width
    detail.freeze_panes = "E4"

    for sheet_name, predicate in [
        ("Judul Mismatch", lambda row: row["Kesesuaian Judul"] == "mismatch"),
        ("Unit Atasan Mismatch", lambda row: row["Kesesuaian Unit/Atasan Org"] == "mismatch"),
        ("Tidak di Production", lambda row: str(row["Status Presence"]).startswith("Tidak")),
    ]:
        subset = [row for row in rows_out if predicate(row)]
        sheet = book.create_sheet(sheet_name)
        sheet["A1"] = f"{sheet_name} ({len(subset)} NIPP)"
        sheet["A1"].fill = PatternFill("solid", fgColor=NAVY)
        sheet["A1"].font = Font(name=BODY, bold=True, color="FFFFFF")
        for col, header in enumerate(headers, start=1):
            cell = sheet.cell(3, col, header)
            cell.fill = PatternFill("solid", fgColor=TEAL)
            cell.font = Font(name=BODY, bold=True, color="FFFFFF")
        for row_no, row in enumerate(subset, start=4):
            for col, header in enumerate(headers, start=1):
                cell = sheet.cell(row_no, col, row[header])
                cell.font = Font(name=BODY, size=8)
                if header == "NIPP":
                    cell.number_format = "@"
        for idx, width in enumerate(widths[: len(headers)], start=1):
            sheet.column_dimensions[get_column_letter(idx)].width = width
        sheet.freeze_panes = "E4"

    output = OUT / "Rekap_Kesesuaian_Organisasi_Roster_vs_Production_20260806.xlsx"
    book.save(output)
    (OUT / "REKAP_KESESUAIAN_ORGANISASI.md").write_text(
        f"""# Rekap Kesesuaian Organisasi — Roster vs Production

## Mekanisme pembacaan atasan dari workbook roster

Workbook `REGIONAL dan SUBHOLDING.xlsx` **tidak** memiliki kolom nama/NIPP atasan orang.

Atasan organisasi dibaca dari kolom **`STEXT_STO`** dengan pemisah `#`:

```text
<Judul Posisi> # <Unit Organisasi / Atasan Org> # <Flag>
```

Contoh: `Senior Officer HSSE # Dinas Perencanaan dan Pengendalian Operasi # Penugasan`

## Angka ringkas (2.705 NIPP)

| Metric | Count |
| --- | ---: |
| Ada di production lookup | {presence['in_production_lookup']} |
| Tidak ada di production lookup | {presence['not_in_production_lookup']} |
| Judul cocok kuat | {title_good} |
| Unit/Atasan org cocok kuat | {org_good} |
| Judul + Unit cocok kuat | {both_good} |
| Judul mismatch | {title_stats.get('mismatch', 0)} |
| Unit/Atasan mismatch | {org_stats.get('mismatch', 0)} |
| Unit roster kosong | {org_stats.get('roster_org_empty', 0)} |

## Artifact
`{output}`
""",
        encoding="utf-8",
    )
    print(
        json.dumps(
            {
                "roster": len(roster),
                "presence": dict(presence),
                "title_stats": dict(title_stats),
                "org_stats": dict(org_stats),
                "title_good": title_good,
                "org_good": org_good,
                "both_good": both_good,
                "output": str(output),
            },
            ensure_ascii=False,
            indent=2,
        )
    )


if __name__ == "__main__":
    main()
