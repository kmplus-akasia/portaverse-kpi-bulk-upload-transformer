#!/usr/bin/env python3
"""Export Subholding mappings that are already considered correct."""

from __future__ import annotations

import json
import re
from collections import Counter
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

OUT = Path("outputs/kamus-group2-subholding-roster-mapping-20260806")
DRAFT = OUT / "subholding_roster_position_first_mapping_20260806.json"
OUTPUT = OUT / "Mapping_Sudah_Benar_Subholding_20260806.xlsx"

NAVY, TEAL, PALE, GREEN = "173651", "138074", "E9F1F8", "D9EAD3"
BODY = "Aptos"


def norm(value: object) -> str:
    return "" if value is None else str(value).strip()


def base(path: object) -> str:
    text = norm(path).replace("\\", "/")
    return text.split("/")[-1].casefold() if text else ""


def sheet_key(value: object) -> str:
    return re.sub(r"[^a-z0-9]+", "", norm(value).casefold())


def same_workbook(left: object, right: object) -> bool:
    a, b = base(left), base(right)
    if not a or not b:
        return False
    if a == b:
        return True

    def simplify(name: str) -> str:
        return name.replace(" - mapping dengan kontrak manajemen.xlsx", ".xlsx").replace(
            " mapping dengan kontrak manajemen.xlsx", ".xlsx"
        )

    sa, sb = simplify(a), simplify(b)
    return sa == sb or sa in sb or sb in sa


def same_sheet(left: object, right: object) -> bool:
    a, b = sheet_key(left), sheet_key(right)
    if not a or not b:
        return False
    return a == b or a in b or b in a


def main() -> None:
    payload = json.loads(DRAFT.read_text(encoding="utf-8"))
    rows = payload.get("rows", [])
    confirmed: list[dict[str, object]] = []

    for row in rows:
        confidence = norm(row.get("Confidence Label"))
        confirm = norm(row.get("Reviewer Confirm Mapping")).upper()
        candidate_workbook = norm(row.get("Candidate Source Workbook"))
        candidate_sheet = norm(row.get("Candidate Worksheet"))
        reviewer_workbook = norm(row.get("Reviewer Source Workbook"))
        reviewer_sheet = norm(row.get("Reviewer Worksheet"))
        reviewed_workbook = norm(row.get("Reviewed Workbook Title"))
        reviewed_sheet = norm(row.get("Reviewed Worksheet Title"))
        status = norm(row.get("Inventory Resolve Status"))
        tag = norm(row.get("Roster Review Tag"))

        high_ok = (
            confidence == "high_confidence"
            and confirm == "YES"
            and bool(candidate_workbook and candidate_sheet)
            and (
                status == "accepted_high_confidence_candidate"
                or (
                    same_workbook(reviewer_workbook, candidate_workbook)
                    and same_sheet(reviewer_sheet, candidate_sheet)
                )
            )
        )
        rw_match_auto = (
            confirm == "YES"
            and bool(candidate_workbook and candidate_sheet and reviewer_workbook and reviewer_sheet)
            and (
                (
                    same_workbook(reviewer_workbook, candidate_workbook)
                    and same_sheet(reviewer_sheet, candidate_sheet)
                )
                or (
                    same_workbook(reviewed_workbook, candidate_workbook)
                    and same_sheet(reviewed_sheet, candidate_sheet)
                )
            )
            and "NEEDS_REVIEW_NEW_56" not in tag
        )

        reasons: list[str] = []
        if high_ok:
            reasons.append("high_confidence_accepted")
        if rw_match_auto and "high_confidence_accepted" not in reasons:
            reasons.append("rw_matches_automatic")
        if not reasons:
            continue

        workbook_path = reviewer_workbook or candidate_workbook
        if "SUBHOLDING" not in workbook_path.upper():
            continue

        confirmed.append(
            {
                "Alasan Termasuk": "; ".join(reasons),
                "Confidence Draft": confidence,
                "Status Pencocokan": status,
                "PMID": norm(row.get("PMID")),
                "PNID": norm(row.get("PNID")),
                "Judul Posisi": norm(row.get("Position Title")),
                "Perusahaan": norm(row.get("Company")),
                "Unit / Group": norm(row.get("Group / Unit")),
                "Jumlah Pegawai": row.get("Active Employees") or 0,
                "NIPP Pegawai": norm(row.get("Active Employee NIPPs")),
                "Nama Pegawai": norm(row.get("Active Employee Names")),
                "File Kamus (dipakai)": reviewer_workbook,
                "Sheet Kamus (dipakai)": reviewer_sheet,
                "Draft Otomatis File": candidate_workbook,
                "Draft Otomatis Sheet": candidate_sheet,
                "File Referensi R&W": reviewed_workbook,
                "Sheet Referensi R&W": reviewed_sheet,
                "Roster Subholding": norm(row.get("Roster Sheet")),
            }
        )

    nipps: set[str] = set()
    for row in confirmed:
        for part in str(row["NIPP Pegawai"]).split(";"):
            nipp = part.strip()
            if nipp:
                nipps.add(nipp)

    book = Workbook()
    summary = book.active
    summary.title = "Ringkasan"
    summary["A1"] = "Daftar Mapping yang Sudah Benar — Subholding Roster"
    summary["A1"].font = Font(name=BODY, size=16, bold=True, color="FFFFFF")
    summary["A1"].fill = PatternFill("solid", fgColor=NAVY)
    summary.merge_cells("A1:B1")
    summary["A2"] = (
        "Kriteria: (1) high_confidence yang diterima, dan/atau "
        "(2) usulan Red & White cocok dengan draft otomatis (file+sheet). "
        "Hanya workbook keluarga KAMUS KPI SUBHOLDING."
    )
    summary["A2"].fill = PatternFill("solid", fgColor=PALE)
    summary.merge_cells("A2:B2")
    meta = [
        ("Baris posisi", len(confirmed)),
        ("Unique NIPP tercakup", len(nipps)),
        ("Sum Jumlah Pegawai (bisa > unique)", sum(int(row["Jumlah Pegawai"] or 0) for row in confirmed)),
        (
            "high_confidence_accepted",
            sum(1 for row in confirmed if "high_confidence_accepted" in str(row["Alasan Termasuk"])),
        ),
        (
            "rw_matches_automatic",
            sum(1 for row in confirmed if row["Alasan Termasuk"] == "rw_matches_automatic"),
        ),
        ("Sumber draft", str(DRAFT)),
    ]
    summary["A4"] = "Item"
    summary["B4"] = "Nilai"
    for cell in (summary["A4"], summary["B4"]):
        cell.fill = PatternFill("solid", fgColor=TEAL)
        cell.font = Font(name=BODY, bold=True, color="FFFFFF")
    for offset, (label, value) in enumerate(meta, start=5):
        summary.cell(offset, 1, label)
        summary.cell(offset, 2, value)
    summary.column_dimensions["A"].width = 42
    summary.column_dimensions["B"].width = 90

    headers = list(confirmed[0].keys()) if confirmed else []
    sheet = book.create_sheet("Mapping Sudah Benar")
    sheet["A1"] = "Baris yang sudah dianggap benar (high confidence dan/atau R&W = otomatis)"
    sheet["A1"].fill = PatternFill("solid", fgColor=NAVY)
    sheet["A1"].font = Font(name=BODY, size=13, bold=True, color="FFFFFF")
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(len(headers), 1))
    for col, header in enumerate(headers, start=1):
        cell = sheet.cell(3, col, header)
        cell.fill = PatternFill("solid", fgColor=TEAL)
        cell.font = Font(name=BODY, bold=True, color="FFFFFF")
        cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
    nipp_col = headers.index("NIPP Pegawai") + 1 if headers else 1
    reason_col = headers.index("Alasan Termasuk") + 1 if headers else 1
    for row_idx, row in enumerate(confirmed, start=4):
        for col_idx, header in enumerate(headers, start=1):
            cell = sheet.cell(row_idx, col_idx, row[header])
            cell.font = Font(name=BODY, size=9)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if col_idx == nipp_col:
                cell.number_format = "@"
        sheet.cell(row_idx, reason_col).fill = PatternFill("solid", fgColor=GREEN)
    end_row = 3 + max(len(confirmed), 1)
    if confirmed:
        table = Table(
            displayName="MappingSudahBenarTable",
            ref=f"A3:{get_column_letter(len(headers))}{end_row}",
        )
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
        sheet.add_table(table)
    widths = [28, 14, 28, 10, 10, 32, 28, 24, 10, 24, 28, 48, 26, 48, 22, 24, 22, 12]
    for idx, width in enumerate(widths[: len(headers)], start=1):
        sheet.column_dimensions[get_column_letter(idx)].width = width
    sheet.freeze_panes = "F4"
    sheet.row_dimensions[3].height = 30

    criteria = book.create_sheet("Kriteria")
    criteria["A1"] = 'Definisi "sudah benar"'
    criteria["A1"].font = Font(name=BODY, size=14, bold=True, color="FFFFFF")
    criteria["A1"].fill = PatternFill("solid", fgColor=NAVY)
    lines = [
        "1. high_confidence_accepted: Confidence Draft = high_confidence, Keputusan YES, dan usulan memakai draft otomatis.",
        "2. rw_matches_automatic: Keputusan YES, dan File+Sheet yang dipakai (atau referensi R&W) cocok dengan Draft Otomatis.",
        "3. Filter tambahan: File Kamus (dipakai) harus path keluarga KAMUS KPI SUBHOLDING.",
        "4. Baris NEEDS_REVIEW_NEW_56 tidak dimasukkan.",
        "5. Unique NIPP = pekerja yang punya ≥1 baris posisi sudah benar; belum tentu semua jabatan ganda orang itu sudah benar.",
    ]
    for offset, line in enumerate(lines, start=3):
        criteria.cell(offset, 1, line)
    criteria.column_dimensions["A"].width = 140

    book.save(OUTPUT)
    print(
        json.dumps(
            {
                "rows": len(confirmed),
                "unique_nipps": len(nipps),
                "reasons": dict(Counter(str(row["Alasan Termasuk"]) for row in confirmed)),
                "output": str(OUTPUT),
            },
            ensure_ascii=False,
            indent=2,
        )
    )


if __name__ == "__main__":
    main()
