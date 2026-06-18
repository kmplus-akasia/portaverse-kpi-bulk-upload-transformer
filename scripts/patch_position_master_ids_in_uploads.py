#!/usr/bin/env python3
"""Patch exact Position Master ID values in generated KPI upload workbooks."""

from __future__ import annotations

import argparse
import csv
from pathlib import Path

import openpyxl


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser()
    parser.add_argument("--output-dir", required=True, type=Path)
    parser.add_argument("--replacement", action="append", required=True)
    parser.add_argument("--audit-output", required=True, type=Path)
    return parser.parse_args()


def parse_replacements(values: list[str]) -> dict[tuple[str, str], str]:
    replacements: dict[tuple[str, str], str] = {}
    for value in values:
        try:
            position_name, old_pmid, new_pmid = value.split("|", 2)
        except ValueError as exc:
            raise SystemExit(
                "--replacement must use 'position name|old pmid|new pmid'"
            ) from exc
        replacements[(position_name, old_pmid)] = new_pmid
    return replacements


def iter_upload_workbooks(output_dir: Path) -> list[Path]:
    return sorted(
        path
        for path in output_dir.rglob("*.xlsx")
        if "Conversion Report" not in path.name and "/upload-ready/" not in str(path)
    )


def main() -> int:
    args = parse_args()
    replacements = parse_replacements(args.replacement)
    audit_rows: list[dict[str, object]] = []

    for path in iter_upload_workbooks(args.output_dir):
        workbook = openpyxl.load_workbook(path)
        if "KPI Template" not in workbook.sheetnames:
            continue
        sheet = workbook["KPI Template"]
        changed = False
        for row_idx in range(2, sheet.max_row + 1):
            position_name = str(sheet.cell(row_idx, 4).value or "").strip()
            pmid = str(sheet.cell(row_idx, 5).value or "").strip()
            new_pmid = replacements.get((position_name, pmid))
            if not new_pmid:
                continue
            sheet.cell(row_idx, 5).value = new_pmid
            audit_rows.append(
                {
                    "workbook": str(path),
                    "row": row_idx,
                    "position_name": position_name,
                    "old_pmid": pmid,
                    "new_pmid": new_pmid,
                }
            )
            changed = True
        if changed:
            workbook.save(path)

    args.audit_output.parent.mkdir(parents=True, exist_ok=True)
    with args.audit_output.open("w", newline="") as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=["workbook", "row", "position_name", "old_pmid", "new_pmid"],
        )
        writer.writeheader()
        writer.writerows(audit_rows)

    print(f"patched_rows={len(audit_rows)}")
    print(f"audit_output={args.audit_output}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
