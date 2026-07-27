#!/usr/bin/env python3
"""Build reviewer-friendly mapping override candidates from conflict reviews."""

from __future__ import annotations

import argparse
from collections import Counter
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


IDENTITY_COLUMNS = ("Source Workbook", "Worksheet")
APPROVAL_OPTIONS = '"YES,NO"'
SCOPE_OPTIONS = '"structural,non_structural"'


def norm(value: Any) -> str:
    return "" if value is None else str(value).strip()


def numeric(value: Any) -> float:
    try:
        return float(value or 0)
    except (TypeError, ValueError):
        return 0.0


def conflict_key(row: dict[str, Any]) -> tuple[str, ...]:
    source = norm(row.get("Source Workbook"))
    worksheet = norm(row.get("Worksheet")) or norm(row.get("Sheet"))
    return source, worksheet


def candidate_identity(row: dict[str, Any]) -> tuple[str, str]:
    scope = norm(row.get("Inferred Scope")) or norm(row.get("Candidate Scope"))
    candidate_id = norm(row.get("Candidate PMID")) if scope == "structural" else norm(row.get("Candidate PNID"))
    return scope, candidate_id


def load_review_rows(path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet_name = None
    for candidate in ["Position Mapping Report", "Review Queue", "Mapping Conflict Review"]:
        if candidate in workbook.sheetnames:
            sheet_name = candidate
            break
    if sheet_name is None:
        raise ValueError(f"{path} does not contain Position Mapping Report")
    worksheet = workbook[sheet_name]
    rows = worksheet.iter_rows(values_only=True)
    headers = [norm(value) for value in next(rows)]
    return [dict(zip(headers, row)) for row in rows]


def status_for(best: dict[str, Any], runner_up: dict[str, Any] | None, duplicate_best_count: int) -> str:
    label = norm(best.get("Confidence Label"))
    if label in {"low_confidence", "scope_uncertain", "no_candidate", "mapping_conflict"}:
        return label
    score = numeric(best.get("Candidate Score"))
    runner_up_score = numeric(runner_up.get("Candidate Score")) if runner_up else 0.0
    if not candidate_identity(best)[1]:
        return "no_candidate"
    if score >= 0.8 and duplicate_best_count == 1 and score - runner_up_score >= 0.15:
        return "review_recommended_high_confidence"
    if score >= 0.65 and duplicate_best_count == 1 and score - runner_up_score >= 0.10:
        return "review_recommended"
    if runner_up and abs(score - runner_up_score) < 0.10:
        return "ambiguous_candidates"
    return "manual_review_required"


def build_override_rows(review_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    grouped: dict[tuple[str, ...], list[dict[str, Any]]] = {}
    for row in review_rows:
        grouped.setdefault(conflict_key(row), []).append(row)

    output_rows: list[dict[str, Any]] = []
    for key, rows in grouped.items():
        sorted_rows = sorted(rows, key=lambda row: numeric(row.get("Candidate Score")), reverse=True)
        best = sorted_rows[0]
        best_score = numeric(best.get("Candidate Score"))
        duplicate_best_count = sum(
            1
            for row in sorted_rows
            if candidate_identity(row) == candidate_identity(best)
            and abs(numeric(row.get("Candidate Score")) - best_score) < 0.0001
        )
        runner_up = next((row for row in sorted_rows[1:] if candidate_identity(row) != candidate_identity(best)), None)
        scope, candidate_id = candidate_identity(best)
        label = norm(best.get("Confidence Label"))
        output_rows.append(
            {
                "Review Status": status_for(best, runner_up, duplicate_best_count),
                "Approved": "",
                "Source Workbook": key[0],
                "Sheet": key[1],
                "Raw Group": norm(best.get("Candidate Group")),
                "Raw Position": norm(best.get("Raw Worksheet Title")) or norm(best.get("Candidate Title")),
                "Direktorat": "",
                "Original Confidence Label": label,
                "Suggested Position Scope": scope if scope in {"structural", "non_structural"} else "",
                "Suggested Position Master ID": candidate_id if scope == "structural" else "",
                "Suggested Position Nomenklatur ID": candidate_id if scope == "non_structural" else "",
                "Reviewer Selected Scope": "",
                "Reviewer Selected Position Master ID": "",
                "Reviewer Selected Position Nomenklatur ID": "",
                "Suggested Position Title": norm(best.get("Candidate Title")),
                "Suggested Group": norm(best.get("Candidate Group")),
                "Suggested Company": norm(best.get("Candidate Company")),
                "Suggested Company Code": norm(best.get("Candidate Company Code")),
                "Candidate Score": best_score,
                "Runner-up Score": numeric(runner_up.get("Candidate Score")) if runner_up else "",
                "Candidate Count": len({candidate_identity(row) for row in sorted_rows if candidate_identity(row)[1]}),
                "Active Employee Name": norm(best.get("Active Employee Name")),
                "Active Employee NIPP": norm(best.get("Active Employee NIPP")),
                "Match Reason": norm(best.get("Confidence Reason")) or norm(best.get("Match Reason")),
                "Reviewer Notes": "",
            }
        )
    return output_rows


def write_sheet(wb: Workbook, title: str, rows: list[dict[str, Any]]) -> None:
    worksheet = wb.create_sheet(title[:31])
    if not rows:
        worksheet.append(["No data"])
        return
    headers = list(rows[0].keys())
    worksheet.append(headers)
    for row in rows:
        worksheet.append([row.get(header, "") for header in headers])
    style_sheet(worksheet)
    add_approval_validation(worksheet, headers)


def add_approval_validation(worksheet: Any, headers: list[str]) -> None:
    if "Approved" not in headers or worksheet.max_row < 2:
        return
    column_index = headers.index("Approved") + 1
    column_letter = get_column_letter(column_index)
    validation = DataValidation(type="list", formula1=APPROVAL_OPTIONS, allow_blank=True)
    validation.error = "Use YES only after the suggested mapping has been reviewed and approved."
    validation.errorTitle = "Invalid approval value"
    validation.prompt = "Choose YES to approve this mapping, or leave blank/NO to skip it."
    validation.promptTitle = "Approval"
    worksheet.add_data_validation(validation)
    validation.add(f"{column_letter}2:{column_letter}{worksheet.max_row}")
    if "Reviewer Selected Scope" in headers:
        scope_column = get_column_letter(headers.index("Reviewer Selected Scope") + 1)
        scope_validation = DataValidation(type="list", formula1=SCOPE_OPTIONS, allow_blank=True)
        scope_validation.error = "Use structural or non_structural when manually resolving scope_uncertain rows."
        scope_validation.errorTitle = "Invalid scope"
        scope_validation.prompt = "Required for scope_uncertain approval when no suggested scope is available."
        scope_validation.promptTitle = "Reviewer selected scope"
        worksheet.add_data_validation(scope_validation)
        scope_validation.add(f"{scope_column}2:{scope_column}{worksheet.max_row}")


def style_sheet(worksheet: Any) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    thin_gray = Side(style="thin", color="D9E2EC")
    border = Border(left=thin_gray, right=thin_gray, top=thin_gray, bottom=thin_gray)
    status_fills = {
        "review_recommended_high_confidence": PatternFill("solid", fgColor="E2F0D9"),
        "review_recommended": PatternFill("solid", fgColor="FFF2CC"),
        "ambiguous_candidates": PatternFill("solid", fgColor="FCE4D6"),
        "manual_review_required": PatternFill("solid", fgColor="F4CCCC"),
        "no_candidate": PatternFill("solid", fgColor="D9EAD3"),
    }
    for cell in worksheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
        status = norm(row[0].value)
        fill = status_fills.get(status)
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if fill:
                cell.fill = fill
    widths = {
        "Review Status": 28,
        "Approved": 12,
        "Source Workbook": 58,
        "Raw Group": 32,
        "Raw Position": 30,
        "Suggested Position Title": 42,
        "Suggested Group": 38,
        "Suggested Company": 38,
        "Match Reason": 28,
        "Reviewer Notes": 34,
    }
    for index in range(1, worksheet.max_column + 1):
        header = norm(worksheet.cell(1, index).value)
        worksheet.column_dimensions[get_column_letter(index)].width = widths.get(header, 18)


def build_workbook(override_rows: list[dict[str, Any]], source_review: Path) -> Workbook:
    workbook = Workbook()
    workbook.remove(workbook.active)
    counts = Counter(row["Review Status"] for row in override_rows)
    summary_rows = [
        {"Metric": "Source Review Workbook", "Value": str(source_review)},
        {"Metric": "Conflict Count", "Value": len(override_rows)},
    ]
    summary_rows.extend({"Metric": f"Status: {status}", "Value": count} for status, count in sorted(counts.items()))
    write_sheet(workbook, "Summary", summary_rows)
    high_confidence = [
        row for row in override_rows if row["Review Status"] == "review_recommended_high_confidence"
    ]
    recommended = [
        row for row in override_rows if row["Review Status"] in {"review_recommended_high_confidence", "review_recommended"}
    ]
    ambiguous = [row for row in override_rows if row["Review Status"] == "ambiguous_candidates"]
    manual = [row for row in override_rows if row["Review Status"] == "manual_review_required"]
    no_candidate = [row for row in override_rows if row["Review Status"] == "no_candidate"]
    write_sheet(workbook, "High Confidence", high_confidence)
    write_sheet(workbook, "All Recommended", recommended)
    write_sheet(workbook, "Ambiguous", ambiguous)
    write_sheet(workbook, "No Candidate", no_candidate)
    write_sheet(workbook, "Manual Review", manual)
    write_sheet(workbook, "Override Candidates", override_rows)
    return workbook


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--review", required=True, type=Path)
    parser.add_argument("--output", required=True, type=Path)
    args = parser.parse_args()

    review_rows = load_review_rows(args.review)
    override_rows = build_override_rows(review_rows)
    workbook = build_workbook(override_rows, args.review)
    args.output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(args.output)
    print(f"conflicts={len(override_rows)}")
    print(f"output={args.output}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
