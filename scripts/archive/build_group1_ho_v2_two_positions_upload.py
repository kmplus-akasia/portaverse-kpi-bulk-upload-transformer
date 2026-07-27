#!/usr/bin/env python3
"""Build upload-ready KPI workbooks for the two reviewed Group 1 HO positions."""

from __future__ import annotations

import csv
import json
import shutil
from pathlib import Path
from zipfile import ZIP_DEFLATED, ZipFile

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
SOURCE_DIR = ROOT / "output/group1_ho_v2_20260709_latest_prod/conversion/upload-ready"
SOURCE_CONFIG = ROOT / "output/group1_ho_v2_20260709_latest_prod/group1_ho_v2_20260709_refreshed.config.json"
OUTPUT_DIR = ROOT / "output/group1_ho_v2_two_positions_upload_20260713"
GENERATED_DIR = OUTPUT_DIR / "conversion"
UPLOAD_DIR = OUTPUT_DIR / "upload-ready"

TARGETS = [
    {
        "source": SOURCE_DIR / "Group Keberlanjutan Korporasi - Unit Pendukung Implementasi dan Pelaporan Keberlanjutan Korporasi 07-09-2026 at 14.05 (2026 v2).xlsx",
        "source_pmid": "35776",
        "pmid": "37582",
        "pmvid": "43599",
        "employee_number": "102882",
        "employee_name": "USMAN SARONI",
        "position_name": "Group Head Keberlanjutan Korporasi",
        "output": "Formulir_Upload_KPI_PMID37582_Group_Head_Keberlanjutan_Korporasi_20260713.xlsx",
    },
    {
        "source": SOURCE_DIR / "Direktorat Wakil Direktur Utama - Group Hubungan Antar Lembaga dan Investor 07-09-2026 at 14.05 (2026 v2).xlsx",
        "source_pmid": "35810",
        "pmid": "35810",
        "pmvid": "40085",
        "employee_number": "103780",
        "employee_name": "PRAMESTIE WULANDARY",
        "position_name": "Group Head Hubungan Antar Lembaga Investor",
        "output": "Formulir_Upload_KPI_PMID35810_Group_Head_Hubungan_Antar_Lembaga_Investor_20260713.xlsx",
    },
]


def text(value: object) -> str:
    return "" if value is None else str(value).strip()


def build_workbook(target: dict[str, str | Path]) -> dict[str, object]:
    source = Path(target["source"])
    destination = GENERATED_DIR / str(target["output"])
    shutil.copy2(source, destination)

    workbook = load_workbook(destination)
    worksheet = workbook["KPI Template"]
    headers = {text(cell.value): cell.column for cell in worksheet[1]}
    pmid_column = headers["Position Master ID (Required)"]
    pmvid_column = headers["Position Master Variant ID (Optional)"]
    pnid_column = headers["Position Nomenklatur ID"]

    kept_rows = []
    for row_number in range(2, worksheet.max_row + 1):
        if text(worksheet.cell(row_number, pmid_column).value) == target["source_pmid"]:
            kept_rows.append(row_number)

    for row_number in range(worksheet.max_row, 1, -1):
        if row_number not in kept_rows:
            worksheet.delete_rows(row_number)

    for row_number in range(2, worksheet.max_row + 1):
        worksheet.cell(row_number, pmid_column).value = str(target["pmid"])
        worksheet.cell(row_number, pmvid_column).value = str(target["pmvid"])
        worksheet.cell(row_number, pnid_column).value = None

    workbook.save(destination)
    return {
        "file": destination.name,
        "pmid": target["pmid"],
        "pmvid": target["pmvid"],
        "employee_number": target["employee_number"],
        "employee_name": target["employee_name"],
        "position_name": target["position_name"],
        "rows": worksheet.max_row - 1,
    }


def main() -> None:
    if OUTPUT_DIR.exists():
        shutil.rmtree(OUTPUT_DIR)
    GENERATED_DIR.mkdir(parents=True)

    summary_rows = [build_workbook(target) for target in TARGETS]
    source_config = json.loads(SOURCE_CONFIG.read_text(encoding="utf-8"))
    selected_config = []
    for target in TARGETS:
        match = next(
            item
            for item in source_config["positions"]
            if text(item.get("position_master_id")) == target["source_pmid"]
            and text(item.get("sheet_name")) in {"Group Head", "GH Hubungan Lembaga-Investor"}
        )
        item = dict(match)
        item["position_master_id"] = str(target["pmid"])
        item["position_master_variant_id"] = str(target["pmvid"])
        item["position_nomenclature_id"] = None
        item["portaverse_position_title"] = str(target["position_name"])
        selected_config.append(item)
    config_path = OUTPUT_DIR / "group1_ho_v2_two_positions_20260713.config.json"
    config_path.write_text(json.dumps({"positions": selected_config}, indent=2), encoding="utf-8")
    manifest_path = OUTPUT_DIR / "upload_manifest.csv"
    with manifest_path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(summary_rows[0]))
        writer.writeheader()
        writer.writerows(summary_rows)

    zip_path = OUTPUT_DIR / "group1_ho_v2_two_positions_upload_ready_20260713.zip"
    with ZipFile(zip_path, "w", ZIP_DEFLATED) as archive:
        for workbook_path in sorted(GENERATED_DIR.glob("*.xlsx")):
            archive.write(workbook_path, workbook_path.name)

    summary = {
        "output_dir": str(OUTPUT_DIR),
        "upload_ready_dir": str(UPLOAD_DIR),
        "zip": str(zip_path),
        "config": str(config_path),
        "workbooks": summary_rows,
        "total_workbooks": len(summary_rows),
        "total_rows": sum(int(row["rows"]) for row in summary_rows),
    }
    (OUTPUT_DIR / "summary.json").write_text(json.dumps(summary, indent=2), encoding="utf-8")
    print(json.dumps(summary, indent=2))


if __name__ == "__main__":
    main()
