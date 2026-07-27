#!/usr/bin/env python3
from __future__ import annotations

import json
import re
import shutil
import zipfile
from collections import Counter
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook, load_workbook


ROOT = Path(__file__).resolve().parents[1]
AUDIT_WORKBOOK = Path(
    "/Users/alfredoteja/Documents/pms-codebase/dashboard-org-kpi-audit/data/output/audit-organisasi-kamus-kpi-ho.xlsx"
)
LATEST_UPLOAD_DIR = ROOT / "output/group1_ho_v2_20260709_latest_prod/conversion/upload-ready"
OUTPUT_DIR = ROOT / "output/group1_ho_v2_followup_upload_20260709"

MISSING_UPLOAD_IDENTITY = ("PMID", "37541")
PROBLEMATIC_IDENTITY = ("PNID", "12474")


def normalize(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def find_column(headers: list[str], *candidates: str) -> int:
    lowered = {header.lower(): idx + 1 for idx, header in enumerate(headers)}
    for candidate in candidates:
        if candidate.lower() in lowered:
            return lowered[candidate.lower()]
    raise KeyError(f"Missing column: {candidates}")


def sheet_rows(ws) -> Iterable[dict[str, object]]:
    row_iter = ws.iter_rows(values_only=True)
    headers = [normalize(value) for value in next(row_iter)]
    for row_idx, values in enumerate(row_iter, start=2):
        row = {headers[col]: values[col] if col < len(values) else None for col in range(len(headers))}
        row["_row"] = row_idx
        yield row


def collect_audit_cases() -> tuple[list[dict[str, object]], list[dict[str, object]]]:
    wb = load_workbook(AUDIT_WORKBOOK, data_only=True, read_only=True)
    ws = wb["Audit Posisi"]
    missing = []
    problematic = []
    for row in sheet_rows(ws):
        status = normalize(row.get("Status Ketersediaan KPI"))
        if status == "Tersedia di Kamus, Belum Terupload":
            missing.append(row)
        elif status == "KPI Parsial - Perlu Review":
            problematic.append(row)
    return missing, problematic


def row_identity(ws, row_idx: int, headers: list[str]) -> tuple[str, str]:
    pmid_col = find_column(headers, "Position Master ID (Required)")
    pnid_col = find_column(headers, "Position Nomenklatur ID")
    pmid = normalize(ws.cell(row_idx, pmid_col).value)
    pnid = normalize(ws.cell(row_idx, pnid_col).value)
    if pmid and not pnid:
        return "PMID", pmid
    if pnid and not pmid:
        return "PNID", pnid
    if pmid and pnid:
        return "DOUBLE", f"{pmid}/{pnid}"
    return "BLANK", ""


def row_identity_from_values(values: tuple[object, ...], headers: list[str]) -> tuple[str, str]:
    pmid_idx = find_column(headers, "Position Master ID (Required)") - 1
    pnid_idx = find_column(headers, "Position Nomenklatur ID") - 1
    pmid = normalize(values[pmid_idx] if pmid_idx < len(values) else None)
    pnid = normalize(values[pnid_idx] if pnid_idx < len(values) else None)
    if pmid and not pnid:
        return "PMID", pmid
    if pnid and not pmid:
        return "PNID", pnid
    if pmid and pnid:
        return "DOUBLE", f"{pmid}/{pnid}"
    return "BLANK", ""


def locate_identity_rows(identity_kind: str, identity_id: str) -> list[tuple[Path, list[int]]]:
    matches = []
    for path in sorted(LATEST_UPLOAD_DIR.glob("*.xlsx")):
        wb = load_workbook(path, data_only=False, read_only=True)
        ws = wb["KPI Template"]
        row_iter = ws.iter_rows(values_only=True)
        headers = [normalize(value) for value in next(row_iter)]
        rows = [
            row_idx
            for row_idx, row_values in enumerate(row_iter, start=2)
            if row_identity_from_values(row_values, headers) == (identity_kind, identity_id)
        ]
        if rows:
            matches.append((path, rows))
    return matches


def contiguous_segments(rows: list[int]) -> list[list[int]]:
    if not rows:
        return []
    segments = [[rows[0]]]
    for row in rows[1:]:
        if row == segments[-1][-1] + 1:
            segments[-1].append(row)
        else:
            segments.append([row])
    return segments


def write_filtered_workbook(source: Path, target: Path, keep_rows: set[int]) -> None:
    shutil.copy2(source, target)
    wb = load_workbook(target)
    ws = wb["KPI Template"]
    for row_idx in range(ws.max_row, 1, -1):
        if row_idx not in keep_rows:
            ws.delete_rows(row_idx, 1)
    wb.save(target)


def validate_upload(path: Path, expected_identity: tuple[str, str]) -> dict[str, object]:
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb["KPI Template"]
    row_iter = ws.iter_rows(values_only=True)
    headers = [normalize(value) for value in next(row_iter)]
    title_idx = find_column(headers, "Title") - 1
    type_idx = find_column(headers, "KPI Type") - 1
    errors = []
    title_counts = Counter()
    type_counts = Counter()
    identity_counts = Counter()
    numeric_titles = []
    for row_idx, row_values in enumerate(row_iter, start=2):
        identity = row_identity_from_values(row_values, headers)
        identity_counts[identity] += 1
        if identity != expected_identity:
            errors.append(f"row {row_idx}: expected {expected_identity}, got {identity}")
        title = normalize(row_values[title_idx] if title_idx < len(row_values) else None)
        title_counts[title] += 1
        if re.fullmatch(r"\d+(?:\.0+)?", title):
            numeric_titles.append((row_idx, title))
        kpi_type = normalize(row_values[type_idx] if type_idx < len(row_values) else None).upper()
        type_counts[kpi_type] += 1
    if numeric_titles:
        errors.append(f"numeric titles found: {numeric_titles[:5]}")
    return {
        "path": str(path),
        "rows": ws.max_row - 1,
        "type_counts": dict(type_counts),
        "identity_counts": {f"{kind}:{value}": count for (kind, value), count in identity_counts.items()},
        "duplicate_titles": {title: count for title, count in title_counts.items() if title and count > 1},
        "numeric_titles": numeric_titles,
        "errors": errors,
    }


def scan_latest_batch_quality() -> dict[str, object]:
    numeric_title_rows = []
    double_or_blank_identity_rows = []
    duplicate_identity_files = []
    for path in sorted(LATEST_UPLOAD_DIR.glob("*.xlsx")):
        wb = load_workbook(path, data_only=True, read_only=True)
        ws = wb["KPI Template"]
        row_iter = ws.iter_rows(values_only=True)
        headers = [normalize(value) for value in next(row_iter)]
        title_idx = find_column(headers, "Title") - 1
        identities = Counter()
        for row_idx, row_values in enumerate(row_iter, start=2):
            title = normalize(row_values[title_idx] if title_idx < len(row_values) else None)
            if re.fullmatch(r"\d+(?:\.0+)?", title):
                numeric_title_rows.append({"file": path.name, "row": row_idx, "title": title})
            identity = row_identity_from_values(row_values, headers)
            if identity[0] in {"DOUBLE", "BLANK"}:
                double_or_blank_identity_rows.append({"file": path.name, "row": row_idx, "identity": identity})
            identities[identity] += 1
        repeated = {f"{kind}:{value}": count for (kind, value), count in identities.items() if count > 40}
        if repeated:
            duplicate_identity_files.append({"file": path.name, "identity_row_counts": repeated})
    return {
        "numeric_title_rows": numeric_title_rows,
        "double_or_blank_identity_rows": double_or_blank_identity_rows,
        "high_row_count_identity_files": duplicate_identity_files,
    }


def write_analysis_workbook(
    missing_cases: list[dict[str, object]],
    problematic_cases: list[dict[str, object]],
    validation_results: list[dict[str, object]],
    selected_rows: dict[str, object],
    quality_scan: dict[str, object],
) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    summary_rows = [
        ("Audit source", str(AUDIT_WORKBOOK)),
        ("Latest conversion source", str(LATEST_UPLOAD_DIR)),
        ("Punya kamus tapi belum terupload", len(missing_cases)),
        ("Upload bermasalah/perlu review", len(problematic_cases)),
        ("Numeric-only title rows in latest 20260709 batch", len(quality_scan["numeric_title_rows"])),
        ("Blank/double identity rows in latest 20260709 batch", len(quality_scan["double_or_blank_identity_rows"])),
    ]
    for row in summary_rows:
        ws.append(row)

    for sheet_name, rows in [("Belum Terupload", missing_cases), ("Perlu Perbaikan", problematic_cases)]:
        sheet = wb.create_sheet(sheet_name)
        headers = [
            "Jenis Identity",
            "ID Identity",
            "Nama Posisi",
            "Unit Organisasi",
            "Nama Perusahaan",
            "Jumlah Pekerja Aktif",
            "NIPP Pekerja Aktif",
            "Nama Pekerja Aktif",
            "Status Ketersediaan KPI",
            "Catatan Audit",
        ]
        sheet.append(headers)
        for row in rows:
            sheet.append([row.get(header) for header in headers])

    sheet = wb.create_sheet("Generated Forms")
    sheet.append(["Case", "Identity", "Source Workbook", "Selected Excel Rows", "Output Workbook", "Rows", "Type Counts", "Errors"])
    for result in validation_results:
        key = result["path"]
        selected = selected_rows.get(key, {})
        sheet.append(
            [
                selected.get("case"),
                selected.get("identity"),
                selected.get("source_workbook"),
                selected.get("selected_rows"),
                key,
                result["rows"],
                json.dumps(result["type_counts"], ensure_ascii=False),
                "; ".join(result["errors"]),
            ]
        )

    sheet = wb.create_sheet("Latest Batch Scan")
    sheet.append(["Finding Type", "File", "Row", "Value"])
    for row in quality_scan["numeric_title_rows"]:
        sheet.append(["numeric_title", row["file"], row["row"], row["title"]])
    for row in quality_scan["double_or_blank_identity_rows"]:
        sheet.append(["identity_issue", row["file"], row["row"], str(row["identity"])])
    if not quality_scan["numeric_title_rows"] and not quality_scan["double_or_blank_identity_rows"]:
        sheet.append(["no_numeric_title_or_identity_issue", "", "", ""])

    for worksheet in wb.worksheets:
        for column in worksheet.columns:
            max_len = max(len(normalize(cell.value)) for cell in column)
            worksheet.column_dimensions[column[0].column_letter].width = min(max(max_len + 2, 12), 80)
        worksheet.freeze_panes = "A2"

    out = OUTPUT_DIR / "analisis_followup_upload_group1_ho_v2_20260709.xlsx"
    wb.save(out)
    return out


def main() -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    missing_cases, problematic_cases = collect_audit_cases()

    missing_matches = locate_identity_rows(*MISSING_UPLOAD_IDENTITY)
    if len(missing_matches) != 1:
        raise SystemExit(f"Expected one source match for {MISSING_UPLOAD_IDENTITY}, got {missing_matches}")
    missing_source, missing_rows = missing_matches[0]

    problematic_matches = locate_identity_rows(*PROBLEMATIC_IDENTITY)
    if len(problematic_matches) != 1:
        raise SystemExit(f"Expected one source match for {PROBLEMATIC_IDENTITY}, got {problematic_matches}")
    problematic_source, problematic_rows = problematic_matches[0]
    segments = contiguous_segments(problematic_rows)
    selected_problematic_rows = max(segments, key=len)

    missing_out = OUTPUT_DIR / "Formulir_Upload_KPI_PunyaKamus_BelumTerupload_PMID_37541_20260709.xlsx"
    problematic_out = OUTPUT_DIR / "Formulir_Upload_KPI_Perbaikan_Upload_Bermasalah_PNID_12474_20260709.xlsx"

    write_filtered_workbook(missing_source, missing_out, set(missing_rows))
    write_filtered_workbook(problematic_source, problematic_out, set(selected_problematic_rows))

    validation_results = [
        validate_upload(missing_out, MISSING_UPLOAD_IDENTITY),
        validate_upload(problematic_out, PROBLEMATIC_IDENTITY),
    ]
    errors = [error for result in validation_results for error in result["errors"]]
    if errors:
        raise SystemExit("\n".join(errors))

    quality_scan = scan_latest_batch_quality()
    selected_rows = {
        str(missing_out): {
            "case": "Punya kamus tapi belum terupload",
            "identity": f"{MISSING_UPLOAD_IDENTITY[0]} {MISSING_UPLOAD_IDENTITY[1]}",
            "source_workbook": missing_source.name,
            "selected_rows": f"{min(missing_rows)}-{max(missing_rows)}",
        },
        str(problematic_out): {
            "case": "Perbaikan upload bermasalah",
            "identity": f"{PROBLEMATIC_IDENTITY[0]} {PROBLEMATIC_IDENTITY[1]}",
            "source_workbook": problematic_source.name,
            "selected_rows": f"{min(selected_problematic_rows)}-{max(selected_problematic_rows)}",
        },
    }
    analysis_out = write_analysis_workbook(
        missing_cases,
        problematic_cases,
        validation_results,
        selected_rows,
        quality_scan,
    )

    zip_path = OUTPUT_DIR / "group1_ho_v2_followup_upload_20260709.zip"
    with zipfile.ZipFile(zip_path, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        for path in [missing_out, problematic_out, analysis_out]:
            zf.write(path, path.name)

    summary = {
        "output_dir": str(OUTPUT_DIR),
        "analysis_workbook": str(analysis_out),
        "zip": str(zip_path),
        "missing_cases": len(missing_cases),
        "problematic_cases": len(problematic_cases),
        "missing_form": validation_results[0],
        "problematic_form": validation_results[1],
        "problematic_source_segments": [f"{min(segment)}-{max(segment)} ({len(segment)} rows)" for segment in segments],
        "latest_batch_quality_scan": {
            "numeric_title_rows": len(quality_scan["numeric_title_rows"]),
            "double_or_blank_identity_rows": len(quality_scan["double_or_blank_identity_rows"]),
            "high_row_count_identity_files": quality_scan["high_row_count_identity_files"],
        },
    }
    (OUTPUT_DIR / "followup_upload_summary_20260709.json").write_text(
        json.dumps(summary, indent=2, ensure_ascii=False),
        encoding="utf-8",
    )
    print(json.dumps(summary, indent=2, ensure_ascii=False))


if __name__ == "__main__":
    main()
