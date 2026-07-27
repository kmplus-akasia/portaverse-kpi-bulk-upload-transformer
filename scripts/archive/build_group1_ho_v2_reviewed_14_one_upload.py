#!/usr/bin/env python3
from __future__ import annotations

import csv
import json
import re
import shutil
import subprocess
import sys
import zipfile
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
REVIEW = Path("/Users/alfredoteja/Downloads/group1_ho_v2_70_identity_manual_mapping_review_20260714.xlsx")
RAW_ROOT = Path("/Users/alfredoteja/Downloads/KAMUS KPI PELINDO GROUP 1 (HO) v2")
OUT = ROOT / "outputs/group1-ho-v2-reviewed-14-one-upload-20260714"
SOURCE_ZIP = OUT / "support/group1_ho_v2_reviewed_sources_20260714.zip"
CONFIG = OUT / "support/group1_ho_v2_reviewed_14.config.json"
CONVERSION_DIR = OUT / "support/conversion"
UPLOAD_READY_DIR = OUT / "upload-ready"
OUTPUT = UPLOAD_READY_DIR / "Formulir_Upload_KPI_Group1_HO_v2_13_Identity_Reviewed_20260714.xlsx"
AUDIT = OUT / "mapping_and_weight_audit.csv"
SUMMARY = OUT / "summary.json"
TEMPLATE = ROOT / "input/KPI Upload Template.xlsx"
REFERENCE = ROOT / "output/group1_ho_v2_20260709_latest_prod/production_position_reference_20260709.json"
HELD_IDENTITIES = {("PNID", "12474")}

UPLOAD_HEADERS = [
    "IDKPI", "Group", "Direktorat", "Posisi", "Position Master ID (Required)",
    "Position Master Variant ID (Optional)", "BSC Perspective", "KPI Type", "Parent KPI ID",
    "Parent KPI Title", "Title", "Description", "Unit", "Polarity", "Period", "Formula",
    "Weight (%)", "Cascading", "Nature Of Work (KAI Only)", "External ID (PKPI)",
    "System KPI ID", "Ownership Type", "Position Nomenklatur ID", "RKM Code ID",
]


def text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def normalized(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", " ", text(value).lower()).strip()


def identity(row: dict[str, Any]) -> tuple[str, str] | None:
    pmid = text(row.get("Position Master ID (Required)"))
    pnid = text(row.get("Position Nomenklatur ID"))
    if pmid and not pnid:
        return "PMID", pmid
    if pnid and not pmid:
        return "PNID", pnid
    return None


def review_rows() -> list[dict[str, Any]]:
    workbook = load_workbook(REVIEW, read_only=True, data_only=True)
    worksheet = workbook["Review 70 Identity"]
    values = worksheet.iter_rows(values_only=True)
    headers = [text(value) for value in next(values)]
    rows = []
    for values_row in values:
        if not values_row or values_row[0] is None:
            continue
        row = {headers[index]: values_row[index] for index in range(len(headers))}
        if text(row.get("Keputusan Manual")) != "SETUJU":
            continue
        rows.append(row)
    return rows


def resolve_source(row: dict[str, Any]) -> tuple[Path, str]:
    workbook_value = text(row.get("Workbook Terpilih")) or text(row.get("Kandidat Workbook Raw"))
    worksheet_value = text(row.get("Worksheet Terpilih")) or text(row.get("Kandidat Worksheet Raw"))
    source = RAW_ROOT / workbook_value
    if not source.exists():
        raise FileNotFoundError(f"Raw workbook tidak ditemukan: {source}")
    workbook = load_workbook(source, read_only=True, data_only=True)
    exact = [name for name in workbook.sheetnames if name == worksheet_value]
    if exact:
        return source, exact[0]
    candidates = [name for name in workbook.sheetnames if normalized(name) == normalized(worksheet_value)]
    if len(candidates) != 1:
        raise ValueError(f"Worksheet ambigu/tidak ditemukan: {worksheet_value!r} in {source.name}; candidates={candidates}")
    return source, candidates[0]


def directorate_and_group(relative: Path) -> tuple[str, str]:
    group_name = relative.parts[0] if len(relative.parts) > 1 else ""
    match = re.search(r"(DIREKTORAT [^-]+)", relative.name, re.IGNORECASE)
    directorate = match.group(1).strip() if match else ""
    return directorate, group_name


def build_config(rows: list[dict[str, Any]]) -> dict[str, Any]:
    positions = []
    sources: set[Path] = set()
    for row in rows:
        source, worksheet = resolve_source(row)
        sources.add(source)
        identity_type = text(row["Jenis Identity"])
        identity_id = text(row["ID Identity"])
        corrected_pmid = text(row.get("PMID Koreksi"))
        corrected_pnid = text(row.get("PNID Koreksi"))
        if corrected_pmid or corrected_pnid:
            if bool(corrected_pmid) == bool(corrected_pnid):
                raise ValueError(f"Koreksi identity tidak valid untuk {identity_type} {identity_id}")
            identity_type = "PMID" if corrected_pmid else "PNID"
            identity_id = corrected_pmid or corrected_pnid
        relative = source.relative_to(RAW_ROOT)
        directorate, group_name = directorate_and_group(relative)
        positions.append(
            {
                "source_workbook": str(relative),
                "sheet_name": worksheet,
                "position_name": text(row["Nomenclature Production"]) or text(row["Nama Posisi Production"]),
                "position_master_id": identity_id if identity_type == "PMID" else None,
                "position_nomenclature_id": identity_id if identity_type == "PNID" else None,
                "position_scope": "structural" if identity_type == "PMID" else "non_structural",
                "portaverse_position_title": text(row["Nama Posisi Production"]),
                "portaverse_group_name": text(row["Unit Organisasi"]),
                "portaverse_company_name": "PT Pelabuhan Indonesia (Persero)",
                "portaverse_company_code": "PLND",
                "cluster_label": text(row["Nomenclature Production"]) or text(row["Nama Posisi Production"]),
                "mapping_confidence_label": "high_confidence",
                "mapping_confidence_reason": "Approved in group1_ho_v2_70_identity_manual_mapping_review_20260714.xlsx.",
                "mapping_review_status": "approved",
                "mapping_override_approved": True,
                "mapping_override_trust_source": "reviewer_manual",
                "candidate_position_master_id": identity_id if identity_type == "PMID" else None,
                "candidate_position_nomenclature_id": identity_id if identity_type == "PNID" else None,
                "group_name": group_name,
                "directorate_name": directorate,
                "expected_impact_count": 10,
                "drop_comment_values": ["Drop"],
            }
        )
    return {"reference_source": str(REFERENCE), "positions": positions, "sources": sorted(str(path) for path in sources)}


def write_source_zip(config: dict[str, Any]) -> None:
    SOURCE_ZIP.parent.mkdir(parents=True, exist_ok=True)
    with zipfile.ZipFile(SOURCE_ZIP, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        for source_value in config["sources"]:
            source = Path(source_value)
            archive.write(source, source.relative_to(RAW_ROOT))


def run_conversion(config: dict[str, Any]) -> None:
    if CONVERSION_DIR.exists():
        shutil.rmtree(CONVERSION_DIR)
    CONVERSION_DIR.mkdir(parents=True)
    serializable = {key: value for key, value in config.items() if key != "sources"}
    CONFIG.parent.mkdir(parents=True, exist_ok=True)
    CONFIG.write_text(json.dumps(serializable, ensure_ascii=False, indent=2), encoding="utf-8")
    command = [
        sys.executable,
        str(ROOT / "scripts/kpi_bulk_transform.py"),
        "--source", str(SOURCE_ZIP),
        "--template", str(TEMPLATE),
        "--mapping", str(REFERENCE),
        "--config", str(CONFIG),
        "--output-dir", str(CONVERSION_DIR),
    ]
    result = subprocess.run(command, cwd=ROOT, text=True, capture_output=True)
    (OUT / "conversion.log").write_text(result.stdout + "\nSTDERR\n" + result.stderr, encoding="utf-8")
    if result.returncode:
        raise RuntimeError(f"Converter gagal ({result.returncode}); lihat {OUT / 'conversion.log'}")


def load_converted_rows() -> list[dict[str, Any]]:
    rows = []
    errors = []
    for report in CONVERSION_DIR.rglob("*.report.csv"):
        with report.open(newline="", encoding="utf-8") as handle:
            for row in csv.DictReader(handle):
                if text(row.get("severity")).lower() == "error":
                    errors.append(f"{report.name}: {row.get('message')}")
    if errors:
        raise RuntimeError("\n".join(errors))
    for path in CONVERSION_DIR.rglob("*.xlsx"):
        workbook = load_workbook(path, read_only=True, data_only=True)
        if "KPI Template" not in workbook.sheetnames:
            continue
        worksheet = workbook["KPI Template"]
        values = worksheet.iter_rows(values_only=True)
        headers = [text(value) for value in next(values)]
        for values_row in values:
            row = {headers[index]: values_row[index] if index < len(values_row) else None for index in range(len(headers))}
            if identity(row):
                rows.append(row)
    return rows


def validate(rows: list[dict[str, Any]], expected: set[tuple[str, str]]) -> tuple[list[dict[str, Any]], list[str]]:
    errors = []
    seen = set()
    deduplicated = []
    for row_number, row in enumerate(rows, 2):
        ident = identity(row)
        if not ident:
            errors.append(f"Baris {row_number}: identity kosong atau ganda")
            continue
        title = text(row.get("Title"))
        if not title:
            errors.append(f"Baris {row_number}: judul KPI kosong")
        key = (ident, text(row.get("KPI Type")).upper(), text(row.get("IDKPI")), title.lower(), text(row.get("Parent KPI ID")))
        if key in seen:
            continue
        seen.add(key)
        deduplicated.append(row)

    actual = {identity(row) for row in deduplicated}
    missing = expected - actual
    unexpected = actual - expected
    if missing:
        errors.append(f"Identity tidak menghasilkan baris: {sorted(missing)}")
    if unexpected:
        errors.append(f"Identity tidak diharapkan: {sorted(unexpected)}")

    weight_groups: defaultdict[tuple[tuple[str, str], str], float] = defaultdict(float)
    for row in deduplicated:
        value = row.get("Weight (%)")
        try:
            weight = float(value or 0)
        except (TypeError, ValueError):
            errors.append(f"Bobot tidak numerik: {identity(row)} {row.get('Title')}={value}")
            continue
        weight_groups[(identity(row), text(row.get("KPI Type")).upper())] += weight
    for (ident, kpi_type), total in sorted(weight_groups.items()):
        if total > 100.0001:
            errors.append(f"Total bobot >100: {ident} {kpi_type}={total:.4f}")
    return deduplicated, errors


def write_output(rows: list[dict[str, Any]]) -> None:
    workbook = load_workbook(TEMPLATE)
    worksheet = workbook["KPI Template"] if "KPI Template" in workbook.sheetnames else workbook.active
    if worksheet.max_row > 1:
        worksheet.delete_rows(2, worksheet.max_row - 1)
    for table_name in list(worksheet.tables.keys()):
        del worksheet.tables[table_name]
    for column, header in enumerate(UPLOAD_HEADERS, 1):
        worksheet.cell(1, column, header)
    for row_number, row in enumerate(rows, 2):
        for column, header in enumerate(UPLOAD_HEADERS, 1):
            worksheet.cell(row_number, column, row.get(header))
    UPLOAD_READY_DIR.mkdir(parents=True, exist_ok=True)
    workbook.save(OUTPUT)


def main() -> None:
    OUT.mkdir(parents=True, exist_ok=True)
    rows = review_rows()
    if len(rows) != 14:
        raise RuntimeError(f"Expected 14 approved rows, found {len(rows)}")
    config = build_config(rows)
    write_source_zip(config)
    run_conversion(config)
    converted = load_converted_rows()
    expected = set()
    for position in config["positions"]:
        if position["position_master_id"]:
            expected.add(("PMID", text(position["position_master_id"])))
        else:
            expected.add(("PNID", text(position["position_nomenclature_id"])))
    converted = [row for row in converted if identity(row) not in HELD_IDENTITIES]
    expected -= HELD_IDENTITIES
    final_rows, errors = validate(converted, expected)
    if errors:
        (OUT / "validation_errors.txt").write_text("\n".join(errors), encoding="utf-8")
        raise RuntimeError(f"Validasi gagal; lihat {OUT / 'validation_errors.txt'}")
    write_output(final_rows)

    weight_totals: defaultdict[tuple[str, str, str], float] = defaultdict(float)
    row_counts: Counter[tuple[str, str]] = Counter()
    for row in final_rows:
        ident = identity(row)
        row_counts[ident] += 1
        weight_totals[(ident[0], ident[1], text(row.get("KPI Type")).upper())] += float(row.get("Weight (%)") or 0)
    with AUDIT.open("w", newline="", encoding="utf-8-sig") as handle:
        headers = ["Jenis Identity", "ID Identity", "KPI Type", "Jumlah Baris Identity", "Total Bobot"]
        writer = csv.DictWriter(handle, fieldnames=headers)
        writer.writeheader()
        for identity_type, identity_id, kpi_type in sorted(weight_totals):
            writer.writerow(
                {
                    "Jenis Identity": identity_type,
                    "ID Identity": identity_id,
                    "KPI Type": kpi_type,
                    "Jumlah Baris Identity": row_counts[(identity_type, identity_id)],
                    "Total Bobot": round(weight_totals[(identity_type, identity_id, kpi_type)], 6),
                }
            )
    summary = {
        "review_workbook": str(REVIEW),
        "approved_mappings": len(rows),
        "included_mappings": len(expected),
        "held_identities": [f"{kind} {value}" for kind, value in sorted(HELD_IDENTITIES)],
        "held_reason": "PNID 12474 menghasilkan total bobot Output/KAI 200% dan sudah memiliki ownership KPI production; ditahan agar upload tidak menggandakan KPI.",
        "output_workbook": str(OUTPUT),
        "output_rows": len(final_rows),
        "output_identities": len({identity(row) for row in final_rows}),
        "identity_type_counts": dict(Counter(identity(row)[0] for row in final_rows)),
        "kpi_type_counts": dict(Counter(text(row.get("KPI Type")).upper() for row in final_rows)),
        "validation_errors": errors,
    }
    SUMMARY.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
