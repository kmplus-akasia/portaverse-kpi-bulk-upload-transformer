#!/usr/bin/env python3
from __future__ import annotations

import json
import re
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook


ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "output/group1_ho_v2_delta_remediation_20260709"
BAD_AUDIT = OUT / "audit_delta_remediation_decisions.xlsx"
SAFE_REPAIR = OUT / "upload_safe_repair_unallocated.xlsx"
SNAPSHOT = OUT / "production_kpi_snapshot_20260709.json"
TEMPLATE = ROOT / "input/KPI Upload Template.xlsx"
OUTPUT = OUT / "upload_all_40_bad_production_positions.xlsx"
AUDIT_OUTPUT = OUT / "upload_all_40_bad_production_positions_audit.xlsx"

HEADERS = [
    "IDKPI",
    "Group",
    "Direktorat",
    "Posisi",
    "Position Master ID (Required)",
    "Position Master Variant ID (Optional)",
    "BSC Perspective",
    "KPI Type",
    "Parent KPI ID",
    "Parent KPI Title",
    "Title",
    "Description",
    "Unit",
    "Polarity",
    "Period",
    "Formula",
    "Weight (%)",
    "Cascading",
    "Nature Of Work (KAI Only)",
    "External ID (PKPI)",
    "System KPI ID",
    "Ownership Type",
    "Position Nomenklatur ID",
    "RKM Code ID",
]


def text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def is_numeric_title(value: Any) -> bool:
    return bool(re.fullmatch(r"\d+(?:\.0+)?", text(value)))


def identity_key(kind: str, value: str) -> tuple[str, str]:
    return text(kind), text(value)


def row_identity(row: dict[str, Any]) -> tuple[str, str] | None:
    pmid = text(row.get("Position Master ID (Required)"))
    pnid = text(row.get("Position Nomenklatur ID"))
    if pmid and not pnid:
        return "PMID", pmid
    if pnid and not pmid:
        return "PNID", pnid
    return None


def load_bad_identities() -> list[dict[str, Any]]:
    wb = load_workbook(BAD_AUDIT, read_only=True, data_only=True)
    ws = wb["Bad Production KPI"]
    rows = ws.iter_rows(values_only=True)
    headers = [text(v) for v in next(rows)]
    result = []
    for values in rows:
        item = {headers[i]: values[i] if i < len(values) else None for i in range(len(headers))}
        if text(item.get("Jenis Identity")) and text(item.get("ID Identity")):
            result.append(item)
    return result


def load_upload_rows(path: Path) -> dict[tuple[str, str], list[dict[str, Any]]]:
    by_identity: dict[tuple[str, str], list[dict[str, Any]]] = defaultdict(list)
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb["KPI Template"]
    rows = ws.iter_rows(values_only=True)
    headers = [text(v) for v in next(rows)]
    for values in rows:
        row = {headers[i]: values[i] if i < len(values) else None for i in range(len(headers))}
        identity = row_identity(row)
        if identity:
            by_identity[identity].append(row)
    return by_identity


def load_prod_rows() -> dict[tuple[str, str], list[dict[str, Any]]]:
    by_identity: dict[tuple[str, str], list[dict[str, Any]]] = defaultdict(list)
    for row in json.loads(SNAPSHOT.read_text(encoding="utf-8"))["rows"]:
        pnid = text(row.get("pnid"))
        pmid = text(row.get("pmid"))
        identity = ("PNID", pnid) if pnid else (("PMID", pmid) if pmid else None)
        if identity:
            by_identity[identity].append(row)
    return by_identity


def reconstruct_from_production(identity: tuple[str, str], prod_rows: list[dict[str, Any]]) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    usable = [
        row
        for row in prod_rows
        if text(row.get("kpi_type")).upper() in {"IMPACT", "OUTPUT", "KAI"}
        and not is_numeric_title(row.get("title"))
    ]
    usable_by_kpi_id = {text(row.get("kpi_id")): row for row in usable}
    excluded_numeric = len(prod_rows) - len(usable)
    # Keep only rows whose parent chain is present after numeric rows are removed.
    stable = []
    for row in usable:
        parent_id = text(row.get("parent_kpi_id"))
        if parent_id and parent_id not in usable_by_kpi_id:
            continue
        stable.append(row)

    order = {"IMPACT": 0, "OUTPUT": 1, "KAI": 2}
    stable.sort(key=lambda r: (order.get(text(r.get("kpi_type")).upper(), 9), text(r.get("kpi_id"))))
    id_by_kpi_id = {text(row.get("kpi_id")): str(900000 + idx) for idx, row in enumerate(stable, start=1)}
    title_by_kpi_id = {text(row.get("kpi_id")): text(row.get("title")) for row in stable}
    out_rows = []
    for row in stable:
        kpi_id = text(row.get("kpi_id"))
        parent_kpi_id = text(row.get("parent_kpi_id"))
        item = {header: "" for header in HEADERS}
        item["IDKPI"] = id_by_kpi_id[kpi_id]
        item["Group"] = ""
        item["Direktorat"] = ""
        item["Posisi"] = ""
        if identity[0] == "PMID":
            item["Position Master ID (Required)"] = identity[1]
        else:
            item["Position Nomenklatur ID"] = identity[1]
        item["BSC Perspective"] = text(row.get("perspective"))
        item["KPI Type"] = text(row.get("kpi_type")).upper()
        item["Parent KPI ID"] = id_by_kpi_id.get(parent_kpi_id, "")
        item["Parent KPI Title"] = title_by_kpi_id.get(parent_kpi_id, "")
        item["Title"] = text(row.get("title"))
        item["Description"] = text(row.get("description"))
        item["Unit"] = text(row.get("target_unit"))
        item["Polarity"] = text(row.get("polarity"))
        item["Period"] = text(row.get("monitoring_period"))
        item["Formula"] = text(row.get("formula"))
        item["Weight (%)"] = row.get("ownership_weight")
        item["Cascading"] = text(row.get("cascading_method"))
        item["Nature Of Work (KAI Only)"] = text(row.get("nature_of_work"))
        item["External ID (PKPI)"] = text(row.get("external_id"))
        item["System KPI ID"] = text(row.get("kpi_id"))
        item["Ownership Type"] = text(row.get("ownership_type") or row.get("kpi_ownership_type"))
        item["RKM Code ID"] = text(row.get("rkm_code_id"))
        out_rows.append(item)
    audit = {
        "source": "production_reconstructed_system_kpi_id",
        "production_rows": len(prod_rows),
        "excluded_numeric_or_non_kpi_rows": excluded_numeric,
        "excluded_missing_parent_rows": len(usable) - len(stable),
        "generated_rows": len(out_rows),
    }
    return out_rows, audit


def write_upload(rows: list[dict[str, Any]]) -> None:
    wb = load_workbook(TEMPLATE)
    ws = wb["KPI Template"] if "KPI Template" in wb.sheetnames else wb.active
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)
    for table_name in list(ws.tables.keys()):
        del ws.tables[table_name]
    for col, header in enumerate(HEADERS, start=1):
        ws.cell(1, col, header)
    for row_idx, row in enumerate(rows, start=2):
        for col, header in enumerate(HEADERS, start=1):
            ws.cell(row_idx, col, row.get(header))
    wb.save(OUTPUT)


def validate(rows: list[dict[str, Any]], expected_identities: set[tuple[str, str]]) -> list[str]:
    errors = []
    got = set()
    for idx, row in enumerate(rows, start=2):
        identity = row_identity(row)
        if identity is None:
            errors.append(f"row {idx}: blank/double identity")
        else:
            got.add(identity)
        if is_numeric_title(row.get("Title")):
            errors.append(f"row {idx}: numeric title {row.get('Title')}")
    missing = expected_identities - got
    if missing:
        errors.append(f"missing identities: {sorted(missing)}")
    return errors


def main() -> None:
    bad = load_bad_identities()
    expected = {identity_key(row["Jenis Identity"], row["ID Identity"]) for row in bad}
    safe_rows = load_upload_rows(SAFE_REPAIR)
    prod_rows = load_prod_rows()

    final_rows: list[dict[str, Any]] = []
    audit_rows = []
    for row in bad:
        identity = identity_key(row["Jenis Identity"], row["ID Identity"])
        if safe_rows.get(identity):
            rows = safe_rows[identity]
            final_rows.extend(rows)
            audit_rows.append(
                {
                    "Identity": f"{identity[0]} {identity[1]}",
                    "Source": "corrected_converter_upload_safe_repair",
                    "Production KPI Rows": row.get("Production KPI Rows"),
                    "Generated Rows": len(rows),
                    "Issue Reasons": row.get("Issue Reasons"),
                    "Notes": "",
                }
            )
            continue
        rows, audit = reconstruct_from_production(identity, prod_rows.get(identity, []))
        final_rows.extend(rows)
        audit_rows.append(
            {
                "Identity": f"{identity[0]} {identity[1]}",
                "Source": audit["source"],
                "Production KPI Rows": audit["production_rows"],
                "Generated Rows": audit["generated_rows"],
                "Issue Reasons": row.get("Issue Reasons"),
                "Notes": (
                    f"Excluded numeric/non-kpi rows={audit['excluded_numeric_or_non_kpi_rows']}; "
                    f"excluded missing-parent rows={audit['excluded_missing_parent_rows']}"
                ),
            }
        )

    errors = validate(final_rows, expected)
    if errors:
        raise SystemExit("\n".join(errors))
    write_upload(final_rows)

    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    summary = [
        ("Bad production identities", len(expected)),
        ("Generated upload rows", len(final_rows)),
        ("Numeric titles in generated form", 0),
        ("Blank/double identity rows", 0),
    ]
    for row in summary:
        ws.append(row)
    ws = wb.create_sheet("Identity Sources")
    headers = ["Identity", "Source", "Production KPI Rows", "Generated Rows", "Issue Reasons", "Notes"]
    ws.append(headers)
    for row in audit_rows:
        ws.append([row.get(header, "") for header in headers])
    wb.save(AUDIT_OUTPUT)
    print(json.dumps({"output": str(OUTPUT), "audit": str(AUDIT_OUTPUT), "identities": len(expected), "rows": len(final_rows)}, indent=2))


if __name__ == "__main__":
    main()
