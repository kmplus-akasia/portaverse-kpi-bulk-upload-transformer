#!/usr/bin/env python3
from __future__ import annotations

from collections import Counter, defaultdict
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "output/group1_ho_v2_unconverted_mapped_20260709"
SOURCE_AUDIT = ROOT / "output/group1_ho_v2_delta_remediation_20260709/worksheet_conversion_upload_gap_audit.xlsx"
MAPPING_AUDIT = OUT / "mapping_audit_unconverted_20260709.xlsx"
ORG_AUDIT = Path(
    "/Users/alfredoteja/Documents/pms-codebase/dashboard-org-kpi-audit/data/output/audit-organisasi-kamus-kpi-ho.xlsx"
)
OUTPUT = OUT / "remaining_37_resolution_audit_20260709.xlsx"


MANUAL_RESOLUTION: dict[str, dict[str, Any]] = {
    "Officer Hukum Komersial": {
        "status": "RESOLVED_ALREADY_IN_PRODUCTION",
        "identity": "PNID 34",
        "prod_title": "Officer Penanganan Perkara Dan Advokasi Hukum",
        "reason": "Audit Posisi menunjukkan PNID 34 sudah KPI Lengkap di Sistem; tidak perlu upload ulang.",
    },
    "Officer Litigasi": {
        "status": "BLOCKED_NO_RELIABLE_IDENTITY",
        "identity": "",
        "prod_title": "",
        "reason": "Tidak ada PNID Officer Litigasi aktif di audit organisasi/production reference; kandidat fuzzy Officer Integrasi ditolak.",
    },
    "Manager Hukum Komersial": {
        "status": "RESOLVED_ALREADY_IN_PRODUCTION",
        "identity": "PMID 166",
        "prod_title": "Department Head Hukum Komersial",
        "reason": "Audit Posisi menunjukkan PMID 166 sudah KPI Lengkap di Sistem; tidak perlu upload ulang.",
    },
    "Officer Sistem Manajemen": {
        "status": "BLOCKED_PROD_REFERENCE_NO_PNID_FOR_TYPE6",
        "identity": "PMID 37576; PMID 37574; PMID 37573",
        "prod_title": "Officer I/Senior Officer I/Senior Officer III Sistem Manajemen",
        "reason": "Audit organisasi punya PMID aktif tipe 6 tanpa PNID; converter menolak mapping sebagai PMID struktural. Perlu PNID production atau keputusan upload manual berbasis PMID.",
    },
    "DH Sistem Manajemen": {
        "status": "RESOLVED_INCLUDED_IN_ONE_UPLOAD",
        "identity": "PMID 37572",
        "prod_title": "Department Head Sistem Manajemen",
        "reason": "Kamus valid dan belum ada KPI production; sudah masuk formulir upload gabungan.",
    },
    "Manager Pengembangan K3": {
        "status": "RESOLVED_INCLUDED_IN_ONE_UPLOAD",
        "identity": "PMID 37577",
        "prod_title": "Manager Pengembangan dan Evaluasi Program K3",
        "reason": "Kamus valid dan belum ada KPI production; sudah masuk formulir upload gabungan.",
    },
    "DH K3": {
        "status": "RESOLVED_ALREADY_IN_PRODUCTION",
        "identity": "PMID 1123",
        "prod_title": "Department Head Keselamatan dan Kesehatan Kerja",
        "reason": "Audit Posisi menunjukkan PMID 1123 sudah KPI Lengkap di Sistem; tidak perlu upload ulang.",
    },
    "Officer MEKA": {
        "status": "RESOLVED_INCLUDED_IN_ONE_UPLOAD",
        "identity": "PNID 12541",
        "prod_title": "Officer Monitoring dan Evaluasi Klaster Ekspansi Korporasi",
        "reason": "MEKA disamakan dengan Department Monitoring dan Evaluasi Klaster Ekspansi Korporasi; sudah masuk formulir upload gabungan.",
    },
    "Officer MEKO": {
        "status": "RESOLVED_INCLUDED_IN_ONE_UPLOAD",
        "identity": "PNID 12553",
        "prod_title": "Officer Monitoring dan Evaluasi Klaster Optimalisasi Korporasi",
        "reason": "MEKO disamakan dengan Department Monitoring dan Evaluasi Klaster Optimalisasi Korporasi; sudah masuk formulir upload gabungan.",
    },
    "PMO": {
        "status": "BLOCKED_NO_RELIABLE_IDENTITY",
        "identity": "",
        "prod_title": "",
        "reason": "Tidak ada identity PMO aktif yang cocok langsung di audit organisasi/production reference; ditahan dari upload.",
    },
    "DH Terminal Petikemas2": {
        "status": "RESOLVED_ALREADY_IN_PRODUCTION",
        "identity": "PMID 1012",
        "prod_title": "Department Head Terminal Petikemas",
        "reason": "Audit Posisi menunjukkan PMID 1012 sudah KPI Lengkap di Sistem; tidak perlu upload ulang.",
    },
    "Manager Pengelolaan Kas": {
        "status": "RESOLVED_ALREADY_IN_PRODUCTION",
        "identity": "PMID 35769",
        "prod_title": "Manager Pengelolaan Aset",
        "reason": "Kandidat production terdekat sudah KPI Lengkap di Sistem; tidak perlu upload ulang dan nama worksheet tidak cukup kuat untuk override.",
    },
    "Officer Employee Service 3": {
        "status": "RESOLVED_INCLUDED_IN_ONE_UPLOAD",
        "identity": "PNID 102",
        "prod_title": "Officer Travel Management 3",
        "reason": "Audit Posisi menempatkan PNID 102 di Kelompok Kerja Employee Service 3 dan belum ada KPI; sudah masuk formulir upload gabungan.",
    },
    "Officer Employee Service 2": {
        "status": "RESOLVED_INCLUDED_IN_ONE_UPLOAD",
        "identity": "PNID 100",
        "prod_title": "Officer Travel Management 2",
        "reason": "Audit Posisi menempatkan PNID 100 di Kelompok Kerja Employee Service 2 dan belum ada KPI; kandidat fuzzy Payroll 2 ditolak.",
    },
    "Officer Employee Service 1": {
        "status": "RESOLVED_INCLUDED_IN_ONE_UPLOAD",
        "identity": "PNID 97; PNID 98",
        "prod_title": "Officer Data Management 1; Officer Travel Management 1",
        "reason": "Audit Posisi menempatkan PNID 97 dan PNID 98 di Kelompok Kerja Employee Service 1 dan belum ada KPI; keduanya masuk formulir upload gabungan.",
    },
    "DH Perencanaan Instalasi dan2": {
        "status": "RESOLVED_ALREADY_IN_PRODUCTION",
        "identity": "PMID 937",
        "prod_title": "Department Head Perencanaan Instalasi dan Alat Apung",
        "reason": "Audit Posisi menunjukkan PMID 937 sudah KPI Lengkap di Sistem; tidak perlu upload ulang.",
    },
    "DH Perencanaan Peralatan2": {
        "status": "RESOLVED_ALREADY_IN_PRODUCTION",
        "identity": "PMID 928",
        "prod_title": "Department Head Perencanaan Peralatan Pelabuhan",
        "reason": "Audit Posisi menunjukkan PMID 928 sudah KPI Lengkap di Sistem; tidak perlu upload ulang.",
    },
    "Senior Expert Auditor": {
        "status": "RESOLVED_ALREADY_IN_PRODUCTION",
        "identity": "PNID 7",
        "prod_title": "Expert Audit",
        "reason": "Audit Posisi menunjukkan PNID 7 sudah KPI Lengkap di Sistem; tidak perlu upload ulang.",
    },
    "Principle Expert Auditor": {
        "status": "RESOLVED_ALREADY_IN_PRODUCTION",
        "identity": "PNID 7",
        "prod_title": "Expert Audit",
        "reason": "Audit Posisi menunjukkan PNID 7 sudah KPI Lengkap di Sistem; tidak perlu upload ulang.",
    },
    "Manager Administrasi dan Hubung": {
        "status": "BLOCKED_NO_RELIABLE_IDENTITY",
        "identity": "",
        "prod_title": "",
        "reason": "Tidak ada nama Administrasi dan Hubungan Antar Lembaga aktif; kandidat mirip Administrasi dan Koordinasi Pengawasan sudah KPI lengkap tetapi bukan match langsung.",
    },
    "DH Manajemen Data": {
        "status": "RESOLVED_INCLUDED_IN_ONE_UPLOAD",
        "identity": "PMID 37539",
        "prod_title": "Department Head Manajemen Data",
        "reason": "Kamus valid dan belum ada KPI production; sudah masuk formulir upload gabungan.",
    },
    "Pimpro Satker Single ERP": {
        "status": "RESOLVED_INCLUDED_IN_ONE_UPLOAD",
        "identity": "PMID 37583",
        "prod_title": "Ketua Satuan Kerja Implementasi Single ERP",
        "reason": "Diminta user untuk upload ulang kamus lengkap walaupun sudah ada 10 KPI Impact; sudah masuk formulir upload gabungan.",
    },
    "Group Head": {
        "status": "RESOLVED_INCLUDED_IN_ONE_UPLOAD",
        "identity": "PMID 37587",
        "prod_title": "Group Head Monitoring Evaluasi Strategi Perusahaan dan Inovasi",
        "reason": "Kamus valid dan belum ada KPI production; sudah masuk formulir upload gabungan.",
    },
    "Manager Key Account": {
        "status": "RESOLVED_INCLUDED_IN_ONE_UPLOAD",
        "identity": "PMID 37528",
        "prod_title": "Plt. Manager Key Account",
        "reason": "Kamus valid dan belum ada KPI production; sudah masuk formulir upload gabungan.",
    },
    "DH Perencanaan & Kualitas Penga": {
        "status": "RESOLVED_ALREADY_IN_PRODUCTION",
        "identity": "PMID 37531",
        "prod_title": "Plt. Department Head Perencanaan & Kualitas Pengawasan Intern",
        "reason": "Audit Posisi menunjukkan PMID 37531 sudah KPI Lengkap di Sistem; tidak perlu upload ulang.",
    },
    "IT Solusi Bisnis": {
        "status": "RESOLVED_ALREADY_IN_PRODUCTION",
        "identity": "PMID 834",
        "prod_title": "Group Head Teknologi Informasi",
        "reason": "Identity production terdekat sudah KPI Lengkap di Sistem; worksheet tidak cukup spesifik untuk override upload.",
    },
    "DH Tata Kelola IT & Pengelolaan": {
        "status": "RESOLVED_INCLUDED_IN_ONE_UPLOAD",
        "identity": "PMID 37540",
        "prod_title": "Department Head Tata Kelola TI dan Pengelolaan Software",
        "reason": "False negative audit untuk NIPP 103412; sudah masuk formulir upload gabungan.",
    },
}


def text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def load_gap_rows() -> list[dict[str, Any]]:
    wb = load_workbook(SOURCE_AUDIT, read_only=True, data_only=True)
    ws = wb["Belum Prod"]
    headers = [text(c.value) for c in next(ws.iter_rows(max_row=1))]
    rows: list[dict[str, Any]] = []
    for values in ws.iter_rows(min_row=2, values_only=True):
        row = {headers[i]: values[i] if i < len(values) else None for i in range(len(headers))}
        if row.get("Operational Bucket") == "BELUM_TERKONVERSI_DAN_BELUM_ADA_DI_PRODUCTION":
            rows.append(row)
    return rows


def load_mapping_rows() -> dict[str, list[dict[str, Any]]]:
    wb = load_workbook(MAPPING_AUDIT, read_only=True, data_only=True)
    ws = wb.active
    headers = [text(c.value) for c in next(ws.iter_rows(max_row=1))]
    by_sheet: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for values in ws.iter_rows(min_row=2, values_only=True):
        row = {headers[i]: values[i] if i < len(values) else None for i in range(len(headers))}
        by_sheet[text(row.get("Worksheet"))].append(row)
    return by_sheet


def load_org_audit_status() -> dict[str, dict[str, Any]]:
    if not ORG_AUDIT.exists():
        return {}
    wb = load_workbook(ORG_AUDIT, read_only=True, data_only=True)
    ws = wb["Audit Posisi"]
    headers = [text(c.value) for c in next(ws.iter_rows(max_row=1))]
    rows = {}
    for values in ws.iter_rows(min_row=2, values_only=True):
        row = {headers[i]: values[i] if i < len(values) else None for i in range(len(headers))}
        key = f"{text(row.get('Jenis Identity'))} {text(row.get('ID Identity'))}"
        rows[key] = row
    return rows


def identity_keys(value: str) -> list[str]:
    return [part.strip() for part in value.split(";") if part.strip()]


def write_sheet(ws, rows: list[dict[str, Any]], headers: list[str]) -> None:
    ws.append(headers)
    for row in rows:
        ws.append([row.get(header, "") for header in headers])
    ws.freeze_panes = "A2"
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
    for column in ws.columns:
        width = max(len(text(cell.value)) for cell in column)
        ws.column_dimensions[get_column_letter(column[0].column)].width = min(max(width + 2, 12), 90)


def main() -> None:
    gap_rows = load_gap_rows()
    mapping_rows = load_mapping_rows()
    org_status = load_org_audit_status()
    resolved_rows: list[dict[str, Any]] = []

    for gap in gap_rows:
        sheet = text(gap.get("Worksheet"))
        if "Group Pengendalian Proyek/" in text(gap.get("Source Workbook")):
            decision = {
                "status": "HELD_PENGENDALIAN_PROYEK",
                "identity": "",
                "prod_title": "",
                "reason": "Ditahan sesuai instruksi user; tidak dimasukkan ke formulir gabungan.",
            }
        else:
            decision = MANUAL_RESOLUTION[sheet]

        upload_rows = 0
        upload_statuses = []
        for audit in mapping_rows.get(sheet, []):
            upload_rows += int(audit.get("Rows") or 0)
            upload_statuses.append(text(audit.get("Status")))

        org_kpi_statuses = []
        active_workers = []
        for key in identity_keys(decision["identity"]):
            row = org_status.get(key)
            if row:
                org_kpi_statuses.append(f"{key}: {text(row.get('Status Ketersediaan KPI'))}")
                active_workers.append(f"{key}: {text(row.get('NIPP Pekerja Aktif'))}")

        resolved_rows.append(
            {
                "Source Workbook": gap.get("Source Workbook"),
                "Worksheet": sheet,
                "Resolution Status": decision["status"],
                "Resolved Identity": decision["identity"],
                "Production Title": decision["prod_title"],
                "Upload Builder Status": "; ".join(upload_statuses),
                "Upload Rows": upload_rows,
                "Audit Org KPI Status": "; ".join(org_kpi_statuses),
                "Active Worker NIPP": "; ".join(active_workers),
                "Decision Reason": decision["reason"],
                "Original Candidate Title": gap.get("Portaverse Position Title"),
                "Original Mapping Confidence": gap.get("Mapping Confidence"),
                "Original Mapping Reason": gap.get("Mapping Reason"),
            }
        )

    counts = Counter(row["Resolution Status"] for row in resolved_rows)
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    write_sheet(
        ws,
        [{"Metric": key, "Value": value} for key, value in [("Total unresolved positions reviewed", len(resolved_rows)), *counts.items()]],
        ["Metric", "Value"],
    )

    ws = wb.create_sheet("37 Resolution")
    headers = [
        "Source Workbook",
        "Worksheet",
        "Resolution Status",
        "Resolved Identity",
        "Production Title",
        "Upload Builder Status",
        "Upload Rows",
        "Audit Org KPI Status",
        "Active Worker NIPP",
        "Decision Reason",
        "Original Candidate Title",
        "Original Mapping Confidence",
        "Original Mapping Reason",
    ]
    write_sheet(ws, resolved_rows, headers)

    wb.save(OUTPUT)
    print(OUTPUT)
    print(dict(counts))


if __name__ == "__main__":
    main()
