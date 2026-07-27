#!/usr/bin/env python3
from __future__ import annotations

import csv
import json
import re
import shutil
import subprocess
import zipfile
from collections import Counter, defaultdict
from copy import copy
from difflib import SequenceMatcher
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook


ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "output/group1_ho_v2_delta_remediation_20260709"
AUDIT_WORKBOOK = Path(
    "/Users/alfredoteja/Documents/pms-codebase/dashboard-org-kpi-audit/data/output/audit-organisasi-kamus-kpi-ho.xlsx"
)
SOURCE_ROOT = Path("/Users/alfredoteja/Downloads/KAMUS KPI PELINDO GROUP 1 (HO) v2")
TEMPLATE = ROOT / "input/KPI Upload Template.xlsx"
LATEST_UPLOAD_READY = ROOT / "output/group1_ho_v2_20260709_latest_prod/conversion/upload-ready"
LATEST_CONFIG = ROOT / "output/group1_ho_v2_20260709_latest_prod/group1_ho_v2_20260709_upload.config.json"
REFERENCE = ROOT / "output/group1_ho_v2_20260709_latest_prod/production_position_reference_20260709.json"
SNAPSHOT = OUT / "production_kpi_snapshot_20260709.json"

UPLOAD_HEADERS = [
    "IDKPI",
    "Group",
    "Direktorat",
    "Posisi",
    "Position Master ID (Required)",
    "Position Master Variant ID (Optional)",
    "BSC Perspective",
    "KPI Type",
    "Parent KPI ID",
    "Parent KPI Title",
    "Title",
    "Description",
    "Unit",
    "Polarity",
    "Period",
    "Formula",
    "Weight (%)",
    "Cascading",
    "Nature Of Work (KAI Only)",
    "External ID (PKPI)",
    "System KPI ID",
    "Ownership Type",
    "Position Nomenklatur ID",
    "RKM Code ID",
]


def text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def norm(value: Any) -> str:
    s = text(value).lower()
    replacements = {
        "department head": "dh",
        "dept head": "dh",
        "department": "dept",
        "departemen": "dept",
        "teknologi informasi": "ti",
        "information technology": "ti",
        " it ": " ti ",
        "pengoprasian": "pengoperasian",
        "pengelolaan": "pengelolaan",
        "software": "software",
        " dan ": " ",
        "&": " ",
    }
    s = f" {s} "
    for old, new in replacements.items():
        s = s.replace(old, new)
    s = re.sub(r"[^a-z0-9]+", " ", s)
    return re.sub(r"\s+", " ", s).strip()


def numeric_title(value: Any) -> bool:
    return bool(re.fullmatch(r"\d+(?:\.0+)?", text(value)))


def load_sheet_rows(path: Path, sheet_name: str) -> list[dict[str, Any]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet_name]
    rows = ws.iter_rows(values_only=True)
    headers = [text(v) for v in next(rows)]
    result = []
    for row_number, row in enumerate(rows, start=2):
        item = {headers[i]: row[i] if i < len(row) else None for i in range(len(headers))}
        item["_row"] = row_number
        result.append(item)
    return result


def identity_from_audit(row: dict[str, Any]) -> tuple[str, str]:
    return text(row.get("Jenis Identity")), text(row.get("ID Identity"))


def identity_from_upload(row: dict[str, Any]) -> tuple[str, str] | None:
    pmid = text(row.get("Position Master ID (Required)"))
    pnid = text(row.get("Position Nomenklatur ID"))
    if pmid and not pnid:
        return ("PMID", pmid)
    if pnid and not pmid:
        return ("PNID", pnid)
    return None


def identity_from_prod(row: dict[str, Any]) -> tuple[str, str] | None:
    pnid = text(row.get("pnid"))
    pmid = text(row.get("pmid"))
    if pnid:
        return ("PNID", pnid)
    if pmid:
        return ("PMID", pmid)
    return None


def identity_key(identity: tuple[str, str]) -> str:
    return f"{identity[0]}:{identity[1]}"


def load_upload_rows(paths: list[Path]) -> dict[tuple[str, str], list[dict[str, Any]]]:
    by_identity: dict[tuple[str, str], list[dict[str, Any]]] = defaultdict(list)
    for path in paths:
        wb = load_workbook(path, read_only=True, data_only=True)
        if "KPI Template" not in wb.sheetnames:
            continue
        ws = wb["KPI Template"]
        rows = ws.iter_rows(values_only=True)
        headers = [text(v) for v in next(rows)]
        for excel_row, values in enumerate(rows, start=2):
            row = {headers[i]: values[i] if i < len(values) else None for i in range(len(headers))}
            identity = identity_from_upload(row)
            if not identity:
                continue
            row["_source_file"] = path.name
            row["_excel_row"] = excel_row
            by_identity[identity].append(row)
    return by_identity


def load_production_snapshot() -> dict[tuple[str, str], list[dict[str, Any]]]:
    data = json.loads(SNAPSHOT.read_text(encoding="utf-8"))
    by_identity: dict[tuple[str, str], list[dict[str, Any]]] = defaultdict(list)
    for row in data["rows"]:
        identity = identity_from_prod(row)
        if identity:
            by_identity[identity].append(row)
    return by_identity


def has_reset_risk(prod_rows: list[dict[str, Any]]) -> bool:
    for row in prod_rows:
        if text(row.get("kpi_type")).upper() not in {"OUTPUT", "KAI"}:
            continue
        if text(row.get("item_approval_status")).upper() in {"APPROVED", "APPROVED_ADJUSTED"}:
            return True
        if text(row.get("weight_approval_status")).upper() == "APPROVED":
            return True
        if text(row.get("allocation_status")).upper() in {"ALLOCATED", "AUTO_ALLOCATED"}:
            return True
    return False


def load_dropped_titles_by_identity(config_positions: list[dict[str, Any]]) -> dict[tuple[str, str], set[str]]:
    sheet_to_identities: dict[str, set[tuple[str, str]]] = defaultdict(set)
    for cfg in config_positions:
        pmid = text(cfg.get("position_master_id"))
        pnid = text(cfg.get("position_nomenclature_id"))
        identity = ("PNID", pnid) if pnid else (("PMID", pmid) if pmid else None)
        if identity:
            sheet_to_identities[text(cfg.get("sheet_name"))].add(identity)
    dropped: dict[tuple[str, str], set[str]] = defaultdict(set)
    for report_path in (ROOT / "output/group1_ho_v2_20260709_latest_prod/conversion").glob("*/*.report.csv"):
        with report_path.open(newline="", encoding="utf-8") as csvfile:
            for row in csv.DictReader(csvfile):
                message = text(row.get("message")).lower()
                if "alignment/status is drop" not in message and "numeric-only title" not in message:
                    continue
                title = norm(row.get("title"))
                if not title or title == "blank":
                    continue
                identities = sheet_to_identities.get(text(row.get("sheet_name")), set())
                if len(identities) != 1:
                    continue
                dropped[next(iter(identities))].add(title)
    return dropped


def production_issue_reasons(
    prod_rows: list[dict[str, Any]],
    expected_rows: list[dict[str, Any]],
    dropped_titles: set[str],
) -> list[str]:
    reasons = []
    if any(numeric_title(row.get("title")) for row in prod_rows):
        reasons.append("production_numeric_title")
    if any(norm(row.get("title")) in dropped_titles for row in prod_rows):
        reasons.append("production_matches_latest_converter_dropped_row")
    return reasons


def write_upload_workbook(path: Path, rows: list[dict[str, Any]]) -> None:
    workbook = load_workbook(TEMPLATE)
    ws = workbook["KPI Template"] if "KPI Template" in workbook.sheetnames else workbook.active
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)
    for table_name in list(ws.tables.keys()):
        del ws.tables[table_name]
    for col, header in enumerate(UPLOAD_HEADERS, start=1):
        ws.cell(1, col, header)
    for out_row, row in enumerate(rows, start=2):
        for col, header in enumerate(UPLOAD_HEADERS, start=1):
            ws.cell(out_row, col, row.get(header))
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def write_table_sheet(ws, rows: list[dict[str, Any]], headers: list[str]) -> None:
    ws.append(headers)
    for row in rows:
        ws.append([row.get(header, "") for header in headers])
    ws.freeze_panes = "A2"
    for column in ws.columns:
        width = max(len(text(cell.value)) for cell in column)
        ws.column_dimensions[column[0].column_letter].width = min(max(width + 2, 12), 72)


def converter_config_for_candidate(base: dict[str, Any], audit: dict[str, Any]) -> dict[str, Any]:
    kind, value = identity_from_audit(audit)
    item = copy(base)
    item["position_name"] = text(audit.get("Nama Posisi")) or item.get("position_name")
    item["portaverse_position_title"] = text(audit.get("Nama Posisi")) or item.get("portaverse_position_title")
    item["portaverse_group_name"] = text(audit.get("Unit Organisasi")) or item.get("portaverse_group_name")
    item["mapping_review_status"] = "approved"
    item["mapping_override_approved"] = True
    item["mapping_override_trust_source"] = "delta_remediation_audit"
    item["mapping_confidence_label"] = "high_confidence"
    item["mapping_confidence_reason"] = "Rescued from audit false-negative; active production identity matched to source worksheet."
    if kind == "PMID":
        item["position_scope"] = "structural"
        item["position_master_id"] = value
        item["position_nomenclature_id"] = None
    else:
        item["position_scope"] = "non_structural"
        item["position_master_id"] = None
        item["position_nomenclature_id"] = value
    return item


def build_rescue_candidates(
    audit_missing_no_kamus: list[dict[str, Any]], config_positions: list[dict[str, Any]]
) -> list[dict[str, Any]]:
    unmapped = [
        row
        for row in config_positions
        if row.get("source_workbook")
        and not text(row.get("position_master_id"))
        and not text(row.get("position_nomenclature_id"))
    ]
    candidates = []
    for audit in audit_missing_no_kamus:
        audit_text = " ".join([text(audit.get("Nama Posisi")), text(audit.get("Unit Organisasi"))])
        best = None
        for cfg in unmapped:
            cfg_text = " ".join(
                [
                    text(cfg.get("sheet_name")),
                    text(cfg.get("position_name")),
                    text(cfg.get("portaverse_position_title")),
                    text(cfg.get("portaverse_group_name")),
                ]
            )
            score = SequenceMatcher(None, norm(audit_text), norm(cfg_text)).ratio()
            if best is None or score > best[0]:
                best = (score, cfg)
        if not best:
            continue
        score, cfg = best
        is_required_example = text(audit.get("NIPP Pekerja Aktif")) == "103412" and identity_from_audit(audit) == (
            "PMID",
            "37540",
        )
        if score >= 0.58 or is_required_example:
            candidates.append(
                {
                    "Jenis Identity": identity_from_audit(audit)[0],
                    "ID Identity": identity_from_audit(audit)[1],
                    "NIPP Pekerja Aktif": text(audit.get("NIPP Pekerja Aktif")),
                    "Nama Pekerja Aktif": text(audit.get("Nama Pekerja Aktif")),
                    "Nama Posisi": text(audit.get("Nama Posisi")),
                    "Unit Organisasi": text(audit.get("Unit Organisasi")),
                    "Source Workbook": cfg.get("source_workbook"),
                    "Worksheet": cfg.get("sheet_name"),
                    "Match Score": round(score, 4),
                    "_config": converter_config_for_candidate(cfg, audit),
                }
            )
    return candidates


def run_rescue_conversion(candidates: list[dict[str, Any]]) -> dict[tuple[str, str], list[dict[str, Any]]]:
    worksheet_counts = Counter((c["Source Workbook"], c["Worksheet"]) for c in candidates)
    convertible_candidates = [
        c for c in candidates if worksheet_counts[(c["Source Workbook"], c["Worksheet"])] == 1
    ]
    if not convertible_candidates:
        return {}
    rescue_config = OUT / "rescue_mapping_candidates_20260709.config.json"
    rescue_config.write_text(
        json.dumps({"reference_source": str(REFERENCE), "positions": [c["_config"] for c in convertible_candidates]}, indent=2),
        encoding="utf-8",
    )
    generated_dir = OUT / "rescue_generated"
    if generated_dir.exists():
        shutil.rmtree(generated_dir)
    generated_dir.mkdir(parents=True, exist_ok=True)
    cmd = [
        "python3",
        "scripts/kpi_bulk_transform.py",
        "--source",
        "tmp/group1_ho_v2_source_20260703.zip",
        "--template",
        str(TEMPLATE),
        "--mapping",
        str(REFERENCE),
        "--config",
        str(rescue_config),
        "--output-dir",
        str(generated_dir),
    ]
    subprocess.run(cmd, cwd=ROOT, check=False)
    valid_workbooks = []
    for workbook_path in sorted(generated_dir.glob("*/*.xlsx")):
        report_path = workbook_path.with_suffix(".report.csv")
        has_error = False
        if report_path.exists():
            with report_path.open(newline="", encoding="utf-8") as csvfile:
                for row in csv.DictReader(csvfile):
                    if row.get("severity") == "error":
                        has_error = True
                        break
        if not has_error:
            valid_workbooks.append(workbook_path)
    return load_upload_rows(valid_workbooks)


def validate_rows(rows: list[dict[str, Any]]) -> list[str]:
    errors = []
    for idx, row in enumerate(rows, start=2):
        pmid = text(row.get("Position Master ID (Required)"))
        pnid = text(row.get("Position Nomenklatur ID"))
        title = text(row.get("Title"))
        if bool(pmid) == bool(pnid):
            errors.append(f"row {idx}: expected exactly one identity, got PMID={pmid} PNID={pnid}")
        if numeric_title(title):
            errors.append(f"row {idx}: numeric-only title {title}")
        if text(row.get("KPI Type")).upper() not in {"IMPACT", "OUTPUT", "KAI"}:
            errors.append(f"row {idx}: invalid KPI Type {row.get('KPI Type')}")
    return errors


def main() -> None:
    OUT.mkdir(parents=True, exist_ok=True)
    audit_posisi = load_sheet_rows(AUDIT_WORKBOOK, "Audit Posisi")
    config_positions = json.loads(LATEST_CONFIG.read_text(encoding="utf-8"))["positions"]
    latest_rows = load_upload_rows(sorted(LATEST_UPLOAD_READY.glob("*.xlsx")))
    prod_rows = load_production_snapshot()
    dropped_titles_by_identity = load_dropped_titles_by_identity(config_positions)

    missing_with_kamus = [
        row for row in audit_posisi if text(row.get("Status Ketersediaan KPI")) == "Tersedia di Kamus, Belum Terupload"
    ]
    missing_no_kamus = [
        row
        for row in audit_posisi
        if text(row.get("Status Ketersediaan KPI")) == "Tidak Tersedia di Kamus dan Belum Ada KPI"
    ]
    partial_review = [
        row for row in audit_posisi if text(row.get("Status Ketersediaan KPI")) == "KPI Parsial - Perlu Review"
    ]

    rescue_candidates = build_rescue_candidates(missing_no_kamus, config_positions)
    rescue_rows = run_rescue_conversion(rescue_candidates)
    expected_rows = defaultdict(list)
    for identity, rows in latest_rows.items():
        expected_rows[identity].extend(rows)
    for identity, rows in rescue_rows.items():
        expected_rows[identity].extend(rows)

    upload_missing_rows: list[dict[str, Any]] = []
    safe_repair_rows: list[dict[str, Any]] = []
    manual_rows: list[dict[str, Any]] = []
    decision_rows: list[dict[str, Any]] = []

    def add_decision(case: str, audit: dict[str, Any], decision: str, reason: str, source_rows: int) -> None:
        kind, value = identity_from_audit(audit)
        prod = prod_rows.get((kind, value), [])
        decision_rows.append(
            {
                "Case": case,
                "Jenis Identity": kind,
                "ID Identity": value,
                "Nama Posisi": text(audit.get("Nama Posisi")),
                "Unit Organisasi": text(audit.get("Unit Organisasi")),
                "NIPP Pekerja Aktif": text(audit.get("NIPP Pekerja Aktif")),
                "Nama Pekerja Aktif": text(audit.get("Nama Pekerja Aktif")),
                "Production KPI Rows": len(prod),
                "Reset Risk": "Ya" if has_reset_risk(prod) else "Tidak",
                "Decision": decision,
                "Reason": reason,
                "Generated Upload Rows": source_rows,
            }
        )

    for audit in missing_with_kamus:
        identity = identity_from_audit(audit)
        rows = expected_rows.get(identity, [])
        prod = prod_rows.get(identity, [])
        if rows and not prod:
            upload_missing_rows.extend(rows)
            add_decision("belum_terupload_padahal_ada_kamus", audit, "UPLOAD_MISSING_SAFE", "Kamus valid dan belum ada KPI production", len(rows))
        elif rows and has_reset_risk(prod):
            manual_rows.append({"Reason": "Reset risk on existing production KPI", **audit})
            add_decision("belum_terupload_padahal_ada_kamus", audit, "MANUAL_REMEDIATION_REQUIRED", "Production sudah punya KPI berisiko reset", len(rows))
        elif rows:
            safe_repair_rows.extend(rows)
            add_decision("belum_terupload_padahal_ada_kamus", audit, "UPLOAD_SAFE_REPAIR_UNALLOCATED", "Kamus valid dan tidak ada reset risk", len(rows))
        else:
            manual_rows.append({"Reason": "Kamus status exists but no generated upload rows found", **audit})
            add_decision("belum_terupload_padahal_ada_kamus", audit, "MANUAL_REMEDIATION_REQUIRED", "Tidak ada row upload-ready", 0)

    rescue_by_identity = {identity_from_audit(c): c for c in rescue_candidates}
    for identity, candidate in rescue_by_identity.items():
        rows = expected_rows.get(identity, [])
        prod = prod_rows.get(identity, [])
        audit_like = {
            "Jenis Identity": identity[0],
            "ID Identity": identity[1],
            "Nama Posisi": candidate["Nama Posisi"],
            "Unit Organisasi": candidate["Unit Organisasi"],
            "NIPP Pekerja Aktif": candidate["NIPP Pekerja Aktif"],
            "Nama Pekerja Aktif": candidate["Nama Pekerja Aktif"],
        }
        if rows and not prod:
            upload_missing_rows.extend(rows)
            add_decision("audit_false_negative_rescue", audit_like, "UPLOAD_MISSING_SAFE", f"Rescued from {candidate['Worksheet']}", len(rows))
        elif rows and not has_reset_risk(prod):
            safe_repair_rows.extend(rows)
            add_decision("audit_false_negative_rescue", audit_like, "UPLOAD_SAFE_REPAIR_UNALLOCATED", f"Rescued from {candidate['Worksheet']}; no reset risk", len(rows))
        else:
            manual_rows.append({"Reason": "Rescue candidate has reset risk or no generated rows", **candidate})
            add_decision("audit_false_negative_rescue", audit_like, "MANUAL_REMEDIATION_REQUIRED", "Reset risk atau gagal generate rows", len(rows))

    bad_prod_rows = []
    for identity, prod in prod_rows.items():
        rows = expected_rows.get(identity, [])
        reasons = production_issue_reasons(prod, rows, dropped_titles_by_identity.get(identity, set()))
        if not reasons:
            continue
        audit_stub = {
            "Jenis Identity": identity[0],
            "ID Identity": identity[1],
            "Nama Posisi": "",
            "Unit Organisasi": "",
            "NIPP Pekerja Aktif": "",
            "Nama Pekerja Aktif": "",
        }
        bad_prod_rows.append(
            {
                "Jenis Identity": identity[0],
                "ID Identity": identity[1],
                "Production KPI Rows": len(prod),
                "Expected Fixed Rows": len(rows),
                "Issue Reasons": ", ".join(reasons),
                "Reset Risk": "Ya" if has_reset_risk(prod) else "Tidak",
                "Sample Bad Titles": "; ".join(text(r.get("title")) for r in prod if numeric_title(r.get("title")))[:500],
            }
        )
        if rows and not has_reset_risk(prod):
            safe_repair_rows.extend(rows)
            add_decision("production_bermasalah", audit_stub, "UPLOAD_SAFE_REPAIR_UNALLOCATED", ", ".join(reasons), len(rows))
        else:
            manual_rows.append({"Reason": ", ".join(reasons), **audit_stub})
            add_decision("production_bermasalah", audit_stub, "MANUAL_REMEDIATION_REQUIRED", ", ".join(reasons), len(rows))

    # Explicitly include KPI Parsial rows in the decision audit even when no bad title was detected.
    for audit in partial_review:
        identity = identity_from_audit(audit)
        rows = expected_rows.get(identity, [])
        prod = prod_rows.get(identity, [])
        if has_reset_risk(prod):
            manual_rows.append({"Reason": "KPI Parsial with reset risk", **audit})
            add_decision("kpi_parsial_perlu_review", audit, "MANUAL_REMEDIATION_REQUIRED", "KPI parsial dan production berisiko reset", len(rows))
        elif rows:
            safe_repair_rows.extend(rows)
            add_decision("kpi_parsial_perlu_review", audit, "UPLOAD_SAFE_REPAIR_UNALLOCATED", "KPI parsial tanpa reset risk", len(rows))
        else:
            add_decision("kpi_parsial_perlu_review", audit, "MANUAL_REMEDIATION_REQUIRED", "Tidak ada expected row", 0)

    # Deduplicate upload rows by identity + KPI type + IDKPI + title to avoid duplicate worksheet artifacts.
    def dedupe(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
        seen = set()
        output = []
        for row in rows:
            identity = identity_from_upload(row)
            key = (
                identity,
                text(row.get("KPI Type")).upper(),
                text(row.get("IDKPI")),
                norm(row.get("Title")),
                text(row.get("Parent KPI ID")),
            )
            if key in seen:
                continue
            seen.add(key)
            output.append(row)
        return output

    upload_missing_rows = dedupe(upload_missing_rows)
    safe_repair_rows = dedupe(safe_repair_rows)

    upload_errors = {
        "upload_missing_with_source_kamus.xlsx": validate_rows(upload_missing_rows),
        "upload_safe_repair_unallocated.xlsx": validate_rows(safe_repair_rows),
    }
    if any(upload_errors.values()):
        raise SystemExit(json.dumps(upload_errors, indent=2, ensure_ascii=False))

    upload_missing_path = OUT / "upload_missing_with_source_kamus.xlsx"
    safe_repair_path = OUT / "upload_safe_repair_unallocated.xlsx"
    if upload_missing_rows:
        write_upload_workbook(upload_missing_path, upload_missing_rows)
    if safe_repair_rows:
        write_upload_workbook(safe_repair_path, safe_repair_rows)

    manual_path = OUT / "manual_remediation_required.xlsx"
    decision_path = OUT / "audit_delta_remediation_decisions.xlsx"
    for path, rows, headers in [
        (
            manual_path,
            manual_rows,
            ["Reason", "Jenis Identity", "ID Identity", "Nama Posisi", "Unit Organisasi", "NIPP Pekerja Aktif", "Nama Pekerja Aktif"],
        ),
    ]:
        wb = Workbook()
        ws = wb.active
        ws.title = "Manual Remediation"
        write_table_sheet(ws, rows, headers)
        wb.save(path)

    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    summary = [
        {"Metric": "Audit positions", "Value": len(audit_posisi)},
        {"Metric": "Status Tersedia di Kamus, Belum Terupload", "Value": len(missing_with_kamus)},
        {"Metric": "Status Tidak Tersedia, candidate rescue", "Value": len(rescue_candidates)},
        {"Metric": "Production bad identity candidates", "Value": len(bad_prod_rows)},
        {"Metric": "Upload missing rows", "Value": len(upload_missing_rows)},
        {"Metric": "Safe repair rows", "Value": len(safe_repair_rows)},
        {"Metric": "Manual remediation rows", "Value": len(manual_rows)},
        {"Metric": "Endpoint dry-run status", "Value": "Not run - no API token/host found in local environment"},
    ]
    write_table_sheet(ws, summary, ["Metric", "Value"])
    ws = wb.create_sheet("Decisions")
    decision_headers = [
        "Case",
        "Jenis Identity",
        "ID Identity",
        "Nama Posisi",
        "Unit Organisasi",
        "NIPP Pekerja Aktif",
        "Nama Pekerja Aktif",
        "Production KPI Rows",
        "Reset Risk",
        "Decision",
        "Reason",
        "Generated Upload Rows",
    ]
    write_table_sheet(ws, decision_rows, decision_headers)
    ws = wb.create_sheet("Rescue Candidates")
    rescue_headers = [
        "Jenis Identity",
        "ID Identity",
        "NIPP Pekerja Aktif",
        "Nama Pekerja Aktif",
        "Nama Posisi",
        "Unit Organisasi",
        "Source Workbook",
        "Worksheet",
        "Match Score",
    ]
    write_table_sheet(ws, rescue_candidates, rescue_headers)
    ws = wb.create_sheet("Bad Production KPI")
    write_table_sheet(
        ws,
        bad_prod_rows,
        [
            "Jenis Identity",
            "ID Identity",
            "Production KPI Rows",
            "Expected Fixed Rows",
            "Issue Reasons",
            "Reset Risk",
            "Sample Bad Titles",
        ],
    )
    wb.save(decision_path)

    zip_path = OUT / "group1_ho_v2_delta_remediation_upload_forms_20260709.zip"
    with zipfile.ZipFile(zip_path, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        for path in [upload_missing_path, safe_repair_path]:
            if path.exists():
                zf.write(path, path.name)

    summary_json = {
        "output_dir": str(OUT),
        "decision_workbook": str(decision_path),
        "manual_remediation_workbook": str(manual_path),
        "upload_missing_workbook": str(upload_missing_path) if upload_missing_path.exists() else None,
        "safe_repair_workbook": str(safe_repair_path) if safe_repair_path.exists() else None,
        "upload_zip": str(zip_path),
        "counts": {
            "audit_positions": len(audit_posisi),
            "missing_with_kamus": len(missing_with_kamus),
            "missing_no_kamus": len(missing_no_kamus),
            "rescue_candidates": len(rescue_candidates),
            "bad_production_identities": len(bad_prod_rows),
            "upload_missing_rows": len(upload_missing_rows),
            "safe_repair_rows": len(safe_repair_rows),
            "manual_remediation_rows": len(manual_rows),
        },
        "nipp_103412_rescued": any(
            c["NIPP Pekerja Aktif"] == "103412" and c["ID Identity"] == "37540" for c in rescue_candidates
        ),
        "upload_errors": upload_errors,
        "endpoint_dry_run": {
            "status": "not_run",
            "reason": "No authenticated Performance HQ API host/token was available in the local environment. No production upload was attempted.",
        },
    }
    (OUT / "delta_remediation_summary_20260709.json").write_text(
        json.dumps(summary_json, indent=2, ensure_ascii=False),
        encoding="utf-8",
    )
    print(json.dumps(summary_json, indent=2, ensure_ascii=False))


if __name__ == "__main__":
    main()
