#!/usr/bin/env python3
from __future__ import annotations

import csv
import json
import re
import shutil
import subprocess
import zipfile
from collections import Counter, defaultdict
from copy import copy
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "output/group1_ho_v2_unconverted_mapped_20260709"
BASE_CONFIG = ROOT / "output/group1_ho_v2_20260709_latest_prod/group1_ho_v2_20260709_upload.config.json"
REFERENCE = ROOT / "output/group1_ho_v2_20260709_latest_prod/production_position_reference_20260709.json"
SNAPSHOT = ROOT / "output/group1_ho_v2_delta_remediation_20260709/production_kpi_snapshot_20260709.json"
TEMPLATE = ROOT / "input/KPI Upload Template.xlsx"
SOURCE_ZIP = ROOT / "tmp/group1_ho_v2_source_20260703.zip"
PREVIOUS_DELTA = ROOT / "output/group1_ho_v2_delta_remediation_20260709/upload_missing_with_source_kamus.xlsx"


UPLOAD_HEADERS = [
    "IDKPI",
    "Group",
    "Direktorat",
    "Posisi",
    "Position Master ID (Required)",
    "Position Master Variant ID (Optional)",
    "BSC Perspective",
    "KPI Type",
    "Parent KPI ID",
    "Parent KPI Title",
    "Title",
    "Description",
    "Unit",
    "Polarity",
    "Period",
    "Formula",
    "Weight (%)",
    "Cascading",
    "Nature Of Work (KAI Only)",
    "External ID (PKPI)",
    "System KPI ID",
    "Ownership Type",
    "Position Nomenklatur ID",
    "RKM Code ID",
]


EXPLICIT_MAPPINGS = [
    # Already produced in the earlier delta package, retained in this one-file handoff.
    {"sheet": "DH Keberlangsungan-Pengoprasian", "kind": "PMID", "id": "37541", "source": "previous_delta"},
    {"sheet": "DH Tata Kelola IT & Pengelolaan", "kind": "PMID", "id": "37540", "source": "previous_delta"},
    {"sheet": "Manager Key Account", "kind": "PMID", "id": "37528", "source": "previous_delta"},
    # New/remaining non-Pengendalian mappings.
    {"sheet": "DH Manajemen Data", "kind": "PMID", "id": "37539"},
    {"sheet": "Group Head", "source_workbook_contains": "Group Monitoring Evaluasi", "kind": "PMID", "id": "37587"},
    {"sheet": "Officer MEKA ", "kind": "PNID", "id": "12541", "title": "Officer Monitoring dan Evaluasi Klaster Ekspansi Korporasi"},
    {"sheet": "Officer MEKO ", "kind": "PNID", "id": "12553", "title": "Officer Monitoring dan Evaluasi Klaster Optimalisasi Korporasi"},
    {"sheet": "DH Sistem Manajemen", "kind": "PMID", "id": "37572"},
    {"sheet": "Manager Pengembangan K3", "kind": "PMID", "id": "37577"},
    {"sheet": "Pimpro Satker Single ERP", "kind": "PMID", "id": "37583", "force_include": True},
    {"sheet": "DH Litigasi", "kind": "PMID", "id": "35886", "title": "Department Head Litigasi"},
    {
        "sheet": "Officer Sistem Manajemen",
        "kind": "PMID",
        "id": "37576",
        "title": "Officer I Sistem Manajemen",
    },
    {
        "sheet": "Officer Sistem Manajemen",
        "kind": "PMID",
        "id": "37574",
        "title": "Senior Officer I Sistem Manajemen",
    },
    {
        "sheet": "Officer Sistem Manajemen",
        "kind": "PMID",
        "id": "37573",
        "title": "Senior Officer III Sistem Manajemen",
    },
    {
        "sheet": "Officer Employee Service 1",
        "kind": "PNID",
        "id": "97",
        "title": "Officer Data Management 1",
    },
    {
        "sheet": "Officer Employee Service 1",
        "kind": "PNID",
        "id": "98",
        "title": "Officer Travel Management 1",
    },
    {
        "sheet": "Officer Employee Service 2",
        "kind": "PNID",
        "id": "100",
        "title": "Officer Travel Management 2",
    },
    {
        "sheet": "Officer Employee Service 3",
        "kind": "PNID",
        "id": "102",
        "title": "Officer Travel Management 3",
    },
]


def text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def identity(row: dict[str, Any]) -> tuple[str, str] | None:
    pmid = text(row.get("Position Master ID (Required)"))
    pnid = text(row.get("Position Nomenklatur ID"))
    if pmid and not pnid:
        return ("PMID", pmid)
    if pnid and not pmid:
        return ("PNID", pnid)
    return None


def load_upload_rows(path: Path) -> dict[tuple[str, str], list[dict[str, Any]]]:
    output: dict[tuple[str, str], list[dict[str, Any]]] = defaultdict(list)
    if not path.exists():
        return output
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb["KPI Template"]
    rows = ws.iter_rows(values_only=True)
    headers = [text(v) for v in next(rows)]
    for values in rows:
        row = {headers[i]: values[i] if i < len(values) else None for i in range(len(headers))}
        ident = identity(row)
        if ident:
            output[ident].append(row)
    return output


def load_production_counts() -> Counter:
    counts: Counter = Counter()
    if not SNAPSHOT.exists():
        return counts
    for row in json.loads(SNAPSHOT.read_text(encoding="utf-8"))["rows"]:
        pnid = text(row.get("pnid"))
        pmid = text(row.get("pmid"))
        if pnid:
            counts[("PNID", pnid)] += 1
        elif pmid:
            counts[("PMID", pmid)] += 1
    return counts


def find_base_config(mapping: dict[str, str], positions: list[dict[str, Any]]) -> dict[str, Any] | None:
    for pos in positions:
        if text(pos.get("sheet_name")) != text(mapping["sheet"]):
            continue
        if mapping.get("source_workbook_contains") and mapping["source_workbook_contains"] not in text(pos.get("source_workbook")):
            continue
        if "Pengendalian Proyek" in text(pos.get("source_workbook")):
            continue
        return pos
    return None


def make_config(base: dict[str, Any], mapping: dict[str, str]) -> dict[str, Any]:
    item = copy(base)
    item["mapping_review_status"] = "approved"
    item["mapping_override_approved"] = True
    item["mapping_override_trust_source"] = "audit_organisasi_kamus_kpi_ho"
    item["mapping_confidence_label"] = "high_confidence"
    item["mapping_confidence_reason"] = "Manual mapping from Audit Posisi and production reference; Pengendalian Proyek excluded."
    if mapping.get("title"):
        item["position_name"] = mapping["title"]
        item["portaverse_position_title"] = mapping["title"]
    if mapping["kind"] == "PMID":
        item["position_scope"] = "structural"
        item["position_master_id"] = mapping["id"]
        item["position_nomenclature_id"] = None
    else:
        item["position_scope"] = "non_structural"
        item["position_master_id"] = None
        item["position_nomenclature_id"] = mapping["id"]
    return item


def run_single_conversion(mapping: dict[str, str], positions: list[dict[str, Any]]) -> tuple[list[dict[str, Any]], str]:
    base = find_base_config(mapping, positions)
    if not base:
        return [], "base_config_not_found"
    cfg = {"reference_source": str(REFERENCE), "positions": [make_config(base, mapping)]}
    safe = re.sub(r"[^A-Za-z0-9]+", "_", f"{mapping['sheet']}_{mapping['kind']}_{mapping['id']}").strip("_")
    cfg_path = OUT / "single_configs" / f"{safe}.config.json"
    cfg_path.parent.mkdir(parents=True, exist_ok=True)
    cfg_path.write_text(json.dumps(cfg, indent=2, ensure_ascii=False), encoding="utf-8")
    output_dir = OUT / "single_conversion" / safe
    if output_dir.exists():
        shutil.rmtree(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    cmd = [
        "python3",
        "scripts/kpi_bulk_transform.py",
        "--source",
        str(SOURCE_ZIP),
        "--template",
        str(TEMPLATE),
        "--mapping",
        str(REFERENCE),
        "--config",
        str(cfg_path),
        "--output-dir",
        str(output_dir),
    ]
    proc = subprocess.run(cmd, cwd=ROOT, text=True, capture_output=True)
    errors = []
    for report in output_dir.glob("*/*.report.csv"):
        with report.open(newline="", encoding="utf-8") as f:
            for row in csv.DictReader(f):
                if row.get("severity") == "error":
                    errors.append(row.get("message", "error"))
    if proc.returncode != 0 or errors:
        return [], "; ".join(errors) or proc.stderr or proc.stdout
    rows_by_identity: dict[tuple[str, str], list[dict[str, Any]]] = defaultdict(list)
    for workbook in output_dir.glob("*/*.xlsx"):
        for ident, rows in load_upload_rows(workbook).items():
            rows_by_identity[ident].extend(rows)
    return rows_by_identity.get((mapping["kind"], mapping["id"]), []), "ok"


def write_upload(path: Path, rows: list[dict[str, Any]]) -> None:
    wb = load_workbook(TEMPLATE)
    ws = wb["KPI Template"] if "KPI Template" in wb.sheetnames else wb.active
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)
    for table_name in list(ws.tables.keys()):
        del ws.tables[table_name]
    for col, header in enumerate(UPLOAD_HEADERS, start=1):
        ws.cell(1, col, header)
    for r_idx, row in enumerate(rows, start=2):
        for c_idx, header in enumerate(UPLOAD_HEADERS, start=1):
            ws.cell(r_idx, c_idx, row.get(header))
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def validate(rows: list[dict[str, Any]]) -> list[str]:
    errors = []
    for idx, row in enumerate(rows, start=2):
        ident = identity(row)
        if not ident:
            errors.append(f"row {idx}: blank/double identity")
        title = text(row.get("Title"))
        if re.fullmatch(r"\d+(?:\.0+)?", title):
            errors.append(f"row {idx}: numeric-only title {title}")
    return errors


def main() -> None:
    OUT.mkdir(parents=True, exist_ok=True)
    positions = json.loads(BASE_CONFIG.read_text(encoding="utf-8"))["positions"]
    previous_rows = load_upload_rows(PREVIOUS_DELTA)
    prod_counts = load_production_counts()
    final_rows: list[dict[str, Any]] = []
    audit_rows: list[dict[str, Any]] = []
    seen = set()
    for mapping in EXPLICIT_MAPPINGS:
        ident = (mapping["kind"], mapping["id"])
        force_include = bool(mapping.get("force_include"))
        if prod_counts[ident] and mapping.get("source") != "previous_delta" and not force_include:
            status = "SKIPPED_ALREADY_HAS_PRODUCTION_KPI"
            rows: list[dict[str, Any]] = []
            detail = f"production rows={prod_counts[ident]}"
        elif mapping.get("source") == "previous_delta":
            rows = previous_rows.get(ident, [])
            status = "INCLUDED_FROM_PREVIOUS_DELTA" if rows else "PREVIOUS_DELTA_ROWS_NOT_FOUND"
            detail = ""
        else:
            rows, detail = run_single_conversion(mapping, positions)
            status = "INCLUDED_CONVERTED" if rows else "BLOCKED_CONVERSION_ERROR"
        for row in rows:
            key = (
                identity(row),
                text(row.get("KPI Type")),
                text(row.get("IDKPI")),
                text(row.get("Title")).lower(),
                text(row.get("Parent KPI ID")),
            )
            if key in seen:
                continue
            seen.add(key)
            final_rows.append(row)
        audit_rows.append(
            {
                "Worksheet": mapping["sheet"],
                "Jenis Identity": mapping["kind"],
                "ID Identity": mapping["id"],
                "Status": status,
                "Rows": len(rows),
                "Production KPI Rows": prod_counts[ident],
                "Detail": detail[:1000],
            }
        )
    errors = validate(final_rows)
    if errors:
        raise SystemExit("\n".join(errors))
    upload_path = OUT / "group1_ho_v2_unconverted_mapped_one_upload_20260709.xlsx"
    write_upload(upload_path, final_rows)
    zip_path = OUT / "group1_ho_v2_unconverted_mapped_one_upload_20260709.zip"
    with zipfile.ZipFile(zip_path, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        zf.write(upload_path, upload_path.name)
    audit_path = OUT / "mapping_audit_unconverted_20260709.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Mapping Audit"
    headers = ["Worksheet", "Jenis Identity", "ID Identity", "Status", "Rows", "Production KPI Rows", "Detail"]
    ws.append(headers)
    for row in audit_rows:
        ws.append([row[h] for h in headers])
    ws.freeze_panes = "A2"
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
    for column in ws.columns:
        width = max(len(text(cell.value)) for cell in column)
        ws.column_dimensions[get_column_letter(column[0].column)].width = min(max(width + 2, 12), 80)
    wb.save(audit_path)
    summary = {
        "output_dir": str(OUT),
        "upload_workbook": str(upload_path),
        "upload_zip": str(zip_path),
        "mapping_audit": str(audit_path),
        "rows": len(final_rows),
        "identities": len({identity(row) for row in final_rows}),
        "type_counts": dict(Counter(text(row.get("KPI Type")).upper() for row in final_rows)),
        "audit_status_counts": dict(Counter(row["Status"] for row in audit_rows)),
        "pengendalian_proyek_excluded": True,
        "validation_errors": errors,
    }
    (OUT / "summary_20260709.json").write_text(json.dumps(summary, indent=2, ensure_ascii=False), encoding="utf-8")
    print(json.dumps(summary, indent=2, ensure_ascii=False))


if __name__ == "__main__":
    main()
