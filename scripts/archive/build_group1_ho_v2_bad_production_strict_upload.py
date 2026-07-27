#!/usr/bin/env python3
from __future__ import annotations

import json
import re
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook


ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "output/group1_ho_v2_delta_remediation_20260709"
BAD_AUDIT = OUT / "audit_delta_remediation_decisions.xlsx"
SAFE_REPAIR = OUT / "upload_safe_repair_unallocated.xlsx"
CANONICAL_IMPACT = ROOT / "output/group1_ho_v2_20260709_latest_prod/missing_ho_impact_only_20260709.xlsx"
TEMPLATE = ROOT / "input/KPI Upload Template.xlsx"
OUTPUT = OUT / "upload_all_40_bad_production_positions_strict_10_impact.xlsx"
AUDIT_OUTPUT = OUT / "upload_all_40_bad_production_positions_strict_10_impact_audit.xlsx"

HEADERS = [
    "IDKPI",
    "Group",
    "Direktorat",
    "Posisi",
    "Position Master ID (Required)",
    "Position Master Variant ID (Optional)",
    "BSC Perspective",
    "KPI Type",
    "Parent KPI ID",
    "Parent KPI Title",
    "Title",
    "Description",
    "Unit",
    "Polarity",
    "Period",
    "Formula",
    "Weight (%)",
    "Cascading",
    "Nature Of Work (KAI Only)",
    "External ID (PKPI)",
    "System KPI ID",
    "Ownership Type",
    "Position Nomenklatur ID",
    "RKM Code ID",
]


def text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def is_numeric_title(value: Any) -> bool:
    return bool(re.fullmatch(r"\d+(?:\.0+)?", text(value)))


def row_identity(row: dict[str, Any]) -> tuple[str, str] | None:
    pmid = text(row.get("Position Master ID (Required)"))
    pnid = text(row.get("Position Nomenklatur ID"))
    if pmid and not pnid:
        return "PMID", pmid
    if pnid and not pmid:
        return "PNID", pnid
    return None


def load_template_rows(path: Path) -> list[dict[str, Any]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb["KPI Template"]
    rows = ws.iter_rows(values_only=True)
    headers = [text(v) for v in next(rows)]
    output = []
    for values in rows:
        output.append({headers[i]: values[i] if i < len(values) else None for i in range(len(headers))})
    return output


def load_bad_identities() -> list[dict[str, Any]]:
    wb = load_workbook(BAD_AUDIT, read_only=True, data_only=True)
    ws = wb["Bad Production KPI"]
    rows = ws.iter_rows(values_only=True)
    headers = [text(v) for v in next(rows)]
    result = []
    for values in rows:
        item = {headers[i]: values[i] if i < len(values) else None for i in range(len(headers))}
        if text(item.get("Jenis Identity")) and text(item.get("ID Identity")):
            result.append(item)
    return result


def canonical_impact_rows() -> list[dict[str, Any]]:
    rows = [r for r in load_template_rows(CANONICAL_IMPACT) if text(r.get("KPI Type")).upper() == "IMPACT"]
    seen = set()
    canonical = []
    for row in rows:
        title = text(row.get("Title"))
        if title in seen:
            continue
        seen.add(title)
        item = {header: row.get(header, "") for header in HEADERS}
        item["Position Master ID (Required)"] = ""
        item["Position Nomenklatur ID"] = ""
        item["Position Master Variant ID (Optional)"] = ""
        item["System KPI ID"] = ""
        item["RKM Code ID"] = ""
        canonical.append(item)
        if len(canonical) == 10:
            return canonical
    raise RuntimeError(f"Expected 10 canonical impact rows, got {len(canonical)}")


def safe_repair_detail_rows() -> dict[tuple[str, str], list[dict[str, Any]]]:
    rows_by_identity: dict[tuple[str, str], list[dict[str, Any]]] = defaultdict(list)
    for row in load_template_rows(SAFE_REPAIR):
        identity = row_identity(row)
        if not identity:
            continue
        if text(row.get("KPI Type")).upper() == "IMPACT":
            continue
        rows_by_identity[identity].append({header: row.get(header, "") for header in HEADERS})
    return rows_by_identity


def clone_impacts_for_identity(identity: tuple[str, str], source: list[dict[str, Any]]) -> list[dict[str, Any]]:
    rows = []
    for index, row in enumerate(source, start=1):
        item = dict(row)
        item["IDKPI"] = str(index)
        item["Position Master ID (Required)"] = identity[1] if identity[0] == "PMID" else ""
        item["Position Nomenklatur ID"] = identity[1] if identity[0] == "PNID" else ""
        rows.append(item)
    return rows


def validate(rows: list[dict[str, Any]], expected: set[tuple[str, str]]) -> list[str]:
    errors = []
    identities = set()
    impact_counts = Counter()
    for idx, row in enumerate(rows, start=2):
        identity = row_identity(row)
        if not identity:
            errors.append(f"row {idx}: blank/double identity")
            continue
        identities.add(identity)
        if is_numeric_title(row.get("Title")):
            errors.append(f"row {idx}: numeric title {row.get('Title')}")
        if text(row.get("KPI Type")).upper() == "IMPACT":
            impact_counts[identity] += 1
    missing = expected - identities
    if missing:
        errors.append(f"missing identities: {sorted(missing)}")
    bad_impact = {identity: count for identity, count in impact_counts.items() if count != 10}
    if bad_impact:
        errors.append(f"bad impact counts: {bad_impact}")
    return errors


def write_upload(rows: list[dict[str, Any]]) -> None:
    wb = load_workbook(TEMPLATE)
    ws = wb["KPI Template"] if "KPI Template" in wb.sheetnames else wb.active
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)
    for table_name in list(ws.tables.keys()):
        del ws.tables[table_name]
    for col, header in enumerate(HEADERS, start=1):
        ws.cell(1, col, header)
    for row_idx, row in enumerate(rows, start=2):
        for col, header in enumerate(HEADERS, start=1):
            ws.cell(row_idx, col, row.get(header))
    wb.save(OUTPUT)


def main() -> None:
    bad = load_bad_identities()
    expected = {(text(row["Jenis Identity"]), text(row["ID Identity"])) for row in bad}
    detail_by_identity = safe_repair_detail_rows()
    impacts = canonical_impact_rows()
    final_rows: list[dict[str, Any]] = []
    audit_rows = []
    for row in bad:
        identity = (text(row["Jenis Identity"]), text(row["ID Identity"]))
        impact_rows = clone_impacts_for_identity(identity, impacts)
        detail_rows = detail_by_identity.get(identity, [])
        final_rows.extend(impact_rows)
        final_rows.extend(detail_rows)
        audit_rows.append(
            {
                "Identity": f"{identity[0]} {identity[1]}",
                "Impact Rows": len(impact_rows),
                "Output/KAI Rows": len(detail_rows),
                "Detail Source": "corrected_converter" if detail_rows else "none_detail_mapping_required",
                "Issue Reasons": row.get("Issue Reasons"),
                "Production KPI Rows": row.get("Production KPI Rows"),
            }
        )
    errors = validate(final_rows, expected)
    if errors:
        raise SystemExit("\n".join(errors))
    write_upload(final_rows)

    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    type_counts = Counter(text(row.get("KPI Type")).upper() for row in final_rows)
    summary = [
        ("Bad production identities", len(expected)),
        ("Generated upload rows", len(final_rows)),
        ("IMPACT rows", type_counts["IMPACT"]),
        ("OUTPUT rows", type_counts["OUTPUT"]),
        ("KAI rows", type_counts["KAI"]),
        ("Identities with corrected detail rows", sum(1 for row in audit_rows if row["Output/KAI Rows"])),
        ("Identities impact-only in this form", sum(1 for row in audit_rows if not row["Output/KAI Rows"])),
    ]
    for item in summary:
        ws.append(item)
    ws = wb.create_sheet("Identity Sources")
    headers = ["Identity", "Impact Rows", "Output/KAI Rows", "Detail Source", "Issue Reasons", "Production KPI Rows"]
    ws.append(headers)
    for row in audit_rows:
        ws.append([row.get(header, "") for header in headers])
    wb.save(AUDIT_OUTPUT)
    print(json.dumps({"output": str(OUTPUT), "audit": str(AUDIT_OUTPUT), "rows": len(final_rows), "types": dict(type_counts)}, indent=2))


if __name__ == "__main__":
    main()
