#!/usr/bin/env python3
from __future__ import annotations

from collections import Counter
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
SOURCE = Path("/Users/alfredoteja/Downloads/position_mapping_report (02:07).xlsx")
OUT_DIR = ROOT / "output/position_mapping_manual_audit_20260710"
OUT_FILE = OUT_DIR / "position_mapping_manual_audit_20260710.xlsx"


BASE_HEADERS = [
    "Source Workbook",
    "Worksheet",
    "Raw Worksheet Title",
    "Normalized Worksheet Title",
    "Inferred Scope",
    "Confidence Label",
    "Confidence Reason",
    "Candidate PMID",
    "Candidate PNID",
    "Candidate Title",
    "Candidate Group",
    "Candidate Company",
    "Candidate Score",
    "Active Variant Count",
    "Active Employee Count",
    "Active Employee Name",
    "Active Employee NIPP",
    "Recommended Action",
]

MANUAL_HEADERS = [
    "Manual Status",
    "Reviewer Confirm Mapping",
    "Reviewer Actual PMID",
    "Reviewer Actual PNID",
    "Reviewer Notes",
]


def text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def load_rows() -> list[dict[str, Any]]:
    wb = load_workbook(SOURCE, read_only=True, data_only=True)
    ws = wb["Position Mapping Report"]
    headers = [text(c.value) for c in next(ws.iter_rows(max_row=1))]
    rows: list[dict[str, Any]] = []
    for values in ws.iter_rows(min_row=2, values_only=True):
        row = {headers[i]: values[i] if i < len(values) else None for i in range(len(headers))}
        if text(row.get("Source Workbook")) and text(row.get("Worksheet")):
            rows.append(row)
    return rows


def has_destination(row: dict[str, Any]) -> bool:
    return any(
        text(row.get(key))
        for key in ["Reviewer Actual PMID", "Reviewer Actual PNID", "Candidate PMID", "Candidate PNID"]
    )


def manual_status(row: dict[str, Any]) -> str:
    if not has_destination(row):
        return "ISI PMID/PNID MANUAL"
    if text(row.get("Confidence Label")) == "mapping_conflict":
        return "PILIH SALAH SATU CANDIDATE"
    if text(row.get("Confidence Label")) == "low_confidence":
        return "REVIEW CANDIDATE"
    if text(row.get("Confidence Label")) == "scope_uncertain":
        return "TENTUKAN SCOPE"
    return "AUTO OK"


def write_sheet(ws, rows: list[dict[str, Any]], headers: list[str]) -> None:
    ws.append(headers)
    for source in rows:
        row = dict(source)
        row.setdefault("Manual Status", manual_status(source))
        row.setdefault("Reviewer Confirm Mapping", source.get("Reviewer Confirm Mapping"))
        row.setdefault("Reviewer Actual PMID", source.get("Reviewer Actual PMID"))
        row.setdefault("Reviewer Actual PNID", source.get("Reviewer Actual PNID"))
        row.setdefault("Reviewer Notes", source.get("Reviewer Notes"))
        ws.append([row.get(header, "") for header in headers])

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
    for column in ws.columns:
        width = max(len(text(cell.value)) for cell in column)
        ws.column_dimensions[get_column_letter(column[0].column)].width = min(max(width + 2, 12), 70)


def main() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    rows = load_rows()
    no_target = [row for row in rows if not has_destination(row)]
    review = [row for row in rows if text(row.get("Confidence Label")) != "high_confidence"]
    confidence_counts = Counter(text(row.get("Confidence Label")) for row in rows)
    scope_counts = Counter(text(row.get("Inferred Scope")) for row in rows)

    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    summary_rows = [
        {"Metric": "Source file", "Value": str(SOURCE)},
        {"Metric": "Worksheet rows audited", "Value": len(rows)},
        {"Metric": "Belum ada tujuan PMID/PNID", "Value": len(no_target)},
        {"Metric": "Perlu review manual", "Value": len(review)},
    ]
    for key, value in confidence_counts.items():
        summary_rows.append({"Metric": f"Confidence - {key}", "Value": value})
    for key, value in scope_counts.items():
        summary_rows.append({"Metric": f"Scope - {key}", "Value": value})
    write_sheet(ws, summary_rows, ["Metric", "Value"])

    manual_headers = BASE_HEADERS + MANUAL_HEADERS
    ws = wb.create_sheet("Belum Ada Tujuan")
    write_sheet(ws, no_target, manual_headers)

    ws = wb.create_sheet("Perlu Review")
    write_sheet(ws, review, manual_headers)

    ws = wb.create_sheet("All Mapping Rows")
    write_sheet(ws, rows, manual_headers)

    wb.save(OUT_FILE)
    print(OUT_FILE)
    print(
        {
            "rows": len(rows),
            "no_target": len(no_target),
            "review": len(review),
            "confidence": dict(confidence_counts),
        }
    )


if __name__ == "__main__":
    main()
