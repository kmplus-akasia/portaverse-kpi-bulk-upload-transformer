#!/usr/bin/env python3
"""Validate generated KPI upload workbooks and create an upload manifest."""

from __future__ import annotations

import argparse
import csv
import json
import re
import shutil
import sys
import zipfile
from pathlib import Path

import openpyxl


EXPECTED_HEADERS = [
    "IDKPI",
    "Group",
    "Direktorat",
    "Posisi",
    "Position Master ID (Required)",
    "Position Master Variant ID (Optional)",
    "BSC Perspective",
    "KPI Type",
    "Parent KPI ID",
    "Parent KPI Title",
    "Title",
    "Description",
    "Unit",
    "Polarity",
    "Period",
    "Formula",
    "Weight (%)",
    "Cascading",
    "Nature Of Work (KAI Only)",
    "External ID (PKPI)",
    "System KPI ID",
    "Ownership Type",
    "Position Nomenklatur ID",
    "RKM Code ID",
]

TRUST_SOURCE_REVIEWER_MANUAL = "reviewer_manual"


def norm(value: object) -> str:
    text = str(value or "").lower()
    text = re.sub(r"\bdh\b", "department head", text)
    text = text.replace("&", " dan ")
    text = re.sub(r"[-_/(),.]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def is_numeric_only_title(value: object) -> bool:
    text = str(value or "").strip()
    return bool(text) and re.fullmatch(r"\d+(?:\.0+)?", text) is not None


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser()
    parser.add_argument("--output-dir", required=True, type=Path)
    parser.add_argument("--config", required=True, type=Path)
    parser.add_argument("--reference", required=True, type=Path)
    parser.add_argument("--fixed-audit", type=Path)
    parser.add_argument("--expected-workbooks", type=int, default=26)
    parser.add_argument("--upload-ready-dir", type=Path)
    parser.add_argument("--zip-output", type=Path)
    return parser.parse_args()


def load_reference_ids(
    reference_path: Path,
) -> tuple[
    set[str],
    set[str],
    dict[str, set[str]],
    dict[str, set[str]],
    dict[str, set[str]],
]:
    with reference_path.open() as handle:
        reference = json.load(handle)
    structural_rows = reference.get("structural_lookup_rows") or reference.get("position_master_rows", [])
    non_structural_rows = reference.get("non_structural_lookup_rows") or reference.get("rows", [])
    master_ids = {str(row["position_master_id"]) for row in structural_rows if row.get("position_master_id")}
    nomenclature_ids = {
        str(row["cluster_id"] or row.get("position_nomenclature_id"))
        for row in non_structural_rows
        if (row.get("cluster_id") or row.get("position_nomenclature_id")) not in (None, "", 0, "0")
    }
    cluster_labels_by_id: dict[str, set[str]] = {}
    master_types_by_id: dict[str, set[str]] = {}
    position_types_by_pnid: dict[str, set[str]] = {}
    for row in structural_rows:
        master_types_by_id.setdefault(str(row["position_master_id"]), set()).add(
            str(row.get("position_master_type_id") or "")
        )
    for row in non_structural_rows:
        cluster_id = row.get("cluster_id") or row.get("position_nomenclature_id")
        if cluster_id in (None, "", 0, "0"):
            continue
        cluster_labels_by_id.setdefault(str(cluster_id), set()).add(norm(row.get("cluster_label")))
        position_types_by_pnid.setdefault(str(cluster_id), set()).add(
            str(row.get("position_master_type_id") or "")
        )
    return (
        master_ids,
        nomenclature_ids,
        cluster_labels_by_id,
        master_types_by_id,
        position_types_by_pnid,
    )


def check_config_scope(
    config_path: Path,
    master_ids: set[str],
    nomenclature_ids: set[str],
    cluster_labels_by_id: dict[str, set[str]],
    master_types_by_id: dict[str, set[str]],
    position_types_by_pnid: dict[str, set[str]],
) -> list[str]:
    errors: list[str] = []
    with config_path.open() as handle:
        config = json.load(handle)
    structural_usage_by_pmid: dict[str, list[str]] = {}
    for pos in config["positions"]:
        scope = (pos.get("position_scope") or "").strip()
        pmid = str(pos.get("position_master_id") or "").strip()
        pnid = str(pos.get("position_nomenclature_id") or "").strip()
        confidence_label = (pos.get("mapping_confidence_label") or "").strip()
        review_status = norm(pos.get("mapping_review_status"))
        override_approved = str(pos.get("mapping_override_approved") or "").lower() in {"1", "true", "yes", "y"}
        review_approved = override_approved or review_status in {"approved", "review approved", "reviewer approved"}
        trusted_reviewer = review_approved and pos.get("mapping_override_trust_source") == TRUST_SOURCE_REVIEWER_MANUAL
        label = f"{pos.get('source_workbook')} :: {pos.get('sheet_name')}"
        if review_status == "needs check":
            continue
        if confidence_label in {"low_confidence", "scope_uncertain", "no_candidate", "mapping_conflict"} and not review_approved:
            errors.append(f"unapproved blocked mapping label {confidence_label}: {label}")
        if scope not in {
            "structural",
            "non_structural",
            "assistant",
            "neglect",
            "mapping_conflict",
            "scope_uncertain",
        }:
            errors.append(f"unsupported position scope {scope or '<blank>'}: {label}")
        if scope in {"mapping_conflict", "scope_uncertain"} and not review_approved:
            if pmid or pnid:
                errors.append(f"{scope} config should not have PMID/PNID before approval: {label}")
            continue
        if scope == "structural":
            if not pmid:
                errors.append(f"structural config missing PMID: {label}")
            if pnid:
                errors.append(f"structural config still has PNID: {label} -> {pnid}")
            if pmid:
                structural_usage_by_pmid.setdefault(pmid, []).append(label)
            production_types = master_types_by_id.get(pmid, set())
            if pmid and production_types != {"5"} and not trusted_reviewer:
                errors.append(
                    f"structural config PMID {pmid} production type "
                    f"{','.join(sorted(production_types)) or 'unknown'} is non-structural: {label}"
                )
        if scope == "assistant":
            pmvid = str(pos.get("position_master_variant_id") or "").strip()
            if pmid != "77":
                errors.append(f"assistant config must use PMID 77: {label} -> {pmid or '<blank>'}")
            if not pmvid:
                errors.append(f"assistant config missing PMVID: {label}")
            if pnid:
                errors.append(f"assistant config still has PNID: {label} -> {pnid}")
        if scope == "non_structural":
            if pmid:
                errors.append(f"non_structural config has PMID populated: {label} -> {pmid}")
            if pnid and pnid not in nomenclature_ids and pnid in master_ids:
                errors.append(
                    f"non_structural config points to structural PMID instead of valid PNID: "
                    f"{label} -> pnid={pnid}"
                )
            production_types = position_types_by_pnid.get(pnid, set())
            if pnid and (not production_types or "5" in production_types) and not trusted_reviewer:
                errors.append(
                    f"non_structural config PNID {pnid} has invalid production types "
                    f"{sorted(production_types)}: {label}"
                )
    for pmid, labels in sorted(structural_usage_by_pmid.items()):
        unique_labels = sorted(set(labels))
        if len(unique_labels) > 1:
            errors.append(
                f"duplicate structural PMID {pmid} mapped by {len(unique_labels)} worksheets: "
                + " | ".join(unique_labels)
            )
    return errors


def load_trusted_reviewer_ids(config_path: Path) -> tuple[set[str], set[str]]:
    with config_path.open() as handle:
        config = json.load(handle)
    trusted_pmids: set[str] = set()
    trusted_pnids: set[str] = set()
    for pos in config.get("positions", []):
        review_status = norm(pos.get("mapping_review_status"))
        override_approved = str(pos.get("mapping_override_approved") or "").lower() in {"1", "true", "yes", "y"}
        review_approved = override_approved or review_status in {"approved", "review approved", "reviewer approved"}
        if not review_approved or pos.get("mapping_override_trust_source") != TRUST_SOURCE_REVIEWER_MANUAL:
            continue
        pmid = str(pos.get("position_master_id") or "").strip()
        pnid = str(pos.get("position_nomenclature_id") or "").strip()
        if pmid and not pnid:
            trusted_pmids.add(pmid)
        if pnid and not pmid:
            trusted_pnids.add(pnid)
    return trusted_pmids, trusted_pnids


def load_fixed_pmids(fixed_audit_path: Path | None) -> set[str]:
    if not fixed_audit_path or not fixed_audit_path.exists():
        return set()
    with fixed_audit_path.open(newline="") as handle:
        rows = list(csv.DictReader(handle))
    return {
        str(row["pmid"]).strip()
        for row in rows
        if str(row.get("pmid") or "").strip()
        and (not row.get("resolved_scope") or row.get("resolved_scope") == "structural")
    }


def load_approved_assistant_identities(config_path: Path) -> set[tuple[str, str]]:
    with config_path.open() as handle:
        config = json.load(handle)
    return {
        (
            str(pos.get("position_master_id") or "").strip(),
            str(pos.get("position_master_variant_id") or "").strip(),
        )
        for pos in config["positions"]
        if (pos.get("position_scope") or "").strip() == "assistant"
        and str(pos.get("mapping_review_status") or "").strip().lower() == "approved"
        and str(pos.get("position_master_id") or "").strip()
        and str(pos.get("position_master_variant_id") or "").strip()
    }


def iter_upload_workbooks(output_dir: Path) -> list[Path]:
    paths = []
    for path in output_dir.rglob("*.xlsx"):
        if path.name.startswith("~$"):
            continue
        if path.parent.name == "upload-ready":
            continue
        if "Conversion Report" in path.name:
            continue
        paths.append(path)
    return sorted(paths)


def check_report_csvs(output_dir: Path) -> list[str]:
    errors: list[str] = []
    for report_path in sorted(output_dir.rglob("*.report.csv")):
        with report_path.open(newline="") as handle:
            for row in csv.DictReader(handle):
                if (row.get("severity") or "").lower() == "error":
                    errors.append(
                        f"{report_path}: row {row.get('source_row')} "
                        f"{row.get('sheet_name')}: {row.get('message')}"
                    )
    return errors


def check_worksheet_rule_coverage(
    path: Path,
    sheet: object,
    last_kpi_row: int,
) -> list[str]:
    errors: list[str] = []
    if last_kpi_row <= 1:
        return errors

    conditional_last_rows = [
        max(cell_range.max_row for cell_range in conditional_format.sqref.ranges)
        for conditional_format in sheet.conditional_formatting
        if any(cell_range.max_row > 1 for cell_range in conditional_format.sqref.ranges)
    ]
    if not conditional_last_rows:
        errors.append(f"{path}: missing KPI-row conditional formatting")
    elif max(conditional_last_rows) < last_kpi_row:
        errors.append(
            f"{path}: conditional formatting ends before final KPI row {last_kpi_row}"
        )

    validation_last_rows = [
        max(cell_range.max_row for cell_range in validation.sqref.ranges)
        for validation in sheet.data_validations.dataValidation
        if any(cell_range.max_row > 1 for cell_range in validation.sqref.ranges)
    ]
    if not validation_last_rows:
        errors.append(f"{path}: missing KPI-row data validation")
    elif min(validation_last_rows) < last_kpi_row:
        errors.append(f"{path}: data validation ends before final KPI row {last_kpi_row}")

    return errors


def validate_workbook(
    path: Path,
    fixed_pmids: set[str],
    master_ids: set[str],
    nomenclature_ids: set[str],
    master_types_by_id: dict[str, set[str]],
    position_types_by_pnid: dict[str, set[str]],
    trusted_pmid_ids: set[str] | None = None,
    trusted_pnid_ids: set[str] | None = None,
    approved_assistant_identities: set[tuple[str, str]] | None = None,
) -> tuple[dict[str, object], list[str], set[str]]:
    errors: list[str] = []
    found_fixed: set[str] = set()
    trusted_pmid_ids = trusted_pmid_ids or set()
    trusted_pnid_ids = trusted_pnid_ids or set()
    approved_assistant_identities = approved_assistant_identities or set()

    with zipfile.ZipFile(path) as archive:
        bad_file = archive.testzip()
        if bad_file:
            errors.append(f"{path}: bad xlsx zip member {bad_file}")

    workbook = openpyxl.load_workbook(path, read_only=False, data_only=False)
    if "KPI Template" not in workbook.sheetnames:
        errors.append(f"{path}: missing KPI Template sheet")
        return {"path": str(path), "rows": 0, "status": "FAIL"}, errors, found_fixed

    sheet = workbook["KPI Template"]
    row_iter = sheet.iter_rows(
        min_row=1,
        max_col=len(EXPECTED_HEADERS),
        values_only=True,
    )
    headers = list(next(row_iter, ()))
    if headers != EXPECTED_HEADERS:
        errors.append(f"{path}: upload header schema changed")

    rows = 0
    structural_rows = 0
    non_structural_rows = 0
    blank_identity_rows = 0
    double_identity_rows = 0
    graph_rows: list[dict[str, object]] = []
    last_kpi_row = 1

    for row_idx, row in enumerate(row_iter, 2):
        title = row[10]
        if title in (None, ""):
            continue
        rows += 1
        last_kpi_row = row_idx
        if is_numeric_only_title(title):
            errors.append(f"{path}: row {row_idx} has numeric-only KPI title={title}")
        pmid = str(row[4] or "").strip()
        pmvid = str(row[5] or "").strip()
        pnid = str(row[22] or "").strip()
        if pmid and pnid:
            double_identity_rows += 1
            errors.append(f"{path}: row {row_idx} has both PMID={pmid} and PNID={pnid}")
        elif not pmid and not pnid:
            blank_identity_rows += 1
            errors.append(f"{path}: row {row_idx} has neither PMID nor PNID")
        elif pmid:
            structural_rows += 1
            trusted_assistant = (pmid, pmvid) in approved_assistant_identities
            trusted = pmid in trusted_pmid_ids or trusted_assistant
            if pmid == "77" and not trusted_assistant:
                errors.append(
                    f"{path}: row {row_idx} assistant PMID=77 has unapproved or missing PMVID={pmvid or '<blank>'}"
                )
            if pmid not in master_ids and not trusted:
                errors.append(f"{path}: row {row_idx} has invalid PMID={pmid}")
            production_types = master_types_by_id.get(pmid, set())
            if production_types != {"5"} and not trusted:
                errors.append(
                    f"{path}: row {row_idx} PMID={pmid} has non-structural "
                    f"production types {sorted(production_types)}"
                )
            if pmid in fixed_pmids:
                found_fixed.add(pmid)
        else:
            non_structural_rows += 1
            trusted = pnid in trusted_pnid_ids
            if pnid not in nomenclature_ids and not trusted:
                errors.append(f"{path}: row {row_idx} has invalid PNID={pnid}")
            production_types = position_types_by_pnid.get(pnid, set())
            if (not production_types or "5" in production_types) and not trusted:
                errors.append(
                    f"{path}: row {row_idx} PNID={pnid} has invalid production "
                    f"types {sorted(production_types)}"
                )
            if pnid in fixed_pmids:
                errors.append(f"{path}: row {row_idx} fixed structural PMID appears in PNID={pnid}")

        id_kpi_raw = row[0]
        parent_id_raw = row[8]
        try:
            id_kpi = int(id_kpi_raw) if id_kpi_raw not in (None, "") else None
        except (TypeError, ValueError):
            id_kpi = None
        try:
            parent_id = int(parent_id_raw) if parent_id_raw not in (None, "") else None
        except (TypeError, ValueError):
            parent_id = None
        identity_label = f"PMID={pmid}" if pmid else f"PNID={pnid}"
        if id_kpi is not None:
            graph_rows.append(
                {
                    "row": row_idx,
                    "identity": identity_label,
                    "id_kpi": id_kpi,
                    "parent_id": parent_id,
                    "kpi_type": str(row[7] or "").strip().upper(),
                }
            )

    errors.extend(check_worksheet_rule_coverage(path, sheet, last_kpi_row))

    global_id_rows: dict[int, dict[str, object]] = {}
    for graph_row in graph_rows:
        id_kpi = int(graph_row["id_kpi"])
        if id_kpi in global_id_rows:
            errors.append(
                f"{path}: row {graph_row['row']} duplicate IDKPI={id_kpi} "
                "dalam satu formulir"
            )
        else:
            global_id_rows[id_kpi] = graph_row

    graph_by_key: dict[tuple[str, int], dict[str, object]] = {}
    for graph_row in graph_rows:
        key = (str(graph_row["identity"]), int(graph_row["id_kpi"]))
        if key in graph_by_key:
            errors.append(
                f"{path}: row {graph_row['row']} duplicate IDKPI={graph_row['id_kpi']} "
                f"pada identity {graph_row['identity']}"
            )
        else:
            graph_by_key[key] = graph_row

    for graph_row in graph_rows:
        parent_id = graph_row["parent_id"]
        if parent_id is None:
            continue
        parent = graph_by_key.get((str(graph_row["identity"]), int(parent_id)))
        if parent is None:
            errors.append(
                f"{path}: row {graph_row['row']} Parent KPI ID={parent_id} "
                f"tidak ditemukan pada identity {graph_row['identity']}"
            )
            continue
        child_type = str(graph_row["kpi_type"])
        parent_type = str(parent["kpi_type"])
        if child_type == "OUTPUT" and parent_type != "IMPACT":
            errors.append(
                f"{path}: row {graph_row['row']} parent OUTPUT harus IMPACT "
                f"pada identity {graph_row['identity']}"
            )
        if child_type == "KAI" and parent_type != "OUTPUT":
            errors.append(
                f"{path}: row {graph_row['row']} parent KAI harus OUTPUT "
                f"pada identity {graph_row['identity']}"
            )

    return (
        {
            "path": str(path),
            "rows": rows,
            "structural_rows": structural_rows,
            "non_structural_rows": non_structural_rows,
            "blank_identity_rows": blank_identity_rows,
            "double_identity_rows": double_identity_rows,
            "status": "FAIL" if errors else "READY",
        },
        errors,
        found_fixed,
    )


def write_manifest(output_dir: Path, records: list[dict[str, object]], upload_ready_dir: Path | None) -> None:
    csv_path = output_dir / "upload_manifest.csv"
    fields = [
        "status",
        "rows",
        "structural_rows",
        "non_structural_rows",
        "blank_identity_rows",
        "double_identity_rows",
        "path",
        "upload_ready_path",
    ]
    with csv_path.open("w", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields)
        writer.writeheader()
        for record in records:
            writer.writerow({field: record.get(field, "") for field in fields})

    md_path = output_dir / "UPLOAD_THESE_FILES.md"
    lines = [
        "# Upload These Files",
        "",
        "Upload the workbooks listed below. Do not upload the conversion report, config, audit, or CSV files.",
        "",
    ]
    for index, record in enumerate(records, 1):
        upload_path = record.get("upload_ready_path") or record["path"]
        lines.append(f"{index}. `{upload_path}`")
    lines.append("")
    md_path.write_text("\n".join(lines))


def prepare_upload_ready(records: list[dict[str, object]], upload_ready_dir: Path) -> None:
    if upload_ready_dir.exists():
        shutil.rmtree(upload_ready_dir)
    upload_ready_dir.mkdir(parents=True, exist_ok=True)
    for record in records:
        source = Path(str(record["path"]))
        target = upload_ready_dir / source.name
        shutil.copy2(source, target)
        record["upload_ready_path"] = str(target)


def write_zip(records: list[dict[str, object]], zip_output: Path) -> None:
    if zip_output.exists():
        zip_output.unlink()
    with zipfile.ZipFile(zip_output, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        for record in records:
            source = Path(str(record.get("upload_ready_path") or record["path"]))
            archive.write(source, arcname=source.name)
    with zipfile.ZipFile(zip_output) as archive:
        bad_file = archive.testzip()
    if bad_file:
        raise RuntimeError(f"bad upload zip member: {bad_file}")


def main() -> int:
    args = parse_args()
    (
        master_ids,
        nomenclature_ids,
        cluster_labels_by_id,
        master_types_by_id,
        position_types_by_pnid,
    ) = load_reference_ids(args.reference)
    fixed_pmids = load_fixed_pmids(args.fixed_audit)
    trusted_pmid_ids, trusted_pnid_ids = load_trusted_reviewer_ids(args.config)
    approved_assistant_identities = load_approved_assistant_identities(args.config)

    errors = check_config_scope(
        args.config,
        master_ids,
        nomenclature_ids,
        cluster_labels_by_id,
        master_types_by_id,
        position_types_by_pnid,
    )
    errors.extend(check_report_csvs(args.output_dir))

    workbook_paths = iter_upload_workbooks(args.output_dir)
    if len(workbook_paths) != args.expected_workbooks:
        errors.append(f"expected {args.expected_workbooks} upload workbooks, found {len(workbook_paths)}")

    records: list[dict[str, object]] = []
    found_fixed_all: set[str] = set()
    for path in workbook_paths:
        record, workbook_errors, found_fixed = validate_workbook(
            path,
            fixed_pmids,
            master_ids,
            nomenclature_ids,
            master_types_by_id,
            position_types_by_pnid,
            trusted_pmid_ids,
            trusted_pnid_ids,
            approved_assistant_identities,
        )
        records.append(record)
        errors.extend(workbook_errors)
        found_fixed_all.update(found_fixed)

    missing_fixed = fixed_pmids - found_fixed_all
    for pmid in sorted(missing_fixed, key=int):
        errors.append(f"fixed structural PMID not found in generated upload rows: {pmid}")

    ready_records = [record for record in records if record["status"] == "READY"]
    if args.upload_ready_dir:
        prepare_upload_ready(ready_records, args.upload_ready_dir)
    if args.zip_output:
        write_zip(ready_records, args.zip_output)

    write_manifest(args.output_dir, records, args.upload_ready_dir)

    total_rows = sum(int(record["rows"]) for record in records)
    structural_rows = sum(int(record["structural_rows"]) for record in records)
    non_structural_rows = sum(int(record["non_structural_rows"]) for record in records)

    print(f"upload_workbooks={len(workbook_paths)}")
    print(f"ready_workbooks={len(ready_records)}")
    print(f"total_rows={total_rows}")
    print(f"structural_rows={structural_rows}")
    print(f"non_structural_rows={non_structural_rows}")
    print(f"fixed_structural_pmids_checked={len(fixed_pmids)}")
    print(f"errors={len(errors)}")

    if errors:
        for error in errors[:100]:
            print(f"ERROR: {error}", file=sys.stderr)
        return 1
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
