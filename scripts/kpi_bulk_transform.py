#!/usr/bin/env python3
"""Transform KPI design workbooks into the official bulk upload template."""

from __future__ import annotations

import argparse
import csv
import json
import re
import sys
import tempfile
import zipfile
import xml.etree.ElementTree as ET
from copy import copy
from collections import Counter, OrderedDict
from dataclasses import dataclass, field
from datetime import datetime, timezone
from enum import Enum
from pathlib import Path
from typing import Any

import position_mapping as strict_position_mapping

try:
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.cell_range import CellRange
except ImportError as exc:  # pragma: no cover - runtime guidance
    raise SystemExit(
        "Missing dependency: openpyxl. Install it with `python3 -m pip install openpyxl`."
    ) from exc


XLSX_NS = {
    "a": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
    "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
}

UPLOAD_HEADERS = [
    "IDKPI",
    "Group",
    "Direktorat",
    "Posisi",
    "Position Master ID (Required)",
    "Position Master Variant ID (Optional)",
    "BSC Perspective",
    "KPI Type",
    "Parent KPI ID",
    "Parent KPI Title",
    "Title",
    "Description",
    "Unit",
    "Polarity",
    "Period",
    "Formula",
    "Weight (%)",
    "Cascading",
    "Nature Of Work (KAI Only)",
    "External ID (PKPI)",
    "System KPI ID",
    "Ownership Type",
    "Position Nomenklatur ID",
    "RKM Code ID",
]

TARGET_COMPANY_ID_DEFAULT = "1"
TARGET_COMPANY_NAME_DEFAULT = "PT Pelabuhan Indonesia (Persero)"
ALLOWED_UPLOAD_POLARITIES = {"POSITIVE", "NEGATIVE", "NEUTRAL"}
ALLOWED_UPLOAD_PERIODS = {"BULANAN", "TRIWULANAN", "TAHUNAN", "SEMESTER", "MONTHLY", "QUARTERLY", "WEEKLY"}
ALLOWED_UPLOAD_CASCADING = {"DIRECT", "INDIRECT", "DUPLICATE"}
ALLOWED_UPLOAD_OWNERSHIP = {"SPECIFIC", "SHARED", "COMMON"}
COLUMN_WIDTHS = {
    "IDKPI": 10,
    "Group": 34,
    "Direktorat": 32,
    "Posisi": 34,
    "Position Master ID (Required)": 18,
    "Position Master Variant ID (Optional)": 18,
    "BSC Perspective": 22,
    "KPI Type": 14,
    "Parent KPI ID": 14,
    "Parent KPI Title": 46,
    "Title": 52,
    "Description": 64,
    "Unit": 12,
    "Polarity": 14,
    "Period": 16,
    "Formula": 72,
    "Weight (%)": 14,
    "Cascading": 16,
    "Nature Of Work (KAI Only)": 24,
    "External ID (PKPI)": 18,
    "System KPI ID": 18,
    "Ownership Type": 18,
    "Position Nomenklatur ID": 20,
    "RKM Code ID": 16,
}


def norm_text(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, str):
        value = value.replace("\r\n", "\n").replace("\r", "\n").strip()
        return value or None
    return str(value)


def normalize_title(value: str | None) -> str:
    value = (norm_text(value) or "").lower()
    value = re.sub(r"[^a-z0-9]+", " ", value)
    return re.sub(r"\s+", " ", value).strip()


def is_placeholder(value: Any) -> bool:
    text = norm_text(value)
    if text is None:
        return True
    return text.strip().lower() in {"(blank)", "blank"}


def to_upper_enum(value: str | None, mapping: dict[str, str]) -> str | None:
    value = norm_text(value)
    if not value or is_placeholder(value):
        return None
    return mapping.get(value, value.upper())


class NormalizationStatus(str, Enum):
    OK = "ok"
    NORMALIZED = "normalized"
    DEFAULTED = "defaulted"
    CROSS_COLUMN = "cross_column"
    AMBIGUOUS = "ambiguous"
    INVALID = "invalid"


@dataclass(frozen=True)
class NormalizedEnum:
    value: str | None
    status: NormalizationStatus
    raw_value: str | None
    message: str


def normalized_key(value: str | None) -> str:
    return normalize_title(value)


def enum_result(
    value: str | None,
    status: NormalizationStatus,
    raw: str | None,
    message: str,
) -> NormalizedEnum:
    return NormalizedEnum(value=value, status=status, raw_value=raw, message=message)


def _coerce_period_value(value: str | None) -> str | None:
    raw = norm_text(value)
    if not raw or is_placeholder(raw):
        return None
    key = normalized_key(raw)
    period_mapping = {
        "triwulan": "TRIWULANAN",
        "triwulanan": "TRIWULANAN",
        "tahunan": "TAHUNAN",
        "tahun": "TAHUNAN",
        "per tahun": "TAHUNAN",
        "per tahunan": "TAHUNAN",
        "semester": "SEMESTER",
        "semesteran": "SEMESTER",
        "per semester": "SEMESTER",
        "bulanan": "BULANAN",
        "monthly": "MONTHLY",
        "quarterly": "QUARTERLY",
        "weekly": "WEEKLY",
    }
    if key in period_mapping:
        return period_mapping[key]
    if raw.upper() in ALLOWED_UPLOAD_PERIODS:
        return raw.upper()
    return None


KNOWN_CROSS_COLUMN_KEYS = {
    normalized_key(raw_value)
    for raw_value in [
        "Positif",
        "positive",
        "pos",
        "Negatif",
        "negative",
        "neg",
        "Netral",
        "neutral",
        "Direct",
        "Indirect",
        "DUPLICATE",
        "Specific",
        "SPESIFIC",
        "Shared",
        "Common",
        "Routine",
        "rutin",
        "Non Routine",
        "Non-Routine",
        "Non routine",
        "non Routine",
        "Non-Rotine",
        "non rotine",
        "Non Rutin",
        "Triwulan",
        "Triwulanan",
        "Tahunan",
        "tahun",
        "per tahun",
        "per tahunan",
        "Semester",
        "Semesteran",
        "per semester",
        "Bulanan",
        "Monthly",
        "Quarterly",
        "Weekly",
    ]
}


def _is_cross_column_key(key: str) -> bool:
    return key in KNOWN_CROSS_COLUMN_KEYS


def _period_tokens(raw: str) -> list[str]:
    return [part.strip() for part in re.split(r"\s*(?:/|\\|;|,|&|\batau\b|\bdan\b)\s*", raw, flags=re.IGNORECASE) if part.strip()]


def normalize_period(value: str | None, fallback: str | None = None) -> NormalizedEnum:
    raw = norm_text(value)
    fallback_value = _coerce_period_value(fallback)
    if not raw or is_placeholder(raw):
        if fallback_value:
            return enum_result(
                fallback_value,
                NormalizationStatus.DEFAULTED,
                raw,
                f"Period missing; defaulted to fallback period {fallback_value}.",
            )
        return enum_result(None, NormalizationStatus.INVALID, raw, "Period missing.")

    raw_key = normalized_key(raw)
    if any(separator in raw for separator in ["/", "\\", ";", ",", "&"]) or re.search(r"\b(?:atau|dan)\b", raw, flags=re.IGNORECASE):
        parts = _period_tokens(raw)
        normalized_parts = [period for period in (_coerce_period_value(part) for part in parts) if period]
        if len(parts) > 1 and len(normalized_parts) == len(parts) and len(set(normalized_parts)) == 1:
            canonical = normalized_parts[0]
            status = NormalizationStatus.OK if raw == canonical else NormalizationStatus.NORMALIZED
            return enum_result(canonical, status, raw, f"Period normalized to {canonical}.")
        if fallback_value:
            return enum_result(
                fallback_value,
                NormalizationStatus.AMBIGUOUS,
                raw,
                f"Ambiguous period defaulted to fallback period {fallback_value}.",
            )
        return enum_result(None, NormalizationStatus.AMBIGUOUS, raw, "Ambiguous period requires review.")

    canonical = _coerce_period_value(raw)
    if canonical:
        status = NormalizationStatus.OK if raw == canonical else NormalizationStatus.NORMALIZED
        return enum_result(canonical, status, raw, f"Period normalized to {canonical}.")

    if _is_cross_column_key(raw_key):
        if fallback_value:
            return enum_result(
                fallback_value,
                NormalizationStatus.CROSS_COLUMN,
                raw,
                f"Period contains cross-column value; defaulted to fallback period {fallback_value}.",
            )
        return enum_result(None, NormalizationStatus.CROSS_COLUMN, raw, "Period contains cross-column value.")
    return enum_result(None, NormalizationStatus.INVALID, raw, "Invalid period.")


def normalize_polarity(value: str | None) -> NormalizedEnum:
    raw = norm_text(value)
    if not raw or is_placeholder(raw):
        return enum_result("POSITIVE", NormalizationStatus.DEFAULTED, raw, "Polarity defaulted to POSITIVE.")

    raw_key = normalized_key(raw)
    polarity_mapping = {
        "positif": "POSITIVE",
        "positive": "POSITIVE",
        "pos": "POSITIVE",
        "negatif": "NEGATIVE",
        "negative": "NEGATIVE",
        "neg": "NEGATIVE",
        "netral": "NEUTRAL",
        "neutral": "NEUTRAL",
    }
    if raw_key in polarity_mapping:
        canonical = polarity_mapping[raw_key]
        status = NormalizationStatus.OK if raw == canonical else NormalizationStatus.NORMALIZED
        return enum_result(canonical, status, raw, f"Polarity normalized to {canonical}.")

    if _is_cross_column_key(raw_key):
        return enum_result("POSITIVE", NormalizationStatus.CROSS_COLUMN, raw, "Polarity contains cross-column value; defaulted to POSITIVE.")
    return enum_result("POSITIVE", NormalizationStatus.INVALID, raw, "Invalid polarity defaulted to POSITIVE.")


BSC_PERSPECTIVE_CANONICAL: dict[str, str] = {
    "financial": "Financial",
    "customer": "Customer",
    "internal business process": "Internal Business Process",
    "internal process": "Internal Business Process",
    "learning growth": "Learning & Growth",
    "learning and growth": "Learning & Growth",
    "learning dan growth": "Learning & Growth",
    "learning & growth": "Learning & Growth",
}


def _bsc_perspective_lookup_keys(raw: str) -> list[str]:
    variants = [
        raw,
        raw.replace("&", "and"),
        re.sub(r"\bdan\b", "and", raw, flags=re.IGNORECASE),
    ]
    keys: list[str] = []
    for variant in variants:
        key = normalized_key(variant)
        if key and key not in keys:
            keys.append(key)
    return keys


def normalize_bsc_perspective(value: str | None) -> NormalizedEnum:
    raw = norm_text(value)
    if not raw or is_placeholder(raw):
        return enum_result(None, NormalizationStatus.INVALID, raw, "BSC Perspective missing.")

    for key in _bsc_perspective_lookup_keys(raw):
        canonical = BSC_PERSPECTIVE_CANONICAL.get(key)
        if canonical:
            status = NormalizationStatus.OK if raw == canonical else NormalizationStatus.NORMALIZED
            return enum_result(canonical, status, raw, f"BSC Perspective normalized to {canonical}.")

    upper = raw.strip().upper()
    enum_mapping = {
        "FINANCIAL": "Financial",
        "CUSTOMER": "Customer",
        "INTERNAL BUSINESS PROCESS": "Internal Business Process",
        "INTERNAL_PROCESS": "Internal Business Process",
        "LEARNING & GROWTH": "Learning & Growth",
        "LEARNING_GROWTH": "Learning & Growth",
    }
    if upper in enum_mapping:
        canonical = enum_mapping[upper]
        status = NormalizationStatus.OK if raw == canonical else NormalizationStatus.NORMALIZED
        return enum_result(canonical, status, raw, f"BSC Perspective normalized to {canonical}.")

    return enum_result(None, NormalizationStatus.INVALID, raw, "BSC Perspective is not a recognized PMS alias.")


def normalize_cascading(value: str | None) -> NormalizedEnum:
    raw = norm_text(value)
    if not raw or is_placeholder(raw):
        return enum_result(None, NormalizationStatus.OK, raw, "Cascading blank.")

    raw_key = normalized_key(raw)
    cascading_mapping = {
        "direct": "DIRECT",
        "indirect": "INDIRECT",
        "duplicate": "DUPLICATE",
    }
    if raw_key in cascading_mapping:
        canonical = cascading_mapping[raw_key]
        status = NormalizationStatus.OK if raw == canonical else NormalizationStatus.NORMALIZED
        return enum_result(canonical, status, raw, f"Cascading normalized to {canonical}.")

    if _is_cross_column_key(raw_key):
        return enum_result("INDIRECT", NormalizationStatus.CROSS_COLUMN, raw, "Cascading contains cross-column value; defaulted to INDIRECT.")
    return enum_result("INDIRECT", NormalizationStatus.INVALID, raw, "Invalid cascading defaulted to INDIRECT.")


def normalize_ownership_type(value: str | None) -> NormalizedEnum:
    raw = norm_text(value)
    if not raw or is_placeholder(raw):
        return enum_result("SPECIFIC", NormalizationStatus.DEFAULTED, raw, "Ownership Type defaulted to SPECIFIC.")

    raw_key = normalized_key(raw)
    ownership_mapping = {
        "specific": "SPECIFIC",
        "spesific": "SPECIFIC",
        "shared": "SHARED",
        "common": "COMMON",
    }
    if raw_key in ownership_mapping:
        canonical = ownership_mapping[raw_key]
        status = NormalizationStatus.OK if raw == canonical else NormalizationStatus.NORMALIZED
        return enum_result(canonical, status, raw, f"Ownership Type normalized to {canonical}.")

    if _is_cross_column_key(raw_key):
        return enum_result("SPECIFIC", NormalizationStatus.CROSS_COLUMN, raw, "Ownership Type contains cross-column value; defaulted to SPECIFIC.")
    return enum_result("SPECIFIC", NormalizationStatus.INVALID, raw, "Invalid Ownership Type defaulted to SPECIFIC.")


def normalize_kai_nature(value: str | None, period: str | None = None) -> NormalizedEnum:
    raw = norm_text(value)
    period_value = normalize_period(period).value
    inferred = "Non Routine" if period_value == "TAHUNAN" else "Routine"
    if not raw or is_placeholder(raw):
        return enum_result(inferred, NormalizationStatus.DEFAULTED, raw, f"KAI Nature inferred as {inferred}.")

    raw_key = normalized_key(raw)
    routine_keys = {"routine", "rutin"}
    non_routine_keys = {"non routine", "non rutin", "non rotine", "non rotin", "non rutine", "non routinee"}
    if raw_key in routine_keys:
        canonical = "Routine"
        status = NormalizationStatus.OK if raw == canonical else NormalizationStatus.NORMALIZED
        return enum_result(canonical, status, raw, "KAI Nature normalized to Routine.")
    if raw_key in non_routine_keys:
        canonical = "Non Routine"
        status = NormalizationStatus.OK if raw == canonical else NormalizationStatus.NORMALIZED
        return enum_result(canonical, status, raw, "KAI Nature normalized to Non Routine.")

    if raw_key in {"pdf", "diunggah"} or re.search(r"(?:https?://|www\.)", raw, flags=re.IGNORECASE):
        return enum_result(inferred, NormalizationStatus.CROSS_COLUMN, raw, f"KAI Nature contains cross-column value; inferred as {inferred}.")

    if _is_cross_column_key(raw_key):
        return enum_result(inferred, NormalizationStatus.CROSS_COLUMN, raw, f"KAI Nature contains cross-column value; inferred as {inferred}.")
    return enum_result(inferred, NormalizationStatus.INVALID, raw, f"Invalid KAI Nature inferred as {inferred}.")


def uploader_period(value: str | None) -> str | None:
    return normalize_period(value).value


def uploader_polarity(value: str | None) -> str | None:
    return normalize_polarity(value).value


def uploader_kai_nature(value: str | None, period: str | None = None) -> str:
    return normalize_kai_nature(value, period).value or "Routine"


def col_to_num(col_ref: str) -> int:
    value = 0
    for ch in col_ref:
        if ch.isalpha():
            value = (value * 26) + (ord(ch.upper()) - 64)
    return value


def read_xlsx_sheet(path: Path, sheet_name: str) -> list[list[Any]]:
    with zipfile.ZipFile(path) as workbook_zip:
        workbook = ET.fromstring(workbook_zip.read("xl/workbook.xml"))
        rels = ET.fromstring(workbook_zip.read("xl/_rels/workbook.xml.rels"))
        relmap = {rel.attrib["Id"]: rel.attrib["Target"] for rel in rels}

        shared_strings: list[str] = []
        if "xl/sharedStrings.xml" in workbook_zip.namelist():
            root = ET.fromstring(workbook_zip.read("xl/sharedStrings.xml"))
            for si in root:
                shared_strings.append(
                    "".join(
                        t.text or ""
                        for t in si.iter("{http://schemas.openxmlformats.org/spreadsheetml/2006/main}t")
                    )
                )

        target = None
        for sheet in workbook.find("a:sheets", XLSX_NS):
            if sheet.attrib["name"] == sheet_name:
                rel_target = relmap[
                    sheet.attrib["{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id"]
                ]
                # Office relationship Targets may be package-absolute ("/xl/worksheets/...")
                # or relative to xl/ ("worksheets/..."). Normalize before joining.
                rel_target = rel_target.lstrip("/")
                target = rel_target if rel_target.startswith("xl/") else f"xl/{rel_target}"
                break
        if not target:
            raise KeyError(f"Sheet '{sheet_name}' not found in {path}")

        root = ET.fromstring(workbook_zip.read(target))
        rows = root.find("a:sheetData", XLSX_NS)
        parsed_rows: list[list[Any]] = []
        for row in rows:
            values: dict[int, Any] = {}
            for cell in row:
                match = re.match(r"([A-Z]+)(\d+)", cell.attrib.get("r", ""))
                if not match:
                    continue
                col_num = col_to_num(match.group(1))
                cell_type = cell.attrib.get("t")
                raw_value = cell.find("a:v", XLSX_NS)
                inline = cell.find("a:is", XLSX_NS)
                value = None
                if cell_type == "s" and raw_value is not None:
                    value = shared_strings[int(raw_value.text)]
                elif cell_type == "inlineStr" and inline is not None:
                    value = "".join(
                        t.text or ""
                        for t in inline.iter("{http://schemas.openxmlformats.org/spreadsheetml/2006/main}t")
                    )
                elif raw_value is not None:
                    value = raw_value.text
                values[col_num] = norm_text(value)
            max_col = max(values) if values else 0
            parsed_rows.append([values.get(i) for i in range(1, max_col + 1)])
        return parsed_rows


@dataclass
class PositionMetadata:
    position_master_id: str
    position_name: str
    organization_name: str | None
    company_name: str | None
    position_type: str | None


@dataclass
class ValidationIssue:
    severity: str
    sheet_name: str
    source_row: int | None
    record_type: str
    title: str | None
    message: str


def append_enum_issue(
    issues: list[ValidationIssue] | None,
    config: PositionConfig,
    source_row: int | None,
    record_type: str,
    title: str | None,
    field_name: str,
    result: NormalizedEnum,
) -> None:
    if issues is None:
        return
    if result.status == NormalizationStatus.OK:
        return

    if result.status == NormalizationStatus.NORMALIZED:
        severity = "info"
    elif result.value is None and result.status in {NormalizationStatus.AMBIGUOUS, NormalizationStatus.INVALID}:
        severity = "error"
    else:
        severity = "warning"

    issues.append(
        ValidationIssue(
            severity=severity,
            sheet_name=config.sheet_name,
            source_row=source_row,
            record_type=record_type,
            title=title,
            message=(
                f"enum_issue category={result.status.value}; field={field_name}; raw={result.raw_value}; "
                f"normalized={result.value}; {result.message}"
            ),
        )
    )


@dataclass
class PositionConfig:
    sheet_name: str
    position_name: str
    group_name: str
    directorate_name: str
    source_workbook: str | None = None
    position_master_id: str | None = None
    position_master_variant_id: str | None = None
    position_nomenclature_id: str | None = None
    position_scope: str | None = None
    portaverse_position_title: str | None = None
    portaverse_group_name: str | None = None
    portaverse_company_name: str | None = None
    portaverse_company_code: str | None = None
    cluster_label: str | None = None
    mapping_confidence_label: str | None = None
    mapping_confidence_reason: str | None = None
    mapping_review_status: str | None = None
    mapping_override_approved: bool = False
    mapping_override_trust_source: str | None = None
    candidate_position_master_id: str | None = None
    candidate_position_nomenclature_id: str | None = None
    candidate_score: float | None = None
    runner_up_position_master_id: str | None = None
    runner_up_position_nomenclature_id: str | None = None
    runner_up_score: float | None = None
    active_variant_count: int | None = None
    active_employee_count: int | None = None
    active_employee_name: str | None = None
    active_employee_nipp: str | None = None
    rkm_code_id: str | None = None
    position_lookup_names: list[str] = field(default_factory=list)
    drop_comment_values: list[str] = field(default_factory=lambda: ["Drop"])
    expected_impact_count: int = 10


@dataclass
class ImpactRecord:
    bsc: str | None
    title: str
    unit: str | None
    period: str | None
    formula: str | None
    polarity: str | None
    weight: str | None
    source_row: int | None = None
    outputs: list[dict[str, Any]] = field(default_factory=list)


@dataclass
class ParsedSheet:
    config: PositionConfig
    metadata: PositionMetadata | None
    impacts: list[ImpactRecord]


def backfill_shared_impact_fields(parsed_sheets: list[ParsedSheet]) -> None:
    canonical: dict[str, dict[str, str | None]] = {}
    for parsed in parsed_sheets:
        for impact in parsed.impacts:
            key = normalize_title(impact.title)
            fields = canonical.setdefault(
                key,
                {
                    "bsc": None,
                    "unit": None,
                    "period": None,
                    "formula": None,
                    "polarity": None,
                    "weight": None,
                },
            )
            for field_name in fields:
                current_value = getattr(impact, field_name)
                if not fields[field_name] and norm_text(current_value) and not is_placeholder(current_value):
                    fields[field_name] = current_value

    for parsed in parsed_sheets:
        for impact in parsed.impacts:
            fields = canonical.get(normalize_title(impact.title), {})
            for field_name, fallback in fields.items():
                current_value = getattr(impact, field_name)
                if (not norm_text(current_value) or is_placeholder(current_value)) and fallback:
                    setattr(impact, field_name, fallback)


class PositionMasterIndex:
    def __init__(self) -> None:
        self.by_id: dict[str, PositionMetadata] = {}
        self.by_title: dict[str, PositionMetadata] = {}

    @classmethod
    def load(cls, root_dir: Path) -> "PositionMasterIndex":
        index = cls()
        files = sorted(root_dir.glob("*.xlsx"))
        for workbook_path in files:
            if workbook_path.name.startswith("~$"):
                continue
            try:
                rows = read_xlsx_sheet(workbook_path, "Master Posisi")
            except zipfile.BadZipFile:
                continue
            if not rows:
                continue
            for row in rows[1:]:
                row = row + [None] * max(0, 8 - len(row))
                position_id = norm_text(row[0])
                position_name = norm_text(row[1])
                if not position_id or not position_name:
                    continue
                metadata = PositionMetadata(
                    position_master_id=position_id,
                    position_name=position_name,
                    organization_name=norm_text(row[2]) if len(row) > 2 else None,
                    company_name=norm_text(row[3]) if len(row) > 3 else None,
                    position_type=norm_text(row[7]) if len(row) > 7 else None,
                )
                index.by_id[position_id] = metadata
                index.by_title[normalize_title(position_name)] = metadata
        return index

    def resolve(self, config: PositionConfig) -> PositionMetadata | None:
        if config.position_master_id:
            metadata = self.by_id.get(str(config.position_master_id))
            if metadata:
                return metadata
        lookup_names = [config.position_name, *config.position_lookup_names]
        for name in lookup_names:
            metadata = self.by_title.get(normalize_title(name))
            if metadata:
                return metadata
        return None


def find_header_row(rows: list[list[Any]]) -> tuple[int, dict[str, int]]:
    for index, row in enumerate(rows, start=1):
        headers = {norm_text(value): idx for idx, value in enumerate(row, start=1) if norm_text(value)}
        if "KPI Impact" in headers and "KPI Output" in headers and "Key Activity Indicator (KAI)" in headers:
            return index, headers
    raise ValueError("Could not find KPI header row")


def row_value(row: list[Any], header_map: dict[str, int], header: str) -> str | None:
    col = header_map.get(header)
    if not col or col - 1 >= len(row):
        return None
    return norm_text(row[col - 1])


def first_row_value(row: list[Any], header_map: dict[str, int], headers: list[str]) -> str | None:
    for header in headers:
        value = row_value(row, header_map, header)
        if value:
            return value
    return None


def is_drop_status(value: Any, drop_values: list[str]) -> bool:
    text = norm_text(value)
    if not text:
        return False
    normalized_drop_values = {normalize_title(item) for item in drop_values}
    normalized_drop_values.update({"drop", "dropped", "hapus", "dihapus", "exclude", "excluded"})
    return normalize_title(text) in normalized_drop_values


def is_numeric_only_title(value: Any) -> bool:
    text = norm_text(value)
    if not text or is_placeholder(text):
        return False
    return re.fullmatch(r"\d+(?:\.0+)?", text) is not None


def parse_block_sheet(
    rows: list[list[Any]],
    config: PositionConfig,
    issues: list[ValidationIssue],
) -> list[ImpactRecord]:
    header_row, header_map = find_header_row(rows)
    drop_status_headers = [
        header
        for header in ["Komentar", "Comment", "Status", "Alignment"]
        if header in header_map
    ]

    impacts: list[ImpactRecord] = []
    current_impact: ImpactRecord | None = None
    current_bsc: str | None = None
    current_impact_defaults: dict[str, str | None] = {}
    current_output_defaults: dict[str, str | None] = {}

    for source_row in range(header_row + 1, len(rows) + 1):
        row = rows[source_row - 1]
        first_cell = norm_text(row[0]) if row else None
        impact_title = row_value(row, header_map, "KPI Impact")
        output_title = row_value(row, header_map, "KPI Output")
        kai_title = row_value(row, header_map, "Key Activity Indicator (KAI)")

        if first_cell == "TOTAL":
            break
        if not any(norm_text(value) for value in row):
            continue

        bsc = row_value(row, header_map, "BSC Perspective")
        if bsc:
            bsc_result = normalize_bsc_perspective(bsc)
            current_bsc = bsc_result.value or bsc

        if impact_title and not is_placeholder(impact_title):
            impact_unit = row_value(row, header_map, "KPI Impact Unit") or current_impact_defaults.get("unit")
            impact_period = row_value(row, header_map, "KPI Impact Frequency") or current_impact_defaults.get("period")
            impact_formula = row_value(row, header_map, "KPI Impact Formula") or current_impact_defaults.get("formula")
            impact_polarity = row_value(row, header_map, "KPI Impact Polarity") or current_impact_defaults.get("polarity")
            impact_weight = row_value(row, header_map, "%Weight (Impact)") or current_impact_defaults.get("weight")
            current_impact = ImpactRecord(
                bsc=current_bsc,
                title=impact_title,
                unit=impact_unit,
                period=impact_period,
                formula=impact_formula,
                polarity=impact_polarity,
                weight=impact_weight,
                source_row=source_row,
            )
            current_impact_defaults = {
                "unit": impact_unit,
                "period": impact_period,
                "formula": impact_formula,
                "polarity": impact_polarity,
                "weight": impact_weight,
            }
            impacts.append(current_impact)
        elif current_impact is None:
            issues.append(
                ValidationIssue(
                    severity="warning",
                    sheet_name=config.sheet_name,
                    source_row=source_row,
                    record_type="row",
                    title=None,
                    message="Row appears before any KPI Impact block and was skipped.",
                )
            )
            continue

        drop_status = first_row_value(row, header_map, drop_status_headers)
        drop_comment = is_drop_status(drop_status, config.drop_comment_values)

        output_weight = row_value(row, header_map, "%Weight (Output)")
        kai_weight = row_value(row, header_map, "%Weight (Activity)")
        numeric_output_title = is_numeric_only_title(output_title)
        numeric_kai_title = is_numeric_only_title(kai_title)

        output_period = (
            row_value(row, header_map, "KPI Output Frequency")
            or current_output_defaults.get("period")
            or (current_impact.period if current_impact else None)
        )
        output_polarity = (
            row_value(row, header_map, "KPI Output Polarity")
            or current_output_defaults.get("polarity")
            or (current_impact.polarity if current_impact else None)
        )
        output_unit = row_value(row, header_map, "KPI Output Unit") or current_output_defaults.get("unit")
        output_formula = row_value(row, header_map, "KPI Output Formula") or current_output_defaults.get("formula")
        output_definition = row_value(row, header_map, "KPI Output Definition") or current_output_defaults.get("description")
        cascading_output = row_value(row, header_map, "Cascading Tagging (KPI Output)") or current_output_defaults.get("cascading")
        coverage_output = row_value(row, header_map, "Coverage KPI Output") or current_output_defaults.get("ownership_type")
        nature_of_work = row_value(row, header_map, "Nature of Work (KAI)") or current_output_defaults.get("nature_of_work")

        keep_output = (
            not is_placeholder(output_title)
            and not drop_comment
            and not numeric_output_title
            and not is_placeholder(output_weight)
            and not is_zero_weight(output_weight)
        )
        keep_kai = (
            keep_output
            and not is_placeholder(kai_title)
            and not drop_comment
            and not numeric_kai_title
            and not is_placeholder(kai_weight)
            and not is_zero_weight(kai_weight)
        )

        if output_title and not keep_output:
            reasons = []
            if drop_comment:
                reasons.append("alignment/status is Drop")
            if numeric_output_title:
                reasons.append("output title is numeric-only title")
            if is_placeholder(output_weight):
                reasons.append("output weight is blank")
            if is_zero_weight(output_weight):
                reasons.append("output weight is 0")
            issues.append(
                ValidationIssue(
                    severity="info",
                    sheet_name=config.sheet_name,
                    source_row=source_row,
                    record_type="output",
                    title=output_title,
                    message=f"Dropped OUTPUT row because {' and '.join(reasons)}.",
                )
            )

        if keep_output:
            current_output_defaults = {
                "period": output_period,
                "polarity": output_polarity,
                "unit": output_unit,
                "formula": output_formula,
                "description": output_definition,
                "cascading": cascading_output,
                "ownership_type": coverage_output,
                "nature_of_work": nature_of_work,
            }
            output_record = {
                "source_row": source_row,
                "title": output_title,
                "description": output_definition,
                "unit": output_unit,
                "period": output_period,
                "formula": output_formula,
                "polarity": output_polarity,
                "weight": output_weight,
                "cascading": cascading_output,
                "ownership_type": coverage_output,
                "kai": None,
            }
            current_impact.outputs.append(output_record)

            if kai_title and not keep_kai:
                reasons = []
                if drop_comment:
                    reasons.append("alignment/status is Drop")
                if numeric_kai_title:
                    reasons.append("KAI title is numeric-only title")
                if is_placeholder(kai_weight):
                    reasons.append("KAI weight is blank")
                if is_zero_weight(kai_weight):
                    reasons.append("KAI weight is 0")
                issues.append(
                    ValidationIssue(
                        severity="info",
                        sheet_name=config.sheet_name,
                        source_row=source_row,
                        record_type="kai",
                        title=kai_title,
                        message=f"Dropped KAI row because {' and '.join(reasons)}.",
                    )
                )

            if keep_kai:
                output_record["kai"] = {
                    "source_row": source_row,
                    "title": kai_title,
                    "description": row_value(row, header_map, "KPI KAI Definition"),
                    "formula": row_value(row, header_map, "KPI KAI Formula"),
                    "weight": kai_weight,
                    "nature_of_work": nature_of_work,
                    "period": output_period,
                    "polarity": output_polarity,
                    "cascading": cascading_output,
                    "ownership_type": coverage_output,
                }
                if output_period:
                    issues.append(
                        ValidationIssue(
                            severity="info",
                            sheet_name=config.sheet_name,
                            source_row=source_row,
                            record_type="kai",
                            title=kai_title,
                            message="KAI period inferred from KPI Output Frequency because the source sheet does not provide a separate KAI period column.",
                        )
                    )
        elif kai_title and not is_placeholder(kai_title):
            issues.append(
                ValidationIssue(
                    severity="warning",
                    sheet_name=config.sheet_name,
                    source_row=source_row,
                    record_type="kai",
                    title=kai_title,
                    message="Skipped KAI because its OUTPUT row was dropped or missing.",
                )
            )

    return impacts


def parse_weight(value: Any) -> float | None:
    text = norm_text(value)
    if not text:
        return None
    text = text.replace("%", "").replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return None


def is_zero_weight(value: Any) -> bool:
    parsed = parse_weight(value)
    return parsed is not None and weights_equal(parsed, 0)


def format_weight(value: float) -> str:
    return str(int(value)) if value.is_integer() else str(value)


def round_weight(value: float) -> float:
    return round(value, 4)


def weights_equal(left: float, right: float, tolerance: float = 0.01) -> bool:
    return abs(left - right) <= tolerance


def is_truthy_flag(value: Any) -> bool:
    return value in (True, 1, "1", "true", "TRUE", "Y", "y")


def is_target_company_row(row: dict[str, Any], target_company_id: str | None) -> bool:
    return not target_company_id or str(row.get("company_id") or "") == str(target_company_id)


def is_active_reference_row(row: dict[str, Any], require_active_org: bool = False) -> bool:
    if row.get("is_company_active") not in (None, "") and not is_truthy_flag(row.get("is_company_active")):
        return False
    if row.get("is_group_active") not in (None, "") and not is_truthy_flag(row.get("is_group_active")):
        return False
    if row.get("is_position_active") not in (None, "") and not is_truthy_flag(row.get("is_position_active")):
        return False
    if require_active_org and row.get("is_position_organization_active") not in (None, ""):
        return is_truthy_flag(row.get("is_position_organization_active"))
    return True


def is_percentage_formula(value: str | None) -> bool:
    text = (norm_text(value) or "").lower()
    if not text:
        return False
    return any(token in text for token in ["%", "/", "target", "realisasi", "progress"])


def merge_duplicate_outputs(
    config: PositionConfig,
    impact: ImpactRecord,
    issues: list[ValidationIssue] | None = None,
) -> list[dict[str, Any]]:
    merged: list[dict[str, Any]] = []
    by_title: dict[str, dict[str, Any]] = {}
    for output in impact.outputs:
        key = normalize_title(output["title"])
        canonical = by_title.get(key)
        if canonical is None:
            by_title[key] = output
            merged.append(output)
            continue

        canonical_weight = parse_weight(canonical.get("weight")) or 0
        duplicate_weight = parse_weight(output.get("weight"))
        if duplicate_weight is None:
            if issues is not None:
                issues.append(
                    ValidationIssue(
                        severity="warning",
                        sheet_name=config.sheet_name,
                        source_row=output.get("source_row"),
                        record_type="OUTPUT",
                        title=output.get("title"),
                        message="Duplicate OUTPUT weight is not numeric and was not included in merged weight.",
                    )
                )
        else:
            canonical["weight"] = format_weight(canonical_weight + duplicate_weight)

        for field_name in [
            "description",
            "unit",
            "period",
            "formula",
            "polarity",
            "cascading",
            "ownership_type",
        ]:
            if not norm_text(canonical.get(field_name)) and norm_text(output.get(field_name)):
                canonical[field_name] = output[field_name]
            elif (
                issues is not None
                and norm_text(canonical.get(field_name))
                and norm_text(output.get(field_name))
                and norm_text(canonical.get(field_name)) != norm_text(output.get(field_name))
            ):
                issues.append(
                    ValidationIssue(
                        severity="warning",
                        sheet_name=config.sheet_name,
                        source_row=output.get("source_row"),
                        record_type="OUTPUT",
                        title=output.get("title"),
                        message=f"Duplicate OUTPUT field conflict for {field_name}; kept first non-empty value.",
                    )
                )

        duplicate_kai = output.get("kai")
        if duplicate_kai:
            canonical.setdefault("duplicate_kai", []).append(duplicate_kai)
        if issues is not None:
            issues.append(
                ValidationIssue(
                    severity="info",
                    sheet_name=config.sheet_name,
                    source_row=output.get("source_row"),
                    record_type="OUTPUT",
                    title=output.get("title"),
                    message="Merged duplicate OUTPUT row into canonical OUTPUT and reparented its KAI child.",
                )
            )
    return merged


def output_kai_records(output: dict[str, Any]) -> list[dict[str, Any]]:
    records: list[dict[str, Any]] = []
    if output.get("kai"):
        records.append(output["kai"])
    records.extend(output.get("duplicate_kai", []))
    return records


def kai_field_matches_output(kai: dict[str, Any], output: dict[str, Any], field_name: str) -> bool:
    kai_value = norm_text(kai.get(field_name))
    if kai_value is None:
        return True
    return kai_value == norm_text(output.get(field_name))


def build_upload_rows(
    config: PositionConfig,
    position_master_id: str,
    impacts: list[ImpactRecord],
    start_id: int,
    issues: list[ValidationIssue] | None = None,
) -> tuple[list[list[Any]], int]:
    position_name = config.position_name
    output_position_master_id, output_position_nomenclature_id = resolve_upload_scope(config, position_master_id)
    output_position_master_variant_id = config.position_master_variant_id

    rows: list[list[Any]] = []
    next_id = start_id
    impact_ids: dict[str, str] = {}

    for impact in impacts:
        impact_id = str(next_id)
        next_id += 1
        impact_ids[impact.title] = impact_id
        impact_polarity_result = normalize_polarity(impact.polarity)
        impact_period_result = normalize_period(impact.period)
        append_enum_issue(issues, config, impact.source_row, "IMPACT", impact.title, "Polarity", impact_polarity_result)
        append_enum_issue(issues, config, impact.source_row, "IMPACT", impact.title, "Period", impact_period_result)
        rows.append(
            [
                impact_id,
                config.group_name,
                config.directorate_name,
                position_name,
                output_position_master_id,
                output_position_master_variant_id,
                impact.bsc,
                "IMPACT",
                None,
                "#N/A",
                impact.title,
                None,
                impact.unit,
                impact_polarity_result.value,
                impact_period_result.value,
                impact.formula,
                impact.weight,
                None,
                None,
                None,
                None,
                None,
                output_position_nomenclature_id,
                config.rkm_code_id,
            ]
        )

    merged_outputs_by_impact: dict[str, list[dict[str, Any]]] = {}
    for impact in impacts:
        merged_outputs_by_impact[impact.title] = merge_duplicate_outputs(config, impact, issues)

    for impact in impacts:
        parent_period_result = normalize_period(impact.period)
        for output in merged_outputs_by_impact[impact.title]:
            output_id = str(next_id)
            next_id += 1
            output["_generated_id"] = output_id
            output_period_result = normalize_period(output.get("period"), parent_period_result.value)
            output["_period_result"] = output_period_result
            output_polarity_result = normalize_polarity(output.get("polarity"))
            output_cascading_result = normalize_cascading(output.get("cascading"))
            output_ownership_result = normalize_ownership_type(output.get("ownership_type"))
            append_enum_issue(
                issues,
                config,
                output.get("source_row"),
                "OUTPUT",
                output.get("title"),
                "Period",
                output_period_result,
            )
            append_enum_issue(
                issues,
                config,
                output.get("source_row"),
                "OUTPUT",
                output.get("title"),
                "Polarity",
                output_polarity_result,
            )
            append_enum_issue(
                issues,
                config,
                output.get("source_row"),
                "OUTPUT",
                output.get("title"),
                "Cascading",
                output_cascading_result,
            )
            append_enum_issue(
                issues,
                config,
                output.get("source_row"),
                "OUTPUT",
                output.get("title"),
                "Ownership Type",
                output_ownership_result,
            )
            rows.append(
                [
                    output_id,
                    config.group_name,
                    config.directorate_name,
                    position_name,
                    output_position_master_id,
                    output_position_master_variant_id,
                    impact.bsc,
                    "OUTPUT",
                    impact_ids[impact.title],
                    impact.title,
                    output["title"],
                    output["description"],
                    output["unit"],
                    output_polarity_result.value,
                    output_period_result.value,
                    output["formula"],
                    output["weight"],
                    output_cascading_result.value,
                    None,
                    None,
                    None,
                    output_ownership_result.value,
                    output_position_nomenclature_id,
                    config.rkm_code_id,
                ]
            )

    for impact in impacts:
        parent_period_result = normalize_period(impact.period)
        for output in merged_outputs_by_impact[impact.title]:
            output_period_result = output.get("_period_result") or normalize_period(output.get("period"), parent_period_result.value)
            for kai in output_kai_records(output):
                kai_period_result = normalize_period(
                    kai.get("period") or output.get("period"),
                    output_period_result.value or parent_period_result.value,
                )
                kai_polarity_result = normalize_polarity(kai.get("polarity"))
                kai_nature_result = normalize_kai_nature(kai.get("nature_of_work"), kai_period_result.value)
                if not kai_field_matches_output(kai, output, "period"):
                    append_enum_issue(
                        issues,
                        config,
                        kai.get("source_row"),
                        "KAI",
                        kai.get("title"),
                        "Period",
                        kai_period_result,
                    )
                if not kai_field_matches_output(kai, output, "polarity"):
                    append_enum_issue(
                        issues,
                        config,
                        kai.get("source_row"),
                        "KAI",
                        kai.get("title"),
                        "Polarity",
                        kai_polarity_result,
                    )
                append_enum_issue(
                    issues,
                    config,
                    kai.get("source_row"),
                    "KAI",
                    kai.get("title"),
                    "Nature Of Work",
                    kai_nature_result,
                )
                if not kai_field_matches_output(kai, output, "cascading"):
                    append_enum_issue(
                        issues,
                        config,
                        kai.get("source_row"),
                        "KAI",
                        kai.get("title"),
                        "Cascading",
                        normalize_cascading(kai.get("cascading")),
                    )
                if not kai_field_matches_output(kai, output, "ownership_type"):
                    append_enum_issue(
                        issues,
                        config,
                        kai.get("source_row"),
                        "KAI",
                        kai.get("title"),
                        "Ownership Type",
                        normalize_ownership_type(kai.get("ownership_type")),
                    )
                if not is_percentage_formula(kai.get("formula")):
                    if issues is not None:
                        issues.append(
                            ValidationIssue(
                                severity="warning",
                                sheet_name=config.sheet_name,
                                source_row=kai.get("source_row"),
                                record_type="KAI",
                                title=kai.get("title"),
                                message="KAI formula may not be percentage-based or is blank.",
                            )
                        )
                rows.append(
                    [
                        str(next_id),
                        config.group_name,
                        config.directorate_name,
                        position_name,
                        output_position_master_id,
                        output_position_master_variant_id,
                        impact.bsc,
                        "KAI",
                        output["_generated_id"],
                        output["title"],
                        kai["title"],
                        kai["description"],
                        "%",
                        kai_polarity_result.value,
                        kai_period_result.value,
                        kai["formula"],
                        kai["weight"],
                        "INDIRECT",
                        kai_nature_result.value,
                        None,
                        None,
                        "SPECIFIC",
                        output_position_nomenclature_id,
                        config.rkm_code_id,
                    ]
                )
                next_id += 1

    return rows, next_id


def add_weight(total_map: dict[tuple[str, ...], float], key: tuple[str, ...], value: Any) -> None:
    parsed = parse_weight(value)
    if parsed is not None:
        total_map[key] = total_map.get(key, 0.0) + parsed


def raw_weight_totals(parsed: ParsedSheet) -> dict[tuple[str, ...], float]:
    totals: dict[tuple[str, ...], float] = {}
    for impact in parsed.impacts:
        add_weight(totals, ("position", "IMPACT", ""), impact.weight)
        add_weight(totals, ("position", "OUTPUT", ""), 0)
        add_weight(totals, ("position", "KAI", ""), 0)
        add_weight(totals, ("impact", "OUTPUT", impact.title), 0)
        for output in impact.outputs:
            add_weight(totals, ("position", "OUTPUT", ""), output.get("weight"))
            add_weight(totals, ("impact", "OUTPUT", impact.title), output.get("weight"))
            add_weight(totals, ("output", "KAI", output.get("title") or ""), 0)
            kai = output.get("kai")
            if kai:
                add_weight(totals, ("position", "KAI", ""), kai.get("weight"))
                add_weight(totals, ("output", "KAI", output.get("title") or ""), kai.get("weight"))
    return totals


def output_weight_totals(rows: list[list[Any]]) -> dict[tuple[str, ...], float]:
    totals: dict[tuple[str, ...], float] = {}
    title_by_id = {str(row[0]): norm_text(row[10]) or "" for row in rows}
    for row in rows:
        row_map = dict(zip(UPLOAD_HEADERS, row))
        kpi_type = norm_text(row_map.get("KPI Type")) or ""
        parent_id = norm_text(row_map.get("Parent KPI ID")) or ""
        parent_title = title_by_id.get(parent_id, norm_text(row_map.get("Parent KPI Title")) or "")
        if kpi_type in {"IMPACT", "OUTPUT", "KAI"}:
            add_weight(totals, ("position", kpi_type, ""), row_map.get("Weight (%)"))
        if kpi_type == "OUTPUT":
            add_weight(totals, ("impact", "OUTPUT", parent_title), row_map.get("Weight (%)"))
        if kpi_type == "KAI":
            add_weight(totals, ("output", "KAI", parent_title), row_map.get("Weight (%)"))
    return totals


def weight_issue_cause(raw_total: float, output_total: float, merged_duplicates: bool) -> str:
    if weights_equal(raw_total, output_total):
        return "Raw Data Issue"
    if merged_duplicates:
        return "Dedupe Adjustment"
    return "Converter Issue"


def append_weight_audit_issues(
    parsed: ParsedSheet,
    rows: list[list[Any]],
    issues: list[ValidationIssue],
) -> None:
    raw_totals = raw_weight_totals(parsed)
    final_totals = output_weight_totals(rows)
    keys = sorted(set(raw_totals) | set(final_totals))
    merged_duplicates = any(
        output.get("duplicate_kai")
        for impact in parsed.impacts
        for output in impact.outputs
    )
    for level, kpi_type, parent_title in keys:
        if level == "position" and kpi_type not in {"IMPACT", "OUTPUT", "KAI"}:
            continue
        if level == "impact" and kpi_type != "OUTPUT":
            continue
        if level == "output" and kpi_type != "KAI":
            continue
        raw_total = round_weight(raw_totals.get((level, kpi_type, parent_title), 0.0))
        output_total = round_weight(final_totals.get((level, kpi_type, parent_title), 0.0))
        if weights_equal(raw_total, 0.0) and weights_equal(output_total, 0.0):
            continue
        if weights_equal(output_total, 100.0):
            continue
        cause = weight_issue_cause(raw_total, output_total, merged_duplicates)
        parent_text = f"; parent={parent_title}" if parent_title else ""
        issues.append(
            ValidationIssue(
                severity="warning",
                sheet_name=parsed.config.sheet_name,
                source_row=None,
                record_type="weight_audit",
                title=parsed.config.position_name,
                message=(
                    f"Weight total not 100: level={level}; kpi_type={kpi_type}{parent_text}; "
                    f"raw_total={raw_total}; output_total={output_total}; cause={cause}."
                ),
            )
        )


def resolve_output_position_master_id(
    config: PositionConfig,
    metadata: PositionMetadata | None,
) -> str | None:
    if is_non_structural_scope(config) and config.position_nomenclature_id:
        return config.position_master_id or (metadata.position_master_id if metadata else None) or "0"
    # Explicit config value has highest precedence for output consistency.
    if config.position_master_id:
        return str(config.position_master_id)
    if metadata and metadata.position_master_id:
        return metadata.position_master_id
    return None


def normalize_position_scope(value: str | None) -> str | None:
    text = normalize_title(value)
    if not text:
        return None
    if text in {"structural", "struktural"}:
        return "structural"
    if text in {"non structural", "non struktural", "nonstruktural", "nonstructural", "general", "fungsional"}:
        return "non_structural"
    return text.replace(" ", "_")


def infer_position_scope(
    position_master_type_id: Any = None,
    type_name: Any = None,
    position_nomenclature_id: Any = None,
) -> str | None:
    if str(position_master_type_id or "") == "5":
        return "structural"
    text_scope = normalize_position_scope(norm_text(type_name))
    if text_scope in {"structural", "non_structural"}:
        return text_scope
    if position_nomenclature_id not in (None, "", 0, "0"):
        return "non_structural"
    return None


def is_non_structural_scope(config: PositionConfig) -> bool:
    return normalize_position_scope(config.position_scope) == "non_structural"


def is_structural_scope(config: PositionConfig) -> bool:
    return normalize_position_scope(config.position_scope) == "structural"


def is_neglect_scope(config: PositionConfig) -> bool:
    return normalize_position_scope(config.position_scope) in {"neglect", "ignored", "skip", "skipped"}


def is_mapping_conflict_scope(config: PositionConfig) -> bool:
    return normalize_position_scope(config.position_scope) == "mapping_conflict"


def is_mapping_review_approved(config: PositionConfig) -> bool:
    status = normalize_title(config.mapping_review_status)
    return bool(config.mapping_override_approved or status in {"approved", "review approved", "reviewer approved"})


def is_mapping_needs_check(config: PositionConfig) -> bool:
    return normalize_title(config.mapping_review_status) == "needs check"


def is_trusted_reviewer_manual_override(config: PositionConfig) -> bool:
    return is_mapping_review_approved(config) and config.mapping_override_trust_source == "reviewer_manual"


def is_blocked_mapping_confidence(config: PositionConfig) -> bool:
    if is_mapping_needs_check(config):
        return True
    label = normalize_position_scope(config.mapping_confidence_label)
    if not label:
        return False
    if label == strict_position_mapping.HIGH_CONFIDENCE:
        return False
    return not is_mapping_review_approved(config)


def resolve_upload_scope(config: PositionConfig, fallback_position_master_id: str | None) -> tuple[str | None, str | None]:
    if is_non_structural_scope(config):
        return None, config.position_nomenclature_id
    if is_structural_scope(config):
        return config.position_master_id or fallback_position_master_id, None
    if config.position_nomenclature_id:
        return None, config.position_nomenclature_id
    return config.position_master_id or fallback_position_master_id, None


def enforce_position_scope_ids(config: PositionConfig) -> None:
    if is_structural_scope(config):
        config.position_nomenclature_id = None
    elif is_non_structural_scope(config):
        config.position_master_id = None


def validate_output_rows(
    config: PositionConfig,
    rows: list[list[Any]],
    issues: list[ValidationIssue],
) -> None:
    all_ids = {str(row[0]) for row in rows}
    for row in rows:
        row_map = dict(zip(UPLOAD_HEADERS, row))
        title = row_map["Title"]
        record_type = row_map["KPI Type"]
        for key in [
            "IDKPI",
            "KPI Type",
            "Title",
            "Polarity",
            "Period",
            "Weight (%)",
        ]:
            if not norm_text(row_map.get(key)):
                issues.append(
                    ValidationIssue(
                        severity="error",
                        sheet_name=config.sheet_name,
                        source_row=None,
                        record_type=record_type or "row",
                        title=title,
                        message=f"Missing required upload field: {key}",
                    )
                )
        polarity = norm_text(row_map.get("Polarity"))
        if polarity and polarity not in ALLOWED_UPLOAD_POLARITIES:
            issues.append(
                ValidationIssue(
                    severity="error",
                    sheet_name=config.sheet_name,
                    source_row=None,
                    record_type=record_type or "row",
                    title=title,
                    message=f"Invalid Polarity enum: {polarity}",
                )
            )
        period = norm_text(row_map.get("Period"))
        if period and period not in ALLOWED_UPLOAD_PERIODS:
            issues.append(
                ValidationIssue(
                    severity="error",
                    sheet_name=config.sheet_name,
                    source_row=None,
                    record_type=record_type or "row",
                    title=title,
                    message=f"Invalid Period enum: {period}",
                )
            )
        cascading = norm_text(row_map.get("Cascading"))
        if cascading and cascading not in ALLOWED_UPLOAD_CASCADING:
            issues.append(
                ValidationIssue(
                    severity="error",
                    sheet_name=config.sheet_name,
                    source_row=None,
                    record_type=record_type or "row",
                    title=title,
                    message=f"Invalid Cascading enum: {cascading}",
                )
            )
        ownership = norm_text(row_map.get("Ownership Type"))
        if ownership and ownership not in ALLOWED_UPLOAD_OWNERSHIP:
            issues.append(
                ValidationIssue(
                    severity="error",
                    sheet_name=config.sheet_name,
                    source_row=None,
                    record_type=record_type or "row",
                    title=title,
                    message=f"Invalid Ownership Type enum: {ownership}",
                )
            )
        nature = norm_text(row_map.get("Nature Of Work (KAI Only)"))
        if record_type == "KAI" and nature and nature not in {"Routine", "Non Routine"}:
            issues.append(
                ValidationIssue(
                    severity="error",
                    sheet_name=config.sheet_name,
                    source_row=None,
                    record_type="KAI",
                    title=title,
                    message=f"Invalid KAI Nature enum: {nature}",
                )
            )
        pmid = norm_text(row_map.get("Position Master ID (Required)"))
        pnid = norm_text(row_map.get("Position Nomenklatur ID"))
        if pmid and pnid:
            issues.append(
                ValidationIssue(
                    severity="error",
                    sheet_name=config.sheet_name,
                    source_row=None,
                    record_type=record_type or "row",
                    title=title,
                    message="Invalid upload scope: row has both PMID and PNID.",
                )
            )
        if not pmid and not pnid:
            issues.append(
                ValidationIssue(
                    severity="error",
                    sheet_name=config.sheet_name,
                    source_row=None,
                    record_type=record_type or "row",
                    title=title,
                    message=(
                        "Missing upload scope: fill Position Master ID (Required) "
                        "or Position Nomenklatur ID."
                    ),
                )
            )
        if record_type in {"IMPACT", "OUTPUT"} and not norm_text(row_map.get("Unit")):
            issues.append(
                ValidationIssue(
                    severity="warning",
                    sheet_name=config.sheet_name,
                    source_row=None,
                    record_type=record_type,
                    title=title,
                    message="KPI item is missing Unit (Satuan).",
                )
            )
        if record_type == "OUTPUT":
            bsc_result = normalize_bsc_perspective(row_map["BSC Perspective"])
            if not bsc_result.value:
                issues.append(
                    ValidationIssue(
                        severity="error",
                        sheet_name=config.sheet_name,
                        source_row=None,
                        record_type="OUTPUT",
                        title=title,
                        message=bsc_result.message or "OUTPUT row is missing BSC Perspective.",
                    )
                )
            if not norm_text(row_map["Parent KPI ID"]):
                issues.append(
                    ValidationIssue(
                        severity="error",
                        sheet_name=config.sheet_name,
                        source_row=None,
                        record_type="OUTPUT",
                        title=title,
                        message="OUTPUT row is missing Parent KPI ID.",
                    )
                )
        if record_type == "KAI":
            if not norm_text(row_map["Nature Of Work (KAI Only)"]):
                issues.append(
                    ValidationIssue(
                        severity="error",
                        sheet_name=config.sheet_name,
                        source_row=None,
                        record_type="KAI",
                        title=title,
                        message="KAI row is missing Nature Of Work.",
                    )
                )
            if not norm_text(row_map["Parent KPI ID"]):
                issues.append(
                    ValidationIssue(
                        severity="error",
                        sheet_name=config.sheet_name,
                        source_row=None,
                        record_type="KAI",
                        title=title,
                        message="KAI row is missing Parent KPI ID.",
                    )
                )
        parent_id = norm_text(row_map["Parent KPI ID"])
        if record_type in {"OUTPUT", "KAI"} and parent_id and parent_id not in all_ids:
            issues.append(
                ValidationIssue(
                    severity="error",
                    sheet_name=config.sheet_name,
                    source_row=None,
                    record_type=record_type,
                    title=title,
                    message=f"Parent KPI ID {parent_id} does not exist in generated rows.",
                )
            )


def excel_cell_value(header: str, value: Any) -> Any:
    if header == "Weight (%)":
        parsed = parse_weight(value)
        return parsed if parsed is not None else value
    return value


def extend_worksheet_rules_to_row(worksheet: Any, source_last_row: int, target_last_row: int) -> None:
    """Extend template conditional formatting and validation through generated rows."""
    if target_last_row <= source_last_row:
        return

    def extended_sqref(sqref: Any) -> str:
        updated_ranges: list[str] = []
        for range_ref in str(sqref).split():
            cell_range = CellRange(range_ref)
            if cell_range.max_row == source_last_row:
                cell_range.max_row = target_last_row
            updated_ranges.append(str(cell_range))
        return " ".join(updated_ranges)

    conditional_rules = OrderedDict()
    for conditional_format, rules in worksheet.conditional_formatting._cf_rules.items():
        updated_format = copy(conditional_format)
        updated_format.sqref = extended_sqref(conditional_format.sqref)
        conditional_rules[updated_format] = rules
    worksheet.conditional_formatting._cf_rules = conditional_rules

    for validation in worksheet.data_validations.dataValidation:
        validation.sqref = extended_sqref(validation.sqref)


def write_output_workbook(template_path: Path, output_path: Path, rows: list[list[Any]]) -> None:
    workbook = load_workbook(template_path)
    worksheet = workbook["KPI Template"] if "KPI Template" in workbook.sheetnames else workbook.active
    template_last_row = worksheet.max_row
    if worksheet.max_column > len(UPLOAD_HEADERS):
        worksheet.delete_cols(len(UPLOAD_HEADERS) + 1, worksheet.max_column - len(UPLOAD_HEADERS))
    for table_name in list(worksheet.tables.keys()):
        del worksheet.tables[table_name]
    header_style = copy(worksheet.cell(row=1, column=1)._style)
    header_font = copy(worksheet.cell(row=1, column=1).font)
    header_fill = copy(worksheet.cell(row=1, column=1).fill)
    header_border = copy(worksheet.cell(row=1, column=1).border)
    for col_index, header in enumerate(UPLOAD_HEADERS, start=1):
        cell = worksheet.cell(row=1, column=col_index, value=header)
        cell._style = copy(header_style)
        cell.font = copy(header_font)
        cell.font = Font(
            name=cell.font.name,
            sz=cell.font.sz,
            b=cell.font.b,
            i=cell.font.i,
            vertAlign=cell.font.vertAlign,
            underline=cell.font.underline,
            strike=cell.font.strike,
            color="000000",
        )
        cell.fill = copy(header_fill)
        cell.border = copy(header_border)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    if worksheet.max_row > 1:
        worksheet.delete_rows(2, worksheet.max_row - 1)
    data_style = copy(worksheet.cell(row=2, column=1)._style) if worksheet.max_row >= 2 else None
    for row_index, row_values in enumerate(rows, start=2):
        for col_index, value in enumerate(row_values, start=1):
            header = UPLOAD_HEADERS[col_index - 1]
            cell = worksheet.cell(row=row_index, column=col_index, value=excel_cell_value(header, value))
            if data_style:
                cell._style = copy(data_style)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if header == "Weight (%)":
                cell.number_format = "0.00"
    for col_index, header in enumerate(UPLOAD_HEADERS, start=1):
        worksheet.column_dimensions[get_column_letter(col_index)].width = COLUMN_WIDTHS.get(header, 18)
    worksheet.row_dimensions[1].height = 36
    worksheet.auto_filter.ref = None
    extend_worksheet_rules_to_row(worksheet, template_last_row, len(rows) + 1)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def write_report(report_path: Path, issues: list[ValidationIssue]) -> None:
    report_path.parent.mkdir(parents=True, exist_ok=True)
    with report_path.open("w", newline="", encoding="utf-8") as csvfile:
        writer = csv.writer(csvfile)
        writer.writerow(["severity", "sheet_name", "source_row", "record_type", "title", "message"])
        for issue in issues:
            writer.writerow(
                [
                    issue.severity,
                    issue.sheet_name,
                    issue.source_row,
                    issue.record_type,
                    issue.title,
                    issue.message,
                ]
            )


def load_config(config_path: Path) -> list[PositionConfig]:
    data = json.loads(config_path.read_text(encoding="utf-8"))
    configs = []
    for item in data["positions"]:
        raw_position_master_id = item.get("position_master_id")
        normalized_position_master_id: str | None
        if raw_position_master_id in (None, "", 0, "0"):
            normalized_position_master_id = None
        else:
            normalized_position_master_id = str(raw_position_master_id)
        raw_position_master_variant_id = item.get("position_master_variant_id")
        normalized_position_master_variant_id = (
            None
            if raw_position_master_variant_id in (None, "", 0, "0")
            else str(raw_position_master_variant_id)
        )
        raw_position_nomenclature_id = item.get("position_nomenclature_id")
        normalized_position_nomenclature_id: str | None
        if raw_position_nomenclature_id in (None, "", 0, "0"):
            normalized_position_nomenclature_id = None
        else:
            normalized_position_nomenclature_id = str(raw_position_nomenclature_id)
        raw_rkm_code_id = item.get("rkm_code_id")
        position_scope = normalize_position_scope(item.get("position_scope"))
        config = PositionConfig(
            sheet_name=item["sheet_name"],
            position_name=item["position_name"],
            group_name=item["group_name"],
            directorate_name=item["directorate_name"],
            source_workbook=item.get("source_workbook"),
            position_master_id=normalized_position_master_id,
            position_master_variant_id=normalized_position_master_variant_id,
            position_nomenclature_id=normalized_position_nomenclature_id,
            position_scope=position_scope,
            portaverse_position_title=item.get("portaverse_position_title"),
            portaverse_group_name=item.get("portaverse_group_name"),
            portaverse_company_name=item.get("portaverse_company_name"),
            portaverse_company_code=item.get("portaverse_company_code"),
            cluster_label=item.get("cluster_label"),
            mapping_confidence_label=item.get("mapping_confidence_label"),
            mapping_confidence_reason=item.get("mapping_confidence_reason"),
            mapping_review_status=item.get("mapping_review_status"),
            mapping_override_approved=is_truthy_flag(item.get("mapping_override_approved")),
            mapping_override_trust_source=item.get("mapping_override_trust_source"),
            candidate_position_master_id=(
                str(item.get("candidate_position_master_id"))
                if item.get("candidate_position_master_id") not in (None, "", 0, "0")
                else None
            ),
            candidate_position_nomenclature_id=(
                str(item.get("candidate_position_nomenclature_id"))
                if item.get("candidate_position_nomenclature_id") not in (None, "", 0, "0")
                else None
            ),
            candidate_score=item.get("candidate_score"),
            runner_up_position_master_id=(
                str(item.get("runner_up_position_master_id"))
                if item.get("runner_up_position_master_id") not in (None, "", 0, "0")
                else None
            ),
            runner_up_position_nomenclature_id=(
                str(item.get("runner_up_position_nomenclature_id"))
                if item.get("runner_up_position_nomenclature_id") not in (None, "", 0, "0")
                else None
            ),
            runner_up_score=item.get("runner_up_score"),
            active_variant_count=item.get("active_variant_count"),
            active_employee_count=item.get("active_employee_count"),
            active_employee_name=item.get("active_employee_name"),
            active_employee_nipp=item.get("active_employee_nipp"),
            rkm_code_id=str(raw_rkm_code_id) if raw_rkm_code_id not in (None, "") else None,
            position_lookup_names=item.get("position_lookup_names", []),
            drop_comment_values=item.get("drop_comment_values", ["Drop"]),
            expected_impact_count=int(item.get("expected_impact_count", 10)),
        )
        enforce_position_scope_ids(config)
        configs.append(config)
    return configs


def safe_path_stem(value: str) -> str:
    stem = Path(value).stem
    stem = re.sub(r"[^A-Za-z0-9._-]+", "-", stem).strip("-._")
    return stem or "workbook"


def clean_output_title_part(value: str | None) -> str:
    text = norm_text(value) or ""
    text = re.sub(r"\([^)]*(done|konfirmasi|kpi)[^)]*\)", "", text, flags=re.IGNORECASE)
    text = re.sub(r"\b(done|selesai)\b.*$", "", text, flags=re.IGNORECASE)
    text = re.sub(r"\s+", " ", text.replace("&", "dan")).strip(" -")
    text = text.title()
    replacements = {
        " Sdm ": " SDM ",
        " Kpi": " KPI",
        " Spi": " SPI",
        " Ssc": " SSC",
        " Anper": " AnPer",
        " Monev": " MonEv",
        " Dan ": " dan ",
        "DIREKTORAT": "Direktorat",
    }
    text = f" {text} "
    for old, new in replacements.items():
        text = text.replace(old, new)
    return text.strip()


def extract_output_title_parts(source_workbook: str, configs: list[PositionConfig]) -> tuple[str, str]:
    stem = Path(source_workbook).stem
    parts = [part.strip() for part in stem.split(" - ") if part.strip()]
    directorate = parts[0] if parts else (configs[0].directorate_name if configs else "Direktorat")
    group = None
    for part in parts[1:]:
        match = re.search(r"\bgroup\b.+", part, flags=re.IGNORECASE)
        if match:
            group = match.group(0)
            break
    if not group and configs:
        candidates = [
            config.portaverse_group_name or config.group_name
            for config in configs
            if config.portaverse_group_name or config.group_name
        ]
        filtered = [
            candidate
            for candidate in candidates
            if normalize_title(candidate) not in {"group head", "manager", "officer"}
            and ":" not in candidate
            and "mandat fungsi" not in normalize_title(candidate)
        ]
        group = Counter(filtered or candidates).most_common(1)[0][0] if candidates else None
    return clean_output_title_part(directorate), clean_output_title_part(group or "Group")


def source_workbook_group_name(source_workbook: str) -> str | None:
    stem = Path(source_workbook).stem
    parts = [part.strip() for part in stem.split(" - ") if part.strip()]
    for part in parts[1:]:
        match = re.search(r"\bgroup\b.+", part, flags=re.IGNORECASE)
        if match:
            return clean_output_title_part(match.group(0))
    if len(parts) >= 3:
        return clean_output_title_part(parts[-1])
    return None


def is_invalid_discovered_group_name(group_name: str | None, position_name: str | None, sheet_name: str) -> bool:
    text = normalize_title(group_name)
    if not text:
        return True
    if text in {normalize_title(position_name), normalize_title(sheet_name), "group head", "manager", "officer"}:
        return True
    return any(token in text for token in ["mandat fungsi", "kpi tidak sesuai", "jenis posisi", "nama posisi"])


def conversion_output_name(
    source_workbook: str,
    configs: list[PositionConfig],
    generated_at: datetime,
    year: int = 2026,
    version: int = 2,
) -> str:
    """Name upload formulir from raw kamus stem + timestamp + pekerja/identity counts.

    Example:
      Kamus KPI SPTP - Mapping dengan Kontrak Manajemen - 20260807_1430 (42 pekerja, 35 identity)
    """
    del year, version  # retained in signature for callers; counts replace version tag
    raw_stem = Path(source_workbook).stem.strip() or "Kamus KPI"
    # Strip trailing " - Mapping dengan Kontrak Manajemen" noise? Keep full raw stem so HO
    # SPTP/SPJM/SPSL/SPMT stay searchable from the original kamus filename.
    timestamp = generated_at.strftime("%Y%m%d_%H%M")
    identity_count = len(configs)
    pekerja_count = 0
    for config in configs:
        try:
            pekerja_count += int(config.active_employee_count or 0)
        except (TypeError, ValueError):
            continue
    if pekerja_count > 0:
        return f"{raw_stem} - {timestamp} ({pekerja_count} pekerja, {identity_count} identity)"
    return f"{raw_stem} - {timestamp} ({identity_count} identity)"


def enforce_default_impact_set(
    impacts: list[ImpactRecord],
    expected_count: int,
    *,
    sheet_name: str,
    position_name: str,
    issues: list[ValidationIssue],
) -> list[ImpactRecord]:
    """Keep one Impact per title (first wins), then cap at expected_count (default 10).

    Duplicate Impact blocks in kamus (same title, multiple rows) are dropped with their
    OUTPUT/KAI children so upload formulir carries the shared Pelindo set only.
    """
    if expected_count <= 0:
        return impacts
    deduped: list[ImpactRecord] = []
    seen_titles: set[str] = set()
    duplicate_count = 0
    for impact in impacts:
        key = normalize_title(impact.title)
        if not key:
            continue
        if key in seen_titles:
            duplicate_count += 1
            continue
        seen_titles.add(key)
        deduped.append(impact)

    truncated = 0
    if len(deduped) > expected_count:
        truncated = len(deduped) - expected_count
        deduped = deduped[:expected_count]

    if duplicate_count or truncated or len(impacts) != len(deduped):
        issues.append(
            ValidationIssue(
                severity="info",
                sheet_name=sheet_name,
                source_row=None,
                record_type="sheet",
                title=position_name,
                message=(
                    f"Enforced default {expected_count} KPI Impact items "
                    f"(parsed={len(impacts)}, unique_kept={len(deduped)}, "
                    f"duplicates_dropped={duplicate_count}, truncated={truncated})."
                ),
            )
        )
    return deduped


def unique_output_name(output_name: str, seen_names: dict[str, int]) -> str:
    count = seen_names.get(output_name, 0) + 1
    seen_names[output_name] = count
    if count == 1:
        return output_name
    return f"{output_name} - {count:02d}"


def config_to_dict(config: PositionConfig) -> dict[str, Any]:
    position_master_id = config.position_master_id
    position_nomenclature_id = config.position_nomenclature_id
    if is_structural_scope(config):
        position_nomenclature_id = None
    elif is_non_structural_scope(config):
        position_master_id = None
    data: dict[str, Any] = {
        "source_workbook": config.source_workbook,
        "sheet_name": config.sheet_name,
        "position_name": config.position_name,
        "position_master_id": position_master_id,
        "position_master_variant_id": config.position_master_variant_id,
        "position_nomenclature_id": position_nomenclature_id,
        "position_scope": config.position_scope,
        "portaverse_position_title": config.portaverse_position_title,
        "portaverse_group_name": config.portaverse_group_name,
        "portaverse_company_name": config.portaverse_company_name,
        "portaverse_company_code": config.portaverse_company_code,
        "cluster_label": config.cluster_label,
        "mapping_confidence_label": config.mapping_confidence_label,
        "mapping_confidence_reason": config.mapping_confidence_reason,
        "mapping_review_status": config.mapping_review_status,
        "mapping_override_approved": config.mapping_override_approved,
        "mapping_override_trust_source": config.mapping_override_trust_source,
        "candidate_position_master_id": config.candidate_position_master_id,
        "candidate_position_nomenclature_id": config.candidate_position_nomenclature_id,
        "candidate_score": config.candidate_score,
        "runner_up_position_master_id": config.runner_up_position_master_id,
        "runner_up_position_nomenclature_id": config.runner_up_position_nomenclature_id,
        "runner_up_score": config.runner_up_score,
        "active_variant_count": config.active_variant_count,
        "active_employee_count": config.active_employee_count,
        "active_employee_name": config.active_employee_name,
        "active_employee_nipp": config.active_employee_nipp,
        "position_lookup_names": config.position_lookup_names,
        "group_name": config.group_name,
        "directorate_name": config.directorate_name,
        "expected_impact_count": config.expected_impact_count,
        "drop_comment_values": config.drop_comment_values,
    }
    if config.rkm_code_id:
        data["rkm_code_id"] = config.rkm_code_id
    return data


def reference_source_for(mapping_path: Path | None) -> dict[str, Any]:
    source: dict[str, Any] = {
        "config_generated_at": datetime.now(timezone.utc).isoformat(timespec="seconds"),
    }
    if not mapping_path:
        return source
    source["file"] = str(mapping_path)
    if mapping_path.exists():
        try:
            payload = json.loads(mapping_path.read_text(encoding="utf-8"))
        except json.JSONDecodeError:
            source["note"] = "Reference metadata could not be parsed."
            return source
        reference_source = payload.get("source", {}) if isinstance(payload, dict) else {}
        source["profile"] = reference_source.get("profile")
        source["database"] = reference_source.get("database")
        source["reference_exported_at"] = reference_source.get("exported_at")
    return source


def write_discovered_config(
    config_path: Path,
    configs: list[PositionConfig],
    mapping_path: Path | None = None,
) -> None:
    config_path.parent.mkdir(parents=True, exist_ok=True)
    payload = {
        "reference_source": reference_source_for(mapping_path),
        "positions": [config_to_dict(config) for config in configs],
    }
    config_path.write_text(json.dumps(payload, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")


def merge_mapping_entry(
    mapping: dict[str, dict[str, Any]],
    key: str | None,
    entry: dict[str, Any],
) -> None:
    normalized = normalize_title(key)
    if not normalized:
        return
    current = mapping.get(normalized)
    if current is None:
        mapping[normalized] = dict(entry)
        mapping[normalized]["_candidates"] = [dict(entry)]
        return
    current.setdefault("_candidates", []).append(dict(entry))
    if mapping_entry_scope(entry) == "structural":
        current.update({key_name: value for key_name, value in entry.items() if value is not None})
        current["position_nomenclature_id"] = None
        current["position_scope"] = "structural"
        return
    if mapping_entry_scope(current) == "structural":
        for key_name, value in entry.items():
            if key_name == "position_nomenclature_id":
                continue
            if current.get(key_name) in (None, "") and value not in (None, ""):
                current[key_name] = value
        current["position_nomenclature_id"] = None
        current["position_scope"] = "structural"
        return
    # PNID rows are more specific for non-structural positions and must win
    # over a generic active PMID row with a similar title.
    if entry.get("position_nomenclature_id"):
        current.update({key_name: value for key_name, value in entry.items() if value is not None})
    else:
        for key_name, value in entry.items():
            if current.get(key_name) in (None, "") and value not in (None, ""):
                current[key_name] = value


def mapping_entry_scope(entry: dict[str, Any]) -> str | None:
    if str(entry.get("position_master_type_id") or "") == "5":
        return "structural"
    return normalize_position_scope(entry.get("position_scope"))


def load_nomenclature_mapping(
    mapping_path: Path | None,
    target_company_id: str | None = TARGET_COMPANY_ID_DEFAULT,
) -> dict[str, dict[str, str | None]]:
    if not mapping_path:
        return {}
    data = json.loads(mapping_path.read_text(encoding="utf-8"))
    position_master_rows = data.get("position_master_rows", []) if isinstance(data, dict) else []
    mapping: dict[str, dict[str, str | None]] = {}
    for row in position_master_rows:
        if not is_target_company_row(row, target_company_id) or not is_active_reference_row(row, require_active_org=True):
            continue
        position_name = norm_text(row.get("position_name"))
        if not position_name:
            continue
        if row.get("position_master_id") in (None, "", 0, "0"):
            continue
        entry = {
            "position_master_id": str(row["position_master_id"]),
            "position_nomenclature_id": None,
            "position_scope": infer_position_scope(row.get("position_master_type_id"), row.get("type_name")),
            "portaverse_position_title": norm_text(row.get("position_name")),
            "portaverse_group_name": norm_text(row.get("group_name")),
            "portaverse_company_name": norm_text(row.get("company_name")),
            "portaverse_company_code": norm_text(row.get("company_code")),
            "cluster_label": None,
            "position_master_type_id": norm_text(row.get("position_master_type_id")),
        }
        merge_mapping_entry(mapping, position_name, entry)

    rows = data.get("rows", data if isinstance(data, list) else [])
    for row in rows:
        if not is_target_company_row(row, target_company_id) or not is_active_reference_row(row):
            continue
        position_name = norm_text(row.get("position_name"))
        cluster_label = norm_text(row.get("cluster_label"))
        if not position_name and not cluster_label:
            continue
        entry: dict[str, str | None] = {
            "position_master_id": None,
            "position_nomenclature_id": None,
            "position_scope": None,
            "portaverse_position_title": norm_text(row.get("position_name")),
            "portaverse_group_name": norm_text(row.get("active_group_name")) or norm_text(row.get("group_name")),
            "portaverse_company_name": norm_text(row.get("active_company_name")) or norm_text(row.get("company_name")),
            "portaverse_company_code": norm_text(row.get("active_company_code")) or norm_text(row.get("company_code")),
            "cluster_label": cluster_label,
            "position_master_type_id": norm_text(row.get("position_master_type_id")),
        }
        if row.get("position_master_id") not in (None, "", 0, "0"):
            entry["position_master_id"] = str(row["position_master_id"])
        if row.get("cluster_id") not in (None, "", 0, "0"):
            entry["position_nomenclature_id"] = str(row["cluster_id"])
        scope = infer_position_scope(
            row.get("position_master_type_id"),
            row.get("type_name"),
            row.get("cluster_id"),
        )
        if scope:
            entry["position_scope"] = scope
        merge_mapping_entry(mapping, position_name, entry)
        merge_mapping_entry(mapping, cluster_label, entry)
    return mapping


def load_position_reference_payload(mapping_path: Path | None) -> dict[str, Any] | list[dict[str, Any]]:
    if not mapping_path:
        return {}
    return json.loads(mapping_path.read_text(encoding="utf-8"))


def load_nomenclature_summary_aliases(summary_path: Path | None) -> dict[str, set[str]]:
    if not summary_path:
        return {}
    workbook = load_workbook(summary_path, read_only=True, data_only=True)
    if "Database Master" not in workbook.sheetnames:
        return {}
    sheet = workbook["Database Master"]
    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = [str(value).strip() if value else "" for value in header_row]
    new_index = next(
        (
            index
            for index, header in enumerate(headers)
            if "Judul Posisi" in header and "Perubahan STO" in header
        ),
        None,
    )
    old_index = headers.index("Kondisi Existing (Before 1 April)") if "Kondisi Existing (Before 1 April)" in headers else None
    aliases: dict[str, set[str]] = {}
    if new_index is None and old_index is None:
        return aliases
    for row in sheet.iter_rows(min_row=2, values_only=True):
        new_title = norm_text(row[new_index]) if new_index is not None and new_index < len(row) else None
        old_title = norm_text(row[old_index]) if old_index is not None and old_index < len(row) else None
        titles = [title for title in [new_title, old_title] if title]
        for title in titles:
            key = normalize_position_lookup(title)
            if not key:
                continue
            aliases.setdefault(key, set()).update(other for other in titles if other != title)
    return aliases


def _sheet_dict_rows(workbook: Any, sheet_name: str) -> list[dict[str, Any]]:
    if sheet_name not in workbook.sheetnames:
        return []
    sheet = workbook[sheet_name]
    iterator = sheet.iter_rows(values_only=True)
    try:
        header_row = next(iterator)
    except StopIteration:
        return []
    headers = [str(value).strip() if value else "" for value in header_row]
    rows: list[dict[str, Any]] = []
    for row in iterator:
        item = {
            headers[index]: row[index]
            for index in range(min(len(headers), len(row)))
            if headers[index]
        }
        if any(value not in (None, "") for value in item.values()):
            rows.append(item)
    return rows


def load_nomenclature_summary_upload_rows(summary_path: Path | None) -> dict[str, dict[str, Any]]:
    if not summary_path:
        return {}
    workbook = load_workbook(summary_path, read_only=True, data_only=True)
    rows = _sheet_dict_rows(workbook, "Upload_Ready")
    by_pmid: dict[str, dict[str, Any]] = {}
    for row in rows:
        pmid = norm_text(row.get("position_master_id"))
        cluster_id = norm_text(row.get("cluster_id"))
        cluster_label = norm_text(row.get("cluster_label"))
        if pmid and cluster_id and cluster_label:
            by_pmid[pmid] = row
    return by_pmid


def apply_nomenclature_summary_aliases(
    payload: dict[str, Any] | list[dict[str, Any]],
    aliases: dict[str, set[str]],
) -> None:
    if not aliases or not isinstance(payload, dict):
        return
    rows: list[dict[str, Any]] = []
    for key in ["structural_lookup_rows", "non_structural_lookup_rows", "position_master_rows", "rows"]:
        rows.extend(row for row in payload.get(key, []) if isinstance(row, dict))
    for row in rows:
        title = norm_text(row.get("cluster_label") or row.get("position_name") or row.get("portaverse_position_title"))
        title_key = normalize_position_lookup(title)
        if not title_key or title_key not in aliases:
            continue
        lookup_keys = list(row.get("normalized_lookup_keys") or row.get("lookup_keys") or [])
        for alias in sorted(aliases[title_key]):
            if alias not in lookup_keys:
                lookup_keys.append(alias)
        row["normalized_lookup_keys"] = lookup_keys


def apply_nomenclature_summary_upload_rows(
    payload: dict[str, Any] | list[dict[str, Any]],
    upload_rows_by_pmid: dict[str, dict[str, Any]],
) -> None:
    if not upload_rows_by_pmid or not isinstance(payload, dict):
        return
    for row in payload.get("non_structural_lookup_rows", []):
        if not isinstance(row, dict):
            continue
        pmid = norm_text(row.get("position_master_id"))
        summary_row = upload_rows_by_pmid.get(pmid or "")
        if not summary_row:
            continue
        cluster_label = norm_text(summary_row.get("cluster_label"))
        cluster_id = norm_text(summary_row.get("cluster_id"))
        if cluster_label:
            row["cluster_label"] = cluster_label
        if cluster_id:
            row["cluster_id"] = cluster_id
            row["position_nomenclature_id"] = cluster_id
        lookup_keys = list(row.get("normalized_lookup_keys") or row.get("lookup_keys") or [])
        for value in [summary_row.get("position_name"), cluster_label, row.get("position_name")]:
            text = norm_text(value)
            if text and text not in lookup_keys:
                lookup_keys.append(text)
        row["normalized_lookup_keys"] = lookup_keys


def load_position_reference(
    mapping_path: Path | None,
    target_company_id: str | None = TARGET_COMPANY_ID_DEFAULT,
    nomenclature_summary_path: Path | None = None,
) -> tuple[dict[str, dict[str, str | None]], strict_position_mapping.LookupIndexes | None]:
    if not mapping_path:
        return {}, None
    payload = load_position_reference_payload(mapping_path)
    apply_nomenclature_summary_aliases(payload, load_nomenclature_summary_aliases(nomenclature_summary_path))
    apply_nomenclature_summary_upload_rows(payload, load_nomenclature_summary_upload_rows(nomenclature_summary_path))
    mapping = load_nomenclature_mapping(mapping_path, target_company_id)
    indexes = strict_position_mapping.build_lookup_indexes(payload, target_company_id)
    return mapping, indexes


def apply_strict_resolution(config: PositionConfig, resolution: strict_position_mapping.MappingResolution) -> None:
    config.position_scope = resolution.inferred_scope
    config.mapping_confidence_label = resolution.confidence_label
    config.mapping_confidence_reason = resolution.confidence_reason
    config.candidate_position_master_id = resolution.position_master_id
    config.candidate_position_nomenclature_id = resolution.position_nomenclature_id
    config.candidate_score = resolution.candidate_score
    config.runner_up_position_master_id = resolution.runner_up_position_master_id
    config.runner_up_position_nomenclature_id = resolution.runner_up_position_nomenclature_id
    config.runner_up_score = resolution.runner_up_score
    config.portaverse_position_title = resolution.candidate_title
    config.portaverse_group_name = resolution.candidate_group
    config.portaverse_company_name = resolution.candidate_company
    config.portaverse_company_code = resolution.candidate_company_code
    config.active_variant_count = resolution.active_variant_count
    config.active_employee_count = resolution.active_employee_count
    config.active_employee_name = resolution.active_employee_name
    config.active_employee_nipp = resolution.active_employee_nipp
    if resolution.upload_allowed:
        config.position_master_id = resolution.position_master_id if resolution.inferred_scope == "structural" else None
        config.position_nomenclature_id = (
            resolution.position_nomenclature_id if resolution.inferred_scope == "non_structural" else None
        )
    elif not is_mapping_review_approved(config):
        config.position_master_id = None
        config.position_nomenclature_id = None
    enforce_position_scope_ids(config)


def apply_validated_override_candidate(
    config: PositionConfig,
    validation: strict_position_mapping.OverrideValidation,
) -> None:
    candidate = validation.candidate
    config.mapping_confidence_reason = validation.reason
    if not candidate:
        return
    config.position_scope = candidate.scope
    config.candidate_position_master_id = candidate.position_master_id
    config.candidate_position_nomenclature_id = candidate.position_nomenclature_id
    config.portaverse_position_title = candidate.title
    config.portaverse_group_name = candidate.group_name
    config.portaverse_company_name = candidate.company_name
    config.portaverse_company_code = candidate.company_code
    config.active_variant_count = candidate.active_variant_count
    config.active_employee_count = candidate.active_employee_count
    config.active_employee_name = "; ".join(candidate.active_employee_names) or None
    config.active_employee_nipp = "; ".join(candidate.active_employee_nipps) or None
    if candidate.scope == "structural":
        config.position_master_id = candidate.position_master_id
        config.position_nomenclature_id = None
    elif candidate.scope == "non_structural":
        config.position_master_id = None
        config.position_nomenclature_id = candidate.position_nomenclature_id
    enforce_position_scope_ids(config)


def apply_best_effort_mapping(configs: list[PositionConfig]) -> None:
    for config in configs:
        label = normalize_position_scope(config.mapping_confidence_label)
        scope = normalize_position_scope(config.position_scope)
        has_candidate = bool(config.candidate_position_master_id or config.candidate_position_nomenclature_id)
        if label in {"low_confidence", "mapping_conflict"} and has_candidate:
            config.mapping_review_status = "auto_approved_best_effort"
            config.mapping_override_approved = True
            if scope == "structural":
                config.position_master_id = config.candidate_position_master_id
                config.position_nomenclature_id = None
            elif scope == "non_structural":
                config.position_master_id = None
                config.position_nomenclature_id = config.candidate_position_nomenclature_id
            config.mapping_confidence_reason = (
                (config.mapping_confidence_reason or "")
                + " Best-effort mode selected the highest-scored active candidate."
            ).strip()
            enforce_position_scope_ids(config)
        elif label in {"no_candidate", "scope_uncertain"} and not has_candidate:
            config.position_scope = "neglect"
            config.mapping_review_status = "auto_skipped_unmapped_best_effort"
            config.mapping_confidence_reason = (
                (config.mapping_confidence_reason or "")
                + " Best-effort mode skipped this worksheet because no active PMID/PNID candidate was available."
            ).strip()


def normalize_position_lookup(value: str | None) -> str:
    """Shared core normalizer plus transform-only expansions used by converter configs."""
    text = strict_position_mapping.normalize_position_lookup(value)
    extras = [
        (r"\bcorpo\b", "corporate"),
        (r"\bcorcomm\b", "corporate communication"),
        (r"\banper\b", "anak perusahaan"),
        (r"\badmin\b", "administrasi"),
        (r"\bperenca\b", "perencanaan"),
        (r"\bpengemb\b", "pengembangan"),
        (r"\bkepatuha\b", "kepatuhan"),
        (r"\bmeko\b", "monitoring evaluasi korporasi"),
        (r"\bmeka\b", "monitoring evaluasi anak perusahaan"),
    ]
    for pattern, replacement in extras:
        text = re.sub(pattern, replacement, text)
    text = re.sub(r"\bcompany\s+\d+\b", "", text)
    return re.sub(r"\s+", " ", text).strip()


def normalize_company_context(value: str | None) -> str:
    text = normalize_title(value)
    text = re.sub(r"\bpt\b", " ", text)
    text = re.sub(r"\bpersero\b", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def unique_mapping_candidate(candidates: list[dict[str, Any]]) -> dict[str, Any]:
    unique: dict[tuple[Any, Any], dict[str, Any]] = {}
    for candidate in candidates:
        identity = (candidate.get("position_master_id"), candidate.get("position_nomenclature_id"))
        unique.setdefault(identity, candidate)
    return next(iter(unique.values())) if len(unique) == 1 else {}


def mapping_candidate_pool(entry: dict[str, Any]) -> list[dict[str, Any]]:
    candidates = entry.get("_candidates")
    if isinstance(candidates, list) and candidates:
        return [candidate for candidate in candidates if isinstance(candidate, dict)]
    return [entry]


def company_context_matches(entry: dict[str, Any], context_hints: list[str] | None) -> bool:
    if not context_hints:
        return False
    company_text = normalize_company_context(
        " ".join(
            [
                norm_text(entry.get("portaverse_company_name")) or "",
                norm_text(entry.get("portaverse_company_code")) or "",
            ]
        )
    )
    if not company_text:
        return False
    return any(hint and hint in company_text for hint in context_hints)


def select_contextual_candidate(entry: dict[str, Any], context_hints: list[str] | None) -> dict[str, Any]:
    if not context_hints:
        return {}
    matches = [
        candidate
        for candidate in mapping_candidate_pool(entry)
        if company_context_matches(candidate, context_hints)
    ]
    return unique_mapping_candidate(matches)


def lookup_position_scope(
    position_name: str,
    nomenclature_mapping: dict[str, dict[str, Any]],
    normalized_mapping: dict[str, dict[str, Any]] | None = None,
    context_hints: list[str] | None = None,
) -> dict[str, Any]:
    direct = nomenclature_mapping.get(normalize_title(position_name))
    if direct:
        contextual = select_contextual_candidate(direct, context_hints)
        if contextual:
            return contextual
    if direct and (
        direct.get("position_nomenclature_id")
        or mapping_entry_scope(direct) == "structural"
    ):
        return direct

    normalized = normalize_position_lookup(position_name)
    is_expanded_lookup = normalized != normalize_title(position_name)
    if not normalized or normalized in {"group head", "department head", "manager", "officer"}:
        return {}

    if normalized_mapping is None:
        normalized_mapping = build_normalized_position_mapping(nomenclature_mapping)
    direct = normalized_mapping.get(normalized)
    if direct:
        contextual = select_contextual_candidate(direct, context_hints)
        if contextual:
            return contextual
        if is_expanded_lookup and context_hints:
            return {}
    if direct:
        direct_unique = unique_mapping_candidate(mapping_candidate_pool(direct))
        if direct_unique:
            return direct_unique
        if not is_expanded_lookup:
            return direct

    exact_without_pnid = nomenclature_mapping.get(normalize_title(position_name))
    if exact_without_pnid:
        return exact_without_pnid

    if len(normalized) < 12:
        return {}
    matches = [
        value
        for key, value in normalized_mapping.items()
        if key and (key.startswith(normalized) or normalized.startswith(key))
    ]
    if context_hints:
        contextual_matches = [
            candidate
            for value in matches
            for candidate in mapping_candidate_pool(value)
            if company_context_matches(candidate, context_hints)
        ]
        contextual = unique_mapping_candidate(contextual_matches)
        if contextual:
            return contextual
        return {}
    flattened_matches = [candidate for value in matches for candidate in mapping_candidate_pool(value)]
    unique = unique_mapping_candidate(flattened_matches)
    if unique:
        return unique
    return {}


def lookup_position_scope_candidates(
    candidates: list[str | None],
    nomenclature_mapping: dict[str, dict[str, Any]],
    normalized_mapping: dict[str, dict[str, Any]] | None = None,
    context_hints: list[str] | None = None,
) -> dict[str, Any]:
    seen: set[str] = set()
    for candidate in candidates:
        text = norm_text(candidate)
        if not text or text in seen:
            continue
        seen.add(text)
        lookup = lookup_position_scope(text, nomenclature_mapping, normalized_mapping, context_hints)
        if lookup:
            return lookup
    return {}


def source_workbook_context_hints(source_workbook: str | None) -> list[str]:
    if not source_workbook:
        return []
    path_parts = [Path(part).stem for part in str(source_workbook).split("/") if part]
    hints: list[str] = []
    generic_hints = {
        "afiliasi",
        "anak perusahaan",
        "bidang",
        "cabang",
        "company",
        "dan",
        "department",
        "director",
        "division",
        "executive",
        "general",
        "group",
        "head",
        "indonesia",
        "kamus",
        "kawasan",
        "kinerja",
        "kontrak manajemen",
        "kpi",
        "manager",
        "mapping dengan kontrak manajemen",
        "officer",
        "pelabuhan",
        "pelindo",
        "persero",
        "pt",
        "regional",
        "subholding",
    }
    for part in path_parts[-3:]:
        text = normalize_company_context(re.sub(r"\bkamus\b|\bkpi\b", " ", part, flags=re.IGNORECASE))
        if text:
            hints.append(text)
        for token in re.findall(r"\b[A-Z0-9]{2,}\b", part):
            hints.append(normalize_company_context(token))
    return [
        hint
        for hint in dict.fromkeys(hints)
        if hint
        and hint not in generic_hints
        and not all(token in generic_hints for token in hint.split())
        and (len(hint) > 2 or hint.isdigit())
    ]


def filter_company_context_hints(
    nomenclature_mapping: dict[str, dict[str, Any]],
    hints: list[str],
) -> list[str]:
    if not hints:
        return []
    cache = getattr(filter_company_context_hints, "_company_text_cache", {})
    cache_key = id(nomenclature_mapping)
    company_texts = cache.get(cache_key)
    if company_texts is None:
        company_texts = [
            normalize_company_context(
                " ".join(
                    [
                        norm_text(candidate.get("portaverse_company_name")) or "",
                        norm_text(candidate.get("portaverse_company_code")) or "",
                    ]
                )
            )
            for value in nomenclature_mapping.values()
            for candidate in mapping_candidate_pool(value)
        ]
        cache[cache_key] = company_texts
        setattr(filter_company_context_hints, "_company_text_cache", cache)
    return [
        hint
        for hint in hints
        if any(hint and hint in company_text for company_text in company_texts)
    ]


def strip_internal_mapping_fields(entry: dict[str, Any]) -> dict[str, Any]:
    return {key: value for key, value in entry.items() if not key.startswith("_")}


def refresh_config_from_mapping(
    config: PositionConfig,
    nomenclature_mapping: dict[str, dict[str, str | None]],
    normalized_mapping: dict[str, dict[str, str | None]] | None = None,
    strict_indexes: strict_position_mapping.LookupIndexes | None = None,
) -> None:
    if not nomenclature_mapping:
        return
    if is_neglect_scope(config):
        config.position_master_id = None
        config.position_nomenclature_id = None
        return
    if config.source_workbook and is_invalid_discovered_group_name(config.group_name, config.position_name, config.sheet_name):
        config.group_name = source_workbook_group_name(config.source_workbook) or config.group_name
    if strict_indexes is not None:
        if is_mapping_review_approved(config):
            if is_trusted_reviewer_manual_override(config):
                enforce_position_scope_ids(config)
                config.mapping_confidence_reason = (
                    config.mapping_confidence_reason
                    or "Reviewer manual override trusted as the mapping source."
                )
                return
            scope = normalize_position_scope(config.position_scope)
            if scope == "scope_uncertain":
                scope = None
            validation = strict_position_mapping.validate_manual_override(
                inferred_scope=scope or strict_position_mapping.infer_worksheet_scope(config.position_name).scope,
                position_master_id=config.position_master_id,
                position_nomenclature_id=config.position_nomenclature_id,
                indexes=strict_indexes,
            )
            if validation.allowed:
                apply_validated_override_candidate(config, validation)
            else:
                config.mapping_confidence_label = config.mapping_confidence_label or "mapping_conflict"
                config.mapping_confidence_reason = validation.reason
                config.position_master_id = None
                config.position_nomenclature_id = None
            return
        resolution = strict_position_mapping.resolve_mapping(
            worksheet=config.sheet_name,
            worksheet_title=config.position_name or config.sheet_name,
            group_name=config.group_name,
            source_workbook=config.source_workbook,
            indexes=strict_indexes,
        )
        apply_strict_resolution(config, resolution)
        return
    group_tail = config.group_name.removeprefix("Group ").removeprefix("Department ") if config.group_name else ""
    context_hints = filter_company_context_hints(
        nomenclature_mapping,
        source_workbook_context_hints(config.source_workbook),
    )
    lookup = lookup_position_scope_candidates(
        [
            config.position_name,
            config.sheet_name,
            *config.position_lookup_names,
            f"{config.position_name} {config.group_name}" if config.group_name else None,
            f"{config.sheet_name} {config.group_name}" if config.group_name else None,
            f"{config.position_name} {group_tail}" if group_tail else None,
            f"{config.sheet_name} {group_tail}" if group_tail else None,
        ],
        nomenclature_mapping,
        normalized_mapping,
        context_hints,
    )
    lookup = strip_internal_mapping_fields(lookup)
    has_manual_position_id = bool(config.position_master_id or config.position_nomenclature_id)
    if not lookup:
        if not has_manual_position_id:
            config.position_master_id = None
            config.position_nomenclature_id = None
            config.position_scope = None
            config.portaverse_position_title = None
            config.portaverse_group_name = None
            config.portaverse_company_name = None
            config.cluster_label = None
        enforce_position_scope_ids(config)
        return

    lookup_company_name = lookup.get("portaverse_company_name")
    has_stale_company_mapping = bool(
        config.portaverse_company_name
        and lookup_company_name
        and normalize_title(config.portaverse_company_name) != normalize_title(lookup_company_name)
    )
    lookup_scope = mapping_entry_scope(lookup)
    if not has_manual_position_id or has_stale_company_mapping:
        config.position_master_id = lookup.get("position_master_id")
        config.position_nomenclature_id = lookup.get("position_nomenclature_id")
        config.position_scope = lookup_scope or lookup.get("position_scope")
    elif not config.position_scope:
        config.position_scope = lookup_scope or lookup.get("position_scope")
    elif lookup_scope == "structural":
        config.position_scope = "structural"
        config.position_master_id = lookup.get("position_master_id") or config.position_master_id or config.position_nomenclature_id
        config.position_nomenclature_id = None
    elif lookup_scope == "non_structural" and not is_structural_scope(config):
        config.position_scope = "non_structural"
        config.position_nomenclature_id = lookup.get("position_nomenclature_id") or config.position_nomenclature_id
        config.position_master_id = None
    config.portaverse_position_title = lookup.get("portaverse_position_title") or config.portaverse_position_title
    config.portaverse_group_name = lookup.get("portaverse_group_name") or config.portaverse_group_name
    config.portaverse_company_name = lookup.get("portaverse_company_name") or config.portaverse_company_name
    config.cluster_label = lookup.get("cluster_label") or config.cluster_label
    enforce_position_scope_ids(config)


def refresh_configs_from_mapping(
    configs: list[PositionConfig],
    nomenclature_mapping: dict[str, dict[str, str | None]],
    strict_indexes: strict_position_mapping.LookupIndexes | None = None,
) -> None:
    if not nomenclature_mapping:
        return
    normalized_mapping = build_normalized_position_mapping(nomenclature_mapping)
    for config in configs:
        refresh_config_from_mapping(config, nomenclature_mapping, normalized_mapping, strict_indexes)


def build_normalized_position_mapping(
    nomenclature_mapping: dict[str, dict[str, str | None]],
) -> dict[str, dict[str, str | None]]:
    normalized_mapping: dict[str, dict[str, Any]] = {}
    for key, value in nomenclature_mapping.items():
        normalized_key = normalize_position_lookup(key)
        if not normalized_key:
            continue
        for candidate in mapping_candidate_pool(value):
            merge_mapping_entry(normalized_mapping, normalized_key, candidate)
    return normalized_mapping


def first_value_after_label(row: list[Any], label: str) -> str | None:
    normalized_label = normalize_title(label)
    for idx, value in enumerate(row):
        if normalize_title(norm_text(value)) != normalized_label:
            continue
        for next_value in row[idx + 1 :]:
            text = norm_text(next_value)
            if text:
                return text
    return None


def discover_sheet_metadata(rows: list[list[Any]]) -> tuple[str | None, str | None]:
    group_name: str | None = None
    position_name: str | None = None
    for row in rows[:12]:
        position_name = position_name or first_value_after_label(row, "Nama Posisi")
        group_name = group_name or first_value_after_label(row, "Posisi")
        if group_name and position_name:
            break
    return group_name, position_name


def should_skip_discovered_sheet(sheet_name: str, position_name: str | None) -> bool:
    text = normalize_title(f"{sheet_name} {position_name or ''}")
    skip_tokens = {
        "jadwal validator",
        "panduan",
        "kpi coverage",
        "kpi existing",
        "new kpi",
        "master data",
        "sheet",
        "manajemen investasi",
        "manajemen portfolio anper",
    }
    if any(token in text for token in skip_tokens):
        return True
    if normalize_title(position_name) in {"all", "jenis posisi", "company name", "kpi output"}:
        return True
    return False


def sheet_tab_color_value(worksheet: Any) -> str | None:
    color = getattr(getattr(worksheet, "sheet_properties", None), "tabColor", None)
    if not color:
        return None
    value = getattr(color, "rgb", None) or getattr(color, "indexed", None) or getattr(color, "theme", None)
    return str(value).upper() if value is not None else None


def is_yellow_tab_color(value: str | None) -> bool:
    if not value:
        return False
    return any(token in value.upper() for token in ["FFFF00", "FFC000", "FFD966", "FFFF99", "FFF2CC", "FFEB3B"])


def is_active_valid_sheet(worksheet: Any) -> bool:
    return getattr(worksheet, "sheet_state", "visible") == "visible" and is_yellow_tab_color(
        sheet_tab_color_value(worksheet)
    )


def active_valid_sheet_names_from_xlsx(workbook_path: Path) -> set[str]:
    workbook_ns = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
    office_rel_ns = "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}"
    package_rel_ns = "{http://schemas.openxmlformats.org/package/2006/relationships}"
    active_names: set[str] = set()
    with zipfile.ZipFile(workbook_path) as archive:
        workbook_root = ET.fromstring(archive.read("xl/workbook.xml"))
        rel_root = ET.fromstring(archive.read("xl/_rels/workbook.xml.rels"))
        rel_targets: dict[str, str] = {}
        for rel in rel_root.findall(f"{package_rel_ns}Relationship"):
            rel_id = rel.attrib.get("Id")
            target = rel.attrib.get("Target")
            if rel_id and target:
                rel_targets[rel_id] = target.lstrip("/")
        for sheet in workbook_root.findall(f"{workbook_ns}sheets/{workbook_ns}sheet"):
            if sheet.attrib.get("state", "visible") != "visible":
                continue
            sheet_name = sheet.attrib.get("name")
            rel_id = sheet.attrib.get(f"{office_rel_ns}id")
            target = rel_targets.get(rel_id or "")
            if not sheet_name or not target:
                continue
            sheet_path = target if target.startswith("xl/") else f"xl/{target}"
            try:
                sheet_root = ET.fromstring(archive.read(sheet_path))
            except KeyError:
                continue
            tab_color = sheet_root.find(f"{workbook_ns}sheetPr/{workbook_ns}tabColor")
            if tab_color is not None and is_yellow_tab_color(tab_color.attrib.get("rgb")):
                active_names.add(sheet_name)
    return active_names


def discover_configs_for_workbook(
    source_workbook: str,
    workbook_path: Path,
    nomenclature_mapping: dict[str, dict[str, str | None]],
    normalized_mapping: dict[str, dict[str, Any]] | None = None,
    strict_indexes: strict_position_mapping.LookupIndexes | None = None,
) -> list[PositionConfig]:
    active_sheet_names = active_valid_sheet_names_from_xlsx(workbook_path)
    require_yellow_tab = bool(active_sheet_names)
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    configs: list[PositionConfig] = []
    directorate_name = Path(source_workbook).stem.split(" - ")[0]
    if normalized_mapping is None:
        normalized_mapping = build_normalized_position_mapping(nomenclature_mapping)
    for worksheet in workbook.worksheets:
        if require_yellow_tab and worksheet.title not in active_sheet_names:
            continue
        if not require_yellow_tab and getattr(worksheet, "sheet_state", "visible") != "visible":
            continue
        rows = [list(row) for row in worksheet.iter_rows(values_only=True)]
        try:
            find_header_row(rows)
        except ValueError:
            continue
        group_name, position_name = discover_sheet_metadata(rows)
        position_name = position_name or worksheet.title
        source_group_name = source_workbook_group_name(source_workbook)
        if is_invalid_discovered_group_name(group_name, position_name, worksheet.title):
            group_name = source_group_name or ""
        else:
            group_name = group_name or ""
        if should_skip_discovered_sheet(worksheet.title, position_name):
            continue
        group_tail = group_name.removeprefix("Group ").removeprefix("Department ") if group_name else ""
        context_hints = filter_company_context_hints(
            nomenclature_mapping,
            source_workbook_context_hints(source_workbook),
        )
        config = PositionConfig(
            source_workbook=source_workbook,
            sheet_name=worksheet.title,
            position_name=position_name,
            group_name=group_name,
            directorate_name=directorate_name,
            position_lookup_names=[position_name, worksheet.title],
        )
        if strict_indexes is not None:
            resolution = strict_position_mapping.resolve_mapping(
                worksheet=worksheet.title,
                worksheet_title=position_name or worksheet.title,
                group_name=group_name,
                source_workbook=source_workbook,
                indexes=strict_indexes,
            )
            apply_strict_resolution(config, resolution)
        else:
            lookup = lookup_position_scope_candidates(
                [
                    position_name,
                    worksheet.title,
                    f"{position_name} {group_name}" if group_name else None,
                    f"{worksheet.title} {group_name}" if group_name else None,
                    f"{position_name} {group_tail}" if group_tail else None,
                    f"{worksheet.title} {group_tail}" if group_tail else None,
                ],
                nomenclature_mapping,
                normalized_mapping,
                context_hints,
            )
            lookup = strip_internal_mapping_fields(lookup)
            config.position_master_id = lookup.get("position_master_id")
            config.position_nomenclature_id = lookup.get("position_nomenclature_id")
            config.position_scope = lookup.get("position_scope") or "mapping_conflict"
            config.portaverse_position_title = lookup.get("portaverse_position_title")
            config.portaverse_group_name = lookup.get("portaverse_group_name")
            config.portaverse_company_name = lookup.get("portaverse_company_name")
            config.cluster_label = lookup.get("cluster_label")
        enforce_position_scope_ids(config)
        configs.append(config)
    return configs


def collect_parsed_sheets(
    source_path: Path,
    positions_dir: Path | None,
    configs: list[PositionConfig],
    issues: list[ValidationIssue],
) -> list[ParsedSheet]:
    master_index = PositionMasterIndex.load(positions_dir) if positions_dir else PositionMasterIndex()
    parsed_sheets: list[ParsedSheet] = []

    for config in configs:
        if is_neglect_scope(config):
            issues.append(
                ValidationIssue(
                    severity="info",
                    sheet_name=config.sheet_name,
                    source_row=None,
                    record_type="sheet",
                    title=config.position_name,
                    message="Position scope is neglect; sheet output was skipped.",
                )
            )
            continue
        if is_mapping_conflict_scope(config):
            issues.append(
                ValidationIssue(
                    severity="error",
                    sheet_name=config.sheet_name,
                    source_row=None,
                    record_type="sheet",
                    title=config.position_name,
                    message=(
                        "mapping_conflict: position identity could not be resolved "
                        "against the production organization reference; sheet output was skipped."
                    ),
                )
            )
            continue
        if is_blocked_mapping_confidence(config):
            label = config.mapping_confidence_label or "unknown"
            reason = config.mapping_confidence_reason or "Mapping requires reviewer approval before upload rows can be generated."
            needs_check = is_mapping_needs_check(config)
            message_suffix = (
                "sheet output was skipped for needs_check review."
                if needs_check
                else "sheet output was skipped."
            )
            issues.append(
                ValidationIssue(
                    severity="warning" if needs_check else "error",
                    sheet_name=config.sheet_name,
                    source_row=None,
                    record_type="sheet",
                    title=config.position_name,
                    message=f"{label}: {reason}; {message_suffix}",
                )
            )
            continue
        metadata = master_index.resolve(config)
        if metadata is None and not (config.position_master_id or config.position_nomenclature_id):
            issues.append(
                ValidationIssue(
                    severity="error",
                    sheet_name=config.sheet_name,
                    source_row=None,
                    record_type="sheet",
                    title=config.position_name,
                    message=(
                        "mapping_conflict: position identity could not be resolved "
                        "against the production organization reference; sheet output was skipped."
                    ),
                )
            )
            continue
        if metadata and config.position_master_id and metadata.position_master_id != str(config.position_master_id):
            issues.append(
                ValidationIssue(
                    severity="warning",
                    sheet_name=config.sheet_name,
                    source_row=None,
                    record_type="sheet",
                    title=config.position_name,
                    message=(
                        f"Config position_master_id={config.position_master_id} differs from master lookup "
                        f"{metadata.position_master_id}; using the config value as output override."
                    ),
                )
            )
        position_master_id = resolve_output_position_master_id(config, metadata)
        if not position_master_id:
            issues.append(
                ValidationIssue(
                    severity="error",
                    sheet_name=config.sheet_name,
                    source_row=None,
                    record_type="sheet",
                    title=config.position_name,
                    message=(
                        "mapping_conflict: position identity could not be resolved "
                        "against the production organization reference; sheet output was skipped."
                    ),
                )
            )
            continue

        try:
            sheet_rows = read_xlsx_sheet(source_path, config.sheet_name)
        except (KeyError, ValueError) as exc:
            issues.append(
                ValidationIssue(
                    severity="error",
                    sheet_name=config.sheet_name,
                    source_row=None,
                    record_type="sheet",
                    title=config.position_name,
                    message=str(exc),
                )
            )
            continue
        impacts = parse_block_sheet(sheet_rows, config, issues)
        raw_impact_count = len(impacts)
        impacts = enforce_default_impact_set(
            impacts,
            config.expected_impact_count,
            sheet_name=config.sheet_name,
            position_name=config.position_name,
            issues=issues,
        )
        if raw_impact_count != config.expected_impact_count:
            issues.append(
                ValidationIssue(
                    severity="warning",
                    sheet_name=config.sheet_name,
                    source_row=None,
                    record_type="sheet",
                    title=config.position_name,
                    message=(
                        f"Parsed {raw_impact_count} KPI Impact rows; enforced "
                        f"{len(impacts)}/{config.expected_impact_count} shared Pelindo impacts."
                    ),
                )
            )
        parsed_sheets.append(ParsedSheet(config=config, metadata=metadata, impacts=impacts))
    return parsed_sheets


def write_transformed_workbook(
    template_path: Path,
    output_path: Path,
    report_path: Path,
    parsed_sheets: list[ParsedSheet],
    issues: list[ValidationIssue],
) -> tuple[int, int, int, int]:
    output_rows: list[list[Any]] = []
    next_global_id = 1
    for parsed in parsed_sheets:
        position_master_id = resolve_output_position_master_id(parsed.config, parsed.metadata)
        if not position_master_id:
            issues.append(
                ValidationIssue(
                    severity="info",
                    sheet_name=parsed.config.sheet_name,
                    source_row=None,
                    record_type="sheet",
                    title=parsed.config.position_name,
                    message=(
                        "Unresolved position mapping; sheet output was skipped."
                    ),
                )
            )
            continue
        rows, next_global_id = build_upload_rows(
            parsed.config,
            position_master_id,
            parsed.impacts,
            next_global_id,
            issues,
        )
        validate_output_rows(parsed.config, rows, issues)
        append_weight_audit_issues(parsed, rows, issues)
        output_rows.extend(rows)

    write_output_workbook(template_path, output_path, output_rows)
    write_report(report_path, issues)

    errors = sum(1 for issue in issues if issue.severity == "error")
    warnings = sum(1 for issue in issues if issue.severity == "warning")
    infos = sum(1 for issue in issues if issue.severity == "info")
    print(f"Wrote workbook: {output_path}")
    print(f"Wrote report: {report_path}")
    print(f"Generated rows: {len(output_rows)}")
    print(f"Issues: errors={errors} warnings={warnings} info={infos}")
    return len(output_rows), errors, warnings, infos


def transform_workbook(
    source_path: Path,
    template_path: Path,
    positions_dir: Path | None,
    configs: list[PositionConfig],
    output_path: Path,
    report_path: Path,
) -> tuple[int, int, int, int]:
    issues: list[ValidationIssue] = []
    parsed_sheets = collect_parsed_sheets(source_path, positions_dir, configs, issues)
    backfill_shared_impact_fields(parsed_sheets)
    return write_transformed_workbook(template_path, output_path, report_path, parsed_sheets, issues)


def should_skip_source_workbook(source_workbook: str) -> bool:
    normalized = normalize_title(source_workbook)
    return "as is" in normalized and "group hukum" in normalized


def extract_zip_workbooks(source_zip: Path, destination: Path) -> dict[str, Path]:
    workbooks: dict[str, Path] = {}
    with zipfile.ZipFile(source_zip) as archive:
        for name in archive.namelist():
            lower = name.lower()
            if not (lower.endswith(".xlsx") or lower.endswith(".xlsm")):
                continue
            if Path(name).name.startswith("~$"):
                continue
            if should_skip_source_workbook(name):
                continue
            target = destination / safe_path_stem(name)
            # Keep original suffix so OOXML readers still open .xlsm payloads.
            target = target.with_suffix(Path(name).suffix.lower())
            target.write_bytes(archive.read(name))
            workbooks[name] = target
    return workbooks


def run_zip_batch(args: argparse.Namespace) -> int:
    if not args.output_dir:
        raise SystemExit("--output-dir is required when --source is a .zip")
    nomenclature_mapping, strict_indexes = load_position_reference(
        args.mapping,
        args.target_company_id,
        args.nomenclature_summary,
    )
    with tempfile.TemporaryDirectory() as tmp:
        extracted = extract_zip_workbooks(args.source, Path(tmp))
        configs = load_config(args.config) if args.config else []
        if not configs:
            normalized_mapping = build_normalized_position_mapping(nomenclature_mapping)
            total_workbooks = len(extracted)
            for index, (source_workbook, workbook_path) in enumerate(extracted.items(), start=1):
                configs.extend(
                    discover_configs_for_workbook(
                        source_workbook,
                        workbook_path,
                        nomenclature_mapping,
                        normalized_mapping,
                        strict_indexes,
                    )
                )
                if index == 1 or index % 25 == 0 or index == total_workbooks:
                    print(
                        f"Discovered configs: {index}/{total_workbooks} workbooks; "
                        f"positions={len(configs)}",
                        flush=True,
                    )
        else:
            refresh_configs_from_mapping(configs, nomenclature_mapping, strict_indexes)
        if args.only_sheet:
            selected = set(args.only_sheet)
            configs = [config for config in configs if config.sheet_name in selected]
        if args.best_effort_mapping:
            apply_best_effort_mapping(configs)
        if args.write_discovered_config:
            write_discovered_config(args.write_discovered_config, configs, args.mapping)

        configs_by_workbook: dict[str, list[PositionConfig]] = {}
        for config in configs:
            if config.source_workbook:
                configs_by_workbook.setdefault(config.source_workbook, []).append(config)

        total_errors = 0
        total_rows = 0
        generated_at = datetime.now()
        batch_items: list[tuple[str, list[ParsedSheet], list[ValidationIssue], Path]] = []
        seen_output_names: dict[str, int] = {}
        for source_workbook, workbook_path in extracted.items():
            workbook_configs = configs_by_workbook.get(source_workbook, [])
            if not workbook_configs:
                continue
            output_name = unique_output_name(
                conversion_output_name(source_workbook, workbook_configs, generated_at),
                seen_output_names,
            )
            output_dir = args.output_dir / output_name
            workbook_issues: list[ValidationIssue] = []
            parsed_sheets = collect_parsed_sheets(workbook_path, args.positions_dir, workbook_configs, workbook_issues)
            batch_items.append((output_name, parsed_sheets, workbook_issues, output_dir))

        backfill_shared_impact_fields(
            [parsed_sheet for _, parsed_sheets, _, _ in batch_items for parsed_sheet in parsed_sheets]
        )

        for output_name, parsed_sheets, workbook_issues, output_dir in batch_items:
            rows, errors, _, _ = write_transformed_workbook(
                args.template,
                output_dir / f"{output_name}.xlsx",
                output_dir / f"{output_name}.report.csv",
                parsed_sheets,
                workbook_issues,
            )
            total_rows += rows
            total_errors += errors
        print(f"ZIP batch generated rows: {total_rows}")
        print(f"ZIP batch errors: {total_errors}")
        return 1 if total_errors else 0


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--source", required=True, type=Path, help="Source KPI workbook (.xlsx) or .zip")
    parser.add_argument("--template", required=True, type=Path, help="Official upload template (.xlsx)")
    parser.add_argument("--positions-dir", type=Path, help="Directory with data_master_posisi xlsx exports")
    parser.add_argument("--config", type=Path, help="JSON config describing sheets to export")
    parser.add_argument(
        "--mapping",
        type=Path,
        help="Offline position reference JSON, e.g. configs/production_position_reference.json",
    )
    parser.add_argument(
        "--target-company-id",
        default=TARGET_COMPANY_ID_DEFAULT,
        help=(
            "Restrict offline position reference lookup to one company_id. "
            f"Default {TARGET_COMPANY_ID_DEFAULT} ({TARGET_COMPANY_NAME_DEFAULT})."
        ),
    )
    parser.add_argument(
        "--nomenclature-summary",
        type=Path,
        help="Workbook containing the newer nomenclature summary, e.g. Unit Kerja Pelindo per April 2026 - edit.xlsx.",
    )
    parser.add_argument("--output", type=Path, help="Output upload workbook (.xlsx), for single-workbook mode")
    parser.add_argument("--report", type=Path, help="Validation report (.csv), for single-workbook mode")
    parser.add_argument("--output-dir", type=Path, help="Output directory, required for ZIP batch mode")
    parser.add_argument("--write-discovered-config", type=Path, help="Write auto-discovered ZIP config to this JSON path")
    parser.add_argument(
        "--only-sheet",
        action="append",
        default=[],
        help="Limit export to a specific sheet name. May be passed multiple times.",
    )
    parser.add_argument(
        "--best-effort-mapping",
        action="store_true",
        help=(
            "Promote the best active candidate for low/conflict mappings and skip still-unmapped "
            "worksheets so the generated upload files contain only usable PMID/PNID rows."
        ),
    )
    args = parser.parse_args()

    if args.source.suffix.lower() == ".zip":
        return run_zip_batch(args)
    if not args.config or not args.output or not args.report:
        raise SystemExit("--config, --output, and --report are required for single-workbook mode")

    configs = load_config(args.config)
    nomenclature_mapping, strict_indexes = load_position_reference(
        args.mapping,
        args.target_company_id,
        args.nomenclature_summary,
    )
    refresh_configs_from_mapping(configs, nomenclature_mapping, strict_indexes)
    if args.only_sheet:
        selected = set(args.only_sheet)
        configs = [cfg for cfg in configs if cfg.sheet_name in selected]

    _, errors, _, _ = transform_workbook(
        args.source,
        args.template,
        args.positions_dir,
        configs,
        args.output,
        args.report,
    )
    return 1 if errors else 0


if __name__ == "__main__":
    raise SystemExit(main())
