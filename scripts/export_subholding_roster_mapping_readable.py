#!/usr/bin/env python3
"""Export Subholding roster position-first mapping in a clearer spreadsheet layout."""

from __future__ import annotations

import argparse
import json
from collections import Counter
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo

NAVY = "173651"
TEAL = "138074"
PALE = "E9F1F8"
GREEN = "D9EAD3"
YELLOW = "FFF2CC"
ORANGE = "FCE4D6"
GRAY = "E7E6E6"
RED = "F4CCCC"
BODY = "Aptos"

STATUS_ID = {
    "NEEDS_REVIEW_NEW_56": "BARU — perlu review",
    "ROSTER_COVERED_REVIEWED": "Sudah ada usulan dari R&W",
    "ROSTER_COVERED_PENDING": "Menunggu keputusan",
}

RESOLVE_ID = {
    "resolved_exact": "File + sheet cocok exact ke inventory",
    "workbook_resolved_sheet_exact": "File inventory ditemukan; sheet cocok exact",
    "workbook_resolved_sheet_fuzzy": "File inventory ditemukan; sheet cocok mirip (fuzzy)",
    "workbook_resolved_sheet_reference": "File inventory ditemukan; nama sheet masih dari R&W (belum exact)",
    "accepted_high_confidence_candidate": "R&W #N/A → memakai draft otomatis (high confidence)",
    "rw_path_as_reference": "Belum ketemu di inventory; nilai R&W dipakai apa adanya",
    "new_roster_worker": "Pekerja baru dari roster (belum ada di review R&W)",
    "no_prior_review_row": "Tidak ada baris review R&W untuk posisi ini",
    "review_blank": "Review R&W kosong",
    "not_reviewed": "Belum direview",
}

MAIN_COLUMNS = [
    ("No.", "Nomor baris"),
    ("Status Baris", "Status ringkas untuk reviewer"),
    ("Keputusan Mapping", "YES = setuju usulan | NEEDS_CHECK = cek dulu | NO = tolak"),
    ("PMID", "Position Master ID (struktural). Kosong jika non-struktural."),
    ("PNID", "Position Nomenclature / cluster ID (non-struktural). Kosong jika struktural."),
    ("Judul Posisi", "Nama jabatan di production"),
    ("Perusahaan", "Nama company production"),
    ("Unit / Group", "Unit organisasi posisi"),
    ("Jumlah Pegawai", "Jumlah NIPP di baris ini (bukan unique lintas baris)"),
    ("NIPP Pegawai", "Daftar NIPP; dipisah ;"),
    ("Nama Pegawai", "Daftar nama; dipisah ;"),
    ("File Kamus (dipakai)", "Path workbook Kamus yang dipakai (utamakan path inventory/config)"),
    ("Sheet Kamus (dipakai)", "Nama worksheet Kamus yang dipakai"),
    ("File Kamus (referensi R&W)", "Workbook yang ditulis Red & White — hanya referensi"),
    ("Sheet Kamus (referensi R&W)", "Worksheet yang ditulis Red & White — hanya referensi"),
    ("Folder Referensi R&W", "Folder path dari review R&W — hanya referensi"),
    ("Status Pencocokan", "Apakah path/sheet sudah dipetakan ke inventory"),
    ("Asal Usulan", "Dari mana usulan file/sheet berasal"),
    ("Confidence Draft", "Label otomatis sebelum/saat review"),
    ("Alasan Confidence", "Penjelasan singkat draft otomatis"),
    ("Draft Otomatis File", "Kandidat resolver otomatis (jika ada)"),
    ("Draft Otomatis Sheet", "Sheet kandidat resolver otomatis (jika ada)"),
    ("Roster Subholding", "Sheet asal di file REGIONAL dan SUBHOLDING (SPTP/SPMT/SPSL/SPJM)"),
    ("Jabatan di Roster", "STEXT_STO dari roster"),
    ("Catatan Reviewer", "Catatan bebas reviewer"),
]


def norm(value: Any) -> str:
    return "" if value is None else str(value).strip()


def map_status(tag: str) -> str:
    primary = (tag or "").split(";")[0]
    return STATUS_ID.get(primary, primary or "(kosong)")


def map_resolve(status: str) -> str:
    return RESOLVE_ID.get(status, status or "(kosong)")


def map_source(source: str) -> str:
    return {
        "rw_reference": "Review R&W (path sebagai referensi)",
        "rw_accept_candidate": "R&W menerima draft otomatis",
        "rw_reviewed": "Review R&W",
        "new_56_resolver": "Pekerja baru + resolver",
        "new_56_stub": "Pekerja baru tanpa identity kuat",
        "resolver_only": "Hanya draft otomatis",
        "unreviewed": "Belum ada keputusan",
    }.get(source, source or "(kosong)")


def to_readable_row(row: dict[str, Any], index: int) -> dict[str, Any]:
    return {
        "No.": index,
        "Status Baris": map_status(norm(row.get("Roster Review Tag"))),
        "Keputusan Mapping": norm(row.get("Reviewer Confirm Mapping")),
        "PMID": norm(row.get("PMID")),
        "PNID": norm(row.get("PNID")),
        "Judul Posisi": norm(row.get("Position Title")),
        "Perusahaan": norm(row.get("Company")),
        "Unit / Group": norm(row.get("Group / Unit")),
        "Jumlah Pegawai": row.get("Active Employees") or 0,
        "NIPP Pegawai": norm(row.get("Active Employee NIPPs")),
        "Nama Pegawai": norm(row.get("Active Employee Names")),
        "File Kamus (dipakai)": norm(row.get("Reviewer Source Workbook")),
        "Sheet Kamus (dipakai)": norm(row.get("Reviewer Worksheet")),
        "File Kamus (referensi R&W)": norm(row.get("Reviewed Workbook Title")),
        "Sheet Kamus (referensi R&W)": norm(row.get("Reviewed Worksheet Title")),
        "Folder Referensi R&W": norm(row.get("Reviewed Folder")),
        "Status Pencocokan": map_resolve(norm(row.get("Inventory Resolve Status"))),
        "Asal Usulan": map_source(norm(row.get("Mapping Source"))),
        "Confidence Draft": norm(row.get("Confidence Label")),
        "Alasan Confidence": norm(row.get("Confidence Reason")),
        "Draft Otomatis File": norm(row.get("Candidate Source Workbook")),
        "Draft Otomatis Sheet": norm(row.get("Candidate Worksheet")),
        "Roster Subholding": norm(row.get("Roster Sheet")),
        "Jabatan di Roster": norm(row.get("Roster Job Title")),
        "Catatan Reviewer": norm(row.get("Reviewer Notes")),
        "_tag": norm(row.get("Roster Review Tag")),
        "_confidence": norm(row.get("Confidence Label")),
    }


def style_header(cell, fill: str = TEAL) -> None:
    cell.fill = PatternFill("solid", fgColor=fill)
    cell.font = Font(name=BODY, bold=True, color="FFFFFF", size=10)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def write_guide(ws) -> None:
    ws.sheet_view.showGridLines = False
    ws["A1"] = "Baca Dulu — Cara memahami file pemetaan ini"
    ws["A1"].font = Font(name=BODY, size=18, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor=NAVY)
    ws.merge_cells("A1:C1")
    ws.row_dimensions[1].height = 32

    blocks = [
        (
            "Apa isi file ini?",
            "Pemetaan posisi pekerja Subholding (roster 2.705 NIPP) ke worksheet Kamus KPI. "
            "Satu baris = satu posisi (PMID atau PNID) beserta pegawai di posisi itu.",
        ),
        (
            "Sheet mana yang perlu dilihat?",
            "1) Ringkasan — angka & metadata\n"
            "2) Pemetaan — tabel utama untuk review\n"
            "3) Perlu Review (56) — pekerja roster yang belum ada di review Red & White\n"
            "4) Antrian Review — baris yang masih perlu perhatian\n"
            "5) Glosarium Kolom — arti setiap kolom",
        ),
        (
            "Bedanya 'dipakai' vs 'referensi R&W'",
            "• File/Sheet Kamus (dipakai) = yang akan dipakai lanjut (path inventory/config jika ketemu).\n"
            "• File/Sheet/Folder (referensi R&W) = apa yang ditulis Red & White; disimpan sebagai jejak, "
            "karena path mereka sering beda dengan nama file di config.",
        ),
        (
            "Cara review cepat",
            "1. Filter Status Baris = 'BARU — perlu review' atau buka sheet Perlu Review (56).\n"
            "2. Cek File Kamus (dipakai) + Sheet Kamus (dipakai).\n"
            "3. Isi Keputusan Mapping: YES / NEEDS_CHECK / NO.\n"
            "4. Jika salah, tulis koreksi di Catatan Reviewer (atau ubah File/Sheet dipakai).",
        ),
        (
            "Catatan penting soal jumlah pegawai",
            "Unique pekerja di seluruh file = 2.705 NIPP.\n"
            "Kolom Jumlah Pegawai per baris bisa dijumlahkan > 2.705 jika ada orang menjabat >1 posisi.",
        ),
    ]
    row = 3
    for title, body in blocks:
        ws.cell(row, 1, title).font = Font(name=BODY, bold=True, color=NAVY, size=12)
        ws.cell(row, 1).fill = PatternFill("solid", fgColor=PALE)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        row += 1
        ws.cell(row, 1, body).alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        ws.row_dimensions[row].height = 70
        row += 2
    ws.column_dimensions["A"].width = 100
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 20


def write_glossary(ws) -> None:
    ws.sheet_view.showGridLines = False
    ws["A1"] = "Glosarium Kolom — sheet Pemetaan"
    ws["A1"].font = Font(name=BODY, size=16, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor=NAVY)
    ws.merge_cells("A1:B1")
    style_header(ws.cell(3, 1, "Nama Kolom"))
    style_header(ws.cell(3, 2, "Arti / Cara pakai"))
    for idx, (name, desc) in enumerate(MAIN_COLUMNS, start=4):
        ws.cell(idx, 1, name).font = Font(name=BODY, bold=True, color=NAVY)
        ws.cell(idx, 2, desc).alignment = Alignment(wrap_text=True)
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 90


def write_summary(ws, meta: dict[str, Any], rows: list[dict[str, Any]]) -> None:
    ws.sheet_view.showGridLines = False
    ws["A1"] = meta.get("title") or "Position First Mapping Subholding"
    ws["A1"].font = Font(name=BODY, size=18, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor=NAVY)
    ws.merge_cells("A1:B1")
    ws.row_dimensions[1].height = 34
    ws["A2"] = (
        "Format ringkas untuk review. Path Red & White hanya referensi; "
        "kolom 'dipakai' mengutamakan path inventory/config."
    )
    ws["A2"].fill = PatternFill("solid", fgColor=PALE)
    ws.merge_cells("A2:B2")

    status_counts = Counter(r["Status Baris"] for r in rows)
    resolve_counts = Counter(r["Status Pencocokan"] for r in rows)
    conf_counts = Counter(r["Confidence Draft"] for r in rows)
    sources = meta.get("sources") or {}

    items = [
        ("Tanggal generate", meta.get("generated_at") or datetime.now().astimezone().isoformat(timespec="seconds")),
        ("Unique pekerja (NIPP)", meta.get("unique_active_employees")),
        ("Jumlah baris posisi", meta.get("position_row_count") or len(rows)),
        ("Keputusan YES", meta.get("reviewer_yes_count")),
        ("Baris BARU perlu review", status_counts.get("BARU — perlu review", 0)),
        ("Baris sudah ada usulan R&W", status_counts.get("Sudah ada usulan dari R&W", 0)),
        ("Roster sumber", (sources.get("roster") or {}).get("path")),
        ("File review R&W", (sources.get("reviewed_mapping") or {}).get("path")),
        ("Inventory worksheet", (sources.get("inventory") or {}).get("path")),
        ("Production reference", (sources.get("production_reference") or {}).get("path")),
        ("Production exported_at", (sources.get("production_reference") or {}).get("exported_at")),
        ("Kebijakan path R&W", meta.get("rw_path_policy")),
    ]
    style_header(ws.cell(4, 1, "Item"))
    style_header(ws.cell(4, 2, "Nilai"))
    for offset, (label, value) in enumerate(items, start=5):
        ws.cell(offset, 1, label).fill = PatternFill("solid", fgColor="F3F6F8")
        ws.cell(offset, 1).font = Font(name=BODY, bold=True, color=NAVY)
        ws.cell(offset, 2, value).alignment = Alignment(wrap_text=True)

    start = 5 + len(items) + 2
    ws.cell(start, 1, "Status Baris").font = Font(name=BODY, bold=True, color="FFFFFF")
    ws.cell(start, 2, "Jumlah").font = Font(name=BODY, bold=True, color="FFFFFF")
    style_header(ws.cell(start, 1))
    style_header(ws.cell(start, 2))
    for offset, (label, count) in enumerate(sorted(status_counts.items()), start=start + 1):
        ws.cell(offset, 1, label)
        ws.cell(offset, 2, count)

    start2 = start + len(status_counts) + 3
    ws.cell(start2, 1, "Status Pencocokan").font = Font(name=BODY, bold=True, color="FFFFFF")
    ws.cell(start2, 2, "Jumlah").font = Font(name=BODY, bold=True, color="FFFFFF")
    style_header(ws.cell(start2, 1))
    style_header(ws.cell(start2, 2))
    for offset, (label, count) in enumerate(sorted(resolve_counts.items(), key=lambda x: -x[1]), start=start2 + 1):
        ws.cell(offset, 1, label)
        ws.cell(offset, 2, count)

    start3 = start2 + len(resolve_counts) + 3
    ws.cell(start3, 1, "Confidence Draft").font = Font(name=BODY, bold=True, color="FFFFFF")
    ws.cell(start3, 2, "Jumlah").font = Font(name=BODY, bold=True, color="FFFFFF")
    style_header(ws.cell(start3, 1))
    style_header(ws.cell(start3, 2))
    for offset, (label, count) in enumerate(sorted(conf_counts.items()), start=start3 + 1):
        ws.cell(offset, 1, label)
        ws.cell(offset, 2, count)

    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 100


def write_table(ws, title: str, rows: list[dict[str, Any]], table_name: str) -> None:
    headers = [name for name, _ in MAIN_COLUMNS]
    ws.sheet_view.showGridLines = False
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws.cell(1, 1, title).fill = PatternFill("solid", fgColor=NAVY)
    ws.cell(1, 1).font = Font(name=BODY, size=14, bold=True, color="FFFFFF")
    ws.row_dimensions[1].height = 28

    # Group color bands in header row 2 as section hints
    ws.cell(2, 1, "Identitas posisi & pegawai | Usulan Kamus yang dipakai | Referensi R&W | Status & draft")
    ws.cell(2, 1).fill = PatternFill("solid", fgColor=PALE)
    ws.cell(2, 1).font = Font(name=BODY, italic=True, color="37556E", size=9)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))

    for col, header in enumerate(headers, start=1):
        style_header(ws.cell(3, col, header))

    nipp_col = headers.index("NIPP Pegawai") + 1
    status_col = headers.index("Status Baris") + 1
    conf_col = headers.index("Confidence Draft") + 1
    for r_idx, row in enumerate(rows, start=4):
        for c_idx, header in enumerate(headers, start=1):
            value = row.get(header, "")
            cell = ws.cell(r_idx, c_idx, value)
            cell.font = Font(name=BODY, size=9, color="263746")
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if c_idx == nipp_col:
                cell.number_format = "@"
        status = row.get("Status Baris")
        if status == "BARU — perlu review":
            ws.cell(r_idx, status_col).fill = PatternFill("solid", fgColor=ORANGE)
        elif status == "Sudah ada usulan dari R&W":
            ws.cell(r_idx, status_col).fill = PatternFill("solid", fgColor=GREEN)
        elif status == "Menunggu keputusan":
            ws.cell(r_idx, status_col).fill = PatternFill("solid", fgColor=YELLOW)
        conf = row.get("Confidence Draft")
        conf_fills = {
            "high_confidence": GREEN,
            "low_confidence": YELLOW,
            "mapping_conflict": RED,
            "no_candidate": GRAY,
        }
        if conf in conf_fills:
            ws.cell(r_idx, conf_col).fill = PatternFill("solid", fgColor=conf_fills[conf])
        ws.row_dimensions[r_idx].height = 20

    end_row = 3 + max(len(rows), 1)
    if rows:
        table = Table(displayName=table_name, ref=f"A3:{get_column_letter(len(headers))}{end_row}")
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
        ws.add_table(table)
        decide_col = get_column_letter(headers.index("Keputusan Mapping") + 1)
        dv = DataValidation(type="list", formula1='"YES,NEEDS_CHECK,NO"', allow_blank=True)
        ws.add_data_validation(dv)
        dv.add(f"{decide_col}4:{decide_col}{end_row}")

    widths = {
        "No.": 6,
        "Status Baris": 22,
        "Keputusan Mapping": 14,
        "PMID": 10,
        "PNID": 10,
        "Judul Posisi": 32,
        "Perusahaan": 28,
        "Unit / Group": 24,
        "Jumlah Pegawai": 10,
        "NIPP Pegawai": 24,
        "Nama Pegawai": 28,
        "File Kamus (dipakai)": 48,
        "Sheet Kamus (dipakai)": 28,
        "File Kamus (referensi R&W)": 28,
        "Sheet Kamus (referensi R&W)": 26,
        "Folder Referensi R&W": 40,
        "Status Pencocokan": 36,
        "Asal Usulan": 28,
        "Confidence Draft": 16,
        "Alasan Confidence": 36,
        "Draft Otomatis File": 40,
        "Draft Otomatis Sheet": 22,
        "Roster Subholding": 12,
        "Jabatan di Roster": 30,
        "Catatan Reviewer": 28,
    }
    for idx, header in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = widths.get(header, 18)
    ws.freeze_panes = "F4"


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--draft",
        type=Path,
        default=Path(
            "outputs/kamus-group2-subholding-roster-mapping-20260806/"
            "subholding_roster_position_first_mapping_20260806.json"
        ),
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=Path(
            "outputs/kamus-group2-subholding-roster-mapping-20260806/"
            "Position_First_Mapping_Subholding_20260806.xlsx"
        ),
    )
    args = parser.parse_args()

    payload = json.loads(args.draft.read_text(encoding="utf-8"))
    meta = payload.get("metadata") or {}
    readable = [to_readable_row(row, idx) for idx, row in enumerate(payload.get("rows", []), start=1)]
    needs_56 = [row for row in readable if row["Status Baris"] == "BARU — perlu review"]
    queue = [
        row
        for row in readable
        if row["Status Baris"] == "BARU — perlu review"
        or row["Keputusan Mapping"] != "YES"
        or "belum exact" in row["Status Pencocokan"]
        or "Belum ketemu" in row["Status Pencocokan"]
    ]

    meta = dict(meta)
    meta["title"] = "Position First Mapping Subholding — 2026-08-06"
    meta["generated_at"] = datetime.now().astimezone().isoformat(timespec="seconds")
    meta["position_row_count"] = len(readable)
    meta["reviewer_yes_count"] = sum(1 for row in readable if row["Keputusan Mapping"] == "YES")

    wb = Workbook()
    guide = wb.active
    guide.title = "Baca Dulu"
    write_guide(guide)
    write_summary(wb.create_sheet("Ringkasan"), meta, readable)
    write_glossary(wb.create_sheet("Glosarium Kolom"))
    write_table(
        wb.create_sheet("Pemetaan"),
        "Pemetaan posisi roster Subholding (2.705 NIPP) → worksheet Kamus KPI",
        readable,
        "PemetaanTable",
    )
    write_table(
        wb.create_sheet("Perlu Review (56)"),
        "56 pekerja roster yang belum ada di review Red & White — wajib direview",
        needs_56,
        "PerluReview56Table",
    )
    write_table(
        wb.create_sheet("Antrian Review"),
        "Antrian: baru / belum YES / sheet masih referensi R&W",
        queue,
        "AntrianReviewTable",
    )

    args.output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(args.output)

    # Side-by-side legend markdown
    legend = args.output.with_name("CARA_BACA_MAPPING.md")
    legend.write_text(
        "\n".join(
            [
                f"# Cara baca — {meta['title']}",
                "",
                f"File: `{args.output}`",
                "",
                "## Sheet",
                "- **Baca Dulu** — panduan singkat",
                "- **Ringkasan** — angka & sumber data",
                "- **Glosarium Kolom** — arti tiap kolom",
                "- **Pemetaan** — tabel utama",
                "- **Perlu Review (56)** — pekerja baru",
                "- **Antrian Review** — yang masih perlu dicek",
                "",
                "## Kolom kunci",
                "| Kolom | Arti |",
                "| --- | --- |",
                *[f"| {name} | {desc} |" for name, desc in MAIN_COLUMNS[:16]],
                "",
            ]
        ),
        encoding="utf-8",
    )
    print(
        json.dumps(
            {
                "output": str(args.output),
                "rows": len(readable),
                "needs_56": len(needs_56),
                "queue": len(queue),
                "legend": str(legend),
            },
            ensure_ascii=False,
            indent=2,
        )
    )


if __name__ == "__main__":
    main()
