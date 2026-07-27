#!/usr/bin/env python3
"""Apply reviewed Group 1 position mapping decisions to a discovered config."""

from __future__ import annotations

import argparse
import json
from collections import Counter
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


TRUST_SOURCE_REVIEWER_MANUAL = "reviewer_manual"


def norm(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def key_for_config(row: dict[str, Any]) -> tuple[str, str]:
    return norm(row.get("source_workbook")), norm(row.get("sheet_name"))


def key_for_review(row: dict[str, Any]) -> tuple[str, str]:
    return norm(row.get("Source Workbook")), norm(row.get("Worksheet") or row.get("Sheet"))


def load_review_rows(review_path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(review_path, read_only=True, data_only=True)
    if "Position Mapping Report" not in workbook.sheetnames:
        raise ValueError(f"{review_path} does not contain Position Mapping Report")
    worksheet = workbook["Position Mapping Report"]
    rows = worksheet.iter_rows(values_only=True)
    headers = [norm(value) for value in next(rows)]
    output: list[dict[str, Any]] = []
    for row in rows:
        item = dict(zip(headers, row))
        if key_for_review(item) == ("", "") and not norm(item.get("Reviewer Confirm Mapping")):
            continue
        output.append(item)
    return output


def indexed_review_rows(rows: list[dict[str, Any]]) -> tuple[dict[tuple[str, str], dict[str, Any]], list[str]]:
    indexed: dict[tuple[str, str], dict[str, Any]] = {}
    errors: list[str] = []
    seen: dict[tuple[str, str], tuple[str, str, str, str, str]] = {}
    for row in rows:
        key = key_for_review(row)
        status = norm(row.get("Reviewer Confirm Mapping")).upper()
        if not key[0] or not key[1] or not status:
            continue
        signature = (
            status,
            norm(row.get("Reviewer Actual PMID")),
            norm(row.get("Reviewer Actual PNID")),
            norm(row.get("Candidate PMID")),
            norm(row.get("Candidate PNID")),
        )
        if key in seen and seen[key] != signature:
            errors.append(f"duplicate review row for {key}")
            continue
        seen[key] = signature
        indexed[key] = row
    return indexed, errors


def apply_yes_review(position: dict[str, Any], review: dict[str, Any], stats: Counter[str]) -> list[str]:
    errors: list[str] = []
    pmid = norm(review.get("Reviewer Actual PMID"))
    pnid = norm(review.get("Reviewer Actual PNID"))
    if pmid and pnid:
        return [f"review row has both Reviewer Actual PMID and PNID: {key_for_review(review)}"]

    position["mapping_review_status"] = "approved"
    position["mapping_override_approved"] = True
    if pmid:
        position["position_scope"] = "structural"
        position["position_master_id"] = pmid
        position["position_nomenclature_id"] = None
        position["candidate_position_master_id"] = pmid
        position["candidate_position_nomenclature_id"] = None
        position["mapping_override_trust_source"] = TRUST_SOURCE_REVIEWER_MANUAL
        stats["manual_override_rows"] += 1
    elif pnid:
        position["position_scope"] = "non_structural"
        position["position_master_id"] = None
        position["position_nomenclature_id"] = pnid
        position["candidate_position_master_id"] = None
        position["candidate_position_nomenclature_id"] = pnid
        position["mapping_override_trust_source"] = TRUST_SOURCE_REVIEWER_MANUAL
        stats["manual_override_rows"] += 1
    else:
        scope = norm(position.get("position_scope"))
        candidate_pmid = norm(position.get("candidate_position_master_id"))
        candidate_pnid = norm(position.get("candidate_position_nomenclature_id"))
        if scope == "structural" and candidate_pmid:
            position["position_master_id"] = candidate_pmid
            position["position_nomenclature_id"] = None
            stats["candidate_approval_rows"] += 1
        elif scope == "non_structural" and candidate_pnid:
            position["position_master_id"] = None
            position["position_nomenclature_id"] = candidate_pnid
            stats["candidate_approval_rows"] += 1
        position.pop("mapping_override_trust_source", None)
    return errors


def apply_needs_check_review(position: dict[str, Any]) -> None:
    position["mapping_review_status"] = "needs_check"
    position["mapping_override_approved"] = False
    position.pop("mapping_override_trust_source", None)


def apply_review_to_config(config_path: Path, review_path: Path, output_path: Path, dry_run: bool = False) -> dict[str, int]:
    config = json.loads(config_path.read_text(encoding="utf-8"))
    review_rows, errors = indexed_review_rows(load_review_rows(review_path))
    config_keys = {key_for_config(position) for position in config.get("positions", [])}
    stats: Counter[str] = Counter()

    for position in config.get("positions", []):
        key = key_for_config(position)
        review = review_rows.get(key)
        if not review:
            continue
        status = norm(review.get("Reviewer Confirm Mapping")).upper()
        if status == "YES":
            stats["review_yes_rows"] += 1
            errors.extend(apply_yes_review(position, review, stats))
        elif status == "NEEDS_CHECK":
            stats["review_needs_check_rows"] += 1
            apply_needs_check_review(position)
        elif status:
            errors.append(f"unsupported review status {status}: {key}")

    unmatched = sorted(key for key in review_rows if key not in config_keys)
    errors.extend(f"review row does not match config: {key}" for key in unmatched)
    if errors:
        raise ValueError("\n".join(errors))

    if not dry_run:
        output_path.parent.mkdir(parents=True, exist_ok=True)
        output_path.write_text(json.dumps(config, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    return dict(stats)


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--config", required=True, type=Path)
    parser.add_argument("--review", required=True, type=Path)
    parser.add_argument("--output", required=True, type=Path)
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()

    stats = apply_review_to_config(args.config, args.review, args.output, args.dry_run)
    for key in sorted(stats):
        print(f"{key}={stats[key]}")
    if args.dry_run:
        print("dry_run=true")
    else:
        print(f"output={args.output}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
