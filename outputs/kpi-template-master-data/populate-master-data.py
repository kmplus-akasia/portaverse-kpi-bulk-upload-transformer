import json
from pathlib import Path

import openpyxl


ROOT = Path("/Users/alfredoteja/Documents/portaverse-kpi-bulk-upload-transformer")
INPUT_WORKBOOK = Path("/Users/alfredoteja/Downloads/KPI_Upload_Template_Portaverse_2026.xlsx")
REFERENCE_PATH = ROOT / "configs/production_position_reference.json"
OUTPUT_DIR = ROOT / "outputs/kpi-template-master-data"
OUTPUT_WORKBOOK = OUTPUT_DIR / "KPI_Upload_Template_Portaverse_2026_with_master_data.xlsx"


def value(v):
    return None if v is None else v


def set_calculation_on_open(workbook):
    calc = getattr(workbook, "calculation", None)
    if calc is not None:
        calc.fullCalcOnLoad = True
        calc.forceFullCalc = True
        calc.calcMode = "auto"


def populate_sheet(ws, headers, rows, freeze_panes="A2"):
    ws.delete_rows(1, ws.max_row)
    ws.append(headers)
    for row in rows:
        ws.append(row)
    ws.freeze_panes = freeze_panes
    end_col = openpyxl.utils.get_column_letter(len(headers))
    ws.auto_filter.ref = f"A1:{end_col}{ws.max_row}"


with REFERENCE_PATH.open("r", encoding="utf-8") as f:
    ref = json.load(f)

nomenclature_headers = [
    "cluster_id (PNID)",
    "mapping_row_id",
    "cluster_label",
    "position_master_id (PMID)",
    "position_name",
    "group_name",
    "type_name",
    "company_name",
    "group_master_id",
    "company_id",
    "job_class_level",
    "is_group_active",
    "is_company_active",
    "company_code",
]

nomenclature_rows = [
    [
        value(r.get("cluster_id")),
        value(r.get("id")),
        value(r.get("cluster_label")),
        value(r.get("position_master_id")),
        value(r.get("position_name")),
        value(r.get("group_name")),
        value(r.get("type_name")),
        value(r.get("company_name") or r.get("active_company_name")),
        value(r.get("group_master_id")),
        value(r.get("company_id")),
        value(r.get("job_class_level")),
        value(r.get("is_group_active")),
        value(r.get("is_company_active")),
        value(r.get("active_company_code")),
    ]
    for r in ref["rows"]
]

position_master_headers = [
    "position_master_id (PMID)",
    "position_name",
    "job_class_level",
    "company_name",
    "group_name",
    "group_master_id",
    "company_id",
    "company_code",
    "position_master_type_id",
    "is_position_active",
    "is_position_organization_active",
    "is_group_active",
    "is_company_active",
    "position_start_date",
    "position_end_date",
]

position_master_source = sorted(
    ref["position_master_rows"],
    key=lambda r: (
        int(r.get("position_master_id") or 0),
        -int(r.get("is_position_active") or 0),
        -int(r.get("is_position_organization_active") or 0),
        -int(r.get("is_group_active") or 0),
        -int(r.get("is_company_active") or 0),
        int(r.get("group_master_id") or 0),
    ),
)

position_master_rows = [
    [
        value(r.get("position_master_id")),
        value(r.get("position_name")),
        value(r.get("job_class_level")),
        value(r.get("company_name")),
        value(r.get("group_name")),
        value(r.get("group_master_id")),
        value(r.get("company_id")),
        value(r.get("company_code")),
        value(r.get("position_master_type_id")),
        value(r.get("is_position_active")),
        value(r.get("is_position_organization_active")),
        value(r.get("is_group_active")),
        value(r.get("is_company_active")),
        value(r.get("position_start_date")),
        value(r.get("position_end_date")),
    ]
    for r in position_master_source
]

wb = openpyxl.load_workbook(INPUT_WORKBOOK)

populate_sheet(
    wb["Master Data Nomenclature"],
    nomenclature_headers,
    nomenclature_rows,
    freeze_panes="A2",
)
populate_sheet(
    wb["Master Data Position Master"],
    position_master_headers,
    position_master_rows,
    freeze_panes="A2",
)

form = wb["Formulir Upload KPI"]
for row in range(5, 505):
    form.cell(row, 4).value = f'=IF(B{row}<>"","",C{row})'
    form.cell(row, 5).value = (
        f'=IF(B{row}<>"",IFERROR(INDEX(\'Master Data Nomenclature\'!$C:$C,'
        f'MATCH(B{row},\'Master Data Nomenclature\'!$A:$A,0)),"Tidak ditemukan"),'
        f'IF(C{row}<>"",IFERROR(INDEX(\'Master Data Position Master\'!$B:$B,'
        f'MATCH(C{row},\'Master Data Position Master\'!$A:$A,0)),"Tidak ditemukan"),""))'
    )
    form.cell(row, 6).value = (
        f'=IF(B{row}<>"",IFERROR(INDEX(\'Master Data Nomenclature\'!$F:$F,'
        f'MATCH(B{row},\'Master Data Nomenclature\'!$A:$A,0)),"Tidak ditemukan"),'
        f'IF(C{row}<>"",IFERROR(INDEX(\'Master Data Position Master\'!$E:$E,'
        f'MATCH(C{row},\'Master Data Position Master\'!$A:$A,0)),"Tidak ditemukan"),""))'
    )
    form.cell(row, 7).value = (
        f'=IF(B{row}<>"",IFERROR(INDEX(\'Master Data Nomenclature\'!$H:$H,'
        f'MATCH(B{row},\'Master Data Nomenclature\'!$A:$A,0)),"Tidak ditemukan"),'
        f'IF(C{row}<>"",IFERROR(INDEX(\'Master Data Position Master\'!$D:$D,'
        f'MATCH(C{row},\'Master Data Position Master\'!$A:$A,0)),"Tidak ditemukan"),""))'
    )

guide = wb["Aturan dan Panduan Pengisian"]
guide["B16"] = (
    "- Catat nilai pada kolom cluster_id (= PNID) di sheet Nomenclature, atau "
    "position_master_id (= PMID) di sheet Position Master."
)
guide["B18"] = (
    "- Relasi: cluster_id pada Nomenclature adalah PNID; position_master_id "
    "menautkan Nomenclature ke Position Master. Untuk PNID, Company/Group "
    "diambil dari baris Nomenclature yang sama."
)
guide["E35"] = "Terisi otomatis dari Master Data via PNID; jika PNID kosong, fallback memakai PMID."

set_calculation_on_open(wb)
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
wb.save(OUTPUT_WORKBOOK)

print(
    json.dumps(
        {
            "output": str(OUTPUT_WORKBOOK),
            "source": ref["source"],
            "nomenclature_rows_written": len(nomenclature_rows),
            "position_master_rows_written": len(position_master_rows),
            "form_formula_rows_updated": 500,
        },
        indent=2,
    )
)
