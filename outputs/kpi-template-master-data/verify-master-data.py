import json
from pathlib import Path

import openpyxl


ROOT = Path("/Users/alfredoteja/Documents/portaverse-kpi-bulk-upload-transformer")
REFERENCE_PATH = ROOT / "configs/production_position_reference.json"
OUTPUT_WORKBOOK = ROOT / "outputs/kpi-template-master-data/KPI_Upload_Template_Portaverse_2026_with_master_data.xlsx"


NOM_HEADERS = [
    "cluster_id (PNID)",
    "mapping_row_id",
    "cluster_label",
    "position_master_id (PMID)",
    "position_name",
    "group_name",
    "type_name",
    "company_name",
    "group_master_id",
    "company_id",
    "job_class_level",
    "is_group_active",
    "is_company_active",
    "company_code",
]

PM_HEADERS = [
    "position_master_id (PMID)",
    "position_name",
    "job_class_level",
    "company_name",
    "group_name",
    "group_master_id",
    "company_id",
    "company_code",
    "position_master_type_id",
    "is_position_active",
    "is_position_organization_active",
    "is_group_active",
    "is_company_active",
    "position_start_date",
    "position_end_date",
]


def val(v):
    return None if v in (None, "") else v


def expected_nomenclature_rows(ref):
    for r in ref["rows"]:
        yield [
            val(r.get("cluster_id")),
            val(r.get("id")),
            val(r.get("cluster_label")),
            val(r.get("position_master_id")),
            val(r.get("position_name")),
            val(r.get("group_name")),
            val(r.get("type_name")),
            val(r.get("company_name") or r.get("active_company_name")),
            val(r.get("group_master_id")),
            val(r.get("company_id")),
            val(r.get("job_class_level")),
            val(r.get("is_group_active")),
            val(r.get("is_company_active")),
            val(r.get("active_company_code")),
        ]


def sorted_position_master_source(ref):
    return sorted(
        ref["position_master_rows"],
        key=lambda r: (
            int(r.get("position_master_id") or 0),
            -int(r.get("is_position_active") or 0),
            -int(r.get("is_position_organization_active") or 0),
            -int(r.get("is_group_active") or 0),
            -int(r.get("is_company_active") or 0),
            int(r.get("group_master_id") or 0),
        ),
    )


def expected_position_master_rows(ref):
    for r in sorted_position_master_source(ref):
        yield [
            val(r.get("position_master_id")),
            val(r.get("position_name")),
            val(r.get("job_class_level")),
            val(r.get("company_name")),
            val(r.get("group_name")),
            val(r.get("group_master_id")),
            val(r.get("company_id")),
            val(r.get("company_code")),
            val(r.get("position_master_type_id")),
            val(r.get("is_position_active")),
            val(r.get("is_position_organization_active")),
            val(r.get("is_group_active")),
            val(r.get("is_company_active")),
            val(r.get("position_start_date")),
            val(r.get("position_end_date")),
        ]


def compare_rows(ws, expected_iter, width):
    checked = 0
    for checked, (actual, expected) in enumerate(
        zip(ws.iter_rows(min_row=2, max_col=width, values_only=True), expected_iter),
        start=1,
    ):
        actual_list = list(actual)
        if actual_list != expected:
            raise AssertionError(
                {
                    "sheet": ws.title,
                    "data_row": checked,
                    "actual": actual_list,
                    "expected": expected,
                }
            )
    return checked


with REFERENCE_PATH.open("r", encoding="utf-8") as f:
    ref = json.load(f)

wb = openpyxl.load_workbook(OUTPUT_WORKBOOK, data_only=False, read_only=False)
expected_sheets = [
    "Aturan dan Panduan Pengisian",
    "KPI Template",
    "Formulir Upload KPI",
    "Master Data Nomenclature",
    "Master Data Position Master",
]
assert wb.sheetnames == expected_sheets, wb.sheetnames

nom = wb["Master Data Nomenclature"]
pm = wb["Master Data Position Master"]
form = wb["Formulir Upload KPI"]
guide = wb["Aturan dan Panduan Pengisian"]

assert nom.max_row == len(ref["rows"]) + 1, nom.max_row
assert pm.max_row == len(ref["position_master_rows"]) + 1, pm.max_row
assert [nom.cell(1, c).value for c in range(1, len(NOM_HEADERS) + 1)] == NOM_HEADERS
assert [pm.cell(1, c).value for c in range(1, len(PM_HEADERS) + 1)] == PM_HEADERS
assert form.freeze_panes == "K409"
assert len(form.data_validations.dataValidation) == 7
assert "cluster_id (= PNID)" in guide["B16"].value
assert guide["E35"].value.startswith("Terisi otomatis dari Master Data via PNID")

for row in range(5, 505):
    assert form.cell(row, 4).value == f'=IF(B{row}<>"","",C{row})'
    assert "'Master Data Nomenclature'!$C:$C" in form.cell(row, 5).value
    assert "'Master Data Nomenclature'!$F:$F" in form.cell(row, 6).value
    assert "'Master Data Nomenclature'!$H:$H" in form.cell(row, 7).value
    assert "'Master Data Position Master'!$B:$B" in form.cell(row, 5).value
    assert "'Master Data Position Master'!$E:$E" in form.cell(row, 6).value
    assert "'Master Data Position Master'!$D:$D" in form.cell(row, 7).value

error_terms = ("#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#N/A")
literal_errors = []
for ws in wb.worksheets:
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and any(term in cell.value for term in error_terms):
                literal_errors.append(f"{ws.title}!{cell.coordinate}={cell.value}")
assert not literal_errors, literal_errors[:10]

wb_ro = openpyxl.load_workbook(OUTPUT_WORKBOOK, data_only=False, read_only=True)
nom_checked = compare_rows(
    wb_ro["Master Data Nomenclature"],
    expected_nomenclature_rows(ref),
    len(NOM_HEADERS),
)
pm_checked = compare_rows(
    wb_ro["Master Data Position Master"],
    expected_position_master_rows(ref),
    len(PM_HEADERS),
)

pnid_76_rows = [r for r in ref["rows"] if r["cluster_id"] == 76]
assert pnid_76_rows
assert len({r["cluster_label"] for r in pnid_76_rows}) == 1
assert len({r["group_name"] for r in pnid_76_rows}) == 1
assert len({r["company_name"] for r in pnid_76_rows}) == 1

pmid_528_rows = [r for r in sorted_position_master_source(ref) if r["position_master_id"] == 528]
assert pmid_528_rows

summary = {
    "workbook": str(OUTPUT_WORKBOOK),
    "source_exported_at": ref["source"]["exported_at"],
    "source_tables": ref["source"]["tables"],
    "nomenclature_rows_verified": nom_checked,
    "position_master_rows_verified": pm_checked,
    "form_formula_rows_verified": 500,
    "data_validations": len(form.data_validations.dataValidation),
    "literal_formula_errors": 0,
    "pnid_76_cluster_label": next(iter({r["cluster_label"] for r in pnid_76_rows})),
    "pmid_528_first_active_row": {
        "position_name": pmid_528_rows[0]["position_name"],
        "group_name": pmid_528_rows[0]["group_name"],
        "company_name": pmid_528_rows[0]["company_name"],
    },
}
print(json.dumps(summary, indent=2))
