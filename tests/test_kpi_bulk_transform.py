import json
import sys
import tempfile
import unittest
import xml.etree.ElementTree as ET
from datetime import datetime
from pathlib import Path
from zipfile import ZipFile

sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "scripts"))

from kpi_bulk_transform import (  # noqa: E402
    ImpactRecord,
    NormalizedEnum,
    NormalizationStatus,
    PositionConfig,
    UPLOAD_HEADERS,
    apply_best_effort_mapping,
    apply_nomenclature_summary_aliases,
    apply_nomenclature_summary_upload_rows,
    append_enum_issue,
    build_upload_rows,
    collect_parsed_sheets,
    conversion_output_name,
    discover_configs_for_workbook,
    is_active_valid_sheet,
    load_config,
    load_nomenclature_summary_aliases,
    load_nomenclature_summary_upload_rows,
    load_nomenclature_mapping,
    merge_mapping_entry,
    parse_block_sheet,
    normalize_cascading,
    normalize_kai_nature,
    normalize_ownership_type,
    normalize_period,
    normalize_polarity,
    normalize_bsc_perspective,
    refresh_configs_from_mapping,
    should_skip_source_workbook,
    source_workbook_context_hints,
    uploader_kai_nature,
    uploader_polarity,
    unique_output_name,
    validate_output_rows,
    write_output_workbook,
)
from position_mapping import build_lookup_indexes  # noqa: E402

from openpyxl import Workbook, load_workbook  # noqa: E402
from openpyxl.styles import Font  # noqa: E402


class KpiBulkTransformTest(unittest.TestCase):
    def test_collect_parsed_sheets_skips_explicit_neglect_before_metadata_fallback(self):
        config = PositionConfig(
            sheet_name="Officer Layanan Pelanggan",
            position_name="Officer Layanan Pelanggan",
            group_name="Group Aliansi Bisnis",
            directorate_name="Direktorat Komersial",
            position_master_id="34580",
            position_scope="neglect",
        )
        issues = []

        parsed = collect_parsed_sheets(
            Path("does-not-exist.xlsx"),
            None,
            [config],
            issues,
        )

        self.assertEqual(parsed, [])
        self.assertFalse(any(issue.severity == "error" for issue in issues))

    def test_collect_parsed_sheets_blocks_unapproved_low_confidence_mapping(self):
        config = PositionConfig(
            sheet_name="Officer Keu",
            position_name="Officer Keu",
            group_name="Group Keuangan",
            directorate_name="Direktorat Keuangan",
            position_scope="non_structural",
            mapping_confidence_label="low_confidence",
            mapping_confidence_reason="Candidate exists, but title match is weak.",
        )
        issues = []

        parsed = collect_parsed_sheets(
            Path("does-not-exist.xlsx"),
            None,
            [config],
            issues,
        )

        self.assertEqual(parsed, [])
        self.assertTrue(any("low_confidence" in issue.message for issue in issues))

    def test_collect_parsed_sheets_holds_needs_check_mapping_without_error(self):
        config = PositionConfig(
            sheet_name="Officer Pending",
            position_name="Officer Pending",
            group_name="Group Keuangan",
            directorate_name="Direktorat Keuangan",
            position_scope="non_structural",
            position_nomenclature_id="888",
            mapping_confidence_label="low_confidence",
            mapping_review_status="needs_check",
        )
        issues = []

        parsed = collect_parsed_sheets(
            Path("does-not-exist.xlsx"),
            None,
            [config],
            issues,
        )

        self.assertEqual(parsed, [])
        self.assertFalse(any(issue.severity == "error" for issue in issues))
        self.assertTrue(any(issue.severity == "warning" and "needs_check" in issue.message for issue in issues))

    def test_approved_scope_uncertain_override_is_validated_against_active_lookup(self):
        config = PositionConfig(
            sheet_name="Kamus KPI Bagian",
            position_name="Kamus KPI Bagian",
            group_name="Group Keuangan",
            directorate_name="Direktorat Keuangan",
            position_scope="non_structural",
            position_nomenclature_id="76",
            mapping_confidence_label="scope_uncertain",
            mapping_override_approved=True,
            mapping_review_status="approved",
        )
        reference = {
            "structural_lookup_rows": [],
            "non_structural_lookup_rows": [
                {
                    "cluster_id": "76",
                    "cluster_label": "Officer Keuangan",
                    "position_master_id": "701",
                    "position_name": "Officer Keuangan",
                    "position_master_type_id": "6",
                    "group_name": "Group Keuangan",
                    "company_id": "1",
                    "company_name": "PT Pelabuhan Indonesia (Persero)",
                    "company_code": "PLD",
                    "active_variant_count": 1,
                    "active_employee_count": 1,
                    "active_employee_names": ["Budi Santoso"],
                    "active_employee_nipps": ["K001"],
                }
            ],
        }
        indexes = build_lookup_indexes(reference, target_company_id="1")

        refresh_configs_from_mapping([config], {"Officer Keuangan": {}}, indexes)

        self.assertEqual(config.position_scope, "non_structural")
        self.assertIsNone(config.position_master_id)
        self.assertEqual(config.position_nomenclature_id, "76")
        self.assertEqual(config.active_employee_name, "Budi Santoso")

    def test_trusted_reviewer_manual_override_survives_strict_refresh(self):
        config = PositionConfig(
            sheet_name="Officer Manual",
            position_name="Officer Manual",
            group_name="Group Keuangan",
            directorate_name="Direktorat Keuangan",
            position_scope="non_structural",
            position_nomenclature_id="888",
            mapping_confidence_label="low_confidence",
            mapping_override_approved=True,
            mapping_review_status="approved",
            mapping_override_trust_source="reviewer_manual",
        )
        indexes = build_lookup_indexes(
            {
                "structural_lookup_rows": [],
                "non_structural_lookup_rows": [],
            },
            target_company_id="1",
        )

        refresh_configs_from_mapping([config], {"Officer Manual": {}}, indexes)

        self.assertEqual(config.position_scope, "non_structural")
        self.assertIsNone(config.position_master_id)
        self.assertEqual(config.position_nomenclature_id, "888")
        self.assertEqual(config.mapping_review_status, "approved")
        self.assertTrue(config.mapping_override_approved)

    def test_latest_upload_headers_include_optional_pnid_columns(self):
        self.assertEqual(len(UPLOAD_HEADERS), 24)
        self.assertEqual(
            UPLOAD_HEADERS[-3:],
            ["Ownership Type", "Position Nomenklatur ID", "RKM Code ID"],
        )
        self.assertEqual(UPLOAD_HEADERS[18], "Nature Of Work (KAI Only)")
        self.assertEqual(UPLOAD_HEADERS[21], "Ownership Type")

    def test_unique_output_name_adds_stable_suffix_for_duplicate_batch_names(self):
        seen = {}

        self.assertEqual(unique_output_name("Direktorat - Group", seen), "Direktorat - Group")
        self.assertEqual(unique_output_name("Direktorat - Group", seen), "Direktorat - Group - 02")
        self.assertEqual(unique_output_name("Direktorat - Group", seen), "Direktorat - Group - 03")

    def test_position_nomenclature_id_overrides_position_master_id_in_output_rows(self):
        config = PositionConfig(
            sheet_name="Officer Kinerja Individu",
            position_name="Officer I Kinerja Individu",
            group_name="Group Pengelolaan SDM",
            directorate_name="Direktorat Sumber Daya Manusia dan Umum",
            position_master_id="528",
            position_nomenclature_id="76",
            position_scope="non_structural",
        )
        impacts = [
            ImpactRecord(
                bsc="Financial",
                title="Net Income",
                unit="Rupiah",
                period="Triwulan",
                formula="Revenue - Cost",
                polarity="Positif",
                weight="15",
            )
        ]

        rows, next_id = build_upload_rows(config, "528", impacts, 1)

        self.assertEqual(next_id, 2)
        self.assertEqual(len(rows[0]), 24)
        row_map = dict(zip(UPLOAD_HEADERS, rows[0]))
        self.assertIsNone(row_map["Position Master ID (Required)"])
        self.assertEqual(row_map["Position Nomenklatur ID"], "76")
        self.assertIsNone(row_map["RKM Code ID"])

    def test_structural_scope_uses_only_position_master_id_even_when_pnid_exists(self):
        config = PositionConfig(
            sheet_name="Group Head",
            position_name="Group Head Pengelolaan SDM",
            group_name="Group Pengelolaan SDM",
            directorate_name="Direktorat Sumber Daya Manusia dan Umum",
            position_master_id="509",
            position_nomenclature_id="74",
            position_scope="structural",
        )
        impact = ImpactRecord(
            bsc="Financial",
            title="Net Income",
            unit="Rupiah",
            period="Triwulan",
            formula="Revenue - Cost",
            polarity="Positif",
            weight="15",
        )

        rows, _ = build_upload_rows(config, "509", [impact], 1)

        row_map = dict(zip(UPLOAD_HEADERS, rows[0]))
        self.assertEqual(row_map["Position Master ID (Required)"], "509")
        self.assertIsNone(row_map["Position Nomenklatur ID"])

    def test_validate_output_rows_rejects_invalid_upload_enums_and_dual_position_ids(self):
        config = PositionConfig(
            sheet_name="Manager Rekrutmen-Karir",
            position_name="Manager Rekrutmen dan Karir",
            group_name="Group Pengelolaan SDM",
            directorate_name="Direktorat SDM & Umum",
        )
        impact = dict.fromkeys(UPLOAD_HEADERS)
        impact.update(
            {
                "IDKPI": "1",
                "Group": "Group Pengelolaan SDM",
                "Direktorat": "Direktorat SDM & Umum",
                "Posisi": "Manager Rekrutmen dan Karir",
                "Position Master ID (Required)": "515",
                "BSC Perspective": "Learning & Growth",
                "KPI Type": "IMPACT",
                "Title": "Pemenuhan formasi",
                "Unit": "%",
                "Polarity": "POSITIVE",
                "Period": "PER TAHUN",
                "Formula": "realisasi/target",
                "Weight (%)": "10",
                "Cascading": "SPECIFIC",
                "Ownership Type": "Non Routine",
                "Position Nomenklatur ID": "515",
            }
        )
        kai = dict(impact)
        kai.update(
            {
                "IDKPI": "2",
                "KPI Type": "KAI",
                "Parent KPI ID": "1",
                "Parent KPI Title": "Pemenuhan formasi",
                "Title": "Follow up rekrutmen",
                "Period": "TAHUNAN",
                "Cascading": "DIRECT",
                "Nature Of Work (KAI Only)": "Kadang",
                "Ownership Type": "SPECIFIC",
                "Position Nomenklatur ID": None,
            }
        )
        issues = []

        validate_output_rows(config, [[row[header] for header in UPLOAD_HEADERS] for row in [impact, kai]], issues)

        messages = [issue.message for issue in issues]
        self.assertIn("Invalid Period enum: PER TAHUN", messages)
        self.assertIn("Invalid Cascading enum: SPECIFIC", messages)
        self.assertIn("Invalid Ownership Type enum: Non Routine", messages)
        self.assertIn("Invalid upload scope: row has both PMID and PNID.", messages)
        self.assertIn("Invalid KAI Nature enum: Kadang", messages)

    def test_load_config_treats_zero_position_master_id_as_missing_and_loads_pnid(self):
        payload = {
            "positions": [
                {
                    "sheet_name": "Officer Kinerja Individu",
                    "position_name": "Officer I Kinerja Individu",
                    "group_name": "Group Pengelolaan SDM",
                    "directorate_name": "Direktorat Sumber Daya Manusia dan Umum",
                    "position_master_id": 0,
                    "position_nomenclature_id": 76,
                    "position_scope": "non_structural",
                }
            ]
        }
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "config.json"
            path.write_text(json.dumps(payload), encoding="utf-8")

            configs = load_config(path)

        self.assertIsNone(configs[0].position_master_id)
        self.assertEqual(configs[0].position_nomenclature_id, "76")
        self.assertEqual(configs[0].position_scope, "non_structural")

    def test_load_config_structural_scope_drops_pnid_and_non_structural_scope_drops_pmid(self):
        payload = {
            "positions": [
                {
                    "sheet_name": "Group Head",
                    "position_name": "Group Head",
                    "group_name": "Group",
                    "directorate_name": "Direktorat",
                    "position_master_id": 946,
                    "position_nomenclature_id": 99,
                    "position_scope": "structural",
                },
                {
                    "sheet_name": "Officer",
                    "position_name": "Officer",
                    "group_name": "Group",
                    "directorate_name": "Direktorat",
                    "position_master_id": 528,
                    "position_nomenclature_id": 82,
                    "position_scope": "non_structural",
                },
            ]
        }
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "config.json"
            path.write_text(json.dumps(payload), encoding="utf-8")

            configs = load_config(path)

        self.assertEqual(configs[0].position_master_id, "946")
        self.assertIsNone(configs[0].position_nomenclature_id)
        self.assertIsNone(configs[1].position_master_id)
        self.assertEqual(configs[1].position_nomenclature_id, "82")

    def test_output_and_kai_inherit_impact_polarity_and_default_kai_nature(self):
        config = PositionConfig(
            sheet_name="Officer Kinerja Individu",
            position_name="Officer I Kinerja Individu",
            group_name="Group Pengelolaan SDM",
            directorate_name="Direktorat Sumber Daya Manusia dan Umum",
        )
        rows = [
            [
                "Jenis Posisi",
                "BSC Perspective",
                "KPI Impact",
                "KPI Impact Unit",
                "KPI Impact Frequency",
                "KPI Impact Formula",
                "%Weight (Impact)",
                "KPI Impact Polarity",
                "KPI Output",
                "%Weight (Output)",
                "KPI Output Definition",
                "KPI Output Unit",
                "KPI Output Frequency",
                "KPI Output Formula",
                "KPI Output Polarity",
                "Coverage KPI Output",
                "Cascading Tagging (KPI Output)",
                "Key Activity Indicator (KAI)",
                "%Weight (Activity)",
            ],
            [
                "Struktural",
                "Financial",
                "Net Income",
                "Rupiah",
                "Triwulan",
                "Revenue - Cost",
                "15",
                "Positif",
                "Penyempurnaan PMS",
                "10",
                "Enhancement PMS",
                "%",
                "Triwulan",
                "realisasi/target",
                None,
                "SPECIFIC",
                "DIRECT",
                "update SOP",
                "5",
            ],
        ]
        issues = []

        impacts = parse_block_sheet(rows, config, issues)

        self.assertEqual(impacts[0].source_row, 2)
        self.assertEqual(impacts[0].outputs[0]["polarity"], "Positif")
        self.assertEqual(impacts[0].outputs[0]["kai"]["polarity"], "Positif")
        self.assertIsNone(impacts[0].outputs[0]["kai"]["nature_of_work"])

    def test_missing_polarity_defaults_to_positive_and_missing_kai_nature_by_period(self):
        config = PositionConfig(
            sheet_name="Officer",
            position_name="Officer",
            group_name="Group Raw",
            directorate_name="Direktorat Raw",
            position_nomenclature_id="333",
            position_scope="non_structural",
        )
        impact = ImpactRecord(
            bsc="Financial",
            title="Revenue",
            unit="Rupiah",
            period="Triwulan",
            formula="Revenue",
            polarity=None,
            weight="100",
            outputs=[
                {
                    "source_row": 2,
                    "title": "Output",
                    "description": "desc",
                    "unit": "%",
                    "period": "Triwulan",
                    "formula": "realisasi/target",
                    "polarity": None,
                    "weight": "100",
                    "cascading": "DIRECT",
                    "ownership_type": "SPECIFIC",
                    "kai": {
                        "source_row": 2,
                        "title": "KAI",
                        "description": "desc",
                        "formula": "realisasi/target",
                        "weight": "100",
                        "nature_of_work": None,
                        "period": "Triwulan",
                        "polarity": None,
                        "cascading": "DIRECT",
                        "ownership_type": "SPECIFIC",
                    },
                }
            ],
        )

        rows, _ = build_upload_rows(config, None, [impact], 1)
        row_maps = [dict(zip(UPLOAD_HEADERS, row)) for row in rows]

        self.assertEqual([row["Polarity"] for row in row_maps], ["POSITIVE", "POSITIVE", "POSITIVE"])
        self.assertEqual(row_maps[2]["Nature Of Work (KAI Only)"], "Routine")

        impact.outputs[0]["kai"]["period"] = "Tahunan"
        rows, _ = build_upload_rows(config, None, [impact], 1)
        yearly_kai = dict(zip(UPLOAD_HEADERS, rows[2]))
        self.assertEqual(yearly_kai["Nature Of Work (KAI Only)"], "Non Routine")

    def test_append_enum_issue_uses_warning_for_defaulted_and_error_for_invalid_without_value(self):
        config = PositionConfig(
            sheet_name="Officer",
            position_name="Officer",
            group_name="Group",
            directorate_name="Direktorat",
        )
        issues = []

        append_enum_issue(
            issues,
            config,
            6,
            "IMPACT",
            "Impact Title",
            "Polarity",
            NormalizedEnum(
                value="POSITIVE",
                status=NormalizationStatus.NORMALIZED,
                raw_value="Positif",
                message="Polarity normalized to POSITIVE.",
            ),
        )
        append_enum_issue(
            issues,
            config,
            7,
            "OUTPUT",
            "Output Title",
            "Period",
            NormalizedEnum(
                value="TRIWULANAN",
                status=NormalizationStatus.DEFAULTED,
                raw_value=None,
                message="Period missing; defaulted to fallback period TRIWULANAN.",
            ),
        )
        append_enum_issue(
            issues,
            config,
            8,
            "OUTPUT",
            "Output Title",
            "Period",
            NormalizedEnum(
                value=None,
                status=NormalizationStatus.INVALID,
                raw_value="bogus",
                message="Invalid period.",
            ),
        )

        self.assertEqual(issues[0].severity, "info")
        self.assertIn(
            "enum_issue category=normalized; field=Polarity; raw=Positif; normalized=POSITIVE; Polarity normalized to POSITIVE.",
            issues[0].message,
        )
        self.assertEqual(issues[1].severity, "warning")
        self.assertIn(
            "enum_issue category=defaulted; field=Period; raw=None; normalized=TRIWULANAN; Period missing; defaulted to fallback period TRIWULANAN.",
            issues[1].message,
        )
        self.assertEqual(issues[2].severity, "error")
        self.assertIn(
            "enum_issue category=invalid; field=Period; raw=bogus; normalized=None; Invalid period.",
            issues[2].message,
        )

    def test_build_upload_rows_uses_impact_source_row_for_normalized_polarity_issue(self):
        config = PositionConfig(
            sheet_name="Officer",
            position_name="Officer",
            group_name="Group Raw",
            directorate_name="Direktorat Raw",
            position_nomenclature_id="333",
            position_scope="non_structural",
        )
        impact = ImpactRecord(
            source_row=12,
            bsc="Financial",
            title="Revenue",
            unit="Rupiah",
            period="TRIWULANAN",
            formula="Revenue",
            polarity="Positif",
            weight="100",
        )
        issues = []

        rows, _ = build_upload_rows(config, None, [impact], 1, issues)

        self.assertEqual(dict(zip(UPLOAD_HEADERS, rows[0]))["Polarity"], "POSITIVE")
        self.assertEqual(len(issues), 1)
        self.assertEqual(issues[0].severity, "info")
        self.assertEqual(issues[0].source_row, 12)
        self.assertEqual(issues[0].record_type, "IMPACT")
        self.assertIn(
            "enum_issue category=normalized; field=Polarity; raw=Positif; normalized=POSITIVE; Polarity normalized to POSITIVE.",
            issues[0].message,
        )

    def test_parsed_kai_inherited_fields_do_not_duplicate_output_enum_issues(self):
        config = PositionConfig(
            sheet_name="Officer Kinerja Individu",
            position_name="Officer I Kinerja Individu",
            group_name="Group Pengelolaan SDM",
            directorate_name="Direktorat Sumber Daya Manusia dan Umum",
        )
        rows = [
            [
                "Jenis Posisi",
                "BSC Perspective",
                "KPI Impact",
                "KPI Impact Unit",
                "KPI Impact Frequency",
                "KPI Impact Formula",
                "%Weight (Impact)",
                "KPI Impact Polarity",
                "KPI Output",
                "%Weight (Output)",
                "KPI Output Definition",
                "KPI Output Unit",
                "KPI Output Frequency",
                "KPI Output Formula",
                "KPI Output Polarity",
                "Coverage KPI Output",
                "Cascading Tagging (KPI Output)",
                "Nature of Work (KAI)",
                "Key Activity Indicator (KAI)",
                "%Weight (Activity)",
            ],
            [
                "Struktural",
                "Financial",
                "Net Income",
                "Rupiah",
                "TRIWULANAN",
                "Revenue - Cost",
                "100",
                "POSITIVE",
                "Penyempurnaan PMS",
                "100",
                "Enhancement PMS",
                "%",
                "Triwulan",
                "realisasi/target",
                "Positif",
                "Non Routine",
                "SPECIFIC",
                "Routine",
                "update SOP",
                "100",
            ],
        ]
        parse_issues = []
        issues = []

        impacts = parse_block_sheet(rows, config, parse_issues)
        self.assertEqual(impacts[0].source_row, 2)

        upload_rows, _ = build_upload_rows(config, "528", impacts, 1, issues)
        row_maps = [dict(zip(UPLOAD_HEADERS, row)) for row in upload_rows]

        self.assertEqual([row["KPI Type"] for row in row_maps], ["IMPACT", "OUTPUT", "KAI"])
        messages = [issue.message for issue in issues]
        self.assertTrue(any("enum_issue category=normalized; field=Period" in message for message in messages))
        self.assertTrue(any("enum_issue category=normalized; field=Polarity" in message for message in messages))
        self.assertTrue(any("enum_issue category=cross_column; field=Cascading" in message for message in messages))
        self.assertTrue(any("enum_issue category=cross_column; field=Ownership Type" in message for message in messages))
        self.assertFalse(
            any(
                issue.record_type == "KAI" and "field=Cascading" in issue.message
                for issue in issues
            )
        )
        self.assertFalse(
            any(
                issue.record_type == "KAI" and "field=Ownership Type" in issue.message
                for issue in issues
            )
        )

    def test_build_upload_rows_reports_ambiguous_output_period(self):
        config = PositionConfig(
            sheet_name="Officer Kinerja Individu",
            position_name="Officer I Kinerja Individu",
            group_name="Group Pengelolaan SDM",
            directorate_name="Direktorat Sumber Daya Manusia dan Umum",
        )
        impact = ImpactRecord(
            bsc="Financial",
            title="Net Income",
            unit="Rupiah",
            period="Triwulan",
            formula="Revenue - Cost",
            polarity="Positif",
            weight="15",
            outputs=[
                {
                    "source_row": 2,
                    "title": "Penyempurnaan PMS",
                    "description": "Enhancement PMS",
                    "unit": "%",
                    "period": "Triwulanan/Tahunan",
                    "formula": "realisasi/target",
                    "polarity": "Positif",
                    "weight": "10",
                    "cascading": "DIRECT",
                    "ownership_type": "SPECIFIC",
                }
            ],
        )
        issues = []

        rows, _ = build_upload_rows(config, "528", [impact], 1, issues)
        output_row = next(
            row for row in (dict(zip(UPLOAD_HEADERS, row)) for row in rows) if row["KPI Type"] == "OUTPUT"
        )

        self.assertEqual(output_row["Period"], "TRIWULANAN")
        matching = [issue for issue in issues if "enum_issue category=ambiguous; field=Period" in issue.message]
        self.assertEqual(len(matching), 1)
        self.assertEqual(matching[0].severity, "warning")

    def test_build_upload_rows_reports_cross_column_output_cascading(self):
        config = PositionConfig(
            sheet_name="Officer Kinerja Individu",
            position_name="Officer I Kinerja Individu",
            group_name="Group Pengelolaan SDM",
            directorate_name="Direktorat Sumber Daya Manusia dan Umum",
        )
        impact = ImpactRecord(
            bsc="Financial",
            title="Net Income",
            unit="Rupiah",
            period="Triwulan",
            formula="Revenue - Cost",
            polarity="Positif",
            weight="15",
            outputs=[
                {
                    "source_row": 2,
                    "title": "Penyempurnaan PMS",
                    "description": "Enhancement PMS",
                    "unit": "%",
                    "period": "Triwulan",
                    "formula": "realisasi/target",
                    "polarity": "Positif",
                    "weight": "10",
                    "cascading": "SPECIFIC",
                    "ownership_type": "SPECIFIC",
                }
            ],
        )
        issues = []

        rows, _ = build_upload_rows(config, "528", [impact], 1, issues)
        output_row = next(
            row for row in (dict(zip(UPLOAD_HEADERS, row)) for row in rows) if row["KPI Type"] == "OUTPUT"
        )

        self.assertEqual(output_row["Cascading"], "INDIRECT")
        matching = [issue for issue in issues if "enum_issue category=cross_column; field=Cascading" in issue.message]
        self.assertEqual(len(matching), 1)
        self.assertEqual(matching[0].severity, "warning")

    def test_build_upload_rows_reports_cross_column_kai_nature(self):
        config = PositionConfig(
            sheet_name="Officer Kinerja Individu",
            position_name="Officer I Kinerja Individu",
            group_name="Group Pengelolaan SDM",
            directorate_name="Direktorat Sumber Daya Manusia dan Umum",
        )
        impact = ImpactRecord(
            bsc="Financial",
            title="Net Income",
            unit="Rupiah",
            period="Triwulan",
            formula="Revenue - Cost",
            polarity="Positif",
            weight="15",
            outputs=[
                {
                    "source_row": 2,
                    "title": "Penyempurnaan PMS",
                    "description": "Enhancement PMS",
                    "unit": "%",
                    "period": "Triwulan",
                    "formula": "realisasi/target",
                    "polarity": "Positif",
                    "weight": "10",
                    "cascading": "DIRECT",
                    "ownership_type": "SPECIFIC",
                    "kai": {
                        "source_row": 2,
                        "title": "update SOP",
                        "description": "desc kai",
                        "formula": "progress x 100%",
                        "weight": "5",
                        "nature_of_work": "Pdf",
                        "period": "Tahunan",
                        "polarity": "Positif",
                    },
                }
            ],
        )
        issues = []

        rows, _ = build_upload_rows(config, "528", [impact], 1, issues)
        kai_row = next(row for row in (dict(zip(UPLOAD_HEADERS, row)) for row in rows) if row["KPI Type"] == "KAI")

        self.assertEqual(kai_row["Period"], "TAHUNAN")
        self.assertEqual(kai_row["Nature Of Work (KAI Only)"], "Non Routine")
        matching = [
            issue for issue in issues if "enum_issue category=cross_column; field=Nature Of Work" in issue.message
        ]
        self.assertEqual(len(matching), 1)
        self.assertEqual(matching[0].severity, "warning")

    def test_unknown_polarity_values_default_to_positive_enum(self):
        self.assertEqual(uploader_polarity("INDIRECT"), "POSITIVE")
        self.assertEqual(uploader_polarity("DUPLICATE"), "POSITIVE")
        self.assertEqual(uploader_polarity("SPECIFIC"), "POSITIVE")
        self.assertEqual(uploader_polarity("Negatif"), "NEGATIVE")

    def test_enum_normalizers_handle_variants_pollution_and_ambiguity(self):
        self.assertEqual(normalize_period("per tahun").value, "TAHUNAN")
        self.assertEqual(normalize_period("Semesteran").value, "SEMESTER")

        combo = normalize_period("Triwulanan/Tahunan")
        self.assertEqual(combo.status, NormalizationStatus.AMBIGUOUS)
        self.assertIsNone(combo.value)

        combo_with_fallback = normalize_period("Triwulanan/Tahunan", fallback="TRIWULANAN")
        self.assertEqual(combo_with_fallback.status, NormalizationStatus.AMBIGUOUS)
        self.assertEqual(combo_with_fallback.value, "TRIWULANAN")

        polarity = normalize_polarity("INDIRECT")
        self.assertEqual(polarity.value, "POSITIVE")
        self.assertEqual(polarity.status, NormalizationStatus.CROSS_COLUMN)
        self.assertEqual(normalize_polarity("Negatif").value, "NEGATIVE")

        self.assertEqual(normalize_bsc_perspective("Financial").value, "Financial")
        self.assertEqual(normalize_bsc_perspective("Learning dan Growth").value, "Learning & Growth")
        self.assertEqual(normalize_bsc_perspective("Learning and Growth").value, "Learning & Growth")
        self.assertEqual(normalize_bsc_perspective("Internal Process").value, "Internal Business Process")
        invalid_bsc = normalize_bsc_perspective("Stakeholder")
        self.assertIsNone(invalid_bsc.value)
        self.assertEqual(invalid_bsc.status, NormalizationStatus.INVALID)

        cascading = normalize_cascading("SPECIFIC")
        self.assertEqual(cascading.value, "INDIRECT")
        self.assertEqual(cascading.status, NormalizationStatus.CROSS_COLUMN)
        self.assertEqual(normalize_cascading("Indirect").value, "INDIRECT")

        ownership_blank = normalize_ownership_type(None)
        self.assertEqual(ownership_blank.value, "SPECIFIC")
        self.assertEqual(ownership_blank.status, NormalizationStatus.DEFAULTED)
        ownership = normalize_ownership_type("Non Routine")
        self.assertEqual(ownership.value, "SPECIFIC")
        self.assertEqual(ownership.status, NormalizationStatus.CROSS_COLUMN)
        self.assertEqual(normalize_ownership_type("SPESIFIC").value, "SPECIFIC")

        self.assertEqual(normalize_kai_nature(None, "TAHUNAN").value, "Non Routine")
        self.assertEqual(normalize_kai_nature(None, "TRIWULANAN").value, "Routine")
        polluted_nature = normalize_kai_nature("INDIRECT", "BULANAN")
        self.assertEqual(polluted_nature.value, "Routine")
        self.assertEqual(polluted_nature.status, NormalizationStatus.CROSS_COLUMN)
        pdf_nature = normalize_kai_nature("Pdf", "TAHUNAN")
        self.assertEqual(pdf_nature.value, "Non Routine")
        self.assertEqual(pdf_nature.status, NormalizationStatus.CROSS_COLUMN)
        uploaded_nature = normalize_kai_nature("Diunggah", "BULANAN")
        self.assertEqual(uploaded_nature.value, "Routine")
        self.assertEqual(uploaded_nature.status, NormalizationStatus.CROSS_COLUMN)
        url_nature = normalize_kai_nature("https://example.com/report.pdf", "BULANAN")
        self.assertEqual(url_nature.status, NormalizationStatus.CROSS_COLUMN)
        self.assertEqual(normalize_kai_nature("Non-Rotine", "BULANAN").value, "Non Routine")

    def test_zero_weight_rows_are_dropped_and_period_inherits_from_parent(self):
        config = PositionConfig(
            sheet_name="Group Head",
            position_name="Group Head",
            group_name="Group Raw",
            directorate_name="Direktorat Raw",
            position_master_id="58",
            position_scope="structural",
        )
        rows = [
            [
                "Jenis Posisi",
                "BSC Perspective",
                "KPI Impact",
                "KPI Impact Unit",
                "KPI Impact Frequency",
                "KPI Impact Formula",
                "%Weight (Impact)",
                "KPI Impact Polarity",
                "KPI Output",
                "%Weight (Output)",
                "KPI Output Definition",
                "KPI Output Unit",
                "KPI Output Frequency",
                "KPI Output Formula",
                "KPI Output Polarity",
                "Coverage KPI Output",
                "Cascading Tagging (KPI Output)",
                "Key Activity Indicator (KAI)",
                "%Weight (Activity)",
            ],
            [
                "Struktural",
                "Financial",
                "Net Income",
                "Rupiah",
                "Triwulan",
                "Revenue - Cost",
                "100",
                "Positif",
                "Zero Output",
                "0",
                "Should be dropped",
                "%",
                None,
                "realisasi/target",
                "INDIRECT",
                "SPECIFIC",
                "DIRECT",
                "Zero KAI",
                "0",
            ],
            [
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                "Valid Output",
                "100",
                "Kept",
                "%",
                None,
                "realisasi/target",
                "DUPLICATE",
                "SPECIFIC",
                "INDIRECT",
                "Valid KAI",
                "100",
            ],
        ]
        issues = []

        impacts = parse_block_sheet(rows, config, issues)
        upload_rows, _ = build_upload_rows(config, "58", impacts, 1)
        row_maps = [dict(zip(UPLOAD_HEADERS, row)) for row in upload_rows]

        self.assertEqual([row["Title"] for row in row_maps], ["Net Income", "Valid Output", "Valid KAI"])
        self.assertEqual([row["Period"] for row in row_maps], ["TRIWULANAN", "TRIWULANAN", "TRIWULANAN"])
        self.assertEqual([row["Polarity"] for row in row_maps], ["POSITIVE", "POSITIVE", "POSITIVE"])

    def test_alignment_drop_rows_are_excluded_from_output_and_kai(self):
        config = PositionConfig(
            sheet_name="Group Head",
            position_name="Group Head",
            group_name="Group Raw",
            directorate_name="Direktorat Raw",
            position_master_id="58",
            position_scope="structural",
        )
        rows = [
            [
                "Jenis Posisi",
                "BSC Perspective",
                "KPI Impact",
                "KPI Impact Unit",
                "KPI Impact Frequency",
                "KPI Impact Formula",
                "%Weight (Impact)",
                "KPI Impact Polarity",
                "KPI Output",
                "%Weight (Output)",
                "KPI Output Definition",
                "KPI Output Unit",
                "KPI Output Frequency",
                "KPI Output Formula",
                "KPI Output Polarity",
                "Coverage KPI Output",
                "Cascading Tagging (KPI Output)",
                "Key Activity Indicator (KAI)",
                "%Weight (Activity)",
                "Alignment",
            ],
            [
                "Struktural",
                "Financial",
                "Net Income",
                "Rupiah",
                "Triwulan",
                "Revenue - Cost",
                "100",
                "Positif",
                "Dropped Output",
                "100",
                "Should not be uploaded",
                "%",
                "Triwulan",
                "realisasi/target",
                "Positif",
                "SPECIFIC",
                "DIRECT",
                "Dropped KAI",
                "100",
                "Drop",
            ],
            [
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                "Valid Output",
                "100",
                "Kept",
                "%",
                "Triwulan",
                "realisasi/target",
                "Positif",
                "SPECIFIC",
                "DIRECT",
                "Valid KAI",
                "100",
                "Diupload",
            ],
        ]
        issues = []

        impacts = parse_block_sheet(rows, config, issues)
        upload_rows, _ = build_upload_rows(config, "58", impacts, 1)
        row_maps = [dict(zip(UPLOAD_HEADERS, row)) for row in upload_rows]

        self.assertEqual([row["Title"] for row in row_maps], ["Net Income", "Valid Output", "Valid KAI"])
        self.assertTrue(any("alignment/status is Drop" in issue.message for issue in issues))

    def test_numeric_output_and_kai_titles_are_excluded_as_summary_rows(self):
        config = PositionConfig(
            sheet_name="Group Head",
            position_name="Group Head",
            group_name="Group Raw",
            directorate_name="Direktorat Raw",
            position_master_id="58",
            position_scope="structural",
        )
        rows = [
            [
                "Jenis Posisi",
                "BSC Perspective",
                "KPI Impact",
                "KPI Impact Unit",
                "KPI Impact Frequency",
                "KPI Impact Formula",
                "%Weight (Impact)",
                "KPI Impact Polarity",
                "KPI Output",
                "%Weight (Output)",
                "KPI Output Definition",
                "KPI Output Unit",
                "KPI Output Frequency",
                "KPI Output Formula",
                "KPI Output Polarity",
                "Coverage KPI Output",
                "Cascading Tagging (KPI Output)",
                "Key Activity Indicator (KAI)",
                "%Weight (Activity)",
            ],
            [
                "Struktural",
                "Learning & Growth",
                "Manpower productivity",
                "Rupiah",
                "Triwulan",
                "Revenue / Employee",
                "100",
                "Positif",
                "8",
                "100",
                "Penyelesaian Temuan Audit",
                "%",
                "Triwulan",
                "realisasi/target",
                "Positif",
                "SPECIFIC",
                "DIRECT",
                "8",
                "100",
            ],
            [
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                "Valid Output 8",
                "100",
                "Kept",
                "%",
                "Triwulan",
                "realisasi/target",
                "Positif",
                "SPECIFIC",
                "DIRECT",
                "Valid KAI 8",
                "100",
            ],
        ]
        issues = []

        impacts = parse_block_sheet(rows, config, issues)
        upload_rows, _ = build_upload_rows(config, "58", impacts, 1)
        row_maps = [dict(zip(UPLOAD_HEADERS, row)) for row in upload_rows]

        self.assertEqual(
            [row["Title"] for row in row_maps],
            ["Manpower productivity", "Valid Output 8", "Valid KAI 8"],
        )
        self.assertTrue(any("numeric-only title" in issue.message for issue in issues))

    def test_kai_nature_normalizes_to_backend_allowed_values(self):
        self.assertEqual(uploader_kai_nature("Routine", "Tahunan"), "Routine")
        self.assertEqual(uploader_kai_nature("routine", "Tahunan"), "Routine")
        self.assertEqual(uploader_kai_nature("Non Routine", "Semester"), "Non Routine")
        self.assertEqual(uploader_kai_nature("Non-Routine", "Semester"), "Non Routine")
        self.assertEqual(uploader_kai_nature("Non-Rotine", "Semester"), "Non Routine")
        self.assertEqual(uploader_kai_nature("(blank)", "Semester"), "Routine")
        self.assertEqual(uploader_kai_nature("Diunggah", "Tahunan"), "Non Routine")

    def test_kai_rows_force_specific_indirect_percent_unit(self):
        config = PositionConfig(
            sheet_name="Officer Kinerja Individu",
            position_name="Officer Kinerja Individu",
            group_name="Group Pengelolaan SDM",
            directorate_name="Direktorat SDM",
            position_nomenclature_id="76",
        )
        impact = ImpactRecord(
            bsc="Learning & Growth",
            title="Manpower productivity",
            unit="IDR",
            period="Triwulan",
            formula="Revenue / Employee",
            polarity="Positif",
            weight="5",
            outputs=[
                {
                    "source_row": 20,
                    "title": "Penyempurnaan PMS",
                    "description": "desc",
                    "unit": "%",
                    "period": "Triwulan",
                    "formula": "realisasi/target",
                    "polarity": "Positif",
                    "weight": "5",
                    "cascading": "DIRECT",
                    "ownership_type": "SPECIFIC",
                    "kai": {
                        "source_row": 20,
                        "title": "update SOP",
                        "description": "desc kai",
                        "formula": "progress x 100%",
                        "weight": "5",
                        "nature_of_work": "Routine",
                        "period": "Triwulan",
                        "polarity": "Positif",
                        "cascading": "SPECIFIC",
                        "ownership_type": "Non Routine",
                    },
                }
            ],
        )
        issues = []
        rows, _ = build_upload_rows(config, "528", [impact], 1, issues)

        kai_row = dict(zip(UPLOAD_HEADERS, rows[2]))
        self.assertEqual(kai_row["KPI Type"], "KAI")
        self.assertEqual(kai_row["Unit"], "%")
        self.assertEqual(kai_row["Cascading"], "INDIRECT")
        self.assertEqual(kai_row["Ownership Type"], "SPECIFIC")
        messages = [issue.message for issue in issues]
        self.assertEqual(
            sum("enum_issue category=cross_column; field=Cascading" in message for message in messages),
            1,
        )
        self.assertEqual(
            sum("enum_issue category=cross_column; field=Ownership Type" in message for message in messages),
            1,
        )

    def test_duplicate_outputs_merge_weight_and_reparent_kai_children(self):
        config = PositionConfig(
            sheet_name="Officer Kinerja Individu",
            position_name="Officer Kinerja Individu",
            group_name="Group Pengelolaan SDM",
            directorate_name="Direktorat SDM",
            position_nomenclature_id="76",
        )
        duplicated_title = "Penyempurnaan Pengelolaan Talent Management berbasis Kompetensi (manajemen kinerja individu)"
        impact = ImpactRecord(
            bsc="Learning & Growth",
            title="Manpower productivity (Revenue per total manpower)",
            unit="IDR",
            period="Triwulan",
            formula="Revenue / Employee",
            polarity="Positif",
            weight="5",
            outputs=[],
        )
        for index, weight in enumerate(["5", "5", "5", "15", "5", "5", "5"], start=1):
            impact.outputs.append(
                {
                    "source_row": 20 + index,
                    "title": duplicated_title,
                    "description": "desc",
                    "unit": "%",
                    "period": "Triwulan",
                    "formula": "realisasi/target",
                    "polarity": "Positif",
                    "weight": weight,
                    "cascading": "DIRECT",
                    "ownership_type": "SPECIFIC",
                    "kai": {
                        "source_row": 20 + index,
                        "title": f"KAI {index}",
                        "description": None,
                        "formula": "progress x 100%",
                        "weight": weight,
                        "nature_of_work": "Routine",
                        "period": "Triwulan",
                        "polarity": "Positif",
                        "cascading": "DIRECT",
                        "ownership_type": "SPECIFIC",
                    },
                }
            )

        rows, _ = build_upload_rows(config, "528", [impact], 1)

        output_rows = [dict(zip(UPLOAD_HEADERS, row)) for row in rows if row[7] == "OUTPUT"]
        kai_rows = [dict(zip(UPLOAD_HEADERS, row)) for row in rows if row[7] == "KAI"]
        self.assertEqual(len(output_rows), 1)
        self.assertEqual(output_rows[0]["Title"], duplicated_title)
        self.assertEqual(output_rows[0]["Weight (%)"], "45")
        self.assertEqual(len(kai_rows), 7)
        self.assertEqual({row["Parent KPI ID"] for row in kai_rows}, {output_rows[0]["IDKPI"]})

    def test_active_valid_sheet_requires_visible_yellow_tab(self):
        class SheetLike:
            sheet_state = "visible"

            class sheet_properties:
                class tabColor:
                    rgb = "FFFFC000"
                    indexed = None
                    theme = None

        self.assertTrue(is_active_valid_sheet(SheetLike()))

        SheetLike.sheet_state = "hidden"
        self.assertFalse(is_active_valid_sheet(SheetLike()))

        SheetLike.sheet_state = "visible"
        SheetLike.sheet_properties.tabColor.rgb = "FF70AD47"
        self.assertFalse(is_active_valid_sheet(SheetLike()))

    def test_discover_configs_falls_back_to_visible_header_sheets_when_no_yellow_tabs_exist(self):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "OFFICER"
        worksheet.append(["Nama Posisi", "Officer"])
        worksheet.append(["Posisi", "Group Operasi"])
        worksheet.append(["BSC Perspective", "KPI Impact", "KPI Output", "Key Activity Indicator (KAI)"])
        mapping = {
            "officer": {
                "position_master_id": "9001",
                "position_nomenclature_id": None,
                "position_scope": "structural",
                "portaverse_position_title": "Officer",
                "portaverse_group_name": "Group Operasi",
                "portaverse_company_name": "PT Pelabuhan Indonesia (Persero)",
                "cluster_label": None,
            }
        }

        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "KAMUS KPI TEST.xlsx"
            workbook.save(path)

            configs = discover_configs_for_workbook("KAMUS KPI TEST.xlsx", path, mapping)

        self.assertEqual(len(configs), 1)
        self.assertEqual(configs[0].sheet_name, "OFFICER")
        self.assertEqual(configs[0].position_master_id, "9001")

    def test_discover_configs_marks_unresolved_position_as_mapping_conflict(self):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "UNKNOWN ROLE"
        worksheet.append(["Nama Posisi", "Unknown Role"])
        worksheet.append(["Posisi", "Group Operasi"])
        worksheet.append(["BSC Perspective", "KPI Impact", "KPI Output", "Key Activity Indicator (KAI)"])

        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "KAMUS KPI TEST.xlsx"
            workbook.save(path)

            configs = discover_configs_for_workbook("KAMUS KPI TEST.xlsx", path, {})

        self.assertEqual(len(configs), 1)
        self.assertEqual(configs[0].position_scope, "mapping_conflict")
        self.assertIsNone(configs[0].position_master_id)
        self.assertIsNone(configs[0].position_nomenclature_id)

    def test_collect_parsed_sheets_blocks_mapping_conflict_scope(self):
        config = PositionConfig(
            sheet_name="Unknown Role",
            position_name="Unknown Role",
            group_name="Group Operasi",
            directorate_name="Direktorat Operasi",
            position_scope="mapping_conflict",
        )
        issues = []

        parsed = collect_parsed_sheets(
            Path("does-not-exist.xlsx"),
            None,
            [config],
            issues,
        )

        self.assertEqual(parsed, [])
        self.assertEqual(issues[0].severity, "error")
        self.assertIn("mapping_conflict", issues[0].message)

    def test_mapping_restricts_lookup_to_target_company_and_uses_cluster_label(self):
        payload = {
            "position_master_rows": [
                {
                    "position_master_id": 111,
                    "position_name": "Officer Perencanaan",
                    "position_master_type_id": 5,
                    "company_id": 136,
                    "company_name": "PT Company Lain",
                    "group_name": "Group Lain",
                    "is_company_active": 1,
                    "is_group_active": 1,
                    "is_position_active": 1,
                    "is_position_organization_active": 1,
                },
                {
                    "position_master_id": 222,
                    "position_name": "Manager Perencanaan",
                    "position_master_type_id": 5,
                    "company_id": 1,
                    "company_name": "PT Pelabuhan Indonesia (Persero)",
                    "group_name": "Group HO",
                    "is_company_active": 1,
                    "is_group_active": 1,
                    "is_position_active": 1,
                    "is_position_organization_active": 1,
                },
            ],
            "rows": [
                {
                    "cluster_id": 333,
                    "cluster_label": "Officer Perencanaan",
                    "position_master_id": 3330,
                    "position_name": "Officer I Perencanaan",
                    "position_master_type_id": 6,
                    "type_name": "General",
                    "company_id": 1,
                    "company_name": "PT Pelabuhan Indonesia (Persero)",
                    "active_company_name": "PT Pelabuhan Indonesia (Persero)",
                    "group_name": "Group HO",
                    "active_group_name": "Group HO",
                    "is_company_active": 1,
                    "is_group_active": 1,
                }
            ],
        }
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "reference.json"
            path.write_text(json.dumps(payload), encoding="utf-8")

            mapping = load_nomenclature_mapping(path, "1")

        self.assertNotIn("officer perencanaan", {k for k, v in mapping.items() if v.get("portaverse_company_name") == "PT Company Lain"})
        self.assertEqual(mapping["officer perencanaan"]["position_nomenclature_id"], "333")
        self.assertEqual(mapping["officer perencanaan"]["position_scope"], "non_structural")
        self.assertEqual(mapping["manager perencanaan"]["position_master_id"], "222")
        self.assertIsNone(mapping["manager perencanaan"]["position_nomenclature_id"])

    def test_discover_config_uses_workbook_company_hint_for_generic_position(self):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "SPV"
        worksheet.append(["Nama Posisi", "SPV"])
        worksheet.append(["Posisi", ""])
        worksheet.append(["BSC Perspective", "KPI Impact", "KPI Output", "Key Activity Indicator (KAI)"])
        mapping = {}
        merge_mapping_entry(
            mapping,
            "Supervisor",
            {
                "position_master_id": "100",
                "position_nomenclature_id": None,
                "position_scope": "structural",
                "portaverse_position_title": "Supervisor SDM PT TEDS",
                "portaverse_group_name": "Divisi SDM PT TEDS",
                "portaverse_company_name": "PT Tanjung Emas Daya Sejahtera",
                "portaverse_company_code": "TEDS",
                "cluster_label": None,
                "position_master_type_id": "5",
            },
        )
        merge_mapping_entry(
            mapping,
            "Supervisor",
            {
                "position_master_id": "200",
                "position_nomenclature_id": None,
                "position_scope": "structural",
                "portaverse_position_title": "Supervisor Sales PT ILCS",
                "portaverse_group_name": "Unit Sales PT ILCS",
                "portaverse_company_name": "PT Integrasi Logistik Cipta Solusi",
                "portaverse_company_code": "ILCS",
                "cluster_label": None,
                "position_master_type_id": "5",
            },
        )

        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "KAMUS KPI PT TEDS.xlsx"
            workbook.save(path)

            configs = discover_configs_for_workbook(
                "KAMUS KPI PT TEDS/KAMUS KPI PT TEDS.xlsx",
                path,
                mapping,
            )

        self.assertEqual(len(configs), 1)
        self.assertEqual(configs[0].position_master_id, "100")
        self.assertEqual(configs[0].position_scope, "structural")

    def test_source_workbook_context_hints_exclude_generic_job_and_org_words(self):
        hints = source_workbook_context_hints(
            "KAMUS KPI PELINDO GROUP 2 (REGIONAL, CABANG DAN SUBHOLDING)/"
            "KAMUS KPI REGIONAL/KAMUS KPI Regional 4/KAMUS KPI Executive Director 4/"
            "Kamus KPI Executive Director Regional 4 - Mapping dengan Kontrak Manajemen.xlsx"
        )

        self.assertNotIn("regional", hints)
        self.assertNotIn("executive director", hints)
        self.assertIn("regional 4", hints)

    def test_discover_config_keeps_ambiguous_abbreviation_as_mapping_conflict(self):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "CORSEC"
        worksheet.append(["Nama Posisi", "CORSEC"])
        worksheet.append(["Posisi", ""])
        worksheet.append(["BSC Perspective", "KPI Impact", "KPI Output", "Key Activity Indicator (KAI)"])
        mapping = {}
        for position_master_id, company_code in [("100", "PMLI"), ("200", "PDS")]:
            merge_mapping_entry(
                mapping,
                "Corporate Secretary",
                {
                    "position_master_id": position_master_id,
                    "position_nomenclature_id": None,
                    "position_scope": "structural",
                    "portaverse_position_title": "Corporate Secretary",
                    "portaverse_group_name": "Divisi Corporate Secretary",
                    "portaverse_company_name": f"PT {company_code}",
                    "portaverse_company_code": company_code,
                    "cluster_label": None,
                    "position_master_type_id": "5",
                },
            )

        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "KAMUS KPI UNKNOWN.xlsx"
            workbook.save(path)

            configs = discover_configs_for_workbook(
                "KAMUS KPI UNKNOWN/KAMUS KPI UNKNOWN.xlsx",
                path,
                mapping,
            )

        self.assertEqual(len(configs), 1)
        self.assertEqual(configs[0].position_scope, "mapping_conflict")
        self.assertIsNone(configs[0].position_master_id)

    def test_discover_config_blocks_unique_abbreviation_with_other_company_hint(self):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "CORSEC"
        worksheet.append(["Nama Posisi", "CORSEC"])
        worksheet.append(["Posisi", ""])
        worksheet.append(["BSC Perspective", "KPI Impact", "KPI Output", "Key Activity Indicator (KAI)"])
        mapping = {}
        merge_mapping_entry(
            mapping,
            "Corporate Secretary",
            {
                "position_master_id": "100",
                "position_nomenclature_id": None,
                "position_scope": "structural",
                "portaverse_position_title": "Corporate Secretary",
                "portaverse_group_name": "Divisi Corporate Secretary",
                "portaverse_company_name": "PT Pendidikan Maritim dan Logistik Indonesia",
                "portaverse_company_code": "PMLI",
                "cluster_label": None,
                "position_master_type_id": "5",
            },
        )
        merge_mapping_entry(
            mapping,
            "Supervisor PDS",
            {
                "position_master_id": "200",
                "position_nomenclature_id": None,
                "position_scope": "structural",
                "portaverse_position_title": "Supervisor PDS",
                "portaverse_group_name": "Divisi Operasi",
                "portaverse_company_name": "PT Pelindo Daya Sejahtera",
                "portaverse_company_code": "PDS",
                "cluster_label": None,
                "position_master_type_id": "5",
            },
        )

        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "KAMUS KPI PT PDS.xlsx"
            workbook.save(path)

            configs = discover_configs_for_workbook(
                "KAMUS KPI AFILIASI NON CLUSTER DAN DAPEN/KAMUS KPI PT PDS/KAMUS KPI PT PDS.xlsx",
                path,
                mapping,
            )

        self.assertEqual(len(configs), 1)
        self.assertEqual(configs[0].position_scope, "mapping_conflict")
        self.assertIsNone(configs[0].position_master_id)

    def test_discover_config_does_not_fuzzy_pick_cross_company_candidate(self):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "REGIONAL MANAGER"
        worksheet.append(["Nama Posisi", "REGIONAL MANAGER"])
        worksheet.append(["Posisi", ""])
        worksheet.append(["BSC Perspective", "KPI Impact", "KPI Output", "Key Activity Indicator (KAI)"])
        mapping = {}
        merge_mapping_entry(
            mapping,
            "Supervisor SDM PT TEDS",
            {
                "position_master_id": "100",
                "position_nomenclature_id": None,
                "position_scope": "structural",
                "portaverse_position_title": "Supervisor SDM PT TEDS",
                "portaverse_group_name": "Divisi SDM PT TEDS",
                "portaverse_company_name": "PT Tanjung Emas Daya Sejahtera",
                "portaverse_company_code": "TEDS",
                "cluster_label": None,
                "position_master_type_id": "5",
            },
        )
        merge_mapping_entry(
            mapping,
            "Regional Manager Jawa PDS",
            {
                "position_master_id": "300",
                "position_nomenclature_id": None,
                "position_scope": "structural",
                "portaverse_position_title": "Regional Manager Jawa PDS",
                "portaverse_group_name": "Regional Jawa PDS",
                "portaverse_company_name": "PT Pelindo Daya Sejahtera",
                "portaverse_company_code": "PDS",
                "cluster_label": None,
                "position_master_type_id": "5",
            },
        )

        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "KAMUS KPI PT TEDS.xlsx"
            workbook.save(path)

            configs = discover_configs_for_workbook(
                "KAMUS KPI PT TEDS/KAMUS KPI PT TEDS.xlsx",
                path,
                mapping,
            )

        self.assertEqual(len(configs), 1)
        self.assertEqual(configs[0].position_scope, "mapping_conflict")
        self.assertIsNone(configs[0].position_master_id)

    def test_refresh_config_from_mapping_clears_old_other_company_id(self):
        mapping = {
            "officer perencanaan": {
                "position_master_id": "3330",
                "position_nomenclature_id": "333",
                "position_scope": "non_structural",
                "portaverse_position_title": "Officer I Perencanaan",
                "portaverse_group_name": "Group HO",
                "portaverse_company_name": "PT Pelabuhan Indonesia (Persero)",
                "cluster_label": "Officer Perencanaan",
            }
        }
        config = PositionConfig(
            sheet_name="Officer Perencanaan",
            position_name="Officer Perencanaan",
            group_name="Group Raw",
            directorate_name="Direktorat Raw",
            position_master_id="111",
            position_scope="structural",
            portaverse_company_name="PT Company Lain",
        )

        refresh_configs_from_mapping([config], mapping)

        self.assertIsNone(config.position_master_id)
        self.assertEqual(config.position_nomenclature_id, "333")
        self.assertEqual(config.position_scope, "non_structural")
        self.assertEqual(config.group_name, "Group Raw")

    def test_refresh_config_from_mapping_uses_structural_scope_when_number_exists_as_pmid_and_pnid(self):
        mapping = {
            "manager rekrutmen karir": {
                "position_master_id": "515",
                "position_nomenclature_id": "515",
                "position_scope": "non_structural",
                "position_master_type_id": "5",
                "portaverse_position_title": "Manager Rekrutmen dan Karir",
                "portaverse_group_name": "Unit Pendukung Rekrutmen dan Karir",
                "portaverse_company_name": "PT Pelabuhan Indonesia (Persero)",
                "cluster_label": "Manager Rekrutmen-Karir",
            }
        }
        config = PositionConfig(
            sheet_name="Manager Rekrutmen-Karir",
            position_name="Manager Rekrutmen-Karir",
            group_name="Group Pengelolaan SDM",
            directorate_name="Direktorat SDM & Umum",
            position_nomenclature_id="515",
            position_scope="non_structural",
        )

        refresh_configs_from_mapping([config], mapping)

        self.assertEqual(config.position_master_id, "515")
        self.assertIsNone(config.position_nomenclature_id)
        self.assertEqual(config.position_scope, "structural")

    def test_load_nomenclature_mapping_structural_title_wins_over_same_number_pnid(self):
        payload = {
            "position_master_rows": [
                {
                    "position_master_id": 515,
                    "position_name": "Manager Rekrutmen dan Karir",
                    "position_master_type_id": 5,
                    "company_id": 1,
                    "company_name": "PT Pelabuhan Indonesia (Persero)",
                    "group_name": "Unit Pendukung Rekrutmen dan Karir",
                    "is_company_active": 1,
                    "is_group_active": 1,
                    "is_position_active": 1,
                    "is_position_organization_active": 1,
                }
            ],
            "rows": [
                {
                    "cluster_id": 515,
                    "cluster_label": "Manager Rekrutmen dan Karir",
                    "position_master_id": 9999,
                    "position_name": "Officer I Rekrutmen dan Karir",
                    "position_master_type_id": 6,
                    "type_name": "General",
                    "company_id": 1,
                    "company_name": "PT Pelabuhan Indonesia (Persero)",
                    "active_company_name": "PT Pelabuhan Indonesia (Persero)",
                    "group_name": "Unit Pendukung Rekrutmen dan Karir",
                    "active_group_name": "Unit Pendukung Rekrutmen dan Karir",
                    "is_company_active": 1,
                    "is_group_active": 1,
                }
            ],
        }
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "reference.json"
            path.write_text(json.dumps(payload), encoding="utf-8")

            mapping = load_nomenclature_mapping(path, "1")

        manager = mapping["manager rekrutmen dan karir"]
        self.assertEqual(manager["position_master_id"], "515")
        self.assertIsNone(manager["position_nomenclature_id"])
        self.assertEqual(manager["position_scope"], "structural")

    def test_refresh_config_from_mapping_preserves_reviewed_manual_pmid(self):
        mapping = {
            "different manager title": {
                "position_master_id": "999",
                "position_nomenclature_id": None,
                "position_scope": "structural",
                "portaverse_position_title": "Different Manager Title",
                "portaverse_group_name": "Group HO",
                "portaverse_company_name": "PT Pelabuhan Indonesia (Persero)",
                "cluster_label": None,
            }
        }
        config = PositionConfig(
            sheet_name="Manager Performa Keuangan",
            position_name="Manager Performa Keuangan",
            group_name="Group Raw",
            directorate_name="Direktorat Raw",
            position_master_id="316",
            position_scope="structural",
        )

        refresh_configs_from_mapping([config], mapping)

        self.assertEqual(config.position_master_id, "316")
        self.assertIsNone(config.position_nomenclature_id)
        self.assertEqual(config.position_scope, "structural")

    def test_refresh_config_prefers_exact_structural_title_before_fuzzy_suffix_match(self):
        mapping = {
            "manager layanan keuangan wilayah timur 1": {
                "position_master_id": "33845",
                "position_nomenclature_id": None,
                "position_scope": "structural",
                "position_master_type_id": "5",
                "portaverse_position_title": "Manager Layanan Keuangan Wilayah Timur 1",
                "portaverse_group_name": "Unit Timur 1",
                "portaverse_company_name": "PT Pelabuhan Indonesia (Persero)",
                "cluster_label": None,
            },
            "manager layanan keuangan wilayah timur 2": {
                "position_master_id": "33854",
                "position_nomenclature_id": None,
                "position_scope": "structural",
                "position_master_type_id": "5",
                "portaverse_position_title": "Manager Layanan Keuangan Wilayah Timur 2",
                "portaverse_group_name": "Unit Timur 2",
                "portaverse_company_name": "PT Pelabuhan Indonesia (Persero)",
                "cluster_label": None,
            },
        }
        config = PositionConfig(
            sheet_name="Manager Layanan Keuangan Timur2",
            position_name="Manager Layanan Keuangan Wilayah Timur 2",
            group_name="Group Pengembangan SSC",
            directorate_name="Direktorat Pengembangan Usaha",
            position_master_id="33854",
            position_scope="structural",
        )

        refresh_configs_from_mapping([config], mapping)

        self.assertEqual(config.position_master_id, "33854")
        self.assertEqual(
            config.portaverse_position_title,
            "Manager Layanan Keuangan Wilayah Timur 2",
        )

    def test_refresh_config_from_mapping_preserves_reviewed_manual_pnid(self):
        mapping = {
            "different officer title": {
                "position_master_id": None,
                "position_nomenclature_id": "999",
                "position_scope": "non_structural",
                "portaverse_position_title": "Different Officer Title",
                "portaverse_group_name": "Group HO",
                "portaverse_company_name": "PT Pelabuhan Indonesia (Persero)",
                "cluster_label": "Different Officer Title",
            }
        }
        config = PositionConfig(
            sheet_name="Officer QA",
            position_name="Officer QA",
            group_name="Group Raw",
            directorate_name="Direktorat Raw",
            position_nomenclature_id="11517",
            position_scope="non_structural",
        )

        refresh_configs_from_mapping([config], mapping)

        self.assertIsNone(config.position_master_id)
        self.assertEqual(config.position_nomenclature_id, "11517")
        self.assertEqual(config.position_scope, "non_structural")

    def test_write_output_workbook_saves_weight_as_numeric_cell(self):
        with tempfile.TemporaryDirectory() as tmp:
            template = Path(tmp) / "template.xlsx"
            output = Path(tmp) / "output.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = "KPI Template"
            ws.append(UPLOAD_HEADERS)
            ws.append([None] * len(UPLOAD_HEADERS))
            wb.save(template)
            row = [
                "1",
                "Group Raw",
                "Direktorat Raw",
                "Officer",
                None,
                None,
                "Financial",
                "IMPACT",
                None,
                "#N/A",
                "Title",
                None,
                "%",
                "POSITIVE",
                "TRIWULANAN",
                "x/y",
                "12.5",
                None,
                None,
                None,
                None,
                None,
                "333",
                None,
            ]

            write_output_workbook(template, output, [row])
            saved = load_workbook(output, data_only=True)
            weight_cell = saved["KPI Template"].cell(row=2, column=17)

        self.assertIsInstance(weight_cell.value, float)
        self.assertEqual(weight_cell.value, 12.5)
        self.assertEqual(saved["KPI Template"].column_dimensions["P"].width, 72)

    def test_write_output_workbook_forces_black_header_text(self):
        with tempfile.TemporaryDirectory() as tmp:
            template = Path(tmp) / "template.xlsx"
            output = Path(tmp) / "output.xlsx"
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "KPI Template"
            worksheet.append(UPLOAD_HEADERS)
            worksheet["A1"].font = Font(bold=True, color="FFFFFF")
            workbook.save(template)

            write_output_workbook(template, output, [])

            saved = load_workbook(output, data_only=True)
            header_font = saved["KPI Template"]["A1"].font

        self.assertEqual(header_font.color.type, "rgb")
        self.assertEqual(header_font.color.rgb, "00000000")

    def test_conversion_output_name_defaults_to_kamus_v2(self):
        output_name = conversion_output_name(
            "Direktorat Komersial - Group Pemasaran dan Aliansi Bisnis.xlsx",
            [],
            datetime(2026, 7, 2, 11, 43),
        )

        self.assertTrue(output_name.endswith("(2026 v2)"))

    def test_group_hukum_as_is_source_workbook_is_skipped_for_v2(self):
        self.assertTrue(
            should_skip_source_workbook("(As-Is) Direktorat Manajemen Risiko - Group Hukum.xlsx")
        )
        self.assertFalse(
            should_skip_source_workbook("(New) Direktorat Manajemen Risiko - Group Hukum.xlsx")
        )
        self.assertFalse(
            should_skip_source_workbook("(As-Is) Direktorat Manajemen Risiko - Group K3.xlsx")
        )

    def test_best_effort_mapping_promotes_candidate_and_skips_unmapped(self):
        low = PositionConfig(
            sheet_name="Officer A",
            position_name="Officer A",
            group_name="Group A",
            directorate_name="Direktorat A",
            position_scope="non_structural",
            mapping_confidence_label="low_confidence",
            candidate_position_nomenclature_id="82",
        )
        missing = PositionConfig(
            sheet_name="PMO",
            position_name="PMO",
            group_name="Group A",
            directorate_name="Direktorat A",
            position_scope="scope_uncertain",
            mapping_confidence_label="scope_uncertain",
        )

        apply_best_effort_mapping([low, missing])

        self.assertTrue(low.mapping_override_approved)
        self.assertEqual(low.position_nomenclature_id, "82")
        self.assertIsNone(low.position_master_id)
        self.assertEqual(missing.position_scope, "neglect")
        self.assertEqual(missing.mapping_review_status, "auto_skipped_unmapped_best_effort")

    def test_nomenclature_summary_aliases_extend_lookup_keys(self):
        with tempfile.TemporaryDirectory() as tmp:
            summary = Path(tmp) / "summary.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Database Master"
            sheet.append([
                "No.",
                "Judul Posisi \n(Perubahan STO 1 April 2026)",
                "Kondisi Existing (Before 1 April)",
            ])
            sheet.append([1, "Department Head Pengelolaan Pelanggan", "DH Layanan Pelanggan"])
            workbook.save(summary)

            aliases = load_nomenclature_summary_aliases(summary)
            payload = {
                "structural_lookup_rows": [
                    {
                        "position_master_id": "1",
                        "position_name": "Department Head Pengelolaan Pelanggan",
                    }
                ]
            }

            apply_nomenclature_summary_aliases(payload, aliases)

        self.assertIn("DH Layanan Pelanggan", payload["structural_lookup_rows"][0]["normalized_lookup_keys"])

    def test_nomenclature_summary_upload_ready_overlays_non_structural_lookup(self):
        with tempfile.TemporaryDirectory() as tmp:
            summary = Path(tmp) / "summary.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Upload_Ready"
            sheet.append(["position_master_id", "position_name", "cluster_id", "cluster_label"])
            sheet.append(["101", "Administrator Operasi Senior Wilayah II", "9001", "Administrator Operasi Wilayah II"])
            workbook.save(summary)

            upload_rows = load_nomenclature_summary_upload_rows(summary)
            payload = {
                "non_structural_lookup_rows": [
                    {
                        "position_master_id": "101",
                        "position_name": "Administrator Operasi Senior Wilayah II",
                        "cluster_id": "1",
                        "cluster_label": "Administrator Operasi",
                    }
                ]
            }

            apply_nomenclature_summary_upload_rows(payload, upload_rows)

        row = payload["non_structural_lookup_rows"][0]
        self.assertEqual(row["cluster_id"], "9001")
        self.assertEqual(row["cluster_label"], "Administrator Operasi Wilayah II")
        self.assertIn("Administrator Operasi Wilayah II", row["normalized_lookup_keys"])

    def test_write_output_workbook_preserves_valid_frozen_pane_view(self):
        with tempfile.TemporaryDirectory() as tmp:
            template = Path(tmp) / "template.xlsx"
            output = Path(tmp) / "output.xlsx"
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "KPI Template"
            worksheet.append(UPLOAD_HEADERS)
            worksheet.append([None] * len(UPLOAD_HEADERS))
            worksheet.freeze_panes = "B2"
            workbook.save(template)

            write_output_workbook(template, output, [])

            with ZipFile(output) as output_zip:
                root = ET.fromstring(output_zip.read("xl/worksheets/sheet1.xml"))

        namespace = {"a": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
        sheet_view = root.find("a:sheetViews/a:sheetView", namespace)
        self.assertIsNotNone(sheet_view)
        pane_selections = [
            selection
            for selection in sheet_view.findall("a:selection", namespace)
            if selection.get("pane")
        ]
        self.assertTrue(pane_selections)
        self.assertIsNotNone(sheet_view.find("a:pane", namespace))


if __name__ == "__main__":
    unittest.main()
