import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation

from scripts import validate_kpi_upload_batch as validator


class ValidateKpiUploadBatchTest(unittest.TestCase):
    def test_config_scope_allows_trusted_reviewer_manual_ids(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            config_path = Path(temp_dir) / "config.json"
            config_path.write_text(
                """
{
  "positions": [
    {
      "source_workbook": "book.xlsx",
      "sheet_name": "Officer",
      "position_scope": "non_structural",
      "position_master_id": null,
      "position_nomenclature_id": "888",
      "mapping_confidence_label": "low_confidence",
      "mapping_review_status": "approved",
      "mapping_override_approved": true,
      "mapping_override_trust_source": "reviewer_manual"
    }
  ]
}
""".strip(),
                encoding="utf-8",
            )

            errors = validator.check_config_scope(
                config_path,
                master_ids=set(),
                nomenclature_ids=set(),
                cluster_labels_by_id={},
                master_types_by_id={},
                position_types_by_pnid={},
            )

        self.assertEqual(errors, [])

    def test_config_scope_allows_trusted_pnid_that_collides_with_pmid_number(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            config_path = Path(temp_dir) / "config.json"
            config_path.write_text(
                """
{
  "positions": [
    {
      "source_workbook": "book.xlsx",
      "sheet_name": "Officer Investasi",
      "position_scope": "non_structural",
      "position_master_id": null,
      "position_nomenclature_id": "12554",
      "mapping_confidence_label": "low_confidence",
      "mapping_review_status": "approved",
      "mapping_override_approved": true,
      "mapping_override_trust_source": "reviewer_manual"
    }
  ]
}
""".strip(),
                encoding="utf-8",
            )

            errors = validator.check_config_scope(
                config_path,
                master_ids={"12554"},
                nomenclature_ids=set(),
                cluster_labels_by_id={},
                master_types_by_id={"12554": {"6"}},
                position_types_by_pnid={},
            )

        self.assertEqual(errors, [])

    def test_config_scope_allows_needs_check_rows_to_remain_held(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            config_path = Path(temp_dir) / "config.json"
            config_path.write_text(
                """
{
  "positions": [
    {
      "source_workbook": "book.xlsx",
      "sheet_name": "Officer Pending",
      "position_scope": "non_structural",
      "position_master_id": null,
      "position_nomenclature_id": "888",
      "mapping_confidence_label": "low_confidence",
      "mapping_review_status": "needs_check",
      "mapping_override_approved": false
    }
  ]
}
""".strip(),
                encoding="utf-8",
            )

            errors = validator.check_config_scope(
                config_path,
                master_ids=set(),
                nomenclature_ids=set(),
                cluster_labels_by_id={},
                master_types_by_id={},
                position_types_by_pnid={},
            )

        self.assertEqual(errors, [])

    def test_config_scope_rejects_untrusted_invalid_ids(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            config_path = Path(temp_dir) / "config.json"
            config_path.write_text(
                """
{
  "positions": [
    {
      "source_workbook": "book.xlsx",
      "sheet_name": "Officer",
      "position_scope": "non_structural",
      "position_master_id": null,
      "position_nomenclature_id": "888",
      "mapping_confidence_label": "low_confidence",
      "mapping_review_status": "approved",
      "mapping_override_approved": true
    }
  ]
}
""".strip(),
                encoding="utf-8",
            )

            errors = validator.check_config_scope(
                config_path,
                master_ids=set(),
                nomenclature_ids=set(),
                cluster_labels_by_id={},
                master_types_by_id={},
                position_types_by_pnid={},
            )

        self.assertTrue(any("invalid production types" in error for error in errors))

    def test_config_scope_rejects_duplicate_structural_pmids(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            config_path = Path(temp_dir) / "config.json"
            config_path.write_text(
                """
{
  "positions": [
    {
      "source_workbook": "book-a.xlsx",
      "sheet_name": "Group Head",
      "position_scope": "structural",
      "position_master_id": "504",
      "position_nomenclature_id": null,
      "mapping_review_status": "approved",
      "mapping_override_approved": true,
      "mapping_override_trust_source": "reviewer_manual"
    },
    {
      "source_workbook": "book-b.xlsx",
      "sheet_name": "Group Head",
      "position_scope": "structural",
      "position_master_id": "504",
      "position_nomenclature_id": null,
      "mapping_review_status": "approved",
      "mapping_override_approved": true,
      "mapping_override_trust_source": "reviewer_manual"
    }
  ]
}
""".strip(),
                encoding="utf-8",
            )

            errors = validator.check_config_scope(
                config_path,
                master_ids={"504"},
                nomenclature_ids=set(),
                cluster_labels_by_id={},
                master_types_by_id={"504": {"5"}},
                position_types_by_pnid={},
            )

        self.assertTrue(any("duplicate structural PMID 504" in error for error in errors))

    def test_workbook_validation_allows_trusted_reviewer_manual_ids(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            workbook_path = Path(temp_dir) / "upload.xlsx"
            self._write_upload_workbook(workbook_path, pnid="888")

            record, errors, _ = validator.validate_workbook(
                workbook_path,
                fixed_pmids=set(),
                master_ids=set(),
                nomenclature_ids=set(),
                master_types_by_id={},
                position_types_by_pnid={},
                trusted_pmid_ids=set(),
                trusted_pnid_ids={"888"},
            )

        self.assertEqual(errors, [])
        self.assertEqual(record["status"], "READY")

    def test_workbook_validation_allows_approved_assistant_identity(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            workbook_path = Path(temp_dir) / "upload.xlsx"
            self._write_upload_workbook(workbook_path, pnid="")
            workbook = validator.openpyxl.load_workbook(workbook_path)
            sheet = workbook["KPI Template"]
            sheet.cell(2, 5).value = "77"
            sheet.cell(2, 6).value = "655"
            workbook.save(workbook_path)

            record, errors, _ = validator.validate_workbook(
                workbook_path,
                fixed_pmids=set(),
                master_ids=set(),
                nomenclature_ids=set(),
                master_types_by_id={},
                position_types_by_pnid={},
                approved_assistant_identities={("77", "655")},
            )

        self.assertEqual(errors, [])
        self.assertEqual(record["status"], "READY")

    def test_workbook_validation_rejects_unapproved_assistant_pmvid(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            workbook_path = Path(temp_dir) / "upload.xlsx"
            self._write_upload_workbook(workbook_path, pnid="")
            workbook = validator.openpyxl.load_workbook(workbook_path)
            sheet = workbook["KPI Template"]
            sheet.cell(2, 5).value = "77"
            sheet.cell(2, 6).value = "999"
            workbook.save(workbook_path)

            _, errors, _ = validator.validate_workbook(
                workbook_path,
                fixed_pmids=set(),
                master_ids=set(),
                nomenclature_ids=set(),
                master_types_by_id={},
                position_types_by_pnid={},
                approved_assistant_identities={("77", "655")},
            )

        self.assertTrue(any("unapproved or missing PMVID=999" in error for error in errors))

    def test_workbook_validation_rejects_untrusted_invalid_ids(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            workbook_path = Path(temp_dir) / "upload.xlsx"
            self._write_upload_workbook(workbook_path, pnid="888")

            _, errors, _ = validator.validate_workbook(
                workbook_path,
                fixed_pmids=set(),
                master_ids=set(),
                nomenclature_ids=set(),
                master_types_by_id={},
                position_types_by_pnid={},
                trusted_pmid_ids=set(),
                trusted_pnid_ids=set(),
            )

        self.assertTrue(any("invalid PNID=888" in error for error in errors))

    def test_workbook_validation_rejects_numeric_only_titles(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            workbook_path = Path(temp_dir) / "upload.xlsx"
            self._write_upload_workbook(workbook_path, pnid="888", title="8")

            _, errors, _ = validator.validate_workbook(
                workbook_path,
                fixed_pmids=set(),
                master_ids=set(),
                nomenclature_ids={"888"},
                master_types_by_id={},
                position_types_by_pnid={"888": {"6"}},
                trusted_pmid_ids=set(),
                trusted_pnid_ids=set(),
            )

        self.assertTrue(any("numeric-only KPI title" in error for error in errors))

    def test_workbook_validation_rejects_rules_ending_before_last_kpi_row(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            workbook_path = Path(temp_dir) / "upload.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "KPI Template"
            sheet.append(validator.EXPECTED_HEADERS)
            for row_number in range(1, 31):
                row = [""] * len(validator.EXPECTED_HEADERS)
                row[0] = row_number
                row[3] = "Officer"
                row[7] = "IMPACT"
                row[10] = f"KPI {row_number}"
                row[22] = "888"
                sheet.append(row)
            sheet.conditional_formatting.add("A2:F25", FormulaRule(formula=["TRUE"]))
            validation = DataValidation(type="list", formula1='"A,B"', allow_blank=True)
            validation.add("G2:G25")
            sheet.add_data_validation(validation)
            workbook.save(workbook_path)

            _, errors, _ = validator.validate_workbook(
                workbook_path,
                fixed_pmids=set(),
                master_ids=set(),
                nomenclature_ids={"888"},
                master_types_by_id={},
                position_types_by_pnid={"888": {"6"}},
            )

        self.assertTrue(any("conditional formatting ends before final KPI row 31" in error for error in errors))
        self.assertTrue(any("data validation ends before final KPI row 31" in error for error in errors))

    def test_workbook_validation_rejects_parent_id_missing_in_same_pnid(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            workbook_path = Path(temp_dir) / "upload.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "KPI Template"
            sheet.append(validator.EXPECTED_HEADERS)

            impact = [""] * len(validator.EXPECTED_HEADERS)
            impact[0] = 1
            impact[3] = "Officer"
            impact[7] = "IMPACT"
            impact[10] = "Impact KPI"
            impact[22] = "888"
            sheet.append(impact)

            output = [""] * len(validator.EXPECTED_HEADERS)
            output[0] = 3
            output[3] = "Officer"
            output[7] = "OUTPUT"
            output[8] = 2
            output[9] = "Impact KPI"
            output[10] = "Output KPI"
            output[22] = "888"
            sheet.append(output)
            workbook.save(workbook_path)

            _, errors, _ = validator.validate_workbook(
                workbook_path,
                fixed_pmids=set(),
                master_ids=set(),
                nomenclature_ids={"888"},
                master_types_by_id={},
                position_types_by_pnid={"888": {"6"}},
                trusted_pmid_ids=set(),
                trusted_pnid_ids=set(),
            )

        self.assertTrue(
            any("Parent KPI ID=2 tidak ditemukan pada identity PNID=888" in error for error in errors)
        )

    def test_workbook_validation_rejects_duplicate_idkpi_across_different_pnids(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            workbook_path = Path(temp_dir) / "upload.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "KPI Template"
            sheet.append(validator.EXPECTED_HEADERS)

            for pnid, title in (("888", "Impact A"), ("999", "Impact B")):
                row = [""] * len(validator.EXPECTED_HEADERS)
                row[0] = 1
                row[3] = "Officer"
                row[7] = "IMPACT"
                row[10] = title
                row[22] = pnid
                sheet.append(row)
            workbook.save(workbook_path)

            _, errors, _ = validator.validate_workbook(
                workbook_path,
                fixed_pmids=set(),
                master_ids=set(),
                nomenclature_ids={"888", "999"},
                master_types_by_id={},
                position_types_by_pnid={"888": {"6"}, "999": {"6"}},
                trusted_pmid_ids=set(),
                trusted_pnid_ids=set(),
            )

        self.assertTrue(any("duplicate IDKPI=1 dalam satu formulir" in error for error in errors))

    def _write_upload_workbook(self, path: Path, pnid: str, title: str = "KPI Title") -> None:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "KPI Template"
        sheet.append(validator.EXPECTED_HEADERS)
        row = [""] * len(validator.EXPECTED_HEADERS)
        row[3] = "Officer"
        row[10] = title
        row[22] = pnid
        sheet.append(row)
        sheet.conditional_formatting.add("A2:F2", FormulaRule(formula=["TRUE"]))
        validation = DataValidation(type="list", formula1='"A,B"', allow_blank=True)
        validation.add("G2:G2")
        sheet.add_data_validation(validation)
        workbook.save(path)


if __name__ == "__main__":
    unittest.main()
